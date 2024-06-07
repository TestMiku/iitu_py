from nomenclature.models import Nomenclature

excel_file = 'nomenclatures.xlsx'


def add_to_db_nomenclature(nomenclature):
    nomenclature = Nomenclature(
        key_product=nomenclature['key_product'],
        name=nomenclature['name']
    )
    nomenclature.save()


def main():
    import openpyxl
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        key_product = row[1]
        name = row[2]

        nomenclature = {
            'key_product': key_product,
            'name': name
        }

        add_to_db_nomenclature(nomenclature)
