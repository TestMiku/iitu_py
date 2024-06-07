import openpyxl
import openpyxl.worksheet.worksheet as openpyxl_worksheet_worksheet
from django.db.models import QuerySet

from finance_module import models
from finance_module.services.unpaid_invoices_service import set_worksheet_columns_width


def set_worksheet_columns(worksheet: openpyxl_worksheet_worksheet.Worksheet) -> None:
    worksheet.append(
        [
            "Подр",
            "У кого занимают",
            "ПМ (ФИО)",
            "Руководитель",
            "Сумма займа",
            "С расчетного счета",
            "Номер р/с",
            "Подр2",
            "Кто занимает",
            "ПМ (ФИО)",
            "Руководитель2",
            "Сумма долга",
            "Дата",
            "Примечание",
            "Письмо ",
            "Дата возврата",
            "Заметки ",
            "Заметки ",
            "Для Изъятия ",
            "Погашение ",
        ]
    )


def change_format_of_cells_debt(worksheet: openpyxl_worksheet_worksheet.Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [13, 16]:
                cell.number_format = "DD.MM.YYYY"
            if cell.column in [5, 12]:
                cell.number_format = "#,##0.00"


def get_debts_for_unloading() -> QuerySet:
    return models.Debt.objects.filter(datetime__gt="2023-11-22")


def set_data_to_worksheet(worksheet: openpyxl_worksheet_worksheet.Worksheet, debts: QuerySet) -> None:
    for debt in debts:
        at = debt.datetime.replace(tzinfo=None) if debt.datetime else ""
        from_whom = debt.from_whom_as_project_region
        to_whom = debt.to_whom_as_project_region
        if not to_whom or not from_whom:
            continue 
        worksheet.append(
            [
                from_whom.subdivision.name if from_whom.subdivision else "",  # TODO: check this field -> "Подр"  # debt.from_whom_as_project_region.subdivision
                from_whom.name,  # TODO: check this field -> "У кого занимают"
                from_whom.name,  # TODO: check this field -> "ПМ (ФИО)"
                f"{from_whom.director.last_name} {from_whom.director.first_name}" if from_whom.director else "",  # TODO: check this field -> "Руководитель"
                debt.sum,  # TODO: check this field -> "Сумма займа"
                "",  # TODO: check this field -> "С расчетного счета"
                "",  # TODO: check this field -> "Номер р/с"
                from_whom.subdivision.name if from_whom.subdivision else "",  # TODO: check this field -> "Подр2"  # debt.to_whom_as_project_region.subdivision if debt.to_whom_as_project_region else "-",
                to_whom.name,  # TODO: check this field -> "Кто занимает"
                to_whom.name,  # TODO: check this field -> "ПМ (ФИО)"
                f"{to_whom.director.last_name} {to_whom.director.first_name}" if to_whom.director else "",  # TODO: check this field -> "Руководитель2"
                debt.sum,  # TODO: check this field -> "Сумма долга"
                at,  # TODO: check this field -> "Дата"
                debt.note,  # TODO: check this field -> "Примечание"
                "",  # TODO: check this field -> "Письмо "
                "",  # TODO: check this field -> "Дата возврата"
                "",  # TODO: check this field -> "Заметки "
                "",  # TODO: check this field -> "Заметки "
                "",  # TODO: check this field -> "Для Изъятия "
                ""  # TODO: check this field -> "Погашение ",
            ]
        )

    stylizing_worksheet(worksheet)
    change_format_of_cells_debt(worksheet)


def stylizing_name_column_worksheet(worksheet: openpyxl_worksheet_worksheet.Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")

            if cell.value in ['У кого занимают', 'С расчетного счета', 'Номер р/с', 'Кто занимает',
                              'Сумма долга', 'Дата', 'Примечание', 'Заметки', 'Для Изъятия', 'Погашение']:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet.row_dimensions[1].height = 30


def stylizing_worksheet(worksheet: openpyxl_worksheet_worksheet.Worksheet) -> None:
    stylizing_name_column_worksheet(worksheet)

    for row_idx, row_data in enumerate(worksheet.iter_rows(min_row=2)):
        for cell in row_data:
            if cell.column in [2, 6, 7, 9, 12, 13, 14, 19, 20] and cell.value:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))

    set_worksheet_columns_width(worksheet, [13])
