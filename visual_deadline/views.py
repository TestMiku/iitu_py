import re
from datetime import datetime, timedelta
from django import template
from django.http import HttpResponse
# from datetime import datetime
# from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from reporter.models import Report
from .models import ExcelData
from django.db import connection
# from django.core import serializers
# from django.http import HttpRequest
from rest_framework.decorators import api_view
from .serializers import ExcelDataSerializer, AllExcelDataSerializer, NamesExcelDataSerializer, ProjectExcelDataSerializer, ProjectManagerExcelDataSerializer
from rest_framework.request import Request
from rest_framework.response import Response
from django.shortcuts import get_object_or_404



def calculate_deadline(end_date: str, start_date: datetime):
    
    deadline = None
    if end_date:
        num = "".join(char for char in end_date if char.isdigit())
        if "або" in end_date:
            current_date = datetime.now()
            deadline = 0
            start_date_for_calc = datetime.combine(start_date, datetime.min.time())
            days_count = 0
            all_days = 0

            while days_count != int(num):
                start_date_for_calc += timedelta(days=1)
                all_days += 1

                # Проверяем, является ли текущий день рабочим
                if start_date_for_calc.weekday() != 5 and start_date_for_calc.weekday() != 6:
                    days_count += 1

            if current_date < start_date_for_calc:
                while current_date.date() != start_date_for_calc.date():
                    current_date += timedelta(days=1)
                    if current_date.weekday() != 0 and current_date.weekday() != 6:
                        deadline += 1
            else:
                while current_date.date() != start_date_for_calc.date():
                    current_date -= timedelta(days=1)
                    deadline += 1 if current_date.weekday() != 0 and current_date.weekday() != 6 else 0
                deadline *= -1
            deadline = str(deadline) + ' р'
        else:
            end_date_deadline = datetime.combine(start_date, datetime.min.time()) + timedelta(days=int(num))
            deadline = str((end_date_deadline - datetime.now()).days)
    return deadline

def upload_file(request):
    
    if request.method == 'POST':
        uploaded_file = request.FILES['file_upload']
        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        ExcelData.objects.all().delete()
        with connection.cursor() as cursor:
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'visual_deadline_exceldata'")
        header_row = next(worksheet.iter_rows(min_row=2, max_row=2))
        headers = {}
        for cell in header_row:
            headers[cell.value] = cell.column
        for row in worksheet.iter_rows(min_row=3):
            get_cell_value = lambda column_name: worksheet.cell(row=row[0].row, column=headers[column_name]).value or None
            
            name = get_cell_value('Покупатель')
            document_number = get_cell_value('Номер заказа')
            start_date = get_cell_value('Дата подписания заказа заказчиком')
            end_date = str(get_cell_value('Срок выполнения работ, дней'))
            provider = get_cell_value('Поставщик')
            max_deadline = get_cell_value('Дата окончания договора') 
            project_manager = get_cell_value('Менеджер проекта')
            project_group = get_cell_value('Группа проектов')
            no_invoice_1C = get_cell_value('№ счет фактуры 1С') or 'None'
            invoice_date = get_cell_value('Дата выставления счёта') 
            order_date = get_cell_value('Дата внесения заказа в систему')
            project = get_cell_value('Проект')             
            deadline = calculate_deadline(end_date, start_date)
            
            responsible_sale = get_cell_value('Ответственный за продажу')
            contract_number = get_cell_value('№ Контракта (Номер ДС)')
            date_document_signed = get_cell_value('Дата подписания контракта')
            order_sum = get_cell_value('Сумма по заказу')
            account_amount = get_cell_value('Сумма по счёту')
            customer_debt = get_cell_value('Долг заказчика по счету')
            
            
            ExcelData.objects.create(  
                name = name,
                document_number = document_number,
                start_date = start_date,
                end_date = end_date,
                max_deadline =  max_deadline if max_deadline is not None else None,
                provider = provider,
                project_manager = project_manager,
                project_group = project_group,
                no_invoice_1C = no_invoice_1C,
                invoice_date = invoice_date,
                order_date = order_date,
                project = project,
                deadline = deadline,
                responsible_sale = responsible_sale,
                contract_number = contract_number,
                date_document_signed = date_document_signed,
                order_sum = order_sum,
                account_amount = account_amount,
                customer_debt = customer_debt,
                )
        Report.objects.create(responsible=request.user, process="deadline - База обновлена", text="deadline.reload_data")
        return redirect("deadline")
    Report.objects.create(responsible=request.user, process="deadline - Открыт главная страница", text="deadline.open")
    return redirect("deadline")

# АрсеорМиттал Темиртау,ИП "Рога",ТОО "Здорова",Сегодня,Прошло, БЦК,ТОО "Привет",ТОО "Тест"
@api_view(['GET'])
def search(request: Request) -> Response:
    query = request.GET.get('query')
    customer = request.GET.get("customer")
    filters = {}
    if query:
        filters["document_number__iregex"] = query
    if customer:
        filters["name"] = customer
   
    data = ExcelData.objects.filter(**filters)
    return Response(ExcelDataSerializer(data, many=True).data)

@api_view(['GET'])
def get_current_data(request: Request) -> Response:
    customer = request.GET.get("customer")
    filters = {}
    if customer:
        customer = customer
        filters["name__in"] = customer.split(",")
        # print(filters)
    
    data = ExcelData.objects.filter(**filters)
    return Response(AllExcelDataSerializer(data, many=True).data)

@api_view(['GET'])
def all_data(request: Request) -> Response:
    id  = request.GET.get("id")
    filters = {}
    serializer = None
    if id:
        filters["id"] = id
        excel_data = ExcelData.objects.filter(**filters)
        serializer = AllExcelDataSerializer(excel_data, many=True)
        return Response(serializer.data)
    else:
        return Response({"error": "Please provide 'id' parameter in the URL."}, status=400)

@api_view(['GET'])
def filtered_data_by_invoice(request: Request) -> Response:
    invoice = request.GET.get('invoice').replace('БН', r'б.*\s*н').replace('onlynum', r'^\d+$') #no_invoice_1C
    filters = {}
    filters["no_invoice_1C__regex"] = invoice
    if not invoice:
        filters["no_invoice_1C__isnull"] = True
    data = ExcelData.objects.filter(**filters).values('name').distinct()
    return Response(NamesExcelDataSerializer(data, many=True).data)


@api_view(['GET'])
def get_all_projects_from_data(request: Request) -> Response:
    invoice = request.GET.get('invoice')
    if invoice:
        invoice = invoice.replace('БН', r'б.*\s*н').replace('onlynum', r'^\d+$')
        filters = {}
        filters["no_invoice_1C__iregex"] = invoice
        data = ExcelData.objects.filter(**filters).values('project').distinct()
    else:
        data = ExcelData.objects.values('project').distinct()
    return Response(ProjectExcelDataSerializer(data, many=True).data)

@api_view(['GET'])
def get_all_project_managers_from_data(request: Request) -> Response:
    invoice = request.GET.get('invoice')
    if invoice:
        invoice = invoice.replace('БН', r'б.*\s*н').replace('onlynum', r'^\d+$')
        filters = {}
        filters["no_invoice_1C__iregex"] = invoice
        data = ExcelData.objects.filter(**filters).values('project_manager').distinct()
    else:
        data = ExcelData.objects.values('project_manager').distinct()
    return Response(ProjectManagerExcelDataSerializer(data, many=True).data)

def main(request):
    Report.objects.create(responsible=request.user, process="deadline - Открыт главная страница", text="deadline.open")
    data_for_calculate = ExcelData.objects.values('start_date', 'end_date')
    for values in data_for_calculate:
        start_date = values.get('start_date')
        end_date = values.get('end_date')
        calculate_deadline(end_date, start_date)
    return render(request, 'visual_deadline/main.html',{'unique_names': ExcelData.objects.values('name').distinct(), 'name_with_deadline': ExcelData.objects.values('name', 'deadline'), 'all_data': ExcelData.objects.all()[:30]})