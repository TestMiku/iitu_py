import dataclasses
import datetime
import decimal
import io
import itertools
import pathlib
import tempfile
import typing
from copy import copy

import openpyxl.styles
import openpyxl.styles.cell_style
import openpyxl.styles.proxy
import openpyxl.utils.cell
import openpyxl.worksheet.worksheet
from django.db.models import Q, QuerySet
from django.db.models.functions import TruncDate
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .. import models
from . import common_service, smart_table_service


def _copy_cell_styles(from_cell: Cell, to_cell: Cell) -> None:
    if from_cell.has_style:
        to_cell.font = copy(from_cell.font)
        to_cell.border = copy(from_cell.border)
        to_cell.fill = copy(from_cell.fill)
        to_cell.number_format = copy(from_cell.number_format)
        to_cell.protection = copy(from_cell.protection)
        to_cell.alignment = copy(from_cell.alignment)


def _cell(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    copy_row: int,
    column: int,
    value: typing.Any,
) -> None:
    _copy_cell_styles(
        worksheet.cell(copy_row, column), worksheet.cell(row, column, value)
    )


paid_invoices_xlsx_1_template = (
    pathlib.Path(__file__).parent / "unload_templates/ДДС Реестр Шаблон.xlsx"
)


def unload_paid_invoices_xlsx_1(
    user, field_values_list, field_type, order
) -> io.BytesIO:
    """Выгрузка ДДС"""
    queryset = models.PaidInvoice.objects.exclude(
        status=models.rejected_paid_invoice_status()
    ).filter(
        project_region__in=common_service.get_project_regions(
            user=user if user.is_authenticated else None
        )
    )
    queryset, field_new_field = smart_table_service.convert_queryset(
        queryset, field_type=field_type
    )
    queryset = smart_table_service.filter_queryset(
        queryset, field_values_list=field_values_list, field_new_field=field_new_field
    )
    queryset = smart_table_service.order_queryset(
        queryset, order=order, field_new_field=field_new_field
    )
    xlsx_1 = tempfile.NamedTemporaryFile("wb+", prefix="ДДС ", suffix=".xlsx")
    workbook = load_workbook(paid_invoices_xlsx_1_template)
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

    for index, paid_invoice in enumerate(queryset, 4):
        worksheet.row_dimensions[index].height = None
        _cell(worksheet, index, 3, 1, paid_invoice.at.date())
        _cell(worksheet, index, 3, 2, paid_invoice.project_region.name)
        _cell(worksheet, index, 3, 3, paid_invoice.number)
        _cell(worksheet, index, 3, 4, paid_invoice.date)
        _cell(worksheet, index, 3, 5, paid_invoice.invoice_number)
        _cell(worksheet, index, 3, 6, paid_invoice.invoice_date)
        _cell(worksheet, index, 3, 7, paid_invoice.project)
        _cell(worksheet, index, 3, 8, paid_invoice.responsible_user_id)
        _cell(worksheet, index, 3, 9, paid_invoice.approver)
        _cell(worksheet, index, 3, 10, paid_invoice.llc)
        _cell(worksheet, index, 3, 11, paid_invoice.contractor)
        _cell(worksheet, index, 3, 12, paid_invoice.comment)
        _cell(worksheet, index, 3, 13, paid_invoice.currency)
        _cell(worksheet, index, 3, 14, paid_invoice.invoice_amount)
        _cell(worksheet, index, 3, 15, paid_invoice.sum)
        _cell(worksheet, index, 3, 16, paid_invoice.invoice_category)
        _cell(worksheet, index, 3, 17, paid_invoice.revenue_expense_articles)
        _cell(worksheet, index, 3, 18, paid_invoice.sales_order)
        _cell(worksheet, index, 3, 19, paid_invoice.bin_or_iin)
        _cell(worksheet, index, 3, 20, paid_invoice.document_amount)
        _cell(worksheet, index, 3, 21, paid_invoice.account.name)
        _cell(worksheet, index, 3, 22, paid_invoice.iic)
        _cell(worksheet, index, 3, 23, paid_invoice.payment_destination_code)
        _cell(worksheet, index, 3, 24, paid_invoice.contract_number)
        _cell(worksheet, index, 3, 25, paid_invoice.invoice_amount)
        _cell(worksheet, index, 3, 26, paid_invoice.paid)
        _cell(worksheet, index, 3, 27, paid_invoice.sum)
    worksheet.move_range(f"A4:AA{queryset.count() + 3}", rows=-1)
    workbook.save(xlsx_1)
    workbook.close()
    xlsx_1.seek(0)
    return xlsx_1


paid_invoices_xlsx_2_template = (
    pathlib.Path(__file__).parent / "unload_templates/1C Шаблон.xlsx"
)


@dataclasses.dataclass(kw_only=True, slots=True)
class _Format:
    font: openpyxl.styles.proxy.StyleProxy
    border: openpyxl.styles.proxy.StyleProxy
    fill: openpyxl.styles.proxy.StyleProxy
    number_format: str
    protection: openpyxl.styles.proxy.StyleProxy
    alignment: openpyxl.styles.proxy.StyleProxy
    quotePrefix: bool = False


def _get_formats(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, row: int
) -> list[_Format]:
    formats = []
    for row_ in worksheet.iter_rows(row, row):
        for cell in row_:
            formats.append(
                _Format(
                    font=copy(cell.font),
                    border=copy(cell.border),
                    fill=copy(cell.fill),
                    number_format=cell.number_format,
                    protection=copy(cell.protection),
                    alignment=copy(cell.alignment),
                    quotePrefix=cell.quotePrefix,
                )
            )
            cell.value = None

    return formats


def _append_values_with_formats(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    formats: list[_Format],
    values: list[typing.Any],
) -> None:
    for column, (format_, value) in enumerate(zip(formats, values), 1):
        cell = worksheet.cell(row, column, value)

        cell.font = format_.font
        cell.alignment = format_.alignment
        cell.number_format = format_.number_format
        cell.fill = format_.fill
        cell.border = format_.border
        cell.protection = format_.protection
        cell.quotePrefix = 0


def unload_paid_invoices_xlsx_2(
    user, field_values_list, field_type, order
) -> io.BytesIO:
    """Выгрузка 1C"""
    queryset = models.PaidInvoice.objects.filter(
        status=models.completed_paid_invoice_status()
    )
    queryset, field_new_field = smart_table_service.convert_queryset(
        queryset, field_type=field_type
    )
    queryset = smart_table_service.filter_queryset(
        queryset, field_values_list=field_values_list, field_new_field=field_new_field
    )
    queryset = smart_table_service.order_queryset(
        queryset, order=order, field_new_field=field_new_field
    )
    return unload_1c(queryset)


def unload_1c(queryset: QuerySet) -> io.BytesIO:
    xlsx_1 = tempfile.NamedTemporaryFile("wb+", prefix="1С ", suffix=".xlsx")
    workbook = load_workbook(paid_invoices_xlsx_2_template)
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
    formats = _get_formats(worksheet, 2)
    formats[19].quotePrefix = False
    for index, paid_invoice in enumerate(queryset, 2):
        worksheet.row_dimensions[index].height = None
        _append_values_with_formats(
            worksheet,
            index,
            formats,
            [
                paid_invoice.number,
                paid_invoice.at.date().strftime("%d.%m.%Y"),
                paid_invoice.invoice_number,
                paid_invoice.invoice_date.strftime("%d.%m.%Y")
                if paid_invoice.invoice_date
                else None,
                paid_invoice.project,
                paid_invoice.responsible_user_id,
                paid_invoice.approver,
                paid_invoice.llc,
                paid_invoice.contractor,
                paid_invoice.comment,
                paid_invoice.currency,
                paid_invoice.sum,
                paid_invoice.invoice_category,
                paid_invoice.revenue_expense_articles,
                paid_invoice.sales_order,
                paid_invoice.bin_or_iin,
                paid_invoice.document_amount,
                paid_invoice.account.number,
                paid_invoice.iic,
                paid_invoice.payment_destination_code,
                paid_invoice.contract_number,
                paid_invoice.invoice_amount,
                paid_invoice.paid_amount_1c,
                paid_invoice.sum,
            ],
        )
    workbook.save(xlsx_1)
    workbook.close()
    xlsx_1.seek(0)
    return xlsx_1


mandatory_payment_seizures_xlsx_1_template = (
    pathlib.Path(__file__).parent / "unload_templates/Изьятие Шаблон.xlsx"
)


def unload_mandatory_payment_seizures_xlsx_1(
    user, field_values_list, field_type, order
) -> io.BytesIO:
    """Выгрузка Изьятий"""
    queryset = models.MandatoryPaymentSeizure.objects.exclude(
        Q(status=models.rejected_mandatory_payment_seizure_status())
        | Q(imported_from_file=True)
    ).filter(
        project_region__in=common_service.get_project_regions(
            user=user if user.is_authenticated else None
        )
    )
    queryset, field_new_field = smart_table_service.convert_queryset(
        queryset, field_type=field_type
    )
    queryset = smart_table_service.filter_queryset(
        queryset, field_values_list=field_values_list, field_new_field=field_new_field
    )
    queryset = smart_table_service.order_queryset(
        queryset, order=order, field_new_field=field_new_field
    )
    xlsx_1 = tempfile.NamedTemporaryFile("wb+", prefix="Изьятие ", suffix=".xlsx")
    workbook = load_workbook(mandatory_payment_seizures_xlsx_1_template)
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

    for index, mandatory_payment_seizure in enumerate(queryset, 4):
        worksheet.row_dimensions[index].height = None
        _cell(worksheet, index, 3, 1, mandatory_payment_seizure.datetime.date())
        _cell(
            worksheet,
            index,
            3,
            2,
            mandatory_payment_seizure.project_region.subdivision
            and mandatory_payment_seizure.project_region.subdivision.name,
        )
        _cell(worksheet, index, 3, 3, mandatory_payment_seizure.project_region.name)
        _cell(worksheet, index, 3, 4, mandatory_payment_seizure.project_region.name)
        _cell(
            worksheet,
            index,
            3,
            5,
            mandatory_payment_seizure.project_region.director
            and mandatory_payment_seizure.project_region.director.get_full_name(),
        )
        _cell(worksheet, index, 3, 6, mandatory_payment_seizure.sum)
        _cell(worksheet, index, 3, 7, mandatory_payment_seizure.mandatory_payment.name)
        _cell(
            worksheet,
            index,
            3,
            8,
            mandatory_payment_seizure.account
            and mandatory_payment_seizure.account.name,
        )
        _cell(
            worksheet,
            index,
            3,
            9,
            mandatory_payment_seizure.account
            and mandatory_payment_seizure.account.number,
        )
        _cell(worksheet, index, 3, 10, None)
        _cell(worksheet, index, 3, 11, None)
        _cell(worksheet, index, 3, 12, None)
    worksheet.move_range(f"A4:L{queryset.count() + 3}", rows=-1)
    workbook.save(xlsx_1)
    workbook.close()
    xlsx_1.seek(0)
    return xlsx_1


interdivisional_debts_xlsx_1_template = (
    pathlib.Path(__file__).parent / "unload_templates/ДМП Шаблон.xlsx"
)


def unload_interdivisional_debts_xlsx_1(
    user, field_values_list, field_type, order
) -> io.BytesIO | None:
    """Выгрузка ДМП"""
    queryset = models.Debt.objects.exclude(
        Q(group__status=models.rejected_debt_translate_group_status())
        | Q(imported_from_file=True)
    ).filter(
        to_whom__in=common_service.get_project_regions(
            user=user if user.is_authenticated else None
        )
        .values_list("name", flat=True)
        .distinct()
    )
    queryset, field_new_field = smart_table_service.convert_queryset(
        queryset, field_type=field_type
    )
    queryset = smart_table_service.filter_queryset(
        queryset, field_values_list=field_values_list, field_new_field=field_new_field
    )
    queryset = smart_table_service.order_queryset(
        queryset, order=order, field_new_field=field_new_field
    )
    xlsx_1 = tempfile.NamedTemporaryFile("wb+", prefix="ДМП ", suffix=".xlsx")
    workbook = load_workbook(interdivisional_debts_xlsx_1_template)
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
    if not queryset:
        return None
    index = 3
    for debt in queryset:
        from_whom = debt.from_whom_as_project_region
        to_whom = debt.to_whom_as_project_region
        if not from_whom or not to_whom:
            continue
        from_whom_subdivision = (
            from_whom.subdivision.name if from_whom.subdivision else ""
        )
        to_whom_subdivision = to_whom.subdivision.name if to_whom.subdivision else ""
        from_whom_director = (
            f"{from_whom.director.last_name} {from_whom.director.first_name}"
            if from_whom.director
            else ""
        )
        to_whom_director = (
            f"{to_whom.director.last_name} {to_whom.director.first_name}"
            if to_whom.director
            else ""
        )
        _cell(worksheet, index, 2, 1, from_whom_subdivision)
        _cell(worksheet, index, 2, 2, debt.from_whom)
        _cell(worksheet, index, 2, 3, debt.from_whom)
        _cell(worksheet, index, 2, 4, from_whom_director)
        _cell(worksheet, index, 2, 5, -debt.sum)
        _cell(worksheet, index, 2, 6, "")
        _cell(worksheet, index, 2, 7, "")
        _cell(worksheet, index, 2, 8, to_whom_subdivision)
        _cell(worksheet, index, 2, 9, debt.to_whom)
        _cell(worksheet, index, 2, 10, debt.to_whom)
        _cell(worksheet, index, 2, 11, to_whom_director)
        _cell(worksheet, index, 2, 12, debt.sum)
        _cell(worksheet, index, 2, 13, "")
        _cell(worksheet, index, 2, 14, "")
        _cell(worksheet, index, 2, 15, debt.datetime.date())
        _cell(worksheet, index, 2, 16, debt.note)
        index += 1
    worksheet.move_range(f"A3:AC{queryset.count() + 2}", rows=-1)
    workbook.save(xlsx_1)
    workbook.close()
    xlsx_1.seek(0)
    return xlsx_1


unload_1_template = pathlib.Path(__file__).parent / "unload_templates/ДДС Шаблон.xlsx"


def unload_1() -> io.BytesIO:
    @dataclasses.dataclass(slots=True, frozen=True, kw_only=True)
    class Row:
        date: datetime.datetime | datetime.date
        operation: typing.Literal["Списание", "Поступление"]
        category: typing.Literal[
            "Расход", "Расход_АДМ", "ДМП", "Переводы между р/с", "Приход"
        ]
        article: str
        project_region: models.ProjectRegion

        @property
        def project_manager(self) -> str:
            return (
                self.project_region.project_manager.get_full_name()
                if self.project_region.project_manager
                else "-"
            )

        @property
        def manager(self) -> str:
            return (
                self.project_region.manager.get_full_name()
                if self.project_region.manager
                else "-"
            )

        @property
        def director(self) -> str:
            return (
                self.project_region.director.get_full_name()
                if self.project_region.director
                else "-"
            )

        @property
        def subdivision(self) -> str:
            return (
                self.project_region.subdivision.name
                if self.project_region.subdivision
                else "-"
            )

        from_account: models.Account | typing.Literal["р/с Отправителя"]
        to_account: models.Account | typing.Literal["р/с Получателя", "р/с Банка"]
        sum: decimal.Decimal

        @property
        def account(self) -> models.Account:
            return (
                self.from_account if self.operation == "Списание" else self.to_account
            )

    rows: list[Row] = []

    for mandatory_payment_seizure in models.MandatoryPaymentSeizure.objects.exclude(
        Q(status__in=[models.rejected_mandatory_payment_seizure_status()])
        | Q(imported_from_file=True)
    ):
        rows.append(
            Row(
                date=mandatory_payment_seizure.datetime.date(),
                operation="Списание",
                category="Расход_АДМ",
                article=mandatory_payment_seizure.mandatory_payment.name,
                from_account=mandatory_payment_seizure.account,
                to_account=common_service.get_administrative_account(),
                sum=mandatory_payment_seizure.sum,
                project_region=mandatory_payment_seizure.project_region,
            )
        )
    for debt_translate_group in models.DebtTranslateGroup.objects.exclude(
        status__in=[models.rejected_debt_translate_group_status()]
    ):
        rows.append(
            Row(
                date=debt_translate_group.datetime.date(),
                operation="Списание",
                category="ДМП",
                article=debt_translate_group.to_whom,
                from_account=debt_translate_group.from_account,
                to_account=debt_translate_group.to_account,
                sum=debt_translate_group.sum,
                project_region=models.ProjectRegion.objects.get(
                    name=debt_translate_group.from_whom
                ),
            )
        )
        rows.append(
            Row(
                date=debt_translate_group.datetime.date(),
                operation="Поступление",
                category="ДМП",
                article=debt_translate_group.from_whom,
                from_account=debt_translate_group.from_account,
                to_account=debt_translate_group.to_account,
                sum=debt_translate_group.sum,
                project_region=models.ProjectRegion.objects.get(
                    name=debt_translate_group.to_whom
                ),
            )
        )

    for transfer in models.Transfer.objects.exclude(
        status__in=[models.rejected_transfer_status()]
    ):
        rows.append(
            Row(
                date=transfer.datetime.date(),
                operation="Списание",
                category="Расход",
                article="Перевод собственных средств",
                from_account=transfer.from_account,
                to_account=transfer.to_account,
                sum=transfer.sum,
                project_region=transfer.project_region,
            )
        )
        rows.append(
            Row(
                date=transfer.datetime.date(),
                operation="Поступление",
                category="Приход",
                article="Перевод собственных средств",
                from_account=transfer.from_account,
                to_account=transfer.to_account,
                sum=transfer.sum,
                project_region=transfer.project_region,
            )
        )
    for (date, account, project_region), paid_invoices in itertools.groupby(
        models.PaidInvoice.objects.annotate(at_date=TruncDate("at")).exclude(
            status__in=[models.rejected_paid_invoice_status()]
        ).order_by("at_date", "account", "project_region"),
        key=lambda paid_invoice: (
            paid_invoice.at_date,
            paid_invoice.account,
            paid_invoice.project_region,
        ),
    ):
        paid_invoices = list(paid_invoices)
        print(paid_invoices)
        rows.append(
            Row(
                date=date,
                operation="Списание",
                category="Расход",
                article="Оплата по Реестру",
                from_account=account,
                to_account="р/с Получателя",
                project_region=project_region,
                sum=sum(paid_invoice.sum for paid_invoice in paid_invoices),
            )
        )
        for paid_invoice in paid_invoices:
            if paid_invoice.commission and paid_invoice.commission_date:
                rows.append(
                    Row(
                        date=paid_invoice.commission_date,
                        operation="Списание",
                        category="Расход",
                        article="Комиссия",
                        from_account=paid_invoice.account,
                        to_account="р/с Банка",
                        project_region=paid_invoice.project_region,
                        sum=paid_invoice.commission,
                    )
                )
    for sutochnye in models.Sutochnye.objects.exclude(
        status__in=[models.rejected_sutochnye_status()]
    ):
        rows.append(
            Row(
                date=sutochnye.created_at.date(),
                operation="Списание",
                category="Расход",
                article="Командировочные",
                from_account=sutochnye.account,
                to_account="р/с Получателя",
                project_region=sutochnye.project_region,
                sum=sutochnye.sum,
            )
        )
    for inflow in models.Inflow.objects.exclude(reserve_percent=0):
        rows.append(
            Row(
                date=inflow.date,
                operation="Поступление",
                category="Приход",
                article="От заказчика (Картел)",
                from_account="р/с Отправителя",
                to_account=inflow.account,
                project_region=inflow.project_region,
                sum=inflow.remainder,
            )
        )
        rows.append(
            Row(
                date=inflow.date,
                operation="Списание",
                category="Расход",
                article=f"Резерв ({inflow.reserve_percent:,.2f}%)",
                from_account=inflow.account,
                to_account=common_service.get_administrative_account(),
                project_region=inflow.project_region,
                sum=inflow.reserve,
            )
        )

    workbook = load_workbook(unload_1_template)
    worksheet = workbook["Таблица"]
    for index, row in enumerate(sorted(rows, key=lambda row: row.date), 2):
        worksheet.cell(index, 1, row.date)
        worksheet.cell(index, 3, row.operation)
        worksheet.cell(index, 4, row.category)
        worksheet.cell(index, 5, row.article)
        worksheet.cell(index, 6, row.project_region.name)
        worksheet.cell(index, 7, row.project_manager)
        worksheet.cell(index, 8, row.manager)
        worksheet.cell(index, 9, row.director)
        worksheet.cell(index, 10, row.subdivision)
        worksheet.cell(index, 11, row.sum)
        worksheet.cell(
            index,
            12,
            row.from_account.name
            if isinstance(row.from_account, models.Account)
            else row.from_account,
        )
        worksheet.cell(
            index,
            13,
            row.from_account.number
            if isinstance(row.from_account, models.Account)
            else row.from_account,
        )
        worksheet.cell(
            index,
            14,
            row.to_account.name
            if isinstance(row.to_account, models.Account)
            else row.to_account,
        )
        worksheet.cell(
            index,
            15,
            row.to_account.number
            if isinstance(row.to_account, models.Account)
            else row.to_account,
        )
    bytes_io = io.BytesIO()
    workbook.save(bytes_io)
    bytes_io.seek(0)
    return bytes_io
