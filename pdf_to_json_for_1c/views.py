import io
import json
import os
import re
from datetime import datetime
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect
from django.conf import settings

import fitz
import base64
import PyPDF2
import openpyxl
import pdfplumber
import requests
# import pdf2image
import pytesseract
from PIL import Image
import xlrd


def main(request: HttpRequest) -> render:
    file_url = request.GET.get('file_url')
    image_path = ''
    html = "pdf_to_json_for_1c/index.html"
    raw_file=''
    _file_type=''
    is_image = False
    if request.method == 'POST':
        raw_file = request.FILES['file_upload']
        _file_type = raw_file.content_type
        if 'excel' in _file_type:
            image_path = 'pdf_to_json_for_1c/temp.xls'
        if 'sheet' in _file_type:
            image_path = 'pdf_to_json_for_1c/temp.xlsx'
        if 'pdf' in _file_type:
            image_path = 'pdf_to_json_for_1c/temp.pdf'
        if 'image' in _file_type:
            is_image = True
            image_path = os.path.join(settings.MEDIA_ROOT, raw_file.name)
            # with open(image_path, 'wb+') as destination:
            #     for chunk in raw_file.chunks():
            #         destination.write(chunk)
        with open(image_path, 'wb+') as temp_file:
            for chunk in raw_file.chunks():
                temp_file.write(chunk)
        print(image_path)
    if file_url:
        try:
            response = requests.get(file_url)
            if response.status_code == 200:
                if 'image' in response.headers['content-type']:
                    is_image = True
                image_path = os.path.join(settings.MEDIA_ROOT, file_url.split('/')[-1])
                with open(image_path, 'wb+') as destination:
                    destination.write(response.content)
                # Чтение данных из PDF
                # else:
                #     with open(temp_pdf, 'wb+') as temp_file:
                #         # Чтение PDF файла по частям (chunks) и запись во временный файл
                #         for chunk in response.iter_content(chunk_size=1024):
                #             temp_file.write(chunk)
                html = "pdf_to_json_for_1c/btrx.html"
        except Exception as e:   
            return HttpResponse("No PDF file URL provided:"+str(e), status=400)
    if image_path:
        # temp_pdf = '/home/portal/avh_portal/avh_portal/pdf_to_json_for_1c/temp.pdf'
        if is_image:
            res = handle_image_file(image_path)
            images = convert_file_to_images(image_path)
            os.remove(image_path)
        elif ('pdf' in image_path.split('/')[-1] or 'pdf' in _file_type) and is_scanned_pdf(image_path):
            res = handle_scan_file(image_path)
        else:
            if 'xlsx' in image_path or 'sheet' in _file_type:
                res = handle_excel_file(image_path)
            elif 'xls' in image_path or 'excel' in _file_type:
                res = handle_excel_file(image_path, True)
            else:
                res = handle_pdf_file(image_path)
        if not is_image:
            images = convert_file_to_images(image_path)
            os.remove(image_path)
            print(images)
        return render(request, html, {'result': res, 'images': images})
    return render(request, html)

def is_scanned_pdf(pdf_path: str) -> bool:
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfFileReader(file)
        for page_num in range(reader.numPages):
            page = reader.getPage(page_num)
            try:
                if not page.extractText():
                    return True
                # Если удалось извлечь текст из страницы, это скорее всего обычный PDF
                return False
            except:
                # Если не удалось извлечь текст, это скорее всего сканированный PDF
                return True


def extract_text_between_keywords(text: str, keyword1: str) -> re.Match | None:
    regex_pattern = re.search(f'{keyword1}([\s\S]*?)итого', text.lower())
    if regex_pattern:
        return regex_pattern
    else:
        return None
    
def convert_file_to_images(pdf_path: str) -> list:
    try:
        images = []
        document = fitz.open(pdf_path)
        for page in document:
            # Получаем изображение текущей страницы в формате RGB
            pixmap = page.get_pixmap(alpha=False, matrix=fitz.Matrix(2.0, 2.0))
            # Преобразуем изображение в формат PNG в виде байтового объекта
            image_bytes = pixmap.tobytes()
            # Кодируем байтовый объект в base64
            image_base64 = base64.b64encode(image_bytes).decode('utf-8')
            # Добавляем изображение в список
            images.append(image_base64)
        return images
    except Exception as e:
        print(e)
        return []

def handle_text(raw_text: str) -> dict[str, str]:
    result = {}
    sender_bin = re.findall(r'БИН.*?(\d+)', raw_text)
    sender_name = re.search(r'(ИП|ТОО|Индивидуальный предприниматель|Товарищество с ограниченной ответственностью)\s*[\"\'\`](.*?)[\"\'\`]', raw_text)
    order_num = re.search(r'на оплату(.*?)\s*от', raw_text) # №\s*(.*?)\s*(от|\d{2}\.)
    order_date = re.search(r'(от|№\d+)\s*(.*?\d[4])\s*г', raw_text)
    schet = re.search(r'(KZ\d.*?)\s+', raw_text)
    table_raw_text = extract_text_between_keywords(raw_text, 'умма') or extract_text_between_keywords(raw_text, 'цена') or extract_text_between_keywords(raw_text, 'аименование')
    
    result['raw_text'] = raw_text
    if sender_bin:
        result['sender_bin'] = sender_bin[0]
        result['byer_bin'] = sender_bin[-1]
    if sender_name:
        result['sender_name'] = sender_name.group(0)
    if order_num:
        result['order_num'] = order_num.group(1).replace('№', '').lstrip()
    if schet:
        result['schet'] = schet.group(1)
    if order_date:
        try:
            raw_date = re.sub(r'[\'\"`<«>»]', '', order_date.group(2).strip())
            if '.' not in raw_date:
                day, month, year = raw_date.split()
                # Получаем числовое значение месяца 
                months = {
                    'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
                    'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
                }
                month_number = months[month]

                # Преобразовываем строку в дату
                date = datetime(int(year), month_number, int(day))
            else:
                date = datetime.strptime(raw_date, '%d.%m.%Y')
            result['order_date'] = date.strftime("%Y-%m-%d")#%d.%m.%Y
        except:
            result['order_date'] = None
    if table_raw_text:
        table_texts = table_raw_text.group(1)
        result['table_text'] = table_texts

    return result
def handle_text_file(path_or_content: str, switch: bool) -> dict[str, str]:
    result = {}
    if switch:
        with open(path_or_content, "r", encoding="utf-8") as file: 
            raw_text = file.read()
            result = handle_text(raw_text)
    else:
        result = handle_text(path_or_content)
    return result


def handle_image_file(image_path):
    with open(f"pdf_to_json_for_1c/templates/scanned_file.txt", "w", encoding="utf-8") as file:
    # Открываем изображение
        with Image.open(image_path) as img:
            # Преобразуем изображение в текст с помощью pytesseract
            result = pytesseract.image_to_string(img, lang='rus')

            # Далее можно выполнить обработку текста, сохранить его или использовать по своему усмотрению
            file.write(result)
    codes = handle_text_file("pdf_to_json_for_1c/templates/scanned_file.txt", True)
    os.remove("pdf_to_json_for_1c/templates/scanned_file.txt")
    return codes
def handle_scan_file(pdf_path: str) -> dict[str, str]:
    # processed_files: dict[str, list[list[str]]] = {}
    if pdf_path:
        # pdf_file_name = str(pdf_path).split(".pdf")[0]
        document = fitz.open(pdf_path)
        with open(f"pdf_to_json_for_1c/templates/scanned_file.txt", "w", encoding="utf-8") as file:
            result = ""
            for index, page in enumerate(document):
                pixmap: fitz.Pixmap = page.get_pixmap(matrix=fitz.Matrix(3.0, 3.0))
                image_path = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
                result += pytesseract.image_to_string(image_path, lang="rus")
                
            file.write(result)
        codes = handle_text_file("pdf_to_json_for_1c/templates/scanned_file.txt", True)
        
        
        
        os.remove("pdf_to_json_for_1c/templates/scanned_file.txt")
    return codes

def get_tables(pdf_path: str) -> list[dict[str, str]]:
    with pdfplumber.open(pdf_path) as pdf:
        tables = []
        headers = ['no', 'code', 'name', 'quantity', 'unit', 'price', 'sum']
        temp_dict = {}
        for page in pdf.pages:
            if any(word in page.extract_table()[0] for word in ['Номенклатура', 'Товар', 'Наименование']):
                if len(page.extract_table()[0]) == 6:
                    headers.pop(1)
                for values in page.extract_table()[1:]:
                    for i, value in enumerate(values):
                        if value and i > len(headers) - 5 and headers[i] != 'unit':
                            value = value.replace(' ', '').replace(',', '.')
                        temp_dict[headers[i]] = value
                    tables.append(temp_dict.copy())
        #Товар
    return tables



def handle_excel_file(excel_path: str, xls=False) -> dict[str, str]:
    if xls:
        workbook = xlrd.open_workbook(excel_path)
        sheet = workbook.sheet_by_index(0)  # Предполагаем, что данные находятся в первом листе
        
        key_positions = {
            'kzt': [],
            'sender': [],
            'byer': [],
            'no': [],
            'code': [],
            'name': [],
            'quantity': [],
            'unit': [],
            'price': [],
            'sum': [],
        }
        
        # Поиск позиций ключевых ячеек
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value == 'ИИК':
                    key_positions['kzt'] = [col_idx, row_idx + 1]
                elif cell_value == 'Поставщик:':
                    key_positions['sender'] = [col_idx + 4, row_idx]
                elif cell_value == 'Покупатель:':
                    key_positions['byer'] = [col_idx + 4, row_idx]
                elif cell_value == '№':
                    key_positions['no'] = [col_idx, row_idx]
                elif cell_value == 'Код':
                    key_positions['code'] = [col_idx, row_idx]
                elif 'Наименование' in str(cell_value):
                    key_positions['name'] = [col_idx, row_idx]
                elif cell_value == 'Кол-во':
                    key_positions['quantity'] = [col_idx, row_idx]
                elif cell_value == 'Ед.':
                    key_positions['unit'] = [col_idx, row_idx]
                elif cell_value == 'Цена':
                    key_positions['price'] = [col_idx, row_idx]
                elif cell_value == 'Сумма':
                    key_positions['sum'] = [col_idx, row_idx]

        print(key_positions)
        # Извлечение данных из ключевых ячеек
        kzt = sheet.cell_value(key_positions['kzt'][1], key_positions['kzt'][0])         
        sender_raw_text = sheet.cell_value(key_positions['sender'][1], key_positions['sender'][0])
        byer_raw_text = sheet.cell_value(key_positions['byer'][1], key_positions['byer'][0])
        
        # Обработка текстовых данных, если необходимо
        sender_data = handle_text_file(sender_raw_text, False)
        byer_data = handle_text_file(byer_raw_text, False)
        sender_data['byer_bin'] = byer_data['byer_bin']
        sender_data['raw_text'] += byer_data['raw_text']
        
        res = sender_data
        res['schet'] = kzt
        table = []
        index = 1
        # Извлечение данных из таблицы
        while True:
            if sheet.cell_value(key_positions['no'][1]+index, key_positions['no'][0]) == '':
                break
            table.append({
                'no': sheet.cell_value(key_positions['no'][1] + index, key_positions['no'][0]),
                'code': sheet.cell_value(key_positions['code'][1] + index, key_positions['code'][0]) if key_positions['code'] else '',
                'name': sheet.cell_value(key_positions['name'][1] + index, key_positions['name'][0]),
                'quantity': sheet.cell_value(key_positions['quantity'][1] + index, key_positions['quantity'][0]),
                'unit': sheet.cell_value(key_positions['unit'][1] + index, key_positions['unit'][0]),
                'price': sheet.cell_value(key_positions['price'][1] + index, key_positions['price'][0]),
                'sum': sheet.cell_value(key_positions['sum'][1] + index, key_positions['sum'][0]),
            })
            index += 1
        
        res['tables'] = table
        return res
    else:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        key_positions = {
            'kzt':[],
            'sender':[],
            'byer':[],
            'no':[],
            'code':[],
            'name':[],
            'quantity':[],
            'unit':[],
            'price':[],
            'sum':[],
        }
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'ИИК':
                    key_positions['kzt'] = [cell.column, cell.row] 
                    
                if cell.value == 'Поставщик:':
                    key_positions['sender'] = [cell.column, cell.row] 
                    
                if cell.value == 'Покупатель:':
                    key_positions['byer'] = [cell.column, cell.row] 
                    
                if cell.value == '№':
                    key_positions['no'] = [cell.column, cell.row] 
                    
                if cell.value == 'Код':
                    key_positions['code'] = [cell.column, cell.row] 
                    
                if 'Наименование' in str(cell.value):
                    key_positions['name'] = [cell.column, cell.row] 
                    
                if cell.value == 'Кол-во':
                    key_positions['quantity'] = [cell.column, cell.row] 
                    
                if cell.value == 'Ед.':
                    key_positions['unit'] = [cell.column, cell.row] 
                    
                if cell.value == 'Цена':
                    key_positions['price'] = [cell.column, cell.row] 
                    
                if cell.value == 'Сумма':
                    key_positions['sum'] = [cell.column, cell.row] 
                    
        kzt = ws.cell(key_positions['kzt'][1]+1, key_positions['kzt'][0]).value         
        sender_raw_text = ws.cell(key_positions['sender'][1], key_positions['sender'][0]+4).value
        byer_raw_text = ws.cell(key_positions['byer'][1], key_positions['byer'][0]+4).value
        sender_data = handle_text_file(sender_raw_text, False)
        byer_data = handle_text_file(byer_raw_text, False)
        sender_data['byer_bin'] = byer_data['byer_bin']
        sender_data['raw_text'] += byer_data['raw_text']
        res = sender_data
        res['schet'] = kzt
        table = []
        index = 1
        while True:
            if ws.cell(key_positions['no'][1]+index, key_positions['no'][0]).value is None:
                break
            table.append({
                'no':ws.cell(key_positions['no'][1]+index, key_positions['no'][0]).value,
                'code':ws.cell(key_positions['code'][1]+index, key_positions['code'][0]).value if key_positions['code'] else '',
                'name':ws.cell(key_positions['name'][1]+index, key_positions['name'][0]).value,
                'quantity':ws.cell(key_positions['quantity'][1]+index, key_positions['quantity'][0]).value,
                'unit':ws.cell(key_positions['unit'][1]+index, key_positions['unit'][0]).value,
                'price':ws.cell(key_positions['price'][1]+index, key_positions['price'][0]).value,
                'sum':ws.cell(key_positions['sum'][1]+index, key_positions['sum'][0]).value,
            })
            index+=1
        res['tables'] = table
        return res
        
    

def handle_pdf_file(pdf_path: str) -> dict[str, str]:
    text = ""
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfFileReader(file)
        for page_num in range(reader.numPages):
            page = reader.getPage(page_num)
            text += page.extractText().replace("\xa0", ' ')
    codes = handle_text_file(text, False)
    codes['tables'] = get_tables(pdf_path)
    
    return codes


def send_to_1c(request: HttpRequest):
    if request.method == 'POST':
        data_received = json.loads(request.body)
        print(data_received)
        login = 'Битрикс24'.encode('utf-8')
        password = 'wnF%AVGg7Xxv'
        # data_received['username'] = login 
        # data_received['password'] = password 
        # session = requests.Session()
        # session.auth = (login, password)
        # print(session.auth)
        response = requests.post('http://10.10.10.166/service83_test/hs/BTX/ExchangeBetweenDatabases1', json=data_received, auth=(login, password))
        print(response.text, '*'*5, response.status_code)
        if response.ok:
            return HttpResponse(status=response.status_code, content=response.json()["Error"]["ru"])
        else:
            error_message = response.json()["Error"]["ru"]
            print(error_message)
            return HttpResponse(status=response.status_code, content=error_message)
    else:
        return HttpResponse(status=400)