import contextlib
import decimal
import io
import logging
import os
import pathlib
from datetime import date, datetime
from tempfile import TemporaryFile

import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.worksheet.worksheet
from django.db.models import (
    Case,
    Exists,
    F,
    OuterRef,
    Q,
    QuerySet,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, TruncDate

from avh_services import read_json
from finance_module import models

from . import common_service


def set_worksheet_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    worksheet.append(
        [
            "ДО",
            "Дата входящего",
            "№ счёта",
            "Дата счёта",
            "Проект",
            "Ответ.",
            "Утверд.",
            "Компания",
            "Контрагенты",
            "Описание",
            "Валюта",
            "Сумма для таблицы",
            "Категория счета",
            "Статьи доходов/расходов",
            "Номер заказа",
            "БИН/ИИН",
            "Сумма документа",
            "Номер р/с",
            "ИИК",
            "КНП",
            "Факт номер договора",
            "Сумма по счету",
            "Оплаченная сумма ранее",
            "Оплачено",
            "Дата оплаты",
            "Регион/ПМ",
            "ПМ (ФИО)",
            "Руководитель",
            "Подр",
            "Расчетный счет",
        ]
    )


def change_format_of_cells_unpaid(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [2, 4, 25]:
                cell.number_format = "DD.MM.YYYY"
            if cell.column in [12, 22, 23, 24]:
                cell.number_format = "#,##0.00"


def get_paid_invoices() -> QuerySet:
    return models.PaidInvoice.objects.filter(
        status=models.completed_paid_invoice_status()
    )


def set_data_to_worksheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, paid_invoices: QuerySet
) -> None:
    for paid_invoice in paid_invoices:
        at = paid_invoice.at.replace(tzinfo=None)

        worksheet.append(
            [
                paid_invoice.number,
                paid_invoice.date,
                paid_invoice.invoice_number,
                paid_invoice.invoice_date,
                paid_invoice.project,
                paid_invoice.responsible_user_id,
                paid_invoice.approver,
                paid_invoice.llc,
                paid_invoice.contractor,
                paid_invoice.comment,
                paid_invoice.currency,
                paid_invoice.sum,
                paid_invoice.invoice_category,
                paid_invoice.revenue_expense_articles,
                paid_invoice.sales_order,
                paid_invoice.bin_or_iin,
                paid_invoice.document_amount,
                paid_invoice.account.number,
                paid_invoice.iic,
                paid_invoice.payment_destination_code,
                paid_invoice.contract_number,
                paid_invoice.invoice_amount,
                paid_invoice.paid_amount_1c,
                paid_invoice.paid,
                at,
                paid_invoice.project_region.name,
                paid_invoice.responsible.get_full_name()
                if paid_invoice.responsible
                else "-",
                paid_invoice.project_region.director.get_full_name()
                if paid_invoice.project_region.director
                else "-",
                (
                    paid_invoice.project_region.subdivision.name
                    if paid_invoice.project_region.subdivision
                    else "-"
                ),
                paid_invoice.account.name,
            ]
        )

    stylizing_worksheet(worksheet)
    change_format_of_cells_unpaid(worksheet)


def stylizing_name_column_worksheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D0CECE", end_color="D0CECE", fill_type="solid"
            )

            if cell.value == "Контрагенты":
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="FF9999", end_color="FF9999", fill_type="solid"
                )
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    worksheet.row_dimensions[1].height = 40


def change_width_of_column(cell, column_idxs: list, max_length) -> int:
    if cell.column in column_idxs:
        max_length = 10
        return max_length

    return max_length


def set_worksheet_columns_width(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, column_idxs
):
    changed_max_length = 0
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

                changed_max_length = change_width_of_column(
                    cell, column_idxs=column_idxs, max_length=max_length
                )
            except TypeError:
                pass

        adjusted_width = changed_max_length + 2
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(column[0].column)
        ].width = adjusted_width


def stylizing_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    stylizing_name_column_worksheet(worksheet)

    for row_idx, row_data in enumerate(worksheet.iter_rows(min_row=2)):
        for cell in row_data:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="76FF4B", end_color="76FF4B", fill_type="solid"
            )
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )

            if cell.column == 24 and cell.value:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
            if cell.column == 25 and cell.value:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                )

    set_worksheet_columns_width(worksheet, [1])


def upload_paid_invoices() -> TemporaryFile:
    workbook = openpyxl.Workbook()
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

    set_worksheet_columns(worksheet)
    set_data_to_worksheet(worksheet, get_paid_invoices())

    temporary_file = TemporaryFile()
    workbook.save(temporary_file)
    temporary_file.seek(0)
    return temporary_file


UNPAID_INVOICES_FILE_PATH: str | pathlib.Path | None = os.getenv(
    "UNPAID_INVOICES_FILE_PATH"
)
if UNPAID_INVOICES_FILE_PATH:
    UNPAID_INVOICES_FILE_PATH = pathlib.Path(UNPAID_INVOICES_FILE_PATH)


def give_right_format_of_date(date: str) -> str | None:
    if date is None:
        return None

    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y"]

    dates = date.split(",")
    formatted_dates = []
    for date in dates:
        date = date.strip()
        for fmt in formats:
            try:
                formatted_date = datetime.strptime(date, fmt).strftime("%Y-%m-%d")
                formatted_dates.append(formatted_date)
                break
            except ValueError:
                continue
    return ", ".join(formatted_dates) if formatted_dates else date


def import_unpaid_invoices(*, bytes_io: io.BytesIO | None = None) -> None:
    models.UnpaidInvoice.objects.all().delete()
    unpaid_invoices = {}
    if not bytes_io:
        assert UNPAID_INVOICES_FILE_PATH is not None
        for row in read_json(UNPAID_INVOICES_FILE_PATH, "colgroup", "quantity"):
            if str(row["blacklist"]).casefold() != "N".casefold():
                continue
            number = row["inv_documentno"]
            if number in unpaid_invoices:
                logging.error(f"Duplicating unpaid invoices number - {number}")
                continue
            unpaid_invoices[number] = models.UnpaidInvoice(
                number=number,
                date=give_right_format_of_date(row["inv_date"]),
                invoice_number=row["invoicename2"] if row["invoicename2"] else None,
                invoice_date=give_right_format_of_date(row["dateinvoiced2"]),
                project=row["c_activity"] if row["c_activity"] else None,
                responsible_user_id=row["uvalue"] if row["uvalue"] else None,
                approver=row["utverditel"] if row["utverditel"] else None,
                llc=row["bpcompany"] if row["bpcompany"] else None,
                contractor=row["inv_bpartner"] if row["inv_bpartner"] else None,
                comment=row["invline_product"] if row["invline_product"] else None,
                currency=row["iso_code"] if row["iso_code"] else None,
                # TODO: check this field -> "Сумма для таблицы"
                invoice_category=row["icname"] if row["icname"] else None,
                revenue_expense_articles=row["chname"] if row["chname"] else None,
                sales_order=row["doc_number"] if row["doc_number"] else None,
                bin_or_iin=row["bin"] if row["bin"] else None,
                document_amount=row["quantity"] or None,
                # TODO: check this field -> "С какого р/с платить"
                iic=row["iik"] if row["iik"] else None,
                contract_number=row["agreement"] if row["agreement"] else None,
                invoice_amount=row["openamt"],
                paid_amount_1c=row["payamt1c"] or None,
                bank=row["bank"] if row["bank"] else None,
                payment_type=(
                    row["typeinvcategory"] if row["typeinvcategory"] else None
                ),
                status=row["i_docstatus"] if row["i_docstatus"] else None,
                creator_user_id=(
                    row["invoicecreator"] if row["invoicecreator"] else None
                ),
                department=row["dcode"] if row["dcode"] else None,
                due_date=None,  # TODO: check this field -> "Дата платежа"
                document_number=(
                    row["docserviceact"] if row["docserviceact"] else None
                ),
                document_date=row["docdate"] if row["docdate"] else None,
                closing_document_amount=row["quantity"] or None,
            )

    else:
        workbook = openpyxl.load_workbook(bytes_io, data_only=True, read_only=True)
        unpaid_invoices_worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook[
            "РЕЕСТР неоплаченные счета"
        ]
        for row in unpaid_invoices_worksheet.iter_rows(
            min_row=2, max_col=39, max_row=10000000, values_only=True
        ):
            if row[0] is None:
                break
            number = row[0]
            if number in unpaid_invoices:
                logging.error(f"Duplicating unpaid invoices number - {number}")
                continue
            unpaid_invoices[number] = models.UnpaidInvoice(
                number=row[0],
                date=row[1],
                invoice_number=row[2],
                invoice_date=row[3] or None,
                project=row[4],
                responsible_user_id=row[5],
                approver=row[6],
                llc=row[7],
                contractor=row[8],
                comment=row[9] or None,
                currency=row[10],
                invoice_category=row[12],
                revenue_expense_articles=row[13],
                sales_order=row[14],
                bin_or_iin=row[15],
                iic=row[18] or None,
                contract_number=row[20],
                invoice_amount=row[21],
                paid_amount_1c=row[22] or None,
                bank=row[26] or None,
                payment_type=row[27] or None,
                status=row[28] or None,
                creator_user_id=row[29] or None,
                department=row[31] or None,
                due_date=row[32] or None,
                document_number=row[33] or None,
                document_date=row[34] or None,
                document_amount=row[35] or None,
                closing_document_amount=row[38] or None,
            )

    models.UnpaidInvoice.objects.bulk_create(unpaid_invoices.values())


def import_unpaid_invoices_work_statuses() -> None:
    document_debt_worksheet = common_service.get_document_debt_worksheet()
    values = document_debt_worksheet.get_values()
    models.UnpaidInvoice.objects.all().update(has_in_document_debts=False, work_status=None)
    for value in values[2:]:
        number = value[0]
        work_status = value[16] or None
        with contextlib.suppress(models.UnpaidInvoice.DoesNotExist):
            unpaid_invoice = models.UnpaidInvoice.objects.get(number=number)
            unpaid_invoice.work_status = work_status
            unpaid_invoice.has_in_document_debts = True
            unpaid_invoice.save()


def import_unpaid_invoices_payment_destination_codes() -> None:
    technical_list_worksheet = common_service.get_technical_list_worksheet()
    for (
        revenue_expense_articles,
        payment_destination_code,
    ) in technical_list_worksheet.get_values("I2:J"):
        unpaid_invoices = models.UnpaidInvoice.objects.filter(
            revenue_expense_articles=revenue_expense_articles
        )
        unpaid_invoices.update(payment_destination_code=payment_destination_code)


def annotate_unpaid_invoices(queryset: QuerySet) -> QuerySet:
    today = date.today()
    return queryset.annotate(
        annotated_has_exception=Exists(
            Subquery(
                models.UnpaidInvoiceException.objects.annotate(
                    created_at_date=TruncDate("created_at")
                ).filter(number=OuterRef("number"), created_at_date=today)
            )
        ),
        annotated_today_paid=Coalesce(
            Subquery(
                models.PaidInvoice.objects.exclude(
                    status=models.rejected_paid_invoice_status()
                )
                .annotate(at_date=TruncDate("at"))
                .filter(
                    number=OuterRef("number"),
                    at_date__lte=today,
                )
                .values("number")
                .annotate(total_sum=Sum("sum"))
                .values("total_sum")
            ),
            decimal.Decimal(),
        ),
        annotated_planned_payment=Coalesce(
            Subquery(
                models.PaidInvoice.objects.exclude(
                    status=models.rejected_paid_invoice_status()
                )
                .annotate(at_date=TruncDate("at"))
                .filter(number=OuterRef("number"), at_date__gt=today)
                .values("number")
                .annotate(total_sum=Sum("sum"))
                .values("total_sum")
            ),
            decimal.Decimal(),
        ),
        annotated_paid=Coalesce(F("paid_amount_1c"), decimal.Decimal())
        + F("annotated_today_paid")
        + F("annotated_planned_payment"),
        annotated_remainder=F("invoice_amount") - F("annotated_paid"),
        annotated_allowed_percent=Case(
            When(annotated_has_exception=True, then=Value(decimal.Decimal(1))),
            When(
                Q(invoice_category__in=("Строительно-монтажные работы", "Обследование", "Оптика", "Электромонтажные работы"))
                & Q(annotated_paid__lt=F("invoice_amount") * decimal.Decimal(0.5)),
                then=Value(decimal.Decimal(0.5)),
            ),
            default=Value(decimal.Decimal(1))
        ),
        annotated_allowed_payment_amount=F("invoice_amount") * F("annotated_allowed_percent") - F("annotated_paid"),
        annotated_pm_sum=Coalesce(
            models.UnpaidInvoicePMSum.objects.filter(
                number=OuterRef("number"), date=today
            ).values("sum"),
            decimal.Decimal(),
        ),
        annotated_payment_decision=Case(
            When(
                Q(invoice_category="ТМЦ") | Q(annotated_has_exception=True),
                then=Value("OK"),
            ),
            When(
                (
                    Q(
                        invoice_category__in=(
                            "Строительно-монтажные работы",
                            "Обследование",
                            "Оптика", "Электромонтажные работы"
                        )
                    )
                    & Q(annotated_paid__gte=F("invoice_amount") * decimal.Decimal(0.5))
                )
                & Q(closing_document_amount__isnull=True),
                then=Value("Предоставьте закрывающий документ!"),
            ),
            When(has_in_document_debts=True, work_status__isnull=True, then=Value("Отказ! Заполнить статус работ")),
            When(
                has_in_document_debts=True,
                work_status__in=(
                    "завершено (проблемный)",
                    "не выполнялось (проблемный)",
                ),
                then=Value("Отказ! проблемный"),
            ),
            default=Value("OK"),
        ),
        annotated_can_pay=Q(annotated_payment_decision="OK"),
    )
