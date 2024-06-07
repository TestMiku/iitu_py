from tempfile import TemporaryFile

import openpyxl
import openpyxl.worksheet.worksheet

from finance_module import models
from finance_module.services.unpaid_invoices_service import set_worksheet_columns_width


# locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')


def set_worksheet_columns_seizure(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    worksheet.append(
        [
            "Дата",
            "Подр",
            "ПМ/Регион",
            "ПМ (ФИО)",
            "Руководитель",
            "Сумма",
            "Контрагент/Статья",
            "С расчетного счета",
            "Номер р/с",
            "На расчетный счет",
            "Номер р/с2",
            "Примечание"
        ]
    )


def get_mandatory_payments_seizure():
    return models.MandatoryPaymentSeizure.objects.filter(imported_from_file=False)


def change_format_of_cells(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [1]:
                cell.number_format = "DD.MM.YYYY"
            if cell.column in [6]:
                cell.number_format = "#,##0.00"


def set_data_to_worksheet_seizure(worksheet: openpyxl.worksheet.worksheet.Worksheet, mandatory_payments):
    for mandatory_payment in mandatory_payments:
        at = mandatory_payment.datetime.date()
        formatted_sum = mandatory_payment.sum

        worksheet.append(
            [
                at,
                mandatory_payment.project_region.subdivision.name if mandatory_payment.project_region.subdivision else "",
                mandatory_payment.project_region.name if mandatory_payment.project_region else "",
                mandatory_payment.project_region.name if mandatory_payment.project_region else "",
                mandatory_payment.project_region.director.get_full_name() if mandatory_payment.project_region.director else "",
                formatted_sum,
                mandatory_payment.mandatory_payment.name if mandatory_payment.mandatory_payment else "",
                mandatory_payment.account.name if mandatory_payment.account else "",
                mandatory_payment.account.number if mandatory_payment.account else "",
                "АДМ",
                "KZ208562203106786042",
                ""
            ]
        )

    stylizing_worksheet(worksheet)
    change_format_of_cells(worksheet)


def stylizing_name_column_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    for cell in worksheet["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.fill = openpyxl.styles.PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")

        if cell.column in [1, 3, 6, 7, 8, 9, 12]:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def stylizing_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    stylizing_name_column_worksheet(worksheet)

    for row_idx, row_data in enumerate(worksheet.iter_rows(min_row=2)):
        for cell in row_data:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style="thin"),
                right=openpyxl.styles.Side(border_style="thin"),
                top=openpyxl.styles.Side(border_style="thin"),
                bottom=openpyxl.styles.Side(border_style="thin"),
            )

            if cell.column in [1, 3, 6, 7, 8, 9, 12]:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    set_worksheet_columns_width(worksheet, [1, 6])


def unloading_seizure() -> TemporaryFile:
    workbook = openpyxl.Workbook()
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

    set_worksheet_columns_seizure(worksheet)
    set_data_to_worksheet_seizure(worksheet, get_mandatory_payments_seizure())

    temporary_file = TemporaryFile()
    workbook.save(temporary_file)
    temporary_file.seek(0)
    return temporary_file
