import os

import django.http
from django.utils import timezone
from decimal import Decimal
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from django.conf import settings
from openpyxl import load_workbook
import re
from openpyxl.styles import Font, Border, Alignment, PatternFill
from openpyxl.styles.borders import Side
from dateutil import parser
from finance_module.models import Inflow, Account, ProjectRegion


def find_cell(sheet, target_phrase):  # Поиск нужного индекса колонны и строки
    for row in sheet.iter_rows(min_row=5, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            if target_phrase in str(cell.value):
                return cell.row, cell.column  # Возвращает индекс строки и столбца, когда находит фразу

    return None, None


def file1_to_dict(file1):  # Читаем первый файл и отправляем нужные данные в словаре
    data = []
    for file in file1:
        wb1 = load_workbook(filename=file)
        ws1 = wb1.active
        check_sender = True

        purpose_row, purpose_column = find_cell(ws1, "Назначение платежа")
        # print("НП: ",purpose_row ,purpose_column)
        credit_row, credit_column = find_cell(ws1, "Кредит")
        # print("Кредит: ",credit_row, credit_column)
        buyer_row, buyer_column = find_cell(ws1, "Корреспондент")
        # print("Корреспондент: ",buyer_row, buyer_column)
        date_row, date_column = find_cell(ws1, "Дата")
        # print("Дата: ",date_row, date_column)
        iik_row, iik_column = find_cell(ws1, "ИИК корр-та")
        print("ИИК: ", iik_row, iik_column)
        if not buyer_row or not buyer_column:
            check_sender = False
            buyer_row, buyer_column = find_cell(ws1, "Отправитель")
            # print("Отправитель: ",buyer_row, buyer_column)

        if purpose_column is None or credit_column is None:
            return None  # Сделать ошибку


        for row in range(purpose_row + 1, ws1.max_row + 1):
            copy_id = ''
            check_id = False
            payment_cell_value = ws1.cell(row=row, column=purpose_column).value

            if payment_cell_value is not None:  # and check_sender:
                payment_cell_value = payment_cell_value.replace('\n', '')
                # Находим номер договора, если будут новые шаблоны добавить сюда
                match = re.search(r'(дог. №|дог №|договор №|договора № |договора №| № №|договору №|Договор от №|согл.дог.№|согл.дог. №|договору |договора |дог. n |дог. n|дог.|дог )(.+?)(?=\s+от)', payment_cell_value.lower(), re.DOTALL)
                if match:

                    # Удаление лишних символов и т.д.
                    id_value = match.group(2).strip().replace('\n', '')
                    id_value = re.sub(r"\s*\([^)]*\)", '', id_value)
                    id_value = re.sub(r"\b усл.кв\b", '', id_value)
                    id_value = re.sub(r"\bот 14.05.2021 Г. № 2313\b", '', id_value)
                    id_value = re.sub(r"\b приложение №7\b", '', id_value)
                    id_value = re.sub(r"№ №", '', id_value)
                else:   # Если не нашли номер договора, тогда записать всю ячейку
                    id_value = payment_cell_value.replace('\n', '')
            # elif payment_cell_value is not None and not check_sender:
            #     id_value = payment_cell_value.replace('\n', '')

            else:
                continue
            iik_value = '-'
            buyer_value = ws1.cell(row=row, column=buyer_column).value
            buyer_value = clean_string(buyer_value, check_sender)
            copy_buyer_value = buyer_value.strip().lower()
            if "аврора" in copy_buyer_value:
                check_id = True
            if check_id:
                copy_id = payment_cell_value
            credit_value = ws1.cell(row=row, column=credit_column).value
            date_value = ws1.cell(row=row, column=date_column).value
            if iik_column is not None:
                iik_value = ws1.cell(row=row, column=iik_column).value
            try:
                dt = parser.parse(date_value)  # Форматируем время
            except TypeError:
                dt = date_value

            # Только если колонна "Кредит" больше 0 и не пустая записываем данные
            if credit_value is not None and int(credit_value) > 0:
                data.append({
                    "id": id_value,
                    "credit": credit_value,
                    "buyer": buyer_value,
                    "date": dt,
                    "check_id": check_id,
                    "copy_id": copy_id,
                    "file_name": os.path.splitext(file.name)[0],
                    "iik": iik_value,
                    })

    return data


def file2_to_dict(file2):
    # Доработка по файлу, что отправила Айдана
    pm_dict = {
        '71-0001': "АДМ на 4",
        '71-0206': "Евгения Богомолова",
        '71-0213': "Александр Азаров",
        '71-0216': "Дмитрий Ваганов",
        '71-0221': "АДМ на 4",
        '71-0222': "Азамат Бейсен",
        '71-0224': "Александр Азаров",
        '71-0226': "Александр Азаров",
        '71-0232': "Дмитрий Ваганов",
        '71-0235': "Дмитрий Ваганов",
        '71-0236': "Александр Азаров",
        '71-0242': "Евгения Богомолова",
        '71-0244': "Евгения Богомолова",
        '71-223': "Азамат Бейсен",
        '71-227': "АДМ на 4",
        }
    wb = load_workbook(file2)
    sheet = wb['7.24']
    sheet.auto_filter.ref = None
    output_data = []
    seen = set()

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[21].value is None:
            continue
        id_value = row[21].value
        recipient_value = row[13].value.replace('ұ', 'у')  # Не читает казахские символы
        recipient_value = recipient_value.replace('қ', 'к')
        recipient_value = recipient_value.split('_')[0]  # Забираем строку ДО "_"
        project_value = '-'.join(row[15].value.split('-')[:2]).strip()  # Забираем строку ДО второй "-"
        for key in pm_dict.keys():
            if project_value == key:  # Если найдено по словарю, что выше, тогда поставить нужное значение
                recipient_value = pm_dict[key]

        buyer_value = row[10].value  # Добавил на будущее, если нужно будет проверять и по покупателям

        record = (id_value, recipient_value, project_value, buyer_value,)
        if record not in seen:
            seen.add(record)
            output_data.append({
                "id": id_value,
                "recipient": recipient_value,
                "project": project_value,
                "buyer_value": buyer_value,  # Сейчас buyer_value нигде не используется
                })

    return output_data


def compare_files(file1, file2):  # Главная функция которая вызывается в views
    dict1 = file1_to_dict(file1)
    # print("dict1 ready:", dict1)
    dict2 = file2_to_dict(file2)
    if dict1 is None:
        return None  # Сделать ошибку
    if dict2 is None:
        return None
    data = []
    inflow_data = []
    counter = 0
    if dict1:
        for item1 in dict1:
            counter += 1
            id = item1['id'].upper()
            date = item1['date'].strftime("%d.%m.%Y")
            buyer = item1['buyer']
            sum = item1['credit']
            reserve = round((item1['credit'] * 0.107), 2)  # По умолчанию резерв 10,7%
            payed = round((item1['credit'] - reserve), 2)
            recipient = "Совпадения не найдены"
            project = "Совпадения не найдены"

            inflow = {  # Создаем словарь чтобы его передавать Самату в InFlow модельку
                'account': item1['file_name'],
                'project_region': 'Не найдено',
                'sum': item1['credit'],
                'reserve_percent': 'Не найдено',
                'imported_from_file': False
            }

            for item2 in dict2:
                if item1['id'].upper() == item2['id'].upper():
                    recipient = item2['recipient']
                    inflow_project = recipient.split()[0]
                    project = item2['project']

                    # cleaned_buyer = buyer.strip().lower()
                    # print("cleaned_buyer: ", cleaned_buyer)
                    if item1['check_id']:  # Если покупатель и поставщик Аврора, тогда резерва нет, идет вся сумма.
                        id = item1['copy_id']
                        reserve_percent = 0
                        reserve = 0
                        payed = item1['credit']
                    else:
                        reserve_percent = 0.107
                    inflow = {  # Дополняем словарь нужными данными
                        'account': item1['file_name'],
                        'project_region': inflow_project,
                        'sum': sum,
                        'reserve_percent': reserve_percent,
                        'imported_from_file': True
                    }
                    break

                elif item1['check_id']:
                    if item1['copy_id'] != '':
                        id = item1['copy_id']
                    recipient = "С Авроры на Аврору"
                    project = "Проект не найден"
                    reserve_percent = 0
                    reserve = 0
                    payed = item1['credit']
                    inflow = {  # Дополняем словарь нужными данными
                        'account': item1['file_name'],
                        'project_region': recipient,
                        'sum': sum,
                        'reserve_percent': reserve_percent,
                        'imported_from_file': True
                    }
                    break

            entry = {  # Основной словарь который показывается на странице
                'num': counter,
                'id': id,
                'date': date,
                'buyer': buyer,
                'sum': sum,
                'reserve': reserve,
                'payed': payed,
                'recipient': recipient,
                'project': project,
                'file_name': item1['file_name'],
                'iik': item1['iik']
            }

            data.append(entry)
            inflow_data.append(inflow)

    save_to_inflow(inflow_data)
    file_path = create_and_save_file(data, inflow_data)
    return data, file_path


def clean_string(s, check):  # Удалить лишний текст в поле "Корреспондент"
    index = s.find("БИН:")
    if index != -1 and check:
        return s[:index]
    else:
        return s


def create_and_save_file(data, inflow_data):  # Создаем и сохраняем excel файл, для отправки
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Приходы другие"
        ws2 = wb.create_sheet("Приходы Аврора")
        # Заголовки файла
        columns = ["№ Контракта | Название", "Дата", "Покупатель", "Вся сумма", "Резерв (10,7%)", "Полученная сумма", "ПМ", "Проект"]
        ws1.append(columns)
        header_color = PatternFill(start_color='F8CBAD', end_color='95B3D7', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, size=12, color='FF000000')
        thin_border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))
        file_name_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        previous_file_name = None
        for item in data:
            current_file_name = item['file_name']
            if current_file_name != previous_file_name:
                # Добавляем объединенную строку с именем файла
                ws1.append([current_file_name])
                ws1.merge_cells(start_row=ws1.max_row, start_column=1, end_row=ws1.max_row, end_column=len(columns))
                for cell in ws1[ws1.max_row]:
                    cell.fill = file_name_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if item['reserve'] > 0:
                row = [
                    item['id'], item['date'], item['buyer'], item['sum'], item['reserve'],
                    item['payed'], item['recipient'], item['project']
                ]
                ws1.append(row)
                if len(str(row[0])) > 165:  # Если слишком длинное слово, увеличить высоту
                    ws1.row_dimensions[ws1.max_row].height = 42
                else:
                    ws1.row_dimensions[ws1.max_row].height = 30
                for cell in ws1[ws1.max_row]:
                    cell.alignment = Alignment(wrap_text=True)
            previous_file_name = current_file_name

        for cell in ws1['1:1']:
            cell.fill = header_color
            cell.font = header_font
        file_names = set(item['file_name'] for item in data)
        for row in ws1.iter_rows():
            for cell in row:
                if cell.value not in file_names and cell.value:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border

        ws1.column_dimensions['A'].width = 95  # Ширина первой колонки и так далее
        ws1.column_dimensions['B'].width = 11
        ws1.column_dimensions['C'].width = 60
        ws1.column_dimensions['D'].width = 19
        ws1.column_dimensions['E'].width = 19
        ws1.column_dimensions['F'].width = 19
        ws1.column_dimensions['G'].width = 40
        ws1.column_dimensions['H'].width = 23

        # Создаем новый лист если отправитель Аврора
        columns = ["Дата", "ИИК корр-та", "Покупатель", "Кредит", "Назначение платежа", "С какого р/с"]
        ws2.append(columns)
        accounts = {
            "KZ228560000007115517": "Аврора Сервис - Евгений-Теле2  7п",
            "KZ338562203101509590": "Аврора Сервис - Евгений-Картел 7п",
            "KZ888562203101509570": "Аврора Сервис - Артур-Картел 7п",
            "KZ768560000004331906": "Аврора Сервис - Алмата КУПРИН7п",
            "KZ9396502F0011892216": "LVE форте - КупиПродай 71п",
            "KZ3196502F0013099333": "LVE форте - СМР 71п",
            "KZ8596502F0012697266": "Аврора 77 форте",
            "KZ818562203111052071": "Аврора Сервис - КупиПродай 71п",
            "KZ798562203102359064": "Аврора Сервис - СМР 71п",
            "KZ258562203109686640": "Аврора77 БЦК 74п",
            "KZ288562203117669492": "Аврора77 БЦК Бисен 71п",
            "KZ528562203122822476": "АС ЦМК 74п",
            "KZ518560000005034248": "Виск 71п",
        }

        p = []
        previous_file_name = None
        for item in data:
            current_file_name = item['file_name']
            if current_file_name != previous_file_name:
                # Добавляем объединенную строку с именем файла
                ws2.append([current_file_name])
                ws2.merge_cells(start_row=ws2.max_row, start_column=1, end_row=ws2.max_row, end_column=len(columns))
                for cell in ws2[ws2.max_row]:
                    cell.fill = file_name_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sender_iik_match = re.search(r"ИИК: (\S+)", item['buyer'])
            if sender_iik_match:
                sender_iik = sender_iik_match.group(1)
            elif item['iik'] != '-':
                sender_iik = item['iik']
            else:
                sender_iik = "Не нашел ИИК в файле"

            sender_name_match = re.search(r"(.*)\nБИН:", item['buyer'])
            if sender_name_match:
                sender_name = sender_name_match.group(1).strip()
            else:
                sender_name = item['buyer']
            # print(sender_iik)
            # print(sender_name)

            from_schet = 'Не известно с какого р/с'
            for key in accounts.keys():
                if sender_iik == key:  # Если найдено по словарю, что выше, тогда поставить нужное значение
                    from_schet = accounts[key]

            if item['reserve'] == 0:
                row = [
                    item['date'], sender_iik, sender_name, item['payed'],
                    item['id'], from_schet
                ]
                ws2.append(row)
                # print("row 4: ", type(row[4]))
                if len(row[4]) > 165:  # Если слишком длинное слово, увеличить высоту
                    ws2.row_dimensions[ws2.max_row].height = 60
                else:
                    ws2.row_dimensions[ws2.max_row].height = 30
                for cell in ws2[ws2.max_row]:
                    cell.alignment = Alignment(wrap_text=True)

                p.append(ws2.cell(row=ws2.max_row+1, column=4, value=item['payed'] * 0.29,).value)
                ws2.cell(row=ws2.max_row, column=5, value='Азамат Бейсен')

                p.append(ws2.cell(row=ws2.max_row + 1, column=4, value=item['payed'] * 0.29, ).value)
                ws2.cell(row=ws2.max_row, column=5, value='Александр Азаров')

                p.append(ws2.cell(row=ws2.max_row + 1, column=4, value=item['payed'] * 0.29, ).value)
                ws2.cell(row=ws2.max_row, column=5, value='Дмитрий Ваганов')

                p.append(ws2.cell(row=ws2.max_row + 1, column=4, value=item['payed'] * 0.13, ).value)
                ws2.cell(row=ws2.max_row, column=5, value='Евгения Богомолова')
                # print(p)
                if item['project'] != "Проект не найден" and item['project'] != "Совпадения не найдены":
                    for item_inflow in inflow_data:
                        try:
                            project_region = ProjectRegion.objects.get(name=item_inflow['project_region'])
                        except ObjectDoesNotExist:
                            continue

                        try:
                            account = Account.objects.get(name=item_inflow['account'])
                        except ObjectDoesNotExist:
                            account = Account.objects.create(name=item_inflow['account'])
                        for entry in p:
                            inflow_exists = Inflow.objects.filter(
                                account=account,
                                project_region=project_region,
                                sum=entry,
                                reserve_percent=item_inflow['reserve_percent'] * 100,
                                date=timezone.localdate(),
                                imported_from_file=True
                            ).exists()

                            if not inflow_exists:
                                # Создаем новый объект Inflow, если такого еще нет
                                new_inflow = Inflow(
                                    account=account,
                                    project_region=project_region,
                                    sum=entry,
                                    reserve_percent=item_inflow['reserve_percent'] * 100,
                                    date=timezone.localdate(),
                                    imported_from_file=True
                                )

                                # Сохраняем объект в базу данных
                                new_inflow.save()
            previous_file_name = current_file_name
                                # print(new_inflow)
                                # print("----------------------------")
                                # print(f"Saved inflow for account {item['account']} and region {item['project_region']} with sum {item['sum']}")

        for cell in ws2['1:1']:
            cell.fill = header_color
            cell.font = header_font
        file_names = set(item['file_name'] for item in data)
        for row in ws2.iter_rows():
            for cell in row:
                if cell.value not in file_names and cell.value:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border

        ws2.column_dimensions['A'].width = 11  # Ширина первой колонки и так далее
        ws2.column_dimensions['B'].width = 23
        ws2.column_dimensions['C'].width = 35
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 60
        ws2.column_dimensions['F'].width = 33

        # Файл каждый раз будет перезаписываться чтобы не занимать много места.
        file_path = os.path.join(settings.BASE_DIR, 'finance_module', 'services/unload_templates', 'income71P.xlsx')
        wb.save(file_path)
        wb.close()
        return file_path
    except Exception as e:
        print(f"Ошибка создания Excel файла: {e}")
        return None


def save_to_inflow(data):  # Для отправки данных в модель Самата
    # print(data)
    for item in data:
        if item['project_region'] != "Проект не найден" and item['project_region'] != "Совпадения не найдены":
            try:
                project_region = ProjectRegion.objects.get(name=item['project_region'])
            except ObjectDoesNotExist:
                continue

            try:
                account = Account.objects.get(name=item['account'])
            except ObjectDoesNotExist:
                account = Account.objects.create(name=item['account'])
        #
            inflow_exists = Inflow.objects.filter(
                account=account,
                project_region=project_region,
                sum=item['sum'],
                reserve_percent=item['reserve_percent'] * 100,
                date=timezone.localdate(),
                imported_from_file=True
            ).exists()

            if not inflow_exists:
                # Создаем новый объект Inflow, если такого еще нет
                new_inflow = Inflow(
                    account=account,
                    project_region=project_region,
                    sum=item['sum'],
                    reserve_percent=item['reserve_percent'] * 100,
                    date=timezone.localdate(),
                    imported_from_file=True
                )

                # Сохраняем объект в базу данных
                new_inflow.save()

                # print(
                #     f"Saved inflow for account {item['account']} and region {item['project_region']} with sum {item['sum']}")
            else:
                pass
                # print(
                #     f"Inflow for account {item['account']} and region {item['project_region']} with sum {item['sum']} already exists and was not saved again.")
