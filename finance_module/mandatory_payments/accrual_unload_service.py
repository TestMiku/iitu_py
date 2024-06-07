import io
from tempfile import TemporaryFile

import openpyxl
import openpyxl.worksheet.worksheet
from django.db.models import QuerySet

from finance_module import models
from finance_module.services.unpaid_invoices_service import set_worksheet_columns_width


def set_worksheet_columns_accrual(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    worksheet.append(
        [
            "Дата",
            "Подр",
            "ПМ/Регион",
            "Cумма",
            "Контрагент/Статья",
            "Примечание",
            "ПМ (ФИО)",
            "Руководитель",
        ]
    )


def get_mandatory_payments_accrual():
    return models.MandatoryPaymentAccrual.objects.filter(imported_from_file=False)


def change_format_of_cells(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [1]:
                cell.number_format = "DD.MM.YYYY"
            if cell.column in [4]:
                cell.number_format = "#,##0.00"


def set_data_to_worksheet_accrual(worksheet: openpyxl.worksheet.worksheet.Worksheet, mandatory_payments: QuerySet):
    for mandatory_payment in mandatory_payments:
        formatted_sum = mandatory_payment.sum

        worksheet.append(
            [
                mandatory_payment.datetime,
                mandatory_payment.project_region.subdivision.name if mandatory_payment.project_region.subdivision else "",
                mandatory_payment.project_region.name if mandatory_payment.project_region else "",
                formatted_sum,
                mandatory_payment.mandatory_payment.name if mandatory_payment.mandatory_payment else "",
                mandatory_payment.note,
                mandatory_payment.project_region.name if mandatory_payment.project_region else "",
                mandatory_payment.project_region.director.get_full_name() if mandatory_payment.project_region.director else "",
            ]
        )

    stylizing_worksheet(worksheet)
    change_format_of_cells(worksheet)


def stylizing_name_column_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    for cell in worksheet["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.fill = openpyxl.styles.PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")


def stylizing_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    stylizing_name_column_worksheet(worksheet)

    for row_idx, row_data in enumerate(worksheet.iter_rows(min_row=2)):
        for cell in row_data:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style="thin"),
                right=openpyxl.styles.Side(border_style="thin"),
                top=openpyxl.styles.Side(border_style="thin"),
                bottom=openpyxl.styles.Side(border_style="thin"),
            )

            if row_idx % 2 == 0:
                cell.fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            else:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    set_worksheet_columns_width(worksheet, [1, 4])


def unloading_accrual() -> TemporaryFile:
    workbook = openpyxl.Workbook()
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active

    set_worksheet_columns_accrual(worksheet)
    set_data_to_worksheet_accrual(worksheet, get_mandatory_payments_accrual())

    temporary_file = TemporaryFile()
    workbook.save(temporary_file)
    temporary_file.seek(0)
    return temporary_file
