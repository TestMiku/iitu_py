import dataclasses
import datetime
import decimal
import io
import itertools
import logging
import typing

import openpyxl
from django.db.models import Case, QuerySet, Sum, When
from django.db.models.fields import DateField
from django.db.models.functions import Cast, Coalesce
from django.utils import timezone

from .. import models

DEFAULT_DATE: typing.Final[datetime.date] = datetime.date(2023, 1, 1)
PROJECT_REGIONS: typing.Final[list[str]] = [
    "Алматы Картел",
    "Алматы Теле2",
    "Алматы Кселл",
    "Алматинская область Картел",
    "Алматинская область Теле2",
    "Алматинская область Кселл",
    "Талдыкорган Картел",
    "Талдыкорган Теле2",
    "Талдыкорган Кселл",
    "ЭМР АО",
    "СКО Картел",
    "СКО Теле2",
    "Костанай Картел",
    "Костанай Теле2",
    "Павлодар Картел",
    "Павлодар Теле2",
    "Семей Картел",
    "Семей Теле2",
    "Оскемен Картел",
    "Оскемен Теле2",
    "Актобе, Уральск Картел",
    "Актобе, Уральск Теле2",
    "Актау, Атырау Картел",
    "Актау, Атырау Теле2",
    "Астана, Акмолинская область Картел, ЗТЕ",
    "Астана, Акмолинская область Теле2",
    "Караганда, Жезказган Картел",
    "Караганда, Жезказган Теле2",
    "Чимкент, Туркестанская область Картел, ЗТЕ",
    "Чимкент, Туркестанская область Теле2",
    "Тараз, Кызылорда Картел",
    "Тараз, Кызылорда Теле2",
    "Дизеление",
    "ВОЛС Картел",
    "ВОЛС Теле2",
    "СКС/FTTB",
    "ЭМР тендера",
    "Аренда, Выкуп Картел",
    "Аренда, Выкуп Теле2",
    "Аренда, Выкуп Кселл",
    "Проектирование Картел",
    "Проектирование Теле2",
    "Проектирование Кселл",
    "74п ЦМК",
    "Аблай Даутбергенов",
    "Азамат",
    "Дмитрий",
    "Александр",
    "Евгения",
    "Антон Пучкин",
    "Сергей Ратников",
]
MANDATORY_PAYMENTS = [
    "ЗП по АВХ",
    "Налоги по АВХ (зарплатные)",
    "Расходы по АВХ",
    "ЗП 01 ",
    "ЗП 230 п",
    "ЗП проектирование Картел",
    "ЗП проектирование Теле2",
    "ЗП проектирование Кселл",
    "Расходы ALLDATA",
    "ЗП по АС ",
    "Налоги зарплатные",
    "Налог по 7П",
    "Сбор W",
    "НДС 3кв 2023",
    "НДС 4кв 2023",
    "Корректировка по налогам",
    "Пеня по налогам",
    "Кредит Х",
    "Лизинг длиномер/прицеп, авто",
    "Лизинг Хюндай",
    "Газели, алматы (новые)",
    "Длинномер новый",
    "Грузовой автомобиль Hyundai ",
    "Автокран",
    "хюндай газели",
    "Лизинг длинномер/автокран",
    "Лизинг хюндай 305",
    "Погрузчик вилочный",
    "Создание мин. остатков на ЦМК ",
    "Листогибочный пресс ",
    "Ленточный станок",
    "Лизинг Камера",
    "Лизинг спецтехника",
    "Лизинг Хюндай 71П",
    "Кредит на сырье",
    "юридические услуги за кредиты ",
    "Расходы по 01 проекту ",
    "Расход АДМ 2023",
    "Срочный сбор расход АДМ",
    "Расход АДМ",
    "Расходы по 230 проекту ",
    "Представительские 14.06",
    "Представительские",
    "Расход по кассе ",
    "Налог 2023",
]


def get_ordered_project_regions(queryset: QuerySet | None = None) -> QuerySet:
    return (
        (queryset or models.ProjectRegion.objects)
        .filter(name__in=PROJECT_REGIONS)
        .order_by(
            Case(
                *(
                    When(name=name, then=index)
                    for index, name in enumerate(PROJECT_REGIONS)
                )
            )
        )
    )


def get_categories(
    mandatory_payments: QuerySet,
) -> tuple[list[list[models.MandatoryPaymentCategory]], int]:
    categories = []
    categories_colspan = 0
    for mandatory_payment in mandatory_payments:
        category = (
            mandatory_payment.category
            or models.MandatoryPaymentCategory.objects.get_or_create(name="Не указан")[
                0
            ]
        )
        parent_categories = []
        colspan = 0
        while category:
            parent_categories.insert(0, category)
            category = category.parent
            colspan += 1
        categories_colspan = max(categories_colspan, colspan)
        categories.append(parent_categories)
    return categories, categories_colspan


def annotate_mandatory_payments_with_levels(
    mandatory_payments: QuerySet,
    categories: list[list[models.MandatoryPaymentCategory]],
) -> None:
    mandatory_payment_levels = []
    for categories_, mandatory_payment in zip(
        zip(*itertools.zip_longest(*categories)), mandatory_payments
    ):
        levels = []

        for index, category in enumerate(categories_):
            i = True
            for mandatory_payment_level in reversed(mandatory_payment_levels):
                if (
                    index >= len(mandatory_payment_level)
                    or mandatory_payment_level[index] is None
                ):
                    continue
                elif mandatory_payment_level[index][0] == category:
                    mandatory_payment_level[index][2] += 1
                    i = False
                break
            if i:
                if category is None:
                    for level in reversed(levels):
                        if level is None:
                            continue
                        level[1] += 1
                        break
                else:
                    levels.append([category, 1, 1])
                    continue
            levels.append(None)
        if levels:
            mandatory_payment.levels = levels
            mandatory_payment_levels.append(levels)


def get_coefficient_1() -> models.MandatoryPaymentAccrualCalculator:
    (
        mandatory_payment_accrual_calculator,
        created,
    ) = models.MandatoryPaymentAccrualCalculator.objects.get_or_create(
        name="Коэффициент 1", type="coefficient"
    )
    if created:
        models.MandatoryPaymentAccrualCalculatorProjectRegion.objects.bulk_create(
            [
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.60,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматы Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматы Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматы Кселл"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматинская область Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматинская область Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Алматинская область Кселл"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Талдыкорган Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Талдыкорган Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Талдыкорган Кселл"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="ЭМР АО"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.90,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="СКО Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.80,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="СКО Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Костанай Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.81,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Костанай Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Павлодар Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.81,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Павлодар Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Семей Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.81,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Семей Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Оскемен Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.81,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Оскемен Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.90,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Актобе, Уральск Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.90,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Актобе, Уральск Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.90,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Актау, Атырау Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.90,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Актау, Атырау Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Астана, Акмолинская область Картел, ЗТЕ"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Астана, Акмолинская область Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Караганда, Жезказган Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Караганда, Жезказган Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Чимкент, Туркестанская область Картел, ЗТЕ"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Чимкент, Туркестанская область Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Тараз, Кызылорда Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Тараз, Кызылорда Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Дизеление"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.60,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="ВОЛС Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="ВОЛС Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=0.60,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="СКС/FTTB"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.71,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="ЭМР тендера"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=3.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Аренда, Выкуп Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.15,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Аренда, Выкуп Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.06,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Проектирование Картел"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.06,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Проектирование Теле2"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.06,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Проектирование Кселл"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=1.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="74п ЦМК"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=7.10,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Аблай Даутбергенов"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=6.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Азамат"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=6.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Дмитрий"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=6.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Александр"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=3.00,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Евгения"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=12.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Антон Пучкин"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
                models.MandatoryPaymentAccrualCalculatorProjectRegion(
                    coefficient=12.50,
                    project_region=models.ProjectRegion.objects.get_or_create(
                        name="Сергей Ратников"
                    )[0],
                    mandatory_payment_accrual_calculator=mandatory_payment_accrual_calculator,
                ),
            ]
        )
    return mandatory_payment_accrual_calculator


_NAME_COLUMN: typing.Final[int] = 4
_DATA_START_COLUMN: typing.Final[int] = 8
_DATA_START_ROW: typing.Final[int] = 12
_DATA_END_ROW: typing.Final[int] = 56
_DATA_END_COLUMN: typing.Final[int] = 56
_DEADLINE_COLUMN: typing.Final[int] = 58
_WORKSHEET_NAME: typing.Final[str] = "Ежемесячные выплаты "
_MANDATORY_PAYMENT_NAME_TRANSLATIONS = {
    "Пеня по Налогам": "Пеня по налогам",
    "ЗП Проектирование Картел": "ЗП проектирование Картел",
    "ЗП Проектирование Кселл": "ЗП проектирование Кселл",
    "ЗП Проектирование Теле2": "ЗП проектирование Теле2",
}  # Ох уж эти Пени.


def _get_project_region_row(column: int) -> int:
    return 9 if column < 49 else 8


def import_mandatory_payments(
    *,
    bytes_io: io.BytesIO,
    import_mandatory_payment_seizures: bool = True,
    import_mandatory_payment_accruals: bool = True,
):

    workbook = openpyxl.load_workbook(bytes_io, read_only=True, data_only=True)
    mandatory_payments = {
        mandatory_payment.name: mandatory_payment
        for mandatory_payment in models.MandatoryPayment.objects.all()
    }
    project_regions = {
        project_region.name: project_region
        for project_region in models.ProjectRegion.objects.all()
    }

    if import_mandatory_payment_accruals:
        models.MandatoryPaymentAccrualGroup.objects.filter(
            imported_from_file=True
        ).delete()
        models.MandatoryPaymentAccrual.objects.filter(imported_from_file=True).delete()
        worksheet = workbook["Начисление"]
        mandatory_payment_accrual_groups: list[models.MandatoryPaymentAccrualGroup] = []
        mandatory_payment_accruals = []
        for index, (date, _, project_region_name, sum, name) in enumerate(
            worksheet.iter_rows(min_row=2, max_col=5, values_only=True), 2
        ):
            if name is None:
                logging.error(
                    f"Импорт насчиление: Обязательный платёж не указан, строка {index}"
                )
                continue
            name = name.strip()
            if name in _MANDATORY_PAYMENT_NAME_TRANSLATIONS:
                name = _MANDATORY_PAYMENT_NAME_TRANSLATIONS[name]
            if project_region_name is None:
                logging.error(
                    f"Импорт насчиление: Проект регион не указан, строка {index}"
                )
                continue
            if sum is None:
                logging.error(f"Импорт насчиление: Сумма не указана, строка {index}")
                continue
            if date is None:
                logging.error(f"Импорт насчиление: Дата не указана, строка {index}")
                continue
            project_region_name = project_region_name.strip()
            if (project_region := project_regions.get(project_region_name)) is None:
                logging.error(
                    f'Импорт насчиление: Проект регион "{project_region_name}" не найден, строка {index}'
                )
                continue
            if (mandatory_payment := mandatory_payments.get(name)) is None:
                logging.error(
                    f'Импорт насчиление: Обязательный платёж "{name}" не найден, строка {index}'
                )
                continue
            sum = decimal.Decimal(sum)
            if (
                not mandatory_payment_accrual_groups
                or mandatory_payment_accrual_groups[-1].datetime != date
                or mandatory_payment_accrual_groups[-1].mandatory_payment
                != mandatory_payment
            ):
                mandatory_payment_accrual_groups.append(
                    models.MandatoryPaymentAccrualGroup(
                        mandatory_payment=mandatory_payment,
                        datetime=date,
                        was_accrued=sum,
                        imported_from_file=True,
                    )
                )
            else:
                mandatory_payment_accrual_groups[-1].was_accrued += sum
            mandatory_payment_accruals.append(
                models.MandatoryPaymentAccrual(
                    mandatory_payment=mandatory_payment,
                    sum=sum,
                    project_region=project_region,
                    datetime=date,
                    imported_from_file=True,
                    deadline=mandatory_payment.get_deadline(date),
                    group=mandatory_payment_accrual_groups[-1],
                )
            )
        models.MandatoryPaymentAccrualGroup.objects.bulk_create(
            mandatory_payment_accrual_groups
        )
        models.MandatoryPaymentAccrual.objects.bulk_create(mandatory_payment_accruals)
    if import_mandatory_payment_seizures:
        models.MandatoryPaymentSeizure.objects.filter(imported_from_file=True).delete()
        worksheet = workbook["Изьятие (новый)"]
        mandatory_payment_seizures = []
        for index, (date, _, project_region_name, _, _, sum, name) in enumerate(
            worksheet.iter_rows(min_row=3, max_col=7, values_only=True), 3
        ):
            if date is None:
                logging.error(f"Импорт изъятие: Дата не указано, строка {index}")
                continue
            if isinstance(project_region_name, str):
                project_region_name = project_region_name.strip()

            if (
                project_region_name := project_regions.get(project_region_name)
            ) is None:
                logging.error(
                    f'Импорт изъятие: Проект регион "{project_region_name}" не найдено, дата {date}, строка {index}'
                )
                continue
            if isinstance(name, str):
                name = name.strip()
            if name in _MANDATORY_PAYMENT_NAME_TRANSLATIONS:
                name = _MANDATORY_PAYMENT_NAME_TRANSLATIONS[name]
            if (mandatory_payment := mandatory_payments.get(name)) is None:
                logging.error(
                    f'Импорт изъятие: Обязательный платёж "{name}" не найдено, дата {date}, строка {index}'
                )
                continue
            mandatory_payment_seizures.append(
                models.MandatoryPaymentSeizure(
                    mandatory_payment=mandatory_payment,
                    project_region=project_region_name,
                    sum=sum,
                    datetime=date,
                    imported_from_file=True,
                    status=models.completed_mandatory_payment_seizure_status(),
                )
            )
        models.MandatoryPaymentSeizure.objects.bulk_create(mandatory_payment_seizures)


PaymentType: typing.TypeAlias = typing.Literal["everything-else", "soon", "overdue"]
@dataclasses.dataclass(slots=True, frozen=True, kw_only=True)
class Payment:
    mandatory_payment: models.MandatoryPayment
    project_region: models.ProjectRegion
    paid_today: bool
    date: datetime.date
    exclude_statuses: list[models.MandatoryPaymentSeizureStatus]
    sum: decimal.Decimal
    min: decimal.Decimal
    deadline: datetime.date | None
    days_left: int | None
    type: PaymentType


def get_payment(
    *,
    mandatory_payment: models.MandatoryPayment | str,
    project_region: models.ProjectRegion | str,
    exclude_statuses: list[models.MandatoryPaymentSeizureStatus] | None = None,
    date: datetime.date = DEFAULT_DATE,
) -> Payment | None:
    if isinstance(mandatory_payment, str):
        try:
            mandatory_payment = models.MandatoryPayment.objects.get(
                name=mandatory_payment
            )
        except models.MandatoryPayment.DoesNotExist:
            return None
    if isinstance(project_region, str):
        try:
            project_region = models.ProjectRegion.objects.get(name=project_region)
        except models.ProjectRegion.DoesNotExist:
            return None
    if exclude_statuses is None:
        exclude_statuses = [models.rejected_mandatory_payment_seizure_status()]
    today = timezone.localdate()
    accruals = (
        models.MandatoryPaymentAccrual.objects.order_by("-datetime")
        .filter(mandatory_payment=mandatory_payment, project_region=project_region)
        .annotate(date=Cast("datetime", DateField()))
    )
    if date:
        accruals = accruals.filter(date__gte=date)
    seizures = (
        models.MandatoryPaymentSeizure.objects.order_by("-datetime")
        .exclude(status__in=exclude_statuses)
        .filter(mandatory_payment=mandatory_payment, project_region=project_region)
        .annotate(date=Cast("datetime", DateField()))
    )
    if date:
        seizures = seizures.filter(date__gte=date)

    paid_today = seizures.filter(date=today).exists()
    seizures_sum = seizures.aggregate(
        seizures_sum=Coalesce(Sum("sum"), decimal.Decimal())
    )["seizures_sum"]
    accruals_sum = accruals.aggregate(
        accruals_sum=Coalesce(Sum("sum"), decimal.Decimal())
    )["accruals_sum"]
    sum_ = accruals_sum - seizures_sum
    type_: PaymentType = "everything-else"
    deadline: datetime.date = None
    unpaid_accruals: list[models.MandatoryPaymentAccrual] = []
    min_ = decimal.Decimal()
    days_left: int | None = None
    rounded_sum = round(sum_)
    if sum_ > 0:
        unpaid_accruals_sum = decimal.Decimal()
        for accrual in accruals:
            if rounded_sum <= round(unpaid_accruals_sum):
                break
            unpaid_accruals_sum += accrual.sum
            unpaid_accruals.append(accrual)
        if unpaid_accruals:
            unpaid_accrual = min(unpaid_accruals, key=lambda unpaid_accrual: unpaid_accrual.deadline)
            deadline = unpaid_accrual.deadline
            days_left = (unpaid_accrual.deadline - today).days
            if days_left < 0:
                type_ = "overdue"
                min_ = sum_
            else:
                type_ = "soon"
                min_ = sum_ / (days_left + 1)
    return Payment(
        sum=rounded_sum,
        min=round(min_),
        paid_today=paid_today,
        mandatory_payment=mandatory_payment,
        project_region=project_region,
        date=date,
        days_left=days_left,
        type=type_,
        deadline=deadline,
        exclude_statuses=exclude_statuses,
    )
