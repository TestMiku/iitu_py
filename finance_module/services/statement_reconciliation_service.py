import collections
import collections.abc
import dataclasses
import datetime
import decimal
import mimetypes
import re
import typing
from collections import defaultdict
from collections.abc import Iterable, Iterator

import openpyxl
import openpyxl.worksheet.worksheet
from bs4 import BeautifulSoup
from django.core.files.uploadedfile import UploadedFile
from django.db import models
from django.db.models.functions import TruncDate
from simple_history.models import HistoricalRecords

from ..models import PaidInvoice as PaidInvoiceModel


class StatementReconciliationVariables(models.Model):
    correspondents_map = models.TextField()
    accounts_map = models.TextField()
    exclude_from_exceptions = models.TextField()
    history = HistoricalRecords()

    @staticmethod
    def _to_dict(x: str, /) -> dict[str, str]:
        dict_ = {}
        for i in x.splitlines():
            if not i:
                continue
            pair = i.split("=", 1)
            if len(pair) == 1:
                raise StatementReconciliationError(
                    f'Вы должный указать контрагента для "{i}" в сопоставление'
                )
            dict_[normalize_spaces(pair[0])] = normalize_spaces(pair[1])
        return dict_

    @property
    def correspondents_map_as_dict(self) -> dict[str, str]:
        return self._to_dict(self.correspondents_map)

    @correspondents_map_as_dict.setter
    def correspondents_map_as_dict(self, value: dict[str, str]) -> None:
        self.correspondents_map = "\n".join(
            f"{key}={value}" for key, value in value.items()
        )

    @property
    def exclude_from_exceptions_as_list(self) -> list[str]:
        return list(map(normalize_spaces, self.exclude_from_exceptions.splitlines()))

    @property
    def accounts_map_as_dict(self) -> dict[str, str]:
        return self._to_dict(self.accounts_map)

    @accounts_map_as_dict.setter
    def accounts_map_as_dict(self, value: dict[str, str]) -> None:
        pass


@dataclasses.dataclass(slots=True)
class Document:
    statement: "Statement"
    id: str
    datetime: datetime.datetime
    correspondent: str
    correspondent_bic: str
    correspondent_iic: str
    debit: decimal.Decimal
    credit: decimal.Decimal
    ppc: str
    correspondent_bank: str
    purpose_of_payment: str
    sender_id: str
    recipient_bin: str
    matched: bool = False

    commission: bool = False  # Является ли документ комиссией.
    documentCommission: typing.Union["Document", None] = (
        None  # Документ привязанный к комиссий.
    )
    commissionDocument: typing.Union["Document", None] = (
        None  # Привязанный к документу комиссия.
    )

    related_paid_invoices: list["PaidInvoice"] | None = None

    @property
    def contractor(self) -> str:
        variables = get_variables()
        return variables.correspondents_map_as_dict.get(
            self.correspondent, self.correspondent
        )

    @property
    def excluded(self) -> bool:
        variables = get_variables()
        return self.contractor in variables.exclude_from_exceptions_as_list

    @property
    def refund(self) -> bool:
        return "возврат" in self.purpose_of_payment.lower()

    def tojson(
        self,
        *,
        include_related_paid_invoices: bool = False,
        include_commission_document: bool = True,
        include_document_commission: bool = True,
    ) -> dict[str, typing.Any]:
        json_ = {
            "id": self.id,
            "datetime": self.datetime.isoformat(),
            "correspondent": self.correspondent,
            "correspondentBic": self.correspondent_bic,
            "correspondentIic": self.correspondent_iic,
            "debit": float(self.debit),
            "credit": float(self.credit),
            "ppc": self.ppc,
            "correspondentBank": self.correspondent_bank,
            "purposeOfPayment": self.purpose_of_payment,
            "senderId": self.sender_id,
            "recipientBin": self.recipient_bin,
            "matched": self.matched,
            "excluded": self.excluded,
            "commission": self.commission,
            "commissionDocument": None,
            "documentCommission": None,
            "refund": self.refund,
            "uploadedFile": self.statement.uploaded_file.name,
        }
        if self.related_paid_invoices and include_related_paid_invoices:
            json_["relatedPaidInvoices"] = [
                paid_invoice.tojson() for paid_invoice in self.related_paid_invoices
            ]
        if include_document_commission and self.documentCommission:
            json_["documentCommission"] = self.documentCommission.tojson(
                include_commission_document=False
            )
        if include_commission_document and self.commissionDocument:
            json_["commissionDocument"] = self.commissionDocument.tojson(
                include_document_commission=False
            )
        return json_


def get_variables() -> StatementReconciliationVariables:
    return StatementReconciliationVariables.objects.get_or_create(id=1)[0]


def parse_number(x: str | int | float, /) -> decimal.Decimal:
    return decimal.Decimal(
        "".join(x.replace(",", ".").split()) or 0 if isinstance(x, str) else x
    )


def normalize_spaces(x: str, /) -> str:
    return " ".join(x.split())


@dataclasses.dataclass(slots=True)
class Statement:
    uploaded_file: UploadedFile
    account: str
    documents: list[Document] = dataclasses.field(default_factory=list)

    @property
    def correspondent_documents(self) -> defaultdict[str, list[Document]]:
        correspondent_documents = defaultdict(list)
        for document in self.documents:
            correspondent_documents[document.contractor].append(document)
        return correspondent_documents

    @staticmethod
    def get_account_from_uploaded_file_name(uploaded_file: UploadedFile) -> str:
        match = re.match(
            r"(.+?)(?:\s+\(\d+\))?\..+", normalize_spaces(uploaded_file.name)
        )
        if not match:
            raise StatementReconciliationError(
                f"Не найдено в имени файла имя расчётного счёта: {uploaded_file.name}"
            )
        variables = get_variables()
        account = variables.accounts_map_as_dict.get(match[1])
        if not account:
            raise StatementReconciliationError(
                f'Добавьте имя расчётного счёта для имени файла "{match[1]}=ИМЯ Р/С" в сопоставление'
            )
        return account

    @classmethod
    def from_html(cls, uploaded_file: UploadedFile) -> "Statement":
        try:
            bs = BeautifulSoup(uploaded_file, features="html.parser")

            self = cls(
                uploaded_file, cls.get_account_from_uploaded_file_name(uploaded_file)
            )
            for tr in bs.select(
                'tr[class]:not([class="R9"]):not([class="R0"]):not(:has( td[colspan="14"]))'
            ):
                tds = tr.select("td")
                id = tds[1].text.strip()
                datetime_ = datetime.datetime.strptime(tds[2].text.strip(), "%d-%m-%Y")
                correspondent_bic = tds[3].text.strip()
                correspondent_iic = tds[4].text.strip()
                sender_id = tds[5].text.strip()
                recipient_bin = tds[6].text.strip()
                ppc = tds[10].text.strip()
                correspondent = normalize_spaces(tds[7].text)
                correspondent_bank = tds[11].text.strip()
                debit = parse_number(tds[8].text)
                credit = parse_number(tds[9].text)
                purpose_of_payment = tds[12].text.strip()
                document = Document(
                    self,
                    id,
                    datetime_,
                    correspondent,
                    correspondent_bic,
                    correspondent_iic,
                    debit,
                    credit,
                    ppc,
                    correspondent_bank,
                    purpose_of_payment,
                    sender_id,
                    recipient_bin,
                )
                self.documents.append(document)
            return self
        except Exception as exception:
            raise StatementReconciliationError(
                f"Неизвестая ошибка при обработке HTML файла {uploaded_file.name}, возможно это не файл выписки, ошибка: {exception}"
            )

    @classmethod
    def from_xlsx(cls, uploaded_file: UploadedFile) -> "Statement":
        try:
            workbook = openpyxl.load_workbook(
                uploaded_file, read_only=True, data_only=True
            )
            worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
            self = cls(
                uploaded_file, cls.get_account_from_uploaded_file_name(uploaded_file)
            )
            if worksheet.cell(16, 1).value == "№":
                for row in worksheet.iter_rows(
                    min_row=17, max_row=10000000, min_col=1, max_col=9, values_only=True
                ):
                    if not row[0]:
                        break

                    datetime_ = datetime.datetime.strptime(row[1], "%d.%m.%Y %H:%M:%S")
                    id = row[2]
                    debit = decimal.Decimal(row[5])
                    credit = decimal.Decimal(row[6])
                    purpose_of_payment = row[7]
                    column4 = row[4]
                    bin_index = column4.index("БИН")
                    correspondent = normalize_spaces(column4[:bin_index])
                    correspondent_bic = ""
                    correspondent_iic = ""
                    correspondent_bank = ""
                    ppc = ""
                    sender_id = ""
                    recipient_bin = ""
                    document = Document(
                        self,
                        id,
                        datetime_,
                        correspondent,
                        correspondent_bic,
                        correspondent_iic,
                        debit,
                        credit,
                        ppc,
                        correspondent_bank,
                        purpose_of_payment,
                        sender_id,
                        recipient_bin,
                    )
                    self.documents.append(document)
            else:
                for row in worksheet.iter_rows(
                    min_row=11,
                    max_row=10000000,
                    min_col=2,
                    max_col=14,
                    values_only=True,
                ):
                    if not row[12]:
                        break

                    id = row[0]
                    datetime_ = datetime.datetime.strptime(row[1], "%d-%m-%Y")
                    debit = round(decimal.Decimal(row[8] or 0))
                    credit = round(decimal.Decimal(row[9] or 0))
                    purpose_of_payment = row[12]
                    correspondent = normalize_spaces(row[7])
                    correspondent_bic = row[2]
                    correspondent_iic = row[3]
                    correspondent_bank = normalize_spaces(row[11])
                    ppc = row[10]
                    sender_id = row[5]
                    recipient_bin = row[6]
                    document = Document(
                        self,
                        id,
                        datetime_,
                        correspondent,
                        correspondent_bic,
                        correspondent_iic,
                        debit,
                        credit,
                        ppc,
                        correspondent_bank,
                        purpose_of_payment,
                        sender_id,
                        recipient_bin,
                    )
                    self.documents.append(document)
            return self
        except Exception as exception:
            raise StatementReconciliationError(
                f"Неизвестая ошибка при обработке XLSX файла {uploaded_file.name}, возможно это не файл выписки, ошибка: {exception}"
            )

    @classmethod
    def from_txt(cls, uploaded_file: UploadedFile) -> "Statement":
        pass


def get_statements(uploaded_files: Iterable[UploadedFile]) -> Iterator[Statement]:
    for uploaded_file in uploaded_files:
        extension = mimetypes.guess_extension(uploaded_file.content_type)
        if extension == ".xlsx":
            statement = Statement.from_xlsx(uploaded_file)
        else:
            continue
        yield statement


@dataclasses.dataclass(slots=True, kw_only=True)
class PaidInvoice:
    id: int | str = None
    number: int
    contractor: str
    paid: decimal.Decimal
    account: str
    commissionDocument: Document | None = None
    related_documents: list[Document] | None = None
    matched: bool = False

    def tojson(
        self, *, include_related_documents: bool = False
    ) -> dict[str, typing.Any]:
        json_ = {
            "id": self.id,
            "number": self.number,
            "contractor": self.contractor,
            "paid": float(self.paid),
            "account": self.account,
            "commissionDocument": (
                self.commissionDocument.tojson() if self.commissionDocument else None
            ),
            "matched": self.matched,
        }
        if self.related_documents and include_related_documents:
            json_["relatedDocuments"] = [
                document.tojson() for document in self.related_documents
            ]
        return json_


class StatementReconciliationError(Exception):
    pass


@dataclasses.dataclass(slots=True)
class StatementReconciliationResult:
    paid_invoices: collections.abc.Iterable[PaidInvoice]
    statements: collections.abc.Iterable[Statement]

    def tojson(self) -> dict[str, typing.Any]:
        return {
            "statements": [
                {
                    "uploadedFile": statement.uploaded_file.name,
                    "account": statement.account,
                    "correspondentDocuments": {
                        correspondent: [
                            document.tojson(include_related_paid_invoices=True)
                            for document in documents
                        ]
                        for correspondent, documents in statement.correspondent_documents.items()
                    },
                }
                for statement in self.statements
            ],
            "paidInvoices": [
                paid_invoice.tojson(include_related_documents=True)
                for paid_invoice in self.paid_invoices
            ],
            "accountCorrespondentPaidInvoices": {
                account: {
                    correspondent: [
                        paid_invoice.tojson(include_related_documents=True)
                        for paid_invoice in paid_invoices
                    ]
                    for correspondent, paid_invoices in correspondent_paid_invoices.items()
                }
                for account, correspondent_paid_invoices in get_account_correspondent_paid_invoices(self.paid_invoices).items()
            }
        }


def get_account_correspondent_paid_invoices(
    paid_invoices: collections.abc.Iterable[PaidInvoice],
) -> collections.defaultdict[str, collections.defaultdict[str, list[PaidInvoice]]]:
    account_correspondent_paid_invoices: collections.defaultdict[
        str, collections.defaultdict[str, list[PaidInvoice]]
    ] = collections.defaultdict(lambda: collections.defaultdict(list))
    for paid_invoice in paid_invoices:
        account_correspondent_paid_invoices[paid_invoice.account][paid_invoice.contractor].append(paid_invoice)
    return account_correspondent_paid_invoices


def statement_reconciliation(
    *,
    uploaded_files: Iterable[UploadedFile],
    date: datetime.date,
    paid_invoices_file: UploadedFile | None = None,
    worksheet_name: str | None = None,
    start_row: int | None = None,
) -> StatementReconciliationResult:
    """Как то дерьмово выглядит, надо изменить"""
    try:
        paid_invoices: list[PaidInvoice] = []
        if paid_invoices_file:
            workbook = openpyxl.open(paid_invoices_file, read_only=True, data_only=True)
            if worksheet_name:
                try:
                    worksheet = workbook[worksheet_name]
                except KeyError:
                    raise StatementReconciliationError(
                        f"Таблица с именем {worksheet_name} не существует в XLSX файле {paid_invoices_file.name}"
                    )
            else:
                worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
            start_row = start_row or 1
            headers = {"до", "контрагенты", "оплачено", "расчетный счет", "дата оплаты"}
            header_indeces = {}
            for row in worksheet.iter_rows(
                min_row=start_row, max_row=start_row, values_only=True
            ):
                for index, header in enumerate(row):
                    if header:
                        header = normalize_spaces(str(header)).lower()
                        if header in headers:
                            header_indeces[header] = index
            if headers != set(header_indeces):
                raise StatementReconciliationError(
                    f"Не найдено столбцов: {', '.join(headers ^ set(header_indeces))}"
                )
            for index, row in enumerate(worksheet.iter_rows(min_row=start_row + 1, values_only=True), start_row + 1):
                row_date = row[header_indeces["дата оплаты"]]
                if isinstance(row_date, datetime.datetime):
                    row_date = row_date.date()
                if row_date == date:
                    account = row[header_indeces["расчетный счет"]]
                    if not account:
                        continue
                    account = normalize_spaces(account)
                    contractor = normalize_spaces(row[header_indeces["контрагенты"]])
                    number = row[header_indeces["до"]]

                    paid_invoice = PaidInvoice(
                        id=f"file-{index}",
                        number=(
                            number
                            if isinstance(number, int)
                            else int("".join(str(number).split()))
                        ),
                        contractor=contractor,
                        paid=round(parse_number(row[header_indeces["оплачено"]])),
                        account=account,
                    )
                    paid_invoices.append(paid_invoice)
        else:
            for paid_invoice in PaidInvoiceModel.objects.annotate(
                at_date=TruncDate("at")
            ).filter(at_date=date):
                if paid_invoice.contractor is None:
                    continue
                paid_invoice_ = PaidInvoice(
                    id=paid_invoice.id,
                    number=paid_invoice.number,
                    contractor=" ".join(paid_invoice.contractor.split()),
                    paid=paid_invoice.sum,
                    account=paid_invoice.account.name,
                )
                paid_invoices.append(paid_invoice_)
        account_correspondent_paid_invoices = get_account_correspondent_paid_invoices(paid_invoices)
        statements = list(get_statements(uploaded_files))
        for statement in statements:
            for correspondent, documents in statement.correspondent_documents.items():
                correspondent_paid_invoices = account_correspondent_paid_invoices[
                    statement.account
                ][correspondent]
                if round(sum(document.debit for document in documents)) == round(
                    sum(
                        paid_invoice.paid
                        for paid_invoice in correspondent_paid_invoices
                    )
                ):
                    for document in documents:
                        document.related_paid_invoices = correspondent_paid_invoices
                        document.matched = True
                    for paid_invoice in correspondent_paid_invoices:
                        paid_invoice.related_documents = documents
                        paid_invoice.matched = True

            for document in statement.documents:
                if match := re.search(
                    r"(?:WWW\.\s+(?:Интернет|Интрнет)(?:-|\s+)банкинг\.|Комиссия\s+за\s+операцию)(?:.+?)Расчетный\s+документ\s+№\s+(\d+)",
                    document.purpose_of_payment,
                    re.IGNORECASE,
                ):
                    document.commission = True
                    id = match[1]
                    finded_document = next(
                        (
                            document
                            for document in statement.documents
                            if document.id == id
                        ),
                        None,
                    )

                    if finded_document:
                        finded_document.commissionDocument = document
                        document.documentCommission = finded_document
                        if finded_document.related_paid_invoices:
                            paid_invoice = next(
                                (
                                    paid_invoice
                                    for paid_invoice in finded_document.related_paid_invoices
                                    if paid_invoice.commissionDocument is None
                                ),
                                None,
                            )
                            if paid_invoice is not None:
                                document.matched = True
                                paid_invoice.commissionDocument = document
        return StatementReconciliationResult(
            statements=statements,
            paid_invoices=paid_invoices,
        )
    except StatementReconciliationError as statement_reconciliation_error:
        raise statement_reconciliation_error
    except Exception as exception:
        raise StatementReconciliationError(f"Неизвестая ошибка: {exception}")
