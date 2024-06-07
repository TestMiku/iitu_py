import collections.abc
import datetime
import decimal
import functools
import io
import logging
import operator
import os
import typing
from tempfile import TemporaryFile

import openpyxl
import openpyxl.formula
from django.contrib.auth import get_user_model
from django.contrib.auth.models import AbstractBaseUser
from django.db import connection, models
from django.db.models import QuerySet, Sum, Q
from django.db.models.functions import Coalesce
from django.db.models.signals import pre_migrate
from django.utils import timezone
from django.dispatch import receiver
from openpyxl.formula.tokenizer import Token

from finance_module.services.common_service import (
    get_account_balance,
    get_project_regions,
)

from ..models import (
    REJECTED_DEBT_TRANSLATE_GROUP_STATUS_NAME,
    Account,
    Debt,
    DebtTranslateGroup,
    ProjectRegion,
    RenewableDebt,
    Setoff,
    Subdivision,
    UnpaidInvoice,
    rejected_debt_translate_group_status,
)
from .mandatory_payments_service import get_coefficient_1
from .unload_debts_service import (
    get_debts_for_unloading,
    set_data_to_worksheet,
    set_worksheet_columns,
)

User = get_user_model()


class TemporaryUnpaidInvoice(models.Model):
    approver = models.CharField(max_length=512)
    balance_due_1c = models.DecimalField(max_digits=19, decimal_places=2)
    type = models.CharField(max_length=512)
    contractor = models.CharField(max_length=512)
    project = models.CharField(max_length=512)
    subdivision = models.CharField(max_length=512)
    account_category = models.CharField(max_length=512)


def get_renewable_debt_sum(renewable_debt: RenewableDebt) -> decimal.Decimal:
    def raise_runtime_error(message: str) -> typing.NoReturn:
        raise RuntimeError(
            f'Error when calculating debt from "{renewable_debt.from_whom}" to "{renewable_debt.to_whom}" with the sql - "{renewable_debt.sql}", {message}'
        )

    try:
        with connection.cursor() as cursor:
            cursor.execute(renewable_debt.sql)
            sum = cursor.fetchone()
    except Exception as exception:
        raise_runtime_error(f"error - {exception}")
    if sum is None or len(sum) > 1 or sum[0] is None:
        raise_runtime_error(f"the sql should return the amount, sql result is {sum}")
    try:
        sum = decimal.Decimal(sum[0])
    except decimal.ConversionSyntax:
        raise_runtime_error(
            f"the amount must be convertible to decimal, sql result is {sum}"
        )
    return sum


CREDITS_FILE_PATH: str | None = os.getenv("CREDITS_FILE_PATH")


def update_debts(
    *, bytes_io: io.BytesIO | None = None, credits_sum: decimal.Decimal | None = None
) -> None:
    TemporaryUnpaidInvoice.objects.all().delete()

    unpaid_invoices = []
    if bytes_io:
        workbook = openpyxl.load_workbook(bytes_io, read_only=True, data_only=True)
        worksheet = workbook["Реестр"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            unpaid_invoices.append(
                TemporaryUnpaidInvoice(
                    approver=row[0],
                    balance_due_1c=decimal.Decimal(row[1]),
                    type=row[2],
                    contractor=row[3],
                    project=row[4],
                    subdivision=row[5],
                    account_category=row[6],
                )
            )
        if not credits_sum:
            worksheet = workbook["Кредиты"]
            credits_sum = decimal.Decimal(worksheet["L2"].value)
    else:
        for unpaid_invoice in UnpaidInvoice.objects.all():
            unpaid_invoices.append(
                TemporaryUnpaidInvoice(
                    approver=unpaid_invoice.approver,
                    balance_due_1c=unpaid_invoice.remainder,
                    type=unpaid_invoice.payment_type,
                    contractor=unpaid_invoice.contractor,
                    project=unpaid_invoice.project,
                    subdivision=unpaid_invoice.department,
                    account_category=unpaid_invoice.invoice_category,
                )
            )

    TemporaryUnpaidInvoice.objects.bulk_create(unpaid_invoices)
    for renewable_debt in RenewableDebt.objects.all():
        Debt.objects.update_or_create(
            from_whom=renewable_debt.from_whom,
            to_whom=renewable_debt.to_whom,
            imported_from_file=False,
            defaults={
                "sum": get_renewable_debt_sum(renewable_debt),
                "renewed": True,
            },
        )

    if not credits_sum:
        workbook = openpyxl.load_workbook(
            CREDITS_FILE_PATH, data_only=True, read_only=True
        )
        worksheet = workbook["Кредиты"]
        credits_sum = decimal.Decimal(worksheet["AJ84"].value)

    mandatory_payment_accrual_calculator = get_coefficient_1()
    for (
        mandatory_payment_accrual_calculator_project_region
    ) in (
        mandatory_payment_accrual_calculator.mandatory_payment_accrual_calculator_project_regions.all()
    ):
        Debt.objects.update_or_create(
            defaults={
                "sum": credits_sum
                * mandatory_payment_accrual_calculator_project_region.coefficient
                / 100
            },
            from_whom="Кредиты и Лизинги",
            to_whom=mandatory_payment_accrual_calculator_project_region.project_region.name,
        )


def unload_debts(queryset: QuerySet | None = None) -> TemporaryFile:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    set_worksheet_columns(worksheet)
    set_data_to_worksheet(worksheet, queryset or get_debts_for_unloading())

    temporary_file = TemporaryFile()
    workbook.save(temporary_file)
    temporary_file.seek(0)
    return temporary_file


def import_debts(*, bytes_io: io.BytesIO) -> None:
    Debt.objects.all().delete()
    workbook = openpyxl.load_workbook(bytes_io, read_only=True, data_only=True)
    worksheet = workbook["ДМП подробно (новый)"]
    debts = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        from_whom, to_whom = row[8], row[1]

        sum1, sum2 = row[4], row[11]
        sum = None
        if sum1:
            sum = sum1
        if sum2:
            sum = -sum2
        if not sum:
            continue
        if sum < 0:
            sum = -sum
            from_whom, to_whom = to_whom, from_whom
        if not from_whom or not to_whom:
            logging.info(f"{from_whom=}, {to_whom=}, {sum=}")
            continue

        date = row[14]
        note = row[15]
        if not isinstance(date, datetime.datetime):
            date = None
        debts.append(
            Debt(
                from_whom=from_whom,
                to_whom=to_whom,
                sum=sum,
                datetime=date and timezone.make_aware(date),
                note=note,
                imported_from_file=True,
            )
        )
    Debt.objects.bulk_create(debts)


unpaid_invoices_table_name: str = "finance_module_temporaryunpaidinvoice"
fields_map: collections.abc.Mapping[str, str] = {
    "Реестр!C5": "project",
    "Реестр!C3": "type",
    "Реестр!C1": "approver",
    "Реестр!C4": "contractor",
    "Реестр!$D:$D": "contractor",
    "Реестр!$A:$A": "approver",
    "Реестр!$E:$E": "project",
    "Реестр!$C:$C": "type",
    "Реестр!$F:$F": "subdivision",
    "'ДМП подробно (новый)'!$I:$I": "to_whom",
    "'ДМП подробно (новый)'!$B:$B": "from_whom",
}


def import_renewable_debts(*, bytes_io: io.BytesIO) -> None:
    """Импорт формул долгов, обычно создаётся SQL запрос к базе данных, оплаты тянется из таблиц `paid_invoices_table_name`"""
    RenewableDebt.objects.all().delete()
    workbook = openpyxl.load_workbook(bytes_io, read_only=True)
    interdivisional_debts_worksheet = workbook["Долги между подразделениями (К)"]
    technical_list_worksheet = workbook["Техлист"]
    renewable_debts = []
    for a, b in interdivisional_debts_worksheet.iter_rows(
        min_col=1, max_col=2, min_row=6, max_row=59
    ):
        project_region1 = b.value or a.value
        for c, d in zip(
            next(
                interdivisional_debts_worksheet.iter_rows(
                    min_col=6, max_col=75, min_row=4, max_row=4
                )
            ),
            next(
                interdivisional_debts_worksheet.iter_rows(
                    min_col=6, max_col=75, min_row=5, max_row=5
                )
            ),
        ):
            project_region2 = d.value or c.value
            if project_region2 not in [
                "74п ЦМК",
                "Счета- Работы: СМР, ЭМР, Обследование",
                "Счета - прочие: сырье, аксессуары для ПК,  дизеление, билеты, проживание, тендерные услуги, аренда офиса,юр услуги, расходы на офис.",
                "Счета - ТМЦ",
            ]:
                continue
            cell = interdivisional_debts_worksheet.cell(a.row, c.column)
            value = cell.value

            if value:
                value = openpyxl.formula.Tokenizer(value)
                sql = ""
                result = collections.defaultdict(set)
                field = None
                table_name = None
                sum_field = None
                for item in value.items:
                    if item.type == Token.FUNC:
                        if item.subtype == Token.CLOSE:
                            where = " AND ".join(
                                f"{field} {operator} ({','.join(value)})"
                                for (operator, field), value in result.items()
                            )
                            sql += f"COALESCE((SELECT SUM({sum_field}) FROM {table_name} WHERE {where}), 0)"
                            result.clear()
                    elif item.type == Token.OPERAND:
                        if item.subtype == Token.RANGE:
                            if item.value == "Реестр!$B:$B":
                                table_name = unpaid_invoices_table_name
                                sum_field = "balance_due_1c"
                            elif item.value == "'ДМП подробно (новый)'!$L:$L":
                                table_name = "finance_module_debt"
                                sum_field = "sum"
                            elif item.value.startswith("Техлист!"):
                                result[("IN", field)].add(
                                    f"'{technical_list_worksheet[item.value.split('!')[1]].value}'"
                                )
                            elif item.value in fields_map:
                                field = fields_map[item.value]
                            else:
                                cell = interdivisional_debts_worksheet[item.value]
                                result[("IN", field)].add(f"'{cell.value}'")
                        elif item.subtype == Token.TEXT:
                            value = item.value[1:-1]
                            type = "IN"
                            if value.startswith("<>"):
                                type = "NOT IN"
                                value = value.removeprefix("<>")
                            result[(type, field)].add(f"'{value}'")
                        elif item.subtype == Token.NUMBER:
                            sql += item.value
                    elif item.type != Token.SEP and item.subtype != Token.ARG:
                        sql += item.value
                renewable_debts.append(
                    RenewableDebt(
                        from_whom=project_region2,
                        to_whom=project_region1,
                        sql=f"SELECT {sql}",
                    )
                )
    RenewableDebt.objects.bulk_create(renewable_debts)


_CREATE_NOT_REJECTED_DEBTS_VIEW_SQL: typing.Final[
    str
] = """--sql
CREATE OR REPLACE VIEW not_rejected_debts AS 
    SELECT t1.*
      FROM finance_module_debt AS t1
           LEFT JOIN finance_module_debttranslategroup AS t2 
           ON t2.id = t1.group_id 
           LEFT JOIN finance_module_debttranslategroupstatus AS t3 
           ON t3.id = t2.status_id
     WHERE t3.name IS NULL OR t3.name != 'Отклонено'
"""
_CREATE_PARTICIPIANTS_VIEW_SQL: typing.Final[
    str
] = """--sql
CREATE OR REPLACE VIEW participants (participant) AS
    WITH participants_cte (name) AS (
        SELECT from_whom 
          FROM not_rejected_debts
         UNION ALL
        SELECT to_whom 
          FROM not_rejected_debts
    )
    SELECT DISTINCT name 
      FROM participants_cte 
"""


def _create_views(cursor) -> None:
    _create_not_rejected_debts_view(cursor)
    _create_participiants_view(cursor)


def _create_not_rejected_debts_view(cursor) -> None:
    cursor.execute(_CREATE_NOT_REJECTED_DEBTS_VIEW_SQL)


def _create_participiants_view(cursor) -> None:
    cursor.execute(
        _CREATE_PARTICIPIANTS_VIEW_SQL,
    )


@receiver(pre_migrate)
def _(*args, **kwargs) -> None:
    with connection.cursor() as cursor:
        cursor.execute("DROP VIEW IF EXISTS participants")
        cursor.execute("DROP VIEW IF EXISTS not_rejected_debts")


def get_participants() -> set[str]:
    debts = Debt.objects.exclude(group__status=rejected_debt_translate_group_status())
    return set(debts.values_list("from_whom", flat=True)) | set(debts.values_list("to_whom", flat=True))

_GET_DEBT_SUM_SQL: typing.Final[
    str
] = """--sql
WITH debts AS (
    SELECT *
      FROM not_rejected_debts
     WHERE '74п ЦМК' NOT IN (from_whom, to_whom) 
        OR NOT imported_from_file
)
SELECT COALESCE((SELECT SUM(sum) FROM debts WHERE from_whom = %(from_whom)s AND to_whom = %(to_whom)s), 0) - COALESCE((SELECT SUM(sum) FROM debts WHERE from_whom=%(to_whom)s AND to_whom=%(from_whom)s), 0)
"""  # Тут вместо 74п ЦМК должен был быть По реестру, но его нету теперь в ТЕЛЕКОМе. Пришлось написать костыль.


def get_debt_sum(
    *, from_whom: str | ProjectRegion, to_whom: str | ProjectRegion
) -> decimal.Decimal:
    if isinstance(from_whom, ProjectRegion):
        from_whom = from_whom.name
    if isinstance(to_whom, ProjectRegion):
        to_whom = to_whom.name
    debts = Debt.objects.exclude(
        Q(group__status=rejected_debt_translate_group_status())
        | Q(from_whom="74п ЦМК")
        | Q(to_whom="74п ЦМК")
    )

    return (
        debts.filter(from_whom=from_whom, to_whom=to_whom).aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
        - debts.filter(from_whom=to_whom, to_whom=from_whom).aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
    )


@typing.overload
def get(
    whom: str | ProjectRegion,
    balance: typing.Literal["plus", "minus"] | None = None,
    type: typing.Literal["creditors", "debtors"] | None = None,
    project_regions_only: typing.Literal[True] = True,
    director: User | None = None,
) -> QuerySet:
    pass


@typing.overload
def get(
    whom: str | ProjectRegion,
    balance: typing.Literal["plus", "minus"] | None = None,
    type: typing.Literal["creditors", "debtors"] | None = None,
    project_regions_only: typing.Literal[False] = False,
) -> set[str]:
    pass


_GET_SQL: typing.Final[
    str
] = """--sql
SELECT CASE 
       WHEN %(project_regions_only)s THEN t2.id::text
       ELSE t1.participant
       END
  FROM participants AS t1
       LEFT JOIN finance_module_projectregion AS t2 
       ON t2.name = t1.participant
 WHERE CASE
       WHEN %(type)s = 'creditors' THEN ROUND(COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = %(whom)s AND to_whom = t1.participant), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = t1.participant AND to_whom = %(whom)s), 0)) < 0
       WHEN %(type)s = 'debtors' THEN ROUND(COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = %(whom)s AND to_whom = t1.participant), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = t1.participant AND to_whom = %(whom)s), 0)) > 0
       ELSE ROUND(COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = %(whom)s AND to_whom = t1.participant), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = t1.participant AND to_whom = %(whom)s), 0)) != 0
       END
   AND (NOT %(project_regions_only)s OR t2.id IS NOT NULL) 
   AND (%(director_id)s IS NULL OR t2.director_id = %(director_id)s)
   AND CASE 
       WHEN %(balance)s = 'minus' THEN ROUND(COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = t1.participant), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE to_whom = t1.participant), 0)) < 0
       WHEN %(balance)s = 'plus' THEN ROUND(COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom = t1.participant), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE to_whom = t1.participant), 0)) > 0
       ELSE 1
       END
"""


def get(
    whom: str | ProjectRegion,
    balance: typing.Literal["plus", "minus"] | None = None,
    type: typing.Literal["creditors", "debtors"] | None = None,
    project_regions_only: bool = True,
    director: User | None = None,
) -> set[str] | QuerySet:
    """
    Получить тех кому `whom` задолжал если `"creditors"`, тех кто задалжал `whom` если `"debtors"`, или всех если `None`, с сальдо `balance`
    >>> get("Алматы Кселл", "plus", "creditors")
    <QuerySet [<ProjectRegion: Костанай Картел>, <ProjectRegion: Павлодар Картел>, <ProjectRegion: Оскемен Картел>, <ProjectRegion: Азамат>, <ProjectRegion: Аблай Даутбергенов>]>
    >>> get("Алматы Кселл", "plus", "creditors", project_regions_only=False)
    {'Счета - прочие: сырье, аксессуары для ПК,  дизеление, билеты, проживание, тендерные услуги, аренда офиса,юр услуги, расходы на офис.', 'Павлодар Картел', 'Счета - ТМЦ', 'Аблай Даутбергенов', 'Оскемен Картел', 'Азамат', 'Счета- Работы: СМР, ЭМР, Обследование', 'Костанай Картел'}
    >>> get("Алматы Кселл", "plus", "creditors", director=User.objects.get(email="20309@avh.kz"))
    <QuerySet [<ProjectRegion: Алматы Картел>, <ProjectRegion: Алматинская область Картел>]>
    """
    if isinstance(whom, ProjectRegion):
        whom = whom.name

    participants: set[str] = set()
    for participant in get_participants():
        debt_sum = round(get_debt_sum(from_whom=whom, to_whom=participant))
        if type == "creditors":
            if debt_sum >= 0:
                continue
        elif type == "debtors":
            if debt_sum <= 0:
                continue
        else:
            if not debt_sum:
                continue
        if balance is not None:
            balance_ = round(get_balance(participant))
            if balance == "minus":
                if balance_ >= 0:
                    continue
            elif balance == "plus":
                if balance_ <= 0:
                    continue
        participants.add(participant)

    if project_regions_only:
        project_regions = ProjectRegion.objects.filter(name__in=participants)
        if director:
            project_regions = project_regions.filter(director=director)
        return project_regions
    return participants


_GET_BALANCE_SQL: typing.Final[
    str
] = """--sql
SELECT COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE from_whom=%(whom)s), 0) - COALESCE((SELECT SUM(sum) FROM not_rejected_debts WHERE to_whom=%(whom)s), 0)
"""


def get_balance(whom: str | ProjectRegion) -> decimal.Decimal:
    if isinstance(whom, ProjectRegion):
        whom = whom.name
    debts = Debt.objects.exclude(group__status=rejected_debt_translate_group_status())
    return (
        debts.filter(from_whom=whom).aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
        - debts.filter(to_whom=whom).aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
    )


def setoff(user: User, subdivision: Subdivision) -> None:
    project_regions = get_project_regions(user=user).filter(subdivision=subdivision)
    accounts = Account.objects.none()
    for project_region in project_regions:
        accounts |= project_region.accounts.all()
    accounts = accounts.filter(
        subdivision=subdivision, is_cash_register=False
    ).distinct()
    setoff = Setoff.objects.create(responsible=user)
    for account in accounts:
        minus = {}
        plus = {}
        for project_region in project_regions:
            balance = round(get_account_balance(account, project_region), 2)
            if balance < 0:
                minus[project_region] = balance
            else:
                plus[project_region] = balance
        for to_whom in minus:
            from_whom = max(plus, key=lambda x: plus[x])
            sum = round(min(-minus[to_whom], plus[from_whom]), 2)
            minus[to_whom] += sum
            plus[from_whom] -= sum
            debts, debt_translate_group = translate(
                from_whom=from_whom,
                from_account=account,
                to_whom=to_whom,
                to_account=account,
                sum=sum,
                responsible=user,
            )
            debt_translate_group.setoff = setoff
            debt_translate_group.save()
            Debt.objects.bulk_create(debts)


def translate(
    *,
    from_whom: ProjectRegion,
    from_whom_balance: decimal.Decimal | None = None,
    from_account: Account,
    to_whom: ProjectRegion,
    to_whom_balance: decimal.Decimal | None = None,
    to_account: Account,
    sum: decimal.Decimal,
    responsible: User | None = None,
    common_director: bool = True,
    full_repayment: bool = False,
    debt_translate_group: DebtTranslateGroup | None = None,
    exclude: set[int] | None = None,
) -> tuple[list[Debt], DebtTranslateGroup]:
    if from_whom_balance is None:
        from_whom_balance = get_balance(from_whom.name)
    if to_whom_balance is None:
        to_whom_balance = get_balance(to_whom.name)

    director = None
    if common_director and from_whom.director == to_whom.director:
        director = from_whom.director
    else:
        common_director = False

    if debt_translate_group is None:
        debt_translate_group = DebtTranslateGroup(
            from_whom=from_whom.name,
            from_account=from_account,
            to_whom=to_whom.name,
            to_account=to_account,
            sum=sum,
            responsible=responsible,
        )
    debt_translate_group.common_director = common_director

    debts = []
    if (
        not from_whom.subdivision
        or from_whom.subdivision.name != "7п"
        or not to_whom.subdivision
        or to_whom.subdivision.name != "7п"
    ):
        debt_translate_group.type = "plus_minus"
        debts.append(
            Debt(
                from_whom=from_whom.name,
                to_whom=to_whom.name,
                sum=sum,
                responsible=responsible,
                group=debt_translate_group,
            )
        )
        return debts, debt_translate_group

    debt_sum = get_debt_sum(to_whom=from_whom, from_whom=to_whom)
    if (
        from_whom_balance > 0
        and to_whom_balance < 0
        and debt_sum > 0
        and (full_repayment or sum > debt_sum)
    ):
        debt_translate_group.type = "minus_plus_and_plus_minus_plus"
        if exclude is None:
            exclude = set()
        if not full_repayment:
            sum -= debt_sum
            debts.append(
                Debt(
                    from_whom=from_whom.name,
                    to_whom=to_whom.name,
                    sum=debt_sum,
                    responsible=responsible,
                    group=debt_translate_group,
                )
            )
            exclude.add(to_whom.id)
        # TODO: Здесь непонятная хрень начинается
        for creditor in get(from_whom, "plus", "creditors", director=director).exclude(
            id__in=exclude
        ):
            debt_sum = get_debt_sum(to_whom=from_whom, from_whom=creditor)
            if sum > debt_sum:
                sum -= debt_sum
                debts.append(
                    Debt(
                        from_whom=from_whom.name,
                        to_whom=creditor.name,
                        sum=debt_sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                exclude.add(creditor.id)
                continue
            for debtor in get(creditor, "minus", "debtors", director=director) & get(
                to_whom, "minus", "debtors", director=director
            ):
                debts.append(
                    Debt(
                        from_whom=from_whom.name,
                        to_whom=creditor.name,
                        sum=sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                debts.append(
                    Debt(
                        from_whom=creditor.name,
                        to_whom=debtor.name,
                        sum=sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                debts.append(
                    Debt(
                        from_whom=debtor.name,
                        to_whom=to_whom.name,
                        sum=sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                return debts, debt_translate_group
        if common_director:
            result = translate(
                from_whom=from_whom,
                to_whom=to_whom,
                sum=sum,
                from_account=from_account,
                to_account=to_account,
                responsible=responsible,
                common_director=False,
                from_whom_balance=from_whom_balance,
                to_whom_balance=to_whom_balance,
                full_repayment=True,
                debt_translate_group=debt_translate_group,
                exclude=exclude,
            )
            debts.extend(result[0])
        else:
            debts.append(
                Debt(
                    from_whom=from_whom.name,
                    to_whom=to_whom.name,
                    sum=sum,
                    responsible=responsible,
                    group=debt_translate_group,
                )
            )
        return debts, debt_translate_group

    if from_whom_balance < 0:
        if to_whom_balance > 0:
            debt_translate_group.type = "minus_plus"
            debts.append(
                Debt(
                    from_whom=from_whom.name,
                    to_whom=to_whom.name,
                    sum=sum,
                    responsible=responsible,
                    group=debt_translate_group,
                )
            )
        else:
            for common_debt_project_region in get(
                from_whom, "plus", "creditors", director=director
            ) & get(to_whom, "plus", "creditors", director=director):
                debt_translate_group.type = "minus_plus_minus"
                debts.append(
                    Debt(
                        from_whom=from_whom.name,
                        to_whom=common_debt_project_region.name,
                        sum=sum,
                        group=debt_translate_group,
                        responsible=responsible,
                    )
                )
                debts.append(
                    Debt(
                        from_whom=common_debt_project_region.name,
                        to_whom=to_whom.name,
                        sum=sum,
                        group=debt_translate_group,
                        responsible=responsible,
                    )
                )
                return debts, debt_translate_group
            for common_debt_project_region in get(
                from_whom, "plus", "creditors", director=director
            ):
                if to_whom not in get(common_debt_project_region, "minus", "creditors"):
                    debt_translate_group.type = "minus_new_plus_to_minus"
                    debts.append(
                        Debt(
                            from_whom=from_whom.name,
                            to_whom=common_debt_project_region.name,
                            sum=sum,
                            group=debt_translate_group,
                            responsible=responsible,
                        )
                    )
                    debts.append(
                        Debt(
                            from_whom=common_debt_project_region.name,
                            to_whom=to_whom.name,
                            sum=sum,
                            responsible=responsible,
                            group=debt_translate_group,
                        )
                    )
                    return debts, debt_translate_group
            if common_director:
                return translate(
                    from_whom=from_whom,
                    to_whom=to_whom,
                    sum=sum,
                    responsible=responsible,
                    common_director=False,
                    from_whom_balance=from_whom_balance,
                    to_whom_balance=to_whom_balance,
                )
    else:
        if to_whom_balance < 0:
            debt_translate_group.type = "plus_minus"
            debts.append(
                Debt(
                    from_whom=from_whom.name,
                    to_whom=to_whom.name,
                    sum=sum,
                    responsible=responsible,
                    group=debt_translate_group,
                )
            )
        else:
            for common_debt_project_region in get(
                from_whom, "minus", "debtors", director=director
            ) & get(to_whom, "minus", "debtors", director=director):
                debt_translate_group.type = "plus_minus_plus"
                debts.append(
                    Debt(
                        from_whom=from_whom.name,
                        to_whom=common_debt_project_region.name,
                        sum=sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                debts.append(
                    Debt(
                        from_whom=common_debt_project_region.name,
                        to_whom=to_whom.name,
                        sum=sum,
                        responsible=responsible,
                        group=debt_translate_group,
                    )
                )
                return debts, debt_translate_group
            for common_debt_project_region in get(
                from_whom, "minus", "debtors", director=director
            ):
                if to_whom not in get(common_debt_project_region, "plus", "debtors"):
                    debt_translate_group.type = "plus_new_minus_to_plus"
                    debts.append(
                        Debt(
                            from_whom=from_whom.name,
                            to_whom=common_debt_project_region.name,
                            sum=sum,
                            responsible=responsible,
                            group=debt_translate_group,
                        )
                    )
                    debts.append(
                        Debt(
                            from_whom=common_debt_project_region.name,
                            to_whom=to_whom.name,
                            sum=sum,
                            responsible=responsible,
                            group=debt_translate_group,
                        )
                    )
                    return debts, debt_translate_group
            if common_director:
                return translate(
                    from_whom=from_whom,
                    to_whom=to_whom,
                    sum=sum,
                    from_account=from_account,
                    to_account=to_account,
                    responsible=responsible,
                    common_director=False,
                    from_whom_balance=from_whom_balance,
                    to_whom_balance=to_whom_balance,
                )

    return debts, debt_translate_group
