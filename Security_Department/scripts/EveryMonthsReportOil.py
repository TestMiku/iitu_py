from openpyxl import load_workbook
import pandas as pd
import requests
import io
from urllib.parse import urlencode
import json
import re
from datetime import datetime

session = requests.Session()

Accounts = {
    '37013@avh.kz': 'Serik1988',
    '30060@avh.kz': 'ErTyUiOp37406810',
}

cardOwners = {
    '7341000000956998': 'Даниев М.',
    '7341000001836256': 'Сатыбалдин Жасулан',
    '7341000001836264': 'Джайлаев Бакытжан',
    '7341000001836272': 'Сурапбергенов Айдос',
    '7341000002191990': 'Разлошко Василий',
    '7341000002192006': 'Бекембаев Талгат',
    '7341000002192014': 'Алибаев Ерлан',
    '7341000002192022': 'Сулиев Сулейман',
    '7341000002192030': 'Тарасевич Андрей',
    '7341000002192048': 'Архипов Александ',
    '7341000002192055': 'Шайкен Нурлан',
    '7341000002192063': 'Абайдуллаев Дильяр',
    '7341000002192071': 'Безымянный',
    '7341000002192089': 'Жексембаев Мурат',
    '7341000002192097': 'Сагынбаев Айбын',
    '7341000002192105': 'Давранов Курванбек',
    '7341000002192113': 'Азизов Алимжан',
    '7341000002192188': 'Давидок Сергей',
    '7341000002192196': 'Анаятов Хиястин',
    '7341000002192287': 'Мансуров Османжан',
    '7341000002192295': 'Безымянный',
    '7342000001502659': 'Курбанов Расим',
    '7342000001502667': 'Матренин Анатолий',
    '7342000001775891': 'Жакиянов Азамат',
    '7342000003490390': 'Беисенбай Рамазан',
    '7342000003491109': 'Кушимов Галымжан',
    '7342000003492081': 'Бутакбаев Бекулан',
    '7342000003492149': 'Тлектесов Айберген',
    '7342000003492248': 'Курбанов Расим',
    '7342000003492594': 'Владимир Викторвич',
    '7342000003492651': 'Керемкулов Канат',
    '7342000003492719': 'Еркын Данияр',
    '7342000003492735': 'Жылысбаев Алимгазы',
    '7342000003533579': 'Жылысбаев Адилгазы',
    
    '200072686': 'Сагындык (Байконыр)',
    '0200073830': 'Сундет',
    '0200073895': 'Есен',
    '0200075914': 'Батыр',
    '0200076598': 'Муханов Ердос',
    '0200076599': 'Алибеков Марат',
    '200079627': 'Маулен',
    '200080240': 'Нариман',
    '200080241': 'Максат',
    '200080242': 'Рома',
    '200080243': 'Акнур',
    '0200080244': 'Нурлан',
    '0200080245': 'Амиралиев Жасулан',
    '0200080247': 'Айсултан',
    '0200080250': 'Томашев Галым',
    '200080251': 'Марат',
    '0200080252': 'Султан',
    '200080253': 'Ернар',
    '200080254': 'Ануар',
    '200080815': 'Айдос',
    '0200081638': 'Шеримбетов Бакытжан',
    '0200083459': 'Абдрахманов Нурсултан',
    '0200086904': 'Бердияр',
    '200091783': 'Тимур',
    '0200093738': 'Дархан',
    '200094922': 'Азамат',
    '0200098216': 'Бауыржан',
    '0200098217': 'Азамат',
    '0200098239': 'Мейрамбек',
    '0200098268': 'Нурсултан ( Кентау)',
    '200099528': 'Темирлан',
    '0200099690': 'Серик',
    '0200099691': 'Олжас',
    '200123270': 'Емберген',
    '0200123271': 'Аян',
    
    '7341000001812935': 'Нариман',
    '7341000001812943': 'Темирлан',
    '7341000001812950': 'Маулен',
    '7341000001813016': 'Максат',
    '7341000001813057': 'Тимур',
    '7341000001816165': 'Куаныш',
    '7341000001836170': 'Султан',
}

def Auth(login='37013@avh.kz', password='Serik1988'):
    url = "https://mycard.petroretail.kz/customers/sign_in?locale=ru"

    response = session.get(url)
    html_content = response.text

    start_index = html_content.find('<meta name="csrf-token"')
    end_index = html_content.find('>', start_index)

    meta_tag_content = html_content[start_index:end_index + 1]

    start_content_index = meta_tag_content.find('content="') + len('content="')
    end_content_index = meta_tag_content.find('"', start_content_index)

    csrf_token = meta_tag_content[start_content_index:end_content_index]

    payload = {
        'authenticity_token': csrf_token,
        'customer[email]': login,
        'customer[password]': password,
        'commit': 'Вход'
    }
    session.post(url, data=payload)
    return csrf_token

def auth_required(func):
    def wrapper(*args, **kwargs):
        Auth()
        return func(*args, **kwargs)
    return wrapper

def encodeFormData(params):
    encodedParams = urlencode({"params": json.dumps(params)})
    return encodedParams

def sendRequest(url, params):
    formattedPayload = encodeFormData(params)
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    response = session.post(url, headers=headers, data=formattedPayload)
    responseData = response.json()

    return responseData

@auth_required
def DoGet(url):
    response = session.get(url)
    return response.text

import aiohttp
import asyncio

async def send_async_request(json_data):
    async with aiohttp.ClientSession() as session:
        async with session.post('https://script.google.com/macros/s/AKfycbxxImC42je0klFYVCMDyYDr5bFHSBrv9Hkga5g4Nn8tFJhMlmFgRno-gxV3oLifcMbs/exec', json=json_data) as response:
            response.raise_for_status()

@auth_required
async def GetReports():
    reportID = takeReportId()
    url = f'https://mycard.petroretail.kz/ru/monthly_reports/{reportID}/download'
    response = session.get(url)
    response.raise_for_status()
    buffer = io.BytesIO(response.content)
    wb = load_workbook(filename=buffer)
    data = []
    number_coord = None 
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        found_number = False
        pokupki_processed = False

        # Ищем строки с номером карты и строку "Всего:"
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell.value, str):
                    if cell.value.startswith("Номер карты:"):
                        card_number = cell.value.split(":")[1].strip()
                        card_number_int = int(card_number) 
                        found_number = True  
                        data.append({int(card_number_int): []})  # Добавляем словарь в список

                        # print(f'Найден номер карты: {cell.value} (Строка: {row_idx}, Столбец: {col_idx})')
                    elif cell.value.startswith("Всего:") and col_idx == 1 and found_number:
                        fifth_col_value = ws.cell(row=row_idx, column=5).value
                        if fifth_col_value is None:
                            value_in_seven_column = ws.cell(row=row_idx, column=7).value
                            value_in__ten_column = ws.cell(row=row_idx, column=10).value
                            date = ws.cell(row=16, column=4).value.split(" - ")[0].split(".")[-2:]
                            month_year_str = ".".join(date)
                            driverName = cardOwners.get(card_number) if cardOwners.get(card_number) is not None else "Нет данных о владельце"
                            data[-1][int(card_number_int)].append(value_in_seven_column)  # Добавляем значение в последний словарь в списке
                            data[-1][int(card_number_int)].append(value_in__ten_column)
                            data[-1][int(card_number_int)].append(month_year_str)
                            data[-1][int(card_number_int)].append(driverName)
                            data[-1][int(card_number_int)].append(ws.title) 
                            # print("sssssssssssss",ws.cell(row=195, column=5).value)
                    
                    elif cell.value.startswith("Покупки:") and col_idx == 5 and not pokupki_processed:
                        print("Cell value:", cell.value)
                        print("Column index:", col_idx)
                        pokupki_litr_value = ws.cell(row=row_idx, column=7).value
                        pokupki_summa_value = ws.cell(row=row_idx, column=10).value
                        print("sssssssssssss",pokupki_litr_value, pokupki_summa_value)
                        if data and data[-1].get(card_number_int):
                            data[-1][card_number_int].append(pokupki_litr_value) 
                            data[-1][card_number_int].append(pokupki_summa_value)
                            pokupki_processed = True
                        else:
                            print("Ошибка: Нет данных для карты")
                            
                            
                            # print(f'Найдено значение "Всего": {cell.value} ({value_in_seven_column})сумма {value_in__ten_column}  (Строка: {row_idx}, Столбец: {col_idx+3})')
    if not found_number:    
        print("Номер карты не найден на листе", sheet_name) 
    if number_coord is None:
        print("Номер карты не найден.")
    
    wb.close()
    json_data = json.dumps(data, ensure_ascii=False)

    await send_async_request(json_data)


# def takeReportId():
#     pattern = r'<a href="/ru/monthly_reports/(.+?)/download">'
#     text = DoGet("https://mycard.petroretail.kz/ru/monthly_reports")
#     match = re.search(pattern, text)

#     if match:
#         return match.group(1)
#     else:
#         return None
    
def takeReportId():
    pattern = r'<a href="/ru/monthly_reports/(.+?)/download">'
    text = DoGet("https://mycard.petroretail.kz/ru/monthly_reports")
    matches = re.findall(pattern, text)

    if len(matches) >= 1:
        return matches[1]
    else:
        return None

def main():
    asyncio.run(GetReports())
    
