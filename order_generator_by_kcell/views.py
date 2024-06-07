from django.http import FileResponse, HttpResponseNotFound, HttpResponseServerError
from django.shortcuts import render, redirect
from django.conf import settings
from django.db import connection

import openpyxl
from openpyxl.styles import Alignment, Border, Side
from reporter.models import Report
from .models import TCPModel


def import_page(request):
    return render(request, 'order_generator_by_kcell/import.html')


def handle_excel_cells(file_path: str) -> str:
    hidden_i = 0
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    kye_values = [['Заказ  №', 'от'], ['Тип Работ:', 'Дата начала выполнения работ:', 'Дата выполнения Интеграции', 'Дата окончания работ:']]
    header_values = []
    table_data = []
    swither = False
    mcol = 6
    tcp_index_from_table = 1
    if not sheet['C2'].value and 'План' in sheet['B13'].value:
        mcol = 8
        hidden_i = 1
        
        kye_values[1][1] = "Планируемая Дата начала выполнения работ:"
        kye_values[1][2] = "Планируемая Дата окончания работ:"
        kye_values[1][3] = "Критическая просрочка(дней):"
        # tcp_index_from_table = 2
    if sheet['H6'].value:
        mcol = 8
    description = ''
    initiator = ''
    for row in sheet.iter_rows(max_col=mcol):
        for i, cell in enumerate(row):
            if cell.value in kye_values[0]:
                header_values.append(row[i+1].value)
            if cell.value in kye_values[1]:
                print(row[i+3].value)
                header_values.append(row[i+3].value)
            if cell.value == '№':
                swither = True
                start_index = i
            if cell.value == 'Описание работ:':
                description = sheet.cell(row=cell.row+1, column=cell.column).value
            if cell.value == 'Инициатор:':
                initiator = row[i+1].value
        if row[0].value and swither:
            table_data.append([row[start_index].value, row[start_index+2-hidden_i].value, row[start_index+1+hidden_i].value, row[start_index+3].value, row[start_index+4].value, str(row[start_index+5].value).replace('.', ',')])
        else:
            swither = False
    if sheet['L4'].value:
        description = sheet['L4'].value
    if sheet['L31'].value:
        initiator = sheet['L31'].value.replace('Контактное лицо: ', '').strip()
    template_excel_path = (
    settings.BASE_DIR / f"order_generator_by_kcell/templates/order_generator_by_kcell/Шаблон.xlsx"
    )
    
    template_wb = openpyxl.load_workbook(template_excel_path)
    template_sheet = template_wb.active
    if not sheet['C2'].value:
        template_sheet['B12'] = kye_values[1][1]
        template_sheet['B13'] = kye_values[1][2]
        template_sheet['B14'] = kye_values[1][3]
    print(header_values)
    header_position = ['E6', 'H6', 'E11', 'E12', 'E13', 'E14']
    for i, value in enumerate(header_position):
        template_sheet[value] = header_values[i]
    table_data.pop(0)
    max_len_table = len(table_data)
    
    template_sheet['N1']=header_values[0]
    template_sheet['L4']=description
    template_sheet['L38']=f'Контактное лицо: {initiator}'
    for i, list_ in enumerate(table_data):
        for j, value in enumerate(list_):
            template_sheet.cell(row=18+i, column=j+1, value=value)
        try:
            founded_data = TCPModel.objects.get(document_number=list_[tcp_index_from_table])
            unit_price_value = founded_data.unit_price
        except TCPModel.DoesNotExist:
            unit_price_value = 0
        template_sheet[f'G{18+i}'] = unit_price_value
            
    
    # Определение стилей
    cell_style = openpyxl.styles.NamedStyle(
        name="custom_style",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
    )
    
    # Применение стилей к заданному диапазону
    for i, row in enumerate(template_sheet.iter_rows(min_row=18, max_row=18 + max_len_table - 1 , min_col=1, max_col=9)):
        for cell in row:
            cell.style = cell_style
        # template_sheet[f'G{i+18}'] = 1
        template_sheet[f'H{i+18}'] = f'=F{i+18}*G{i+18}'
        template_sheet[f'I{i+18}'] = f'=H{i+18}*1.12'
        template_sheet[f'G{i+18}'].number_format = '#,##0.00'
        template_sheet[f'H{i+18}'].number_format = '#,##0.00'
        template_sheet[f'I{i+18}'].number_format = '#,##0.00'
    #----------------Вырезка облости с стилями-------------------   
    start_row = 77
    end_row = 91
    start_col = 1
    end_col = 9
    

    # Определяем место, куда нужно вставить данные
    target_row = 18 + max_len_table
    target_col = 1

    

    # Копирование данных
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell_copy = template_sheet.cell(row=row_idx, column=col_idx)
            cell_target = template_sheet.cell(row=target_row + row_idx - start_row, column=target_col + col_idx - start_col)
            
            # Копирование значений для верхнего левого угла каждой объединенной области
            merged_ranges = template_sheet.merged_cells.ranges
            for merged_range in merged_ranges:
                min_row, min_col, max_row, max_col = merged_range.bounds
                if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
                    cell_target.value = cell_copy.value
                    break
            else:
                cell_target.value = cell_copy.value

            # Копирование стилей
            if cell_copy.has_style:
                cell_target.font = cell_copy.font.copy()
                cell_target.border = cell_copy.border.copy()
                cell_target.fill = cell_copy.fill.copy()

    # Очистка данных в заданном диапазоне
    for row in template_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            try:
                cell.value = None
            except:pass
            cell.font = openpyxl.styles.Font()
            cell.border = openpyxl.styles.Border()
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
    
    # Очистка данных и форматирования в заданном диапазоне
    for merged_cells_range in template_sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_cells_range.bounds
        if start_row <= min_row <= end_row and start_col <= min_col <= end_col:
            # Если объединенная область перекрывается с заданным диапазоном, разъединяем ее
            template_sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    
    #-----------------------------------------------------------------------
    
    template_sheet[f'I{target_row+1}'] = f'=SUM(I18:I{target_row-1})'
    template_sheet[f'I{target_row+1}'].number_format = '#,##0.00'
    
    template_sheet[f'I{target_row+2}'] = f'=(I{target_row+1}/112)*12'
    template_sheet[f'I{target_row+2}'].number_format = '#,##0.00'

    template_sheet[f'F{target_row+4}'] = f'=I{target_row+1}'
    template_sheet[f'F{target_row+4}'].number_format = '#,##0.00'
    
    template_wb.save(f'order_generator_by_kcell/tempfile.xlsx')
    handled_excel_path = (
        settings.BASE_DIR / f"order_generator_by_kcell/tempfile.xlsx"
    )
    filename = f'{header_values[0]} {table_data[0][2]}.xlsx'
    return handled_excel_path, filename



def main(request):
    Report.objects.create(responsible=request.user, process="og_by_kcell - Открыт главная страница", text="og_by_kcell.open")
    return render(request, 'order_generator_by_kcell/index.html')

def import_data(request):
    if request.method == 'POST':
        file_path = request.FILES['file_import']
        TCPModel.objects.all().delete()
        with connection.cursor() as cursor:
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'order_generator_by_kcell_tcpmodel'")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_col=2, max_col=7, min_row=2):
            TCPModel.objects.create(document_number=row[0].value, unit_price=row[-1].value)
    return redirect('order_generator_by_kcell')

def handle_uploaded_file(request):
    # Report.objects.create(responsible=request.user, process="Генератор заказа KCELL Обработка", text="og_by_kcell.handle")
    if request.method == 'POST':
        # print(request.FILES, )
        file_path = request.FILES['file_upload']
        print(file_path)
        handled_excel_path, file_name= handle_excel_cells(file_path)
        Report.objects.create(responsible=request.user, process="og_by_kcell - Обработано", text="og_by_kcell.open")
        # handled_excel_path = f'{header_values[0]}.xlsx'
        try:
        # Отправляем файл пользователю
            response = FileResponse(open(handled_excel_path, "rb"), filename=file_name)
            return response
        except FileNotFoundError:
            # Возвращаем 404, если файл не найден
            return HttpResponseNotFound()
        except Exception as e:
            print(f"Error: {e}")
            # Возвращаем 500 в случае других ошибок
            return HttpResponseServerError()
    return redirect('order_generator_by_kcell')