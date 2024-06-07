import tempfile
import io
import openpyxl.worksheet.worksheet
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from django.db.models import QuerySet

from finance_module.services.unpaid_invoices_service import set_worksheet_columns_width, get_paid_invoices


def set_worksheet_columns_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    worksheet.append(
        [
            "ДО",
            "Дата",
            "№ счёта",
            "Дата счёта",
            "Проект",
            "Ответственный",
            "Утвердитель",
            "ТОО",
            "Контрагент",
            "Комментарий",
            "Валюта",
            "Сумма для таблицы ",
            "Категория счёта",
            "Статьи доходов/расходов",
            "Заказ на продажу",
            "БИН/ИИН",
            "Сумма документа",
            "С какого р/с платить",
            "ИИК",
            "КНП",
            "Фактический номер договора ",
            "Сумма по счёту",
            "Оплаченная ранее сумма (1С)",
            "Сумма от ПМ"
        ]
    )


def change_format_of_cells_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [2, 4]:
                cell.number_format = "DD.MM.YYYY"
            if cell.column in [12, 24]:
                cell.number_format = "#,##0.00"


def set_data_to_worksheet_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet, paid_invoices: QuerySet) -> None:
    for paid_invoice in paid_invoices:
        at = paid_invoice.at.replace(tzinfo=None)

        responsible_user_id = paid_invoice.responsible.avh_user_id_from_email if paid_invoice.responsible else None

        worksheet.append(
            [
                paid_invoice.number,
                at,
                paid_invoice.invoice_number,
                paid_invoice.invoice_date,
                paid_invoice.project,
                responsible_user_id,
                paid_invoice.approver,
                paid_invoice.llc,
                paid_invoice.contractor,
                paid_invoice.comment,
                paid_invoice.currency,
                paid_invoice.sum,
                paid_invoice.invoice_category,
                paid_invoice.revenue_expense_articles,
                paid_invoice.sales_order,
                paid_invoice.bin_or_iin,
                paid_invoice.document_amount,
                paid_invoice.account.name,
                paid_invoice.iic,
                paid_invoice.payment_destination_code,
                paid_invoice.contract_number,
                paid_invoice.invoice_amount,
                paid_invoice.paid_amount_1c,
                paid_invoice.sum
            ]
        )

    stylizing_worksheet_logistic(worksheet)
    change_format_of_cells_logistic(worksheet)


def stylizing_name_column_worksheet_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(name='Times New Roman', bold=True, size=8)
    worksheet.row_dimensions[1].height = 41.5


def set_worksheet_row_height_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.font = Font(name='Times New Roman', bold=False, size=8)
        worksheet.row_dimensions[row[0].row].height = 53.5
    for col in worksheet.columns:
        worksheet.column_dimensions[col[0].column_letter].width = 9


def stylizing_worksheet_logistic(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    stylizing_name_column_worksheet_logistic(worksheet)

    for row_idx, row_data in enumerate(worksheet.iter_rows(min_row=1)):
        for cell in row_data:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    set_worksheet_columns_width(worksheet, [1, 2])
    set_worksheet_row_height_logistic(worksheet)


def unloading_logistic() -> io.BytesIO | None:
    workbook = openpyxl.Workbook()
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
    xlsx_1 = tempfile.NamedTemporaryFile("wb+", prefix="1С ", suffix=".xlsx")
    set_worksheet_columns_logistic(worksheet)
    set_data_to_worksheet_logistic(worksheet, get_paid_invoices())

    workbook.save(xlsx_1)
    workbook.close()
    xlsx_1.seek(0)
    return xlsx_1

