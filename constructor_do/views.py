import openpyxl
from datetime import datetime
import os
import json
from io import BytesIO
import random

from django.conf import settings
from django.db.models import Q
from django.shortcuts import render, HttpResponseRedirect
from django.core import serializers
from django.http import FileResponse, HttpResponseNotFound, HttpResponseServerError, StreamingHttpResponse, JsonResponse

from calculator_emr.models import dataBS, dataTCP, data1922
from calculator.models import HeaderPosition, Data7112

order_dict = {}
import_data = {
    "import_data": [
        {
            "Спецификация счёта": []
        }
    ]
}
tcps, unique_bs = [], []
order_numbers = []
previous_orders = []
show_previous = ''
project_managers = ['798 - Маратов Олжас Бауыржанович_СД_798_Инженер проектировщик', '21929 - Головкова Кристина Александровна', '21297 - Краснюков Илья Михайлович', '20277 - Иващенко Андрей_700_20277_Менеджер по лизингу', 'from google sheet']
result_file_name = ''
filtered_data_by_pm = None


def main(request):
    global order_dict, show_previous, result_file_name, unique_bs, filtered_data_by_pm

    if filtered_data_by_pm is None:
        filtered_data_by_pm = Data7112.objects.filter(
            Q(project_manager__in=project_managers) | Q(iprovider__in=project_managers) | Q(
                iconfirmer__in=project_managers)).exclude(odescription='-')
        unique_bs = list(set([row.odescription for row in filtered_data_by_pm]))

    bs_name = request.POST.get("bs_name") if request.POST.get("bs_name") else request.GET.get("bs_name")
    order_numer = request.POST.get("on_name")
    product_name = request.POST.get("product_name")
    search_key = request.POST.get("searchKey")
    name_tcp = request.POST.get("nameTCP")
    unit_count = request.POST.get("unitCount")
    unit_price = request.POST.get("unitPrice")
    total_price = request.POST.get("totalPrice")
    unit_name = request.POST.get("unitName")
    tax_included = request.POST.get("taxIncluded")
    show_previous = request.POST.get("showPrevious") if request.POST.get("showPrevious") is not None else show_previous

    if bs_name and order_numer:
        if all(order_numer not in value or product_name not in value for value in order_dict.values()):
            order_dict[order_numer+str(random.randint(100000, 999999))] = [bs_name, order_numer, product_name]
            if result_file_name == '':
                result_file_name = f'{bs_name.strip()} - {order_numer.split(" ")[0]}'
    if order_numer:
        filtered_data_by_order = dataBS.objects.filter(order_number=order_numer)
        return render(request, "constructor_do/index.html", {"order_numer_": order_numer, "all_orders": order_dict.values(), "previous_orders": previous_orders, "show_previous": show_previous, "dataToParse": serializers.serialize("json", filtered_data_by_order), "data_tcp": dataTCP.objects.all(), "data_BS": dataBS.objects.all(), "unique_bs": unique_bs})
    elif bs_name:
        filtered_data_by_bs = filtered_data_by_pm.filter(odescription=bs_name).distinct('order_number', 'odateordered', 'productname')
        return render(request, "constructor_do/index.html", {"filtered_data_bs": filtered_data_by_bs, "data_BS": dataBS.objects.values('project').distinct(), "unique_bs": unique_bs, "data_tcp": dataTCP.objects.all(), "all_orders": order_dict, "previous_orders": previous_orders, "show_previous": show_previous})
    elif search_key:
        fill_xlsx([search_key, name_tcp, unit_count, unit_price, total_price, tax_included], order_dict)
        fill_html([search_key, name_tcp, unit_count, unit_price, total_price, unit_name, tax_included], order_dict)
        fill_previous_orders(order_dict)
        order_dict = {}
        return JsonResponse({"previous_orders": previous_orders})
    return render(request, "constructor_do/index.html", {"data_BS": dataBS.objects.values('project').distinct(), "unique_bs": unique_bs, "data_tcp": dataTCP.objects.all(), "all_orders": order_dict, "previous_orders": previous_orders, "show_previous": show_previous},)


def fill_xlsx(data_tcp, orders):
    input_file = 'constructor_do/input.xlsx'
    result_file = 'constructor_do/result.xlsx'

    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    try:
        workbook = openpyxl.load_workbook(result_file)
        sheet = workbook.active
    except FileNotFoundError:
        print("File not found")

    current_row = sheet.max_row if sheet.cell(row=sheet.max_row, column=33).value is None else sheet.max_row + 1

    non_empty_count = 0

    for cell in sheet['V']:
        if cell.value is not None:
            non_empty_count += 1

    index_num = f'{non_empty_count - 1}0'

    comment_cell = sheet.cell(row=3, column=4)
    comment_cell.value = comment_cell.value + f', {data_tcp[1]}' if comment_cell.value else data_tcp[1]
    sheet.cell(row=3, column=2).value = datetime.now().date().strftime("%d.%m.%Y")

    sheet.cell(row=current_row, column=22).value = index_num
    sheet.cell(row=current_row, column=23).value = data_tcp[1]
    sheet.cell(row=current_row, column=24).value = data_tcp[2]
    sheet.cell(row=current_row, column=25).value = data_tcp[2]
    sheet.cell(row=current_row, column=26).value = 'Обследование'
    sheet.cell(row=current_row, column=27).value = data_tcp[4]
    sheet.cell(row=current_row, column=28).value = '7_7п_Телеком'
    sheet.cell(row=current_row, column=29).value = 'Телекоммуникации'
    sheet.cell(row=current_row, column=31).value = 'НДС 12%' if data_tcp[5] == 'true' else "Без налога"

    for key, value in orders.items():
        order_data = data1922.objects.filter(order_number=value[1])

        sheet.cell(row=current_row, column=33).value = index_num
        sheet.cell(row=current_row, column=34).value = value[1].split(' ')[0]
        sheet.cell(row=current_row, column=35).value = order_data[0].projects_group if order_data else None
        sheet.cell(row=current_row, column=36).value = value[0]
        sheet.cell(row=current_row, column=37).value = 1.0
        sheet.cell(row=current_row, column=37).number_format = '#,##0.00'
        sheet.cell(row=current_row, column=38).value = data_tcp[3]
        sheet.cell(row=current_row, column=39).value = '7_7п_Телеком'
        current_row = current_row + 1

    workbook.save('constructor_do/result.xlsx')


def download_excel(request):
    global result_file_name
    file_path = (
        settings.BASE_DIR / "constructor_do/result.xlsx"
    )
    try:
        # Define a file stream generator
        def file_stream(file_path):
            with open(file_path, "rb") as file:
                yield from file
            # Delete the file after yielding its content
            os.remove(file_path)

        # Create a streaming response
        response = StreamingHttpResponse(file_stream(file_path),
                                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="result.xlsx"'
        previous_orders.clear()
        return response
    except FileNotFoundError:
        # Возвращаем 404, если файл не найден
        return HttpResponseNotFound()
    except Exception as e:
        print(f"Error: {e}")
        # Возвращаем 500 в случае других ошибок
        return HttpResponseServerError()


def fill_html(data_tcp, orders):
    global import_data, tcps, order_numbers
    order_nums = ''.join(order[1] if orders[next(reversed(orders.keys()))] == order else f'{order[1]}, ' for order in orders.values())
    tcps.append(data_tcp)
    order_numbers.append(order_nums)
    data = {}
    data["Номер строки"] = (len(import_data["import_data"][0]["Спецификация счёта"]) + 1) * 10
    data["Ключ поиска"] = data_tcp[0]
    data["Количество (в счете)"] = int(data_tcp[2])
    data["Введённая цена"] = float(data_tcp[4])
    data["Налог"] = "НДС 12%" if data_tcp[6] == 'true' else "Без налога"
    data["Связь заказ/счёт"] = [{"Количество введённое": 1.0, "Заказ": order[1].split(' ')[0], "Спецификация заказа": 10, "Итоговая сумма": float(data_tcp[3])} for order in orders.values()]
    import_data["import_data"][0]["Спецификация счёта"].append(data)


def download_html(request):
    global result_file_name, filtered_data_by_pm
    json_data = json.dumps(import_data, ensure_ascii=False, indent=4)
    context = {
        'tcp_data': tcps,
        'order_data': order_numbers,
        'import_data': json_data,
        'current_date': datetime.now().date().strftime("%d.%m.%Y")
    }

    try:
        # Render the template with context
        response = render(request, "constructor_do/result.html", context)

        # Create a BytesIO object from the rendered content
        file = BytesIO(response.content)

        # Return a FileResponse, setting as_attachment to True to prompt download
        file_response = FileResponse(file, as_attachment=True, filename=f"{result_file_name}.html")

        # Clear lists after returning FileResponse
        tcps.clear()
        order_numbers.clear()
        import_data["import_data"][0]["Спецификация счёта"].clear()
        previous_orders.clear()
        result_file_name = ''
        filtered_data_by_pm = None

        return file_response

    except Exception as e:
        print(f"Error: {e}")


def delete_order(request, index):
    try:
        order_dict.pop(index)
    except Exception as e:
        print(e)
        return HttpResponseRedirect(request.META['HTTP_REFERER'])
    return HttpResponseRedirect(request.META['HTTP_REFERER'])


def delete_all(request):
    try:
        order_dict.clear()
    except Exception as e:
        print(e)
        return HttpResponseRedirect(request.META['HTTP_REFERER'])
    return HttpResponseRedirect(request.META['HTTP_REFERER'])


def fill_previous_orders(orders):
    global previous_orders
    for value in order_dict.values():
        previous_orders.append(value)
