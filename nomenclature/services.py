from openpyxl.reader.excel import load_workbook
from nomenclature.models import Nomenclature


class ParseXLSXDataToNomenclatureModel:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def __parse_xlsx(self):
        wb = load_workbook(self.file_path)
        sheet = wb.active
        return sheet

    def __parse_xlsx_data(self):
        sheet = self.__parse_xlsx()
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            data.append(row)
        return data

    def __get_count_of_rows(self):
        sheet = self.__parse_xlsx()
        return sheet.max_row

    @staticmethod
    def __get_count_of_nomenclature_objects():
        return Nomenclature.objects.count()

    @staticmethod
    def create_nomenclature(data):
        for row in data:
            if Nomenclature.objects.filter(key_product=row[1]).exists():
                continue
            if not (row[1] and row[2] and row[9] and row[20]):
                continue

            nomenclature = Nomenclature(
                key_product=row[1],
                name=row[2],
                unit=row[9],
                expense_item=row[20],
            )
            nomenclature.save()

    def parse(self):
        data = self.__parse_xlsx_data()
        self.create_nomenclature(data)
        return True
