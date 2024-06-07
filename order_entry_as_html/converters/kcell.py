import dataclasses
import datetime
import enum
import io
import json
import mimetypes
import typing
from collections.abc import Iterator

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from avh_services.formatters import (
    format_count,
    format_date,
    format_number,
)
from avh_services.keys import tcp_key
from tester_atp_avr.models import TCP

from ..models import OrderRegion
from .converter import Converter
from .data import Data, TotalMixin
from .exceptions import ConvertError

TCP_FILE_NAME: typing.Final[str] = "Кселл"


@dataclasses.dataclass
class ExcelRow:
    sites: list[str]

    tcp: str
    name: str
    measuring_unit: str
    count: float

    @property
    def work_cost(self) -> int | str:
        tcp = self.tcp.split(".")
        if len(tcp) != 2:
            raise ConvertError("Пункт ТЦП неправильный")
        tcp_category_id, tcp_id = tcp
        try:
            return (
                TCP.objects.get(
                    tcp_id=tcp_id,
                    tcp_category__tcp_category_id=tcp_category_id,
                    tcp_category__tcp_file__name=TCP_FILE_NAME,
                ).price
                * self.count
            )
        except TCP.DoesNotExist:
            raise ConvertError("Пункт ТЦП не найден")


@dataclasses.dataclass
class ExcelData(Data, TotalMixin):
    contractor_agreement: str
    order_number: str
    order_date: datetime.date
    region: int
    customer: str
    contactor: str
    work_type: str
    start_work_date: datetime.date
    end_work_date: datetime.date
    rows: list[ExcelRow]
    contract_date: datetime.date | None
    critical_delinquency: datetime.timedelta | None = None
    integration_date: datetime.date | None = None
    c_activityfield: int = 1000008
    c_activitytype: int = 1

    @property
    def json(self) -> str:
        if "А-АВР-П" in self.order_number:
            self.c_activitytype = 1000014
        else:
            self.c_activitytype = 1000000
        return json.dumps(
            {
                "import_data": [
                    {
                        "BS_name": self.contractor_agreement,
                        "datesignagreement2": format_date(self.contract_date),
                        "datesignagreement3": format_date(self.order_date),
                        "agreement3": self.order_number,
                        "region": self.region,
                        "address": "",
                        "table": [
                            {
                                "number": index,
                                "C_Activityfield_ID": self.c_activityfield,
                                "C_ActivityType_ID": self.c_activitytype,
                                "value": f"кс.{tcp}",
                                "number_TCP": tcp,
                                "name": name,
                                "measuer": measuring_unit,
                                "quantity": count,
                                "date": self.completion_date.days,
                                "summ": work_cost,
                            }
                            for index, sites, tcp, name, measuring_unit, count, work_cost, row in self
                        ],
                    }
                ]
            },
            ensure_ascii=False,
        )

    @property
    def completion_date(self) -> datetime.timedelta:
        return self.end_work_date - self.start_work_date

    @property
    def total(self) -> float:
        return sum(row.work_cost for row in self.rows)

    def __iter__(self) -> Iterator[tuple[int, str, str, str, str, str, ExcelRow]]:
        for index, row in enumerate(sorted(self.rows, key=lambda x: tcp_key(x.tcp)), 1):
            try:
                work_cost = format_number(row.work_cost)
            except ConvertError as error:
                work_cost = str(error)

            yield (
                index,
                ", ".join(row.sites),
                row.tcp,
                row.name,
                (
                    row.measuring_unit.lower()
                    if row.measuring_unit == "Пролет"
                    else row.measuring_unit
                ),
                format_count(row.count),
                work_cost,
                row,
            )


class ExcelConverter(Converter):
    def to_data_list(
        self,
        bytes_io: io.BytesIO,
        /,
        *,
        filename: str | None = None,
        region: OrderRegion | None = None,
        contract_date: datetime.date | None = None,
        **kwargs,
    ) -> list[ExcelData]:
        try:
            if region is None:
                raise ConvertError("Выберете регион")
            if contract_date is None:
                raise ConvertError("Выберете дата подписание контракта")

            workbook = openpyxl.load_workbook(bytes_io, data_only=True)
            worksheet: Worksheet = workbook.active
            is_old_format = worksheet["B6"].value == "Заказ  №"

            def get_value(name: str, /) -> typing.Any:
                for name_, *_, value in worksheet.iter_rows(
                    min_col=2, min_row=8, max_row=17, max_col=5, values_only=True
                ):
                    if not name_:
                        continue
                    if " ".join(name_.strip().split()).casefold() == " ".join(name.strip().split()).casefold():
                        return value
                return None

            critical_delinquency = get_value("Критическая просрочка(дней):")
            if is_old_format:
                order_number = worksheet["C6"].value
                order_date = worksheet["F6"].value
                start_work_date = worksheet["E12"].value
                integration_date = worksheet["E13"].value
                end_work_date = worksheet["E14"].value
                if critical_delinquency:
                    critical_delinquency = datetime.timedelta(
                        days=int(critical_delinquency)
                    )
            else:
                order_number =  worksheet["E6"].value
                order_date = worksheet["H6"].value
                start_work_date = get_value("Планируемая Дата начала выполнения работ:")
                integration_date = get_value("Планируемая Дата интеграции:")
                end_work_date = get_value("Планируемая Дата окончания работ:")
                if critical_delinquency:
                    critical_delinquency = datetime.timedelta(
                        days=int(critical_delinquency.split(" ", 1)[0])
                    )
            order_date = datetime.datetime.strptime(order_date, "%d.%m.%Y")
            customer = get_value("Заказчик:")
            contactor = get_value("Подрядчик:")
            if filename and "А-АВР-П" in filename:
                work_type = "проектирование"
            else:
                work_type = "смр"
            start_work_date = datetime.datetime.strptime(start_work_date, "%d.%m.%Y")
            if integration_date:
                integration_date = datetime.datetime.strptime(
                    integration_date, "%d.%m.%Y"
                )
            end_work_date = datetime.datetime.strptime(end_work_date, "%d.%m.%Y")

            rows = []
            contractor_agreement: str | None = None
            for row in worksheet.iter_rows(min_row=18, values_only=True):
                if row[0] == "№":
                    continue
                if row[0] is None:
                    break
                try:
                    if is_old_format:
                        sites = row[1].split(",")
                        tcp = row[2].strip()
                    else:
                        sites = row[2].split(",")
                        tcp = row[1].strip()
                    if contractor_agreement is None:
                        contractor_agreement = sites[0]
                    excel_row = ExcelRow(
                        sites,
                        tcp,
                        row[3].split(" ", 1)[-1],
                        row[4].strip(),
                        float(row[5].strip()),
                    )
                    rows.append(excel_row)
                except KeyError:
                    raise ConvertError("Данные не верные")
            return [
                ExcelData(
                    rows=rows,
                    order_number=order_number,
                    order_date=order_date,
                    customer=customer,
                    contactor=contactor,
                    work_type=work_type,
                    region=region.code,
                    start_work_date=start_work_date,
                    integration_date=integration_date,
                    end_work_date=end_work_date,
                    critical_delinquency=critical_delinquency,
                    contractor_agreement=contractor_agreement,
                    contract_date=contract_date,
                )
            ]
        except ConvertError as convert_error:
            raise convert_error from None
        except Exception as exception:
            raise ConvertError(f"Неизвестная ошибка: {exception}")


converters = {
    mimetypes.types_map[".xlsx"]: ExcelConverter(
        "order_entry_as_html/order_entry_template_kcell.html"
    )
}
