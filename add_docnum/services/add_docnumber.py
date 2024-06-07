import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, FileResponse
from zoneinfo import ZoneInfo
from django.http import JsonResponse
from document_debts.models import DebtDocument
from django.db.models import Q
from io import BytesIO
import re
import pandas as pd
from django.db import transaction

def add_document_number(file):
    try:
        # Проверка расширения файла
        if not file.name.endswith('.xlsx'):
            return HttpResponse("Необходимо загрузить Excel файл (.xlsx).", status=400)

        # Загрузка файла для обработки
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        header_color = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        tocheck_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, size=11, color='FF000000')
        value_font = Font(name='Calibri', bold=False, size=11)
        thin_border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))

        # Поиск индексов нужных столбцов
        header_row = ws[1]
        index_contragent = None
        index_total = None
        index_reg_number = None
        for cell in header_row:
            if cell.value == "Контрагент":
                index_contragent = cell.column
            elif cell.value == "Итого, с учетом косвенных налогов":
                index_total = cell.column
            elif cell.value == "Регистрационный номер":
                index_reg_number = cell.column

        if index_contragent is None or index_total is None:
            return HttpResponse("Не найдены необходимые столбцы.", status=400)

        last_column = ws.max_column
        # Сопоставление и добавление номера документа
        ws.cell(row=1, column=last_column + 1, value="Отв.логист")
        ws.cell(row=1, column=last_column + 2, value="ДО")
        ws.cell(row=1, column=last_column + 3, value="Проект")
        ws.cell(row=1, column=last_column + 4, value="Дата")
        ws.cell(row=1, column=last_column + 5, value="Внесено")

        for row in range(2, ws.max_row + 1):
            contragent = ws.cell(row=row, column=index_contragent).value
            total = ws.cell(row=row, column=index_total).value
            if contragent is None or total is None:
                continue

            contragent = normalize_string(contragent)
            pattern = f'.*{contragent}.*'
            documents = DebtDocument.objects.filter(
                Q(totallineskzt=total, postavshik__iregex=pattern, actdocno=None) |
                Q(postavshik__iregex=pattern, nscheta__regex=r'[-/]\d$', actdocno=None)
            )
            # print(documents.count())
            if documents:
                count = 0
                for doc in documents:
                    total = ws.cell(row=row, column=index_total).value
                    # Проверяем наличие специфических символов в nscheta
                    if re.search(r'[-/]', doc.nscheta or ""):
                        is_composite = True
                    else:
                        is_composite = False
                    # Для обычных выписок
                    if total == doc.totallineskzt and (not is_composite):
                        otvzakup = doc.otvzakup.split()
                        project = doc.gruppa_proekrov.split()
                        cell_otvzakup = ws.cell(row=row, column=last_column + 1, value=otvzakup[0])
                        if count == 0:
                            cell_docnum = ws.cell(row=row, column=last_column + 2, value=doc.documentno)
                        else:
                            current_value = ws.cell(row=row, column=last_column + 2).value
                            new_value = f"{current_value}, {doc.documentno}"
                            cell_docnum = ws.cell(row=row, column=last_column + 2, value=new_value)
                            cell_docnum.fill = tocheck_color
                        cell_project = ws.cell(row=row, column=last_column + 3, value=project[0])
                        cell_otvzakup.font = value_font
                        cell_docnum.font = value_font
                        cell_project.font = value_font

                        count += 1
                    # Для суммированных выписок
                    elif is_composite:
                        otvzakup = doc.otvzakup.split()
                        project = doc.gruppa_proekrov.split()
                        cell_otvzakup = ws.cell(row=row, column=last_column + 1, value=otvzakup[0])

                        if ws.cell(row=row, column=last_column + 2).value is None:
                            cell_docnum = ws.cell(row=row, column=last_column + 2, value=doc.documentno)
                        else:
                            current_value = ws.cell(row=row, column=last_column + 2).value
                            new_value = f"{current_value}, {doc.documentno}"
                            cell_docnum = ws.cell(row=row, column=last_column + 2, value=new_value)

                        cell_project = ws.cell(row=row, column=last_column + 3, value=project[0])
                        cell_otvzakup.font = value_font
                        cell_docnum.font = value_font
                        cell_project.font = value_font

        # Окрашиваем первую строку в персиковый цвет как в шаблоне
        for cell in ws['1:1']:
            cell.fill = header_color
            cell.font = header_font

        # Установка высоты строк в 30 и центрирование всего текста
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 30
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # Авто-определение ширины столбцов
        for column in ws.columns:
            max_length = 8
            column_letter = get_column_letter(column[0].column)
            if column[0].column != index_total and column[0].column != index_reg_number:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value)) + 2
                    except:
                        pass

                ws.column_dimensions[column_letter].width = max_length
            else:
                ws.column_dimensions[column_letter].width = 18
                for cell in ws[column_letter]:
                    cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        mark_duplicates(ws, [last_column + 1, last_column + 2, last_column + 3])

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        wb.close()
        virtual_workbook.seek(0)

        # Отправка файла пользователю
        response = FileResponse(virtual_workbook, as_attachment=True, filename='updated_document.xlsx')
        return response

    except Exception as e:
        return HttpResponse(f"Произошла ошибка: {e}", status=500)


def normalize_string(s):
    s = s.lower()  # Привести к нижнему регистру
    s = s.replace(',', '')  # Удалить запятые
    s = s.replace('"', '')  # Удалить двойные кавычки
    s = re.sub(r"\bип\b", '', s)  # Удалить "ИП" как отдельное слово
    s = re.sub(r"\bтоо\b", '', s)
    s = re.sub(r"\btoo\b", '', s)
    s = re.sub(r"\bооо\b", '', s)
    s = re.sub(r"\bиндивидуальный предприниматель\b", '', s)
    s = re.sub(r"\bтоварищество с ограниченной ответственностью\b", '', s)
    s = re.sub(r"\bао\b", '', s)
    s = re.sub(r"шепирова с.б.", 'шепирова', s)
    s = re.sub(r"кыдыралинов a.c.", 'шепирова', s)
    s = re.sub(r"nurazia \(нуразия\)", 'nurazia', s)
    s = re.sub(r"\s*ltd.\s*", '', s)
    s = re.sub(r"\bltd\b", '', s)
    s = re.sub(r"\s+", ' ', s)  # Удалить лишние пробелы
    s = s.strip()  # Удалить пробелы в начале и конце строки
    return s


def mark_duplicates(ws, col_indices):
    # Словарь для отслеживания комбинаций значений
    value_dict = {}
    duplicate_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Синий цвет

    # Сбор всех значений по строкам
    for row in range(2, ws.max_row + 1):
        values = tuple(ws.cell(row=row, column=col).value for col in col_indices)
        if None not in values:  # Игнорировать строки с пустыми значениями
            if values in value_dict:
                value_dict[values].append(row)
            else:
                value_dict[values] = [row]

    # Применение синего цвета к дубликатам
    for rows in value_dict.values():
        if len(rows) > 1:
            for row in rows:
                ws.cell(row=row, column=ws.max_column-3).fill = duplicate_fill


def update_19_20(file):
    if not file.name.endswith('.xlsx'):
        return HttpResponse("Необходимо загрузить Excel файл (.xlsx).", status=400)

    time_zone = ZoneInfo("Asia/Almaty")
    df = pd.read_excel(file)
    df['Дата счета'] = pd.to_datetime(df['Дата счета'], errors='coerce')
    df = df[df['Дата счета'].dt.year >= 2024]
    date_columns = ['Дата ДО', 'Дата счета', 'Дата оплаты в 1С', 'Дата возврата', 'Дата документа',
                    'Дата фактической сдачи документов']
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.tz_localize(None).dt.tz_localize(time_zone)
    print(len(df))
    filtered_df = df[df.iloc[:, 12].notna()].copy()
    df = None
    print(len(filtered_df))
    for column in filtered_df.columns:
        filtered_df[column] = filtered_df[column].astype(object).where(filtered_df[column].notnull(), None)

    result_dict = {}
    for index, row in filtered_df.iterrows():
        key = row.iloc[0]
        if key in result_dict:
            result_dict[key].append(index)
        else:
            result_dict[key] = [index]
    print(len(result_dict))

    existing_documents = set(DebtDocument.objects.filter(documentno__in=list(result_dict.keys())).values_list('documentno', flat=True))
    existing_documents = set(map(int, existing_documents))
    print(len(existing_documents))
    counter = 0
    with transaction.atomic():
        for documentno, indices in result_dict.items():
            if documentno not in existing_documents:
                for index in indices:
                    row = filtered_df.loc[index]
                    DebtDocument.objects.create(
                        documentno=documentno,
                        dateinvoiced=row['Дата ДО'],
                        nscheta=row['№ Счета'],
                        datascheta=row['Дата счета'],
                        gruppa_proekrov=row['Группа проектов'],
                        otvzakup=row['Ответственный закупщик'],
                        utverditel=row['Утвердитель'],
                        too=row['Заказчик'],
                        bin=row['ИИН/БИН'],
                        postavshik=row['Поставщик'],
                        coment=row['Комментарий'],
                        valyuta=row['Валюта'],
                        totallines=row['Общая стоимость'],
                        sumpaid=row['Оплаченная сумма 1С'],
                        notpayamt1c=row['Неоплаченная сумма 1С'],
                        icname=row['Категория ДО'],
                        chname=row['Статьи доходов/расходов'],
                        doc_number=row['Заказ на продажу'],
                        bank=row['Банк'],
                        accountno=row['Расчетный счет (ИИК)'],
                        security_agreed=row['Согласование с СБ'],
                        site=row['Объект (сайт)'],
                        name=row['Номенклатура'],
                        status=row['Статус оплаты'],
                        paydate1c=row['Дата оплаты в 1С'],
                        docstatus=row['Статус ДО'],
                        dname=row['Подразделение'],
                        createdby=row['Создатель ДО'],
                        factnumdoc=row['Фактический номер договора'],
                        totallineskzt=row['Общая стоимость в KZT'],
                        payamt1ckzt=row['Оплаченная сумма(1С) в KZT'],
                        notpayamt1ckzt=row['Неоплаченная сумма(1С) в KZT'],
                        daterefund=row['Дата возврата'],
                        refundamt=row['Сумма вовзрата'],
                        refundamtkzt=row['Сумма вовзрата (KZT)'],
                        actdocno=row['Системный номер документа'],
                        docserviceact=row['Номер документа'],
                        docdate=row['Дата документа'],
                        dateprocessed=row['Дата фактической сдачи документов'],
                        quantity=row['Сумма документа'],
                        amount=row['Закрывающий документ представлен на'],
                        unclosedbalance=row['Незакрытый остаток'],
                        c_invoice_id=row['ID счёте(служебное поле)'],
                        region=row['Регион ПМ']
                     )
                    # print(f"Added new document with documentno: {documentno}")
                    counter += 1

    if counter > 0:
        return JsonResponse({"message": "База данных успешно обновлена!"}, status=200)
        print('check')
    else:
        return JsonResponse({"message": "База данных актуальна, обновление не требуется"}, status=200)
