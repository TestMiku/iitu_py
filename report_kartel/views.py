import datetime
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpRequest, HttpResponse, HttpResponseBadRequest, HttpResponseRedirect, JsonResponse, QueryDict
from django.db.models import Max, Sum, Q
from .forms import UploadFileForm
from .models import MainModel, PMModel, PlansModel, ReserveModel
import pandas as pd
from pandas import NaT
import json
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.exceptions import ObjectDoesNotExist
from django.utils.timezone import now
from datetime import timedelta
from datetime import datetime
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from django.urls import reverse
from openpyxl import Workbook
from django.http import HttpResponse
import io
import datetime


def display_data(request):
    current_user = None
    pms = []
    if request.user.is_authenticated:
        current_user = f'{request.user.first_name} {request.user.last_name}'
    else: 
        current_user = "not defined[not logged in]"

    if request.user.is_superuser or request.user.email == '20585@avh.kz' or request.user.email == '20114@avh.kz' or request.user.email == '20849@avh.kz':
        queryset = MainModel.objects.using('report_kartel_db').all()
        pms = MainModel.objects.using('report_kartel_db').values_list('pm', flat=True).distinct().order_by('pm')

        search_order = request.GET.get('order_number')
        search_activity_type = request.GET.get('activity_type')
        search_region = request.GET.get('region')
        search_onwho = request.GET.get('on_who')
        pm = request.GET.get('pm')

        # Check if reset parameter is included in the request
        reset_filter = request.GET.get('reset_filter')

        if reset_filter:  # If reset filter is requested, reset all filter criteria
            search_order = None
            search_activity_type = None
            search_region = None
            search_onwho = None
            pm = None
        else:  # Apply filter criteria if not resetting
            if search_order:

                queryset = queryset.filter(order_number__iregex=search_order)
            if search_activity_type:
                queryset = queryset.filter(activity_type__iregex=search_activity_type)
            if search_region:
                queryset = queryset.filter(region__iregex=search_region)
            if search_onwho:
                queryset = queryset.filter(on_who__iregex=search_onwho)

            if pm:
                queryset = queryset.filter(pm=pm)

    else:
        queryset = MainModel.objects.using('report_kartel_db').filter(pm=current_user)

        search_order = request.GET.get('order_number')
        search_activity_type = request.GET.get('activity_type')
        search_region = request.GET.get('region')
        search_onwho = request.GET.get('on_who')
        pm = request.GET.get('pm')

        # Check if reset parameter is included in the request
        reset_filter = request.GET.get('reset_filter')

        if reset_filter:  # If reset filter is requested, reset all filter criteria
            search_order = None
            search_activity_type = None
            search_region = None
            search_onwho = None
            pm = None
        else:  # Apply filter criteria if not resetting
            if search_order:
                queryset = queryset.filter(order_number__iregex=search_order)
            if search_activity_type:
                queryset = queryset.filter(activity_type__iregex=search_activity_type)
            if search_region:
                queryset = queryset.filter(region__iregex=search_region)
            if search_onwho:
                queryset = queryset.filter(on_who__iregex=search_onwho)
            if pm:
                queryset = queryset.filter(pm=pm)

    dropdown_options = [
        {'value': ' ', 'label': ' '},
        {'value': 'ПМ', 'label': 'ПМ'},
        {'value': 'Сметный отдел', 'label': 'Сметный отдел'},
        {'value': 'Подписан', 'label': 'Подписан'},
        {'value': 'Проектирование', 'label': 'Проектирование'},
        {'value': 'Заказчик', 'label': 'Заказчик'},
        {'value': 'Заморожен', 'label': 'Заморожен'},
        {'value': 'Аннулировать', 'label': 'Аннулировать'},
        {'value': 'Аннулирован', 'label': 'Аннулирован'},
        {'value': 'В процессе аннулирования', 'label': 'В процессе аннулирования'},
        
    ]

    dropdown_month = [
        {'value': ' ', 'label': ' '},
        {'value': 'Январь', 'label': 'Январь'},
        {'value': 'Февраль', 'label': 'Февраль'},
        {'value': 'Март', 'label': 'Март'},
        {'value': 'Апрель', 'label': 'Апрель'},
        {'value': 'Май', 'label': 'Май'},
        {'value': 'Июнь', 'label': 'Июнь'},
        {'value': 'Июль', 'label': 'Июль'},
        {'value': 'Август', 'label': 'Август'},
        {'value': 'Сентябрь', 'label': 'Сентябрь'},
        {'value': 'Октябрь', 'label': 'Октябрь'},
        {'value': 'Ноябрь', 'label': 'Ноябрь'},
        {'value': 'Декабрь', 'label': 'Декабрь'},
    ]

    dropdown_month_plan = [
        {'value': ' ', 'label': ' '},
        {'value': 'заморожен', 'label': 'заморожен'},
        {'value': 'закрыт', 'label': 'закрыт'},
        {'value': 'нет плана', 'label': 'нет плана'},
        {'value': 'аннулирован', 'label': 'аннулирован'},
        {'value': 'исключен', 'label': 'исключен'},
        {'value': 'Январь', 'label': 'Январь'},
        {'value': 'Февраль', 'label': 'Февраль'},
        {'value': 'Март', 'label': 'Март'},
        {'value': 'Апрель', 'label': 'Апрель'},
        {'value': 'Май', 'label': 'Май'},
        {'value': 'Июнь', 'label': 'Июнь'},
        {'value': 'Июль', 'label': 'Июль'},
        {'value': 'Август', 'label': 'Август'},
        {'value': 'Сентябрь', 'label': 'Сентябрь'},
        {'value': 'Октябрь', 'label': 'Октябрь'},
        {'value': 'Ноябрь', 'label': 'Ноябрь'},
        {'value': 'Декабрь', 'label': 'Декабрь'},
    ]

    dropdown_year = [ 
        {'value': ' ', 'label': ' '},
        {'value': '2024', 'label': '2024'},
        {'value': '2025', 'label': '2025'},
        {'value': '2026', 'label': '2026'},
        {'value': '2027', 'label': '2027'},
        {'value': '2028', 'label': '2028'},
        {'value': '2029', 'label': '2029'},
        {'value': '2030', 'label': '2030'},
        {'value': '2031', 'label': '2031'},
        {'value': '2032', 'label': '2032'},
        {'value': '2033', 'label': '2033'},
        {'value': '2034', 'label': '2034'},
    ]

    dropdown_year_plan = [
        {'value': ' ', 'label': ' '},
        {'value': 'заморожен', 'label': 'заморожен'},
        {'value': 'закрыт', 'label': 'закрыт'},
        {'value': 'нет плана', 'label': 'нет плана'},
        {'value': 'аннулирован', 'label': 'аннулирован'},
        {'value': 'исключен', 'label': 'исключен'},
        {'value': '2024', 'label': '2024'},
        {'value': '2025', 'label': '2025'},
        {'value': '2026', 'label': '2026'},
        {'value': '2027', 'label': '2027'},
        {'value': '2028', 'label': '2028'},
        {'value': '2029', 'label': '2029'},
        {'value': '2030', 'label': '2030'},
        {'value': '2031', 'label': '2031'},
        {'value': '2032', 'label': '2032'},
        {'value': '2033', 'label': '2033'},
        {'value': '2034', 'label': '2034'},
    ]

    dropdown_status = [
        {'value' : ' ', 'label' : ' '},
        {'value' : 'в процессе', 'label' : 'в процессе'},
        {'value' : 'выполнено', 'label' : 'выполнено'},
        {'value' : 'не начато', 'label' : 'не начато'},
        {'value' : 'аннулировать', 'label' : 'аннулировать'},
        {'value' : 'аннулирован', 'label' : 'аннулирован'},
        {'value' : 'заморожен', 'label' : 'заморожен'},
    ]

    dropdown_sign_status = [
        {'value' : ' ', 'label' : ' '},
        {'value' : 'отправлено', 'label' : 'отправлено'},
        {'value' : 'не отправлено', 'label' : 'не отправлено'},

    ]
    # Number of objects to display per page
    items_per_page = 50

    paginator = Paginator(queryset, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        queryset = paginator.get_page(page_number)
    except PageNotAnInteger:
        queryset = paginator.get_page(1)
    except EmptyPage:
        queryset = paginator.get_page(paginator.num_pages)

    if request.user.first_name == "Алёна" and request.user.last_name == "Носикова" or request.user.first_name == "Анастасия" and request.user.last_name == "Пастухова":
        sum_vat = MainModel.objects.using('report_kartel_db').aggregate(Sum('order_sum_vat'))['order_sum_vat__sum']
        account_sum = MainModel.objects.using('report_kartel_db').aggregate(Sum('sum_by_invoice'))['sum_by_invoice__sum']
        avans_sum = MainModel.objects.using('report_kartel_db').aggregate(Sum('avans_sum'))['avans_sum__sum']
    else: 
        sum_vat = MainModel.objects.using('report_kartel_db').filter(pm=current_user).aggregate(Sum('order_sum_vat'))['order_sum_vat__sum']
        account_sum = MainModel.objects.using('report_kartel_db').filter(pm=current_user).aggregate(Sum('sum_by_invoice'))['sum_by_invoice__sum']
        avans_sum = MainModel.objects.using('report_kartel_db').filter(pm=current_user).aggregate(Sum('avans_sum'))['avans_sum__sum']

    total_sum_vat = 0
    total_account_sum = 0
    total_avans_sum = 0

    if sum_vat:
        format_sum_vat = "{:,.2f}".format(round(float(sum_vat), 2)).replace(',', ' ')
        total_sum_vat = format_sum_vat.replace('.', ',')
    if account_sum: 
        format_account_sum = "{:,.2f}".format(round(float(account_sum), 2)).replace(',', ' ')
        total_account_sum = format_account_sum.replace('.', ',')    
    if avans_sum:
        format_avans_sum = "{:,.2f}".format(round(float(avans_sum), 2)).replace(',', ' ')
        total_avans_sum = format_avans_sum.replace('.', ',')

    context = {
        'data': queryset,
        'pms': pms,
        'dropdown_options': dropdown_options, 
        'dropdown_month': dropdown_month, 
        'dropdown_year': dropdown_year,
        'dropdown_status': dropdown_status,
        'dropdown_sign_status': dropdown_sign_status,
        'dropdown_year_plan' : dropdown_year_plan,
        'dropdown_month_plan' : dropdown_month_plan,
        'total_sum_vat': total_sum_vat, 
        'total_account_sum': total_account_sum,
        'total_avans_sum': total_avans_sum,
    }
    

    return render(request, 'report_kartel/display_table.html', context)


def upload_data(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            process_excel_file(excel_file)
            return HttpResponseRedirect('/report-kartel/display/') 
    else:
        form = UploadFileForm()
    return render(request, r'report_kartel/upload_file.html', {'form': form})

def upload_google(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            process_google_file(excel_file)
            return HttpResponseRedirect('/report-kartel/display/') 
    else:
        form = UploadFileForm()
    return render(request, r'report_kartel/upload_google.html', {'form': form})


def new_test(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            reserve_db()
            excel_to_db(excel_file)
            update_main_with_reserve()


            return HttpResponseRedirect('/report-kartel/display/')
    else:
        form = UploadFileForm()
    return render(request, r'report_kartel/test.html', {'form': form})



def process_excel_file(excel_file):
    df = pd.read_excel(excel_file, skiprows=[1], dtype=str, keep_default_na=False)
    df.replace({pd.NaT: None})
    df.fillna('', inplace=True)

    max_id = MainModel.objects.using('report_kartel_db').all().aggregate(Max('id'))['id__max']
    starting_id = 1

    pm_data = PMModel.objects.using('report_kartel_db').all()

    for index, row in df.iterrows():
        order_number = row['Номер заказа']
        region = row['Регион']
        activity_field = row['Сфера деятельности']
        activity_type = row['Вид деятельности']
        project_type = row['Тип проекта(работ)']
        customer = row['Покупатель']
        plan_month = row['План на месяц']
        plan_year = row['План на месяц (ГОД)']
        
        if row['Статус согласования'] == 'исключен' or row['Статус согласования'] == 'Исключен':
            plan_month = 'исключен'
            plan_year = 'исключен' 

        if row['Статус согласования'] == 'аннулирован' or row['Статус согласования'] == 'Аннулирован':
            plan_month = 'аннулирован'
            plan_year = 'аннулирован'        
        
        if row['Вид деятельности'] == 'Транспортировка' or row['Вид деятельности'] == 'Проектирование':
            plan_month, plan_year = get_plans(order_number)
        else:
            if plan_month=='' or plan_year=='':
                plan_month = 'нет плана'
                plan_year = 'нет плана'

        if plan_month is not None and plan_month in ('Закрыт', 'Аннулирован', 'Исключен', 'Заморожен'):
            plan_month = plan_month.lower()
        else: 
            plan_month = plan_month

        if plan_year is not None:
            plan_year = plan_year.lower()

        pm_match = pm_data.filter(region=region, activity_field=activity_field, activity_type=activity_type,
                                  project_type=project_type, customer=customer).first()
        pm_value = pm_match.pm if pm_match else 'N/A'

        existing_row = MainModel.objects.using('report_kartel_db').filter(order_number=order_number, activity_type=activity_type).first()

        if existing_row:
            for field, value in row.items():
                setattr(existing_row, field, value)
            existing_row.pm = pm_value
            existing_row.plan_month = plan_month  # Обновление планов
            existing_row.plan_year = plan_year 
            existing_row.plan_month_new = plan_month
            existing_row.plan_year_new = plan_year   # Обновление планов
            existing_row.save()                     
        else:
            # Create new row
            MainModel.objects.using('report_kartel_db').create(
                order_number=order_number,
                activity_type=activity_type,
                pm=pm_value,
                plan_month = plan_month,  # Добавление планов
                plan_year = plan_year, 
                plan_month_new = plan_month,
                plan_year_new = plan_year,
                order_entered_date=row['Дата внесения заказа в систему'],
                order_entered_date_month=row['Дата внесения заказа в систему (месяц)'],
                order_entered_date_year=row['Дата внесения заказа в систему (год)'],
                podryad_transfer=row['Дата передачи в подряд'],
                podryad_transfer_month=row['Дата передачи в подряд (МЕСЯЦ)'],
                podryad_transfer_year=row['Дата передачи в подряд (ГОД)'],
                order_number_for_work=row['№ заказа на работу'],
                signed_order_date=row['Дата подписания заказа заказчиком'],
                signed_order_date_month=row['Дата подписания заказа заказчиком (МЕСЯЦ)'],
                signed_order_date_year=row['Дата подписания заказа заказчиком (ГОД)'],
                work_period_days=row['Срок выполнения работ, дней'],
                finish_date_plan=row['Дата выполнения работ по плану'],
                left_days_to_finish=row['Дней осталось до планируемой даты'],
                agreement_status=row['Статус согласования'],
                customer=row['Покупатель'],
                provider=row['Поставщик'],
                project=row['Проект'],
                agreement_attachment=row['Приложение к договору'],
                partition=row['Подразделение'],
                project_group=row['Группа проектов'],
                region=row['Регион'],
                activity_field=row['Сфера деятельности'],
                service_range=row['Номенклатура   услуг'],
                comment=row['Комментарий'],
                contract_number=row['№ Контракта (Номер ДС)'],
                contract_sign_date=row['Дата подписания контракта'],
                order_sum_vat=row['Сумма по заказу, с НДС'],
                comment_finance=row['Комментарий фин. отдела'],
                # comment_pm = row['!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'],
                account_number_avans=row['№ счета авансового платежа, с НДС'],
                date_invoice=row['Дата выставления счёта '],
                date_avans=row['Дата выставления счёта авансового платежа (МЕСЯЦ)'],
                date_avans_year=row['Дата выставления счёта авансового платежа (ГОД)'],
                avans_sum=row['Сумма по счёту авансового платежа, с НДС'],
                account_number=row['Номер счёта №'],
                invoice_faktura_number=row['№ счет фактуры 1С'],
                date_invoice_release=row['Дата выставления счёта'],
                date_invoice_release_month=row['Дата выставления счёта (МЕСЯЦ)'],
                date_invoice_release_year=row['Дата выставления счёта (ГОД)'],
                sum_by_invoice=row['Сумма по счёту'],
                work_type=row['Тип проекта(работ)'],
                # on_who = row['!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'],
                work_status=row['Статус работ'],
                work_finish_month=row['Месяц выполнения работ'],
                work_finish_year=row['Год выполнения работ'],
                # sign_status = row['!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'],
                order_sent_date=row['Дата отправки заказа заказчику'],
                invoice_payment_plan_date=row['Планируемая дата платежа счёт-фактуры  '],
                invoice_payment_real_date=row['Фактическая дата платежа счёт-фактуры'],
                date_factoring=row['Дата факторинга'],
            )

        starting_id += 1


def reserve_db():
    """ Function to copy data from MainModel to ReserveModel """
    data_main = MainModel.objects.using('report_kartel_db').all()
    
    for data in data_main:
        ReserveModel.objects.using('report_kartel_db').create(
            order_number = data.order_number,
            order_entered_date = data.order_entered_date,
            order_entered_date_month = data.order_entered_date_month,
            order_entered_date_year = data.order_entered_date_year,
            podryad_transfer = data.podryad_transfer,
            podryad_transfer_month = data.podryad_transfer_month,
            podryad_transfer_year = data.podryad_transfer_year,
            order_number_for_work = data.order_number_for_work,
            signed_order_date = data.signed_order_date,
            signed_order_date_month = data.signed_order_date_month,
            signed_order_date_year = data.signed_order_date_year,
            work_period_days = data.work_period_days,
            finish_date_plan = data.finish_date_plan,
            left_days_to_finish = data.left_days_to_finish,
            agreement_status = data.agreement_status,
            customer = data.customer,
            provider = data.provider,
            pm = data.pm,
            project = data.project,
            agreement_attachment = data.agreement_attachment,
            partition = data.partition,
            project_group = data.project_group,
            region = data.region,
            activity_field = data.activity_field,
            activity_type = data.activity_type,
            service_range = data.service_range,
            comment = data.comment,
            contract_number = data.contract_number,
            contract_sign_date = data.contract_sign_date,
            order_sum_vat = data.order_sum_vat,
            plan_month = data.plan_month_new,
            plan_year = data.plan_year_new,
            plan_month_new = data.plan_month_new,
            plan_year_new = data.plan_year_new,
            comment_finance = data.comment_finance,
            comment_pm = data.comment_pm,
            account_number_avans = data.account_number_avans,
            date_invoice = data.date_invoice,
            date_avans = data.date_avans,
            date_avans_year = data.date_avans_year,
            avans_sum = data.avans_sum,
            account_number = data.account_number,
            invoice_faktura_number = data.invoice_faktura_number,
            date_invoice_release = data.date_invoice_release,
            date_invoice_release_month = data.date_invoice_release_month,
            date_invoice_release_year = data.date_invoice_release_year,
            sum_by_invoice = data.sum_by_invoice,
            work_type = data.work_type,
            on_who = data.on_who,
            work_status = data.work_status,
            work_finish_month = data.work_finish_month,
            work_finish_year = data.work_finish_year,
            sign_status = data.sign_status,
            order_sent_date = data.order_sent_date,
            invoice_payment_plan_date = data.invoice_payment_plan_date,
            invoice_payment_real_date = data.invoice_payment_real_date,
            date_factoring = data.date_factoring,    
        )
    MainModel.objects.using('report_kartel_db').all().delete()


def excel_to_db(excel_file):
    df = pd.read_excel(excel_file, skiprows=[1], dtype=str, keep_default_na=False)
    df.replace({pd.NaT: None})
    df.fillna('', inplace=True)

    max_id = MainModel.objects.using('report_kartel_db').all().aggregate(Max('id'))['id__max']
    starting_id = 1

    pm_data = PMModel.objects.using('report_kartel_db').all()

    for index, row in df.iterrows():
        order_number = row['Номер заказа']
        region = row['Регион']
        activity_field = row['Сфера деятельности']
        activity_type = row['Вид деятельности']
        project_type = row['Тип проекта(работ)']
        customer = row['Покупатель']
        plan_month = row['План на месяц']
        plan_year = row['План на месяц (ГОД)']
        
        if row['Статус согласования'] in ('исключен', 'Исключен', 'аннулирован', 'Аннулирован'):
            plan_month = row['Статус согласования'].lower()
            plan_year = row['Статус согласования'].lower()

        if row['Вид деятельности'] in ('Транспортировка', 'Проектирование'):
            plan_month, plan_year = get_plans(order_number)
        else:
            if plan_month == '' or plan_year == '':
                plan_month = 'нет плана'
                plan_year = 'нет плана'

        if plan_month is not None and plan_month in ('Закрыт', 'Аннулирован', 'Исключен', 'Заморожен'):
            plan_month = plan_month.lower()

        if plan_year is not None:
            plan_year = plan_year.lower()

        pm_match = pm_data.filter(region=region, activity_field=activity_field, activity_type=activity_type,
                                  project_type=project_type, customer=customer).first()
        pm_value = pm_match.pm if pm_match else 'N/A'

        MainModel.objects.using('report_kartel_db').create(
            order_number=order_number,
            activity_type=activity_type,
            pm=pm_value,
            plan_month=plan_month,
            plan_year=plan_year,
            plan_month_new=plan_month,
            plan_year_new=plan_year,
            order_entered_date=row['Дата внесения заказа в систему'],
            order_entered_date_month=row['Дата внесения заказа в систему (месяц)'],
            order_entered_date_year=row['Дата внесения заказа в систему (год)'],
            podryad_transfer=row['Дата передачи в подряд'],
            podryad_transfer_month=row['Дата передачи в подряд (МЕСЯЦ)'],
            podryad_transfer_year=row['Дата передачи в подряд (ГОД)'],
            order_number_for_work=row['№ заказа на работу'],
            signed_order_date=row['Дата подписания заказа заказчиком'],
            signed_order_date_month=row['Дата подписания заказа заказчиком (МЕСЯЦ)'],
            signed_order_date_year=row['Дата подписания заказа заказчиком (ГОД)'],
            work_period_days=row['Срок выполнения работ, дней'],
            finish_date_plan=row['Дата выполнения работ по плану'],
            left_days_to_finish=row['Дней осталось до планируемой даты'],
            agreement_status=row['Статус согласования'],
            customer=row['Покупатель'],
            provider=row['Поставщик'],
            project=row['Проект'],
            agreement_attachment=row['Приложение к договору'],
            partition=row['Подразделение'],
            project_group=row['Группа проектов'],
            region=row['Регион'],
            activity_field=row['Сфера деятельности'],
            service_range=row['Номенклатура   услуг'],
            comment=row['Комментарий'],
            contract_number=row['№ Контракта (Номер ДС)'],
            contract_sign_date=row['Дата подписания контракта'],
            order_sum_vat=row['Сумма по заказу, с НДС'],
            comment_finance=row['Комментарий фин. отдела'],
            # comment_pm=row['Тип проекта(работ)'],
            account_number_avans=row['№ счета авансового платежа, с НДС'],
            date_invoice=row['Дата выставления счёта '],
            date_avans=row['Дата выставления счёта авансового платежа (МЕСЯЦ)'],
            date_avans_year=row['Дата выставления счёта авансового платежа (ГОД)'],
            avans_sum=row['Сумма по счёту авансового платежа, с НДС'],
            account_number=row['Номер счёта №'],
            invoice_faktura_number=row['№ счет фактуры 1С'],
            date_invoice_release=row['Дата выставления счёта'],
            date_invoice_release_month=row['Дата выставления счёта (МЕСЯЦ)'],
            date_invoice_release_year=row['Дата выставления счёта (ГОД)'],
            sum_by_invoice=row['Сумма по счёту'],
            work_type=row['Тип проекта(работ)'],
            work_status=row['Статус работ'],
            work_finish_month=row['Месяц выполнения работ'],
            work_finish_year=row['Год выполнения работ'],
            order_sent_date=row['Дата отправки заказа заказчику'],
            invoice_payment_plan_date=row['Планируемая дата платежа счёт-фактуры  '],
            invoice_payment_real_date=row['Фактическая дата платежа счёт-фактуры'],
            date_factoring=row['Дата факторинга'],
        )

    starting_id += 1



from django.db import transaction
from django.db.models import F
from .models import MainModel, ReserveModel

def update_main_with_reserve():
    data_reserve = ReserveModel.objects.using('report_kartel_db').all()
    batch_size = 1000  # Adjust batch size as needed
    total_records = data_reserve.count()
    processed_records = 0

    main_instances = MainModel.objects.using('report_kartel_db').all()

    with transaction.atomic():
        for main_instance in main_instances:
            for reserve_data in data_reserve:
                if (
                    reserve_data.order_number == main_instance.order_number and
                    reserve_data.activity_field == main_instance.activity_field and 
                    reserve_data.activity_type == main_instance.activity_type and 
                    reserve_data.region == main_instance.region and
                    reserve_data.pm == main_instance.pm
                ):
                    main_instance.order_number = reserve_data.order_number
                    main_instance.order_entered_date = reserve_data.order_entered_date
                    main_instance.order_entered_date_month = reserve_data.order_entered_date_month
                    main_instance.order_entered_date_year = reserve_data.order_entered_date_year
                    main_instance.podryad_transfer = reserve_data.podryad_transfer
                    main_instance.podryad_transfer_month = reserve_data.podryad_transfer_month
                    main_instance.podryad_transfer_year = reserve_data.podryad_transfer_year
                    main_instance.order_number_for_work = reserve_data.order_number_for_work
                    main_instance.signed_order_date = reserve_data.signed_order_date
                    main_instance.signed_order_date_month = reserve_data.signed_order_date_month
                    main_instance.signed_order_date_year = reserve_data.signed_order_date_year
                    main_instance.work_period_days = reserve_data.work_period_days
                    main_instance.finish_date_plan = reserve_data.finish_date_plan
                    main_instance.left_days_to_finish = reserve_data.left_days_to_finish
                    main_instance.agreement_status = reserve_data.agreement_status
                    main_instance.customer = reserve_data.customer
                    main_instance.provider = reserve_data.provider
                    main_instance.pm = reserve_data.pm
                    main_instance.project = reserve_data.project
                    main_instance.agreement_attachment = reserve_data.agreement_attachment
                    main_instance.partition = reserve_data.partition
                    main_instance.project_group = reserve_data.project_group
                    main_instance.region = reserve_data.region
                    main_instance.activity_field = reserve_data.activity_field
                    main_instance.activity_type = reserve_data.activity_type
                    main_instance.service_range = reserve_data.service_range
                    main_instance.comment = reserve_data.comment
                    main_instance.contract_number = reserve_data.contract_number
                    main_instance.contract_sign_date = reserve_data.contract_sign_date
                    main_instance.order_sum_vat = reserve_data.order_sum_vat
                    main_instance.plan_month = reserve_data.plan_month
                    main_instance.plan_year = reserve_data.plan_year
                    main_instance.plan_month_new = reserve_data.plan_month_new
                    main_instance.plan_year_new = reserve_data.plan_year_new
                    main_instance.comment_finance = reserve_data.comment_finance
                    main_instance.comment_pm = reserve_data.comment_pm
                    main_instance.account_number_avans = reserve_data.account_number_avans
                    main_instance.date_invoice = reserve_data.date_invoice
                    main_instance.date_avans = reserve_data.date_avans
                    main_instance.date_avans_year = reserve_data.date_avans_year
                    main_instance.avans_sum = reserve_data.avans_sum
                    main_instance.account_number = reserve_data.account_number
                    main_instance.invoice_faktura_number = reserve_data.invoice_faktura_number
                    main_instance.date_invoice_release = reserve_data.date_invoice_release
                    main_instance.date_invoice_release_month = reserve_data.date_invoice_release_month
                    main_instance.date_invoice_release_year = reserve_data.date_invoice_release_year
                    main_instance.sum_by_invoice = reserve_data.sum_by_invoice
                    main_instance.work_type = reserve_data.work_type
                    main_instance.on_who = reserve_data.on_who
                    main_instance.work_status = reserve_data.work_status
                    main_instance.work_finish_month = reserve_data.work_finish_month
                    main_instance.work_finish_year = reserve_data.work_finish_year
                    main_instance.sign_status = reserve_data.sign_status
                    main_instance.order_sent_date = reserve_data.order_sent_date
                    main_instance.invoice_payment_plan_date = reserve_data.invoice_payment_plan_date
                    main_instance.invoice_payment_real_date = reserve_data.invoice_payment_real_date
                    main_instance.date_factoring = reserve_data.date_factoring
                    main_instance.save()

    ReserveModel.objects.using('report_kartel_db').all().delete()
    # Optionally, you can commit the transaction here if needed


def process_google_file(excel_file):
    df = pd.read_excel(excel_file, skiprows=[1], dtype=str, keep_default_na=False)
    df.replace({pd.NaT: None})
    df.fillna('', inplace=True)

    max_id = MainModel.objects.using('report_kartel_db').all().aggregate(Max('id'))['id__max']
    starting_id = 1

    for index, row in df.iterrows():
        order_number = row['Номер заказа']
        region = row['Регион']
        activity_field = row['Сфера деятельности']
        activity_type = row['Вид деятельности']
        project_type = row['Тип проекта(работ)']
        customer = row['Покупатель']
        plan_month = row['План на месяц (новый)']
        plan_year = row['План на год (новый)']
        
        
        if row['ПМ'] == '':
            pm = 'N/A'
        else:
            pm = row['ПМ']

        if plan_month is not None and plan_month in ('Закрыт', 'Аннулирован', 'Исключен', 'Заморожен'):
            plan_month = plan_month.lower()
        else: 
            plan_month = plan_month

        if plan_year is not None:
            plan_year = plan_year.lower()

        existing_row = MainModel.objects.using('report_kartel_db').filter(order_number=order_number, activity_type=activity_type).first()

        if existing_row:
            for field, value in row.items():
                setattr(existing_row, field, value)
            existing_row.plan_month = plan_month  # Обновление планов
            existing_row.plan_year = plan_year 
            existing_row.plan_month_new = plan_month
            existing_row.plan_year_new = plan_year   # Обновление планов
            existing_row.save()                     
        else:
            # Create new row
            MainModel.objects.using('report_kartel_db').create(
                order_number=order_number,
                activity_type=activity_type,
                pm=pm,
                plan_month = plan_month,  # Добавление планов
                plan_year = plan_year, 
                plan_month_new = plan_month,
                plan_year_new = plan_year,
                order_entered_date=row['Дата внесения заказа в систему'],
                order_entered_date_month=row['Дата внесения заказа в систему (месяц)'],
                order_entered_date_year=row['Дата внесения заказа в систему (год)'],
                podryad_transfer=row['Дата передачи в подряд'],
                podryad_transfer_month=row['Дата передачи в подряд (МЕСЯЦ)'],
                podryad_transfer_year=row['Дата передачи в подряд (ГОД)'],
                order_number_for_work=row['№ заказа на работу'],
                signed_order_date=row['Дата подписания заказа заказчиком'],
                signed_order_date_month=row['Дата подписания заказа заказчиком (МЕСЯЦ)'],
                signed_order_date_year=row['Дата подписания заказа заказчиком (ГОД)'],
                work_period_days=row['Срок выполнения работ, дней'],
                finish_date_plan=row['Дата выполнения работ по плану'],
                left_days_to_finish=row['Дней осталось до планируемой даты'],
                agreement_status=row['Статус согласования'],
                customer=row['Покупатель'],
                provider=row['Поставщик'],
                project=row['Проект'],
                agreement_attachment=row['Приложение к договору'],
                partition=row['Подразделение'],
                project_group=row['Группа проектов'],
                region=row['Регион'],
                activity_field=row['Сфера деятельности'],
                service_range=row['Номенклатура   услуг'],
                comment=row['Комментарий'],
                contract_number=row['№ Контракта (Номер ДС)'],
                contract_sign_date=row['Дата подписания контракта'],
                order_sum_vat=row['Сумма по заказу, с НДС'],
                comment_finance=row['Комментарий фин. отдела'],
                comment_pm = row['Комментарий ПМ-а'],
                account_number_avans=row['№ счета авансового платежа, с НДС'],
                date_invoice=row['Дата выставления счёта '],
                date_avans=row['Дата выставления счёта авансового платежа (МЕСЯЦ)'],
                date_avans_year=row['Дата выставления счёта авансового платежа (ГОД)'],
                avans_sum=row['Сумма по счёту авансового платежа, с НДС'],
                account_number=row['Номер счёта №'],
                invoice_faktura_number=row['№ счет фактуры 1С'],
                date_invoice_release=row['Дата выставления счёта'],
                date_invoice_release_month=row['Дата выставления счёта (МЕСЯЦ)'],
                date_invoice_release_year=row['Дата выставления счёта (ГОД)'],
                sum_by_invoice=row['Сумма по счёту'],
                work_type=row['Тип проекта(работ)'],
                on_who = row['На ком висит заказ:'],
                work_status=row['Статус работ'].lower(),
                work_finish_month=row['Месяц выполнения работ'],
                work_finish_year=row['Год выполнения работ'],
                sign_status = row['Статус подписания (отправлено/не отправлено) для неподписанных заказов'].lower(),
                order_sent_date=row['Дата отправки заказа заказчику'],
                invoice_payment_plan_date=row['Планируемая дата платежа счёт-фактуры  '],
                invoice_payment_real_date=row['Фактическая дата платежа счёт-фактуры'],
                date_factoring=row['Дата факторинга'],
            )

        starting_id += 1


def get_plans(order_number):
    plan_model_data = PlansModel.objects.using('report_kartel_db').filter(order_number=order_number).first()
    if plan_model_data:
        return plan_model_data.plan_month, plan_model_data.plan_year
    else:
        return "нет плана", "нет плана"
    
@csrf_exempt
def save_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            for row_idx, row in enumerate(data):
                for col_idx, value in enumerate(row):
                    MainModel.objects.using('report_kartel_db').create(row=row_idx+1, col=col_idx+1, value=value)
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

@require_POST
def update_data(request: HttpRequest) -> JsonResponse:
    data = json.loads(request.body)
    field = data.get("field")
    value = data.get("value")
    id = data.get("modelId")
    
    months = {
            'январь': 1,
            'февраль': 2,
            'март': 3,
            'апрель': 4,
            'май': 5,
            'июнь': 6,
            'июль': 7,
            'август': 8,
            'сентябрь': 9,
            'октябрь': 10,
            'ноябрь': 11,
            'декабрь': 12,
        }


    if not field or not value or not id:
        return JsonResponse({
            "detail": "field and value and id is required"
        }, status=400)
    try:
        main_model = MainModel.objects.using('report_kartel_db').get(id=id)
    except ObjectDoesNotExist:
        return JsonResponse({
            "detail": f"main_model with id {id} does not exists"
        }, status=404)
    try:
        if field in ('plan_month_new', 'plan_year_new') and months.get(str.lower(main_model.plan_month or "")) == datetime.date.today().month:
            return JsonResponse({'detail': "Дохуя хитрый?"}, status=418)
    except TypeError:
        return JsonResponse({"detail" : "You are comparing not month type of Data"})
    setattr(main_model, field, value)
    main_model.save()
    return JsonResponse({
        "detail": f"Field {field} updated with value {value}"
    })


def update_pm(request):
    pms_with_regions = MainModel.objects.using('report_kartel_db').exclude(pm='N/A').values_list('pm', 'region').distinct().order_by('pm')
    dropdown_activity_type = MainModel.objects.using('report_kartel_db').values_list('activity_type').distinct().order_by('activity_type')
    dropdown_activity_field = MainModel.objects.using('report_kartel_db').values_list('activity_field').distinct().order_by('activity_field')
    dropdown_work_type = MainModel.objects.using('report_kartel_db').values_list('work_type').distinct().order_by('work_type')
    dropdown_customer = MainModel.objects.using('report_kartel_db').values_list('customer').distinct().order_by('customer')
    
    if request.method == 'POST':
        old_pm = request.POST.get('old_pm')
        new_pm = request.POST.get('new_pm')
        
        # Check if old_pm is 'N/A', if so, return a bad request response
        if old_pm == 'N/A':
            return HttpResponseBadRequest("Cannot change 'N/A' PMs.")
        
        # Retrieve values for all compliance fields corresponding to old_pm
        compliance_fields = MainModel.objects.using('report_kartel_db').filter(pm=old_pm).first()

        if compliance_fields is not None:
            # Unpack compliance fields
            region = compliance_fields.region
            activity_type = compliance_fields.activity_type
            activity_field = compliance_fields.activity_field
            customer = compliance_fields.customer
            work_type = compliance_fields.work_type

            # Update only the 'pm' field in rows where all compliance fields match old_pm
            MainModel.objects.using('report_kartel_db').filter(
                pm=old_pm,
                region=region,
                activity_type=activity_type,
                activity_field=activity_field,
                customer=customer,
                work_type=work_type
            ).update(pm=new_pm)
            
            return redirect('/report-kartel/display')  # Redirect to a success page or back to the main page
        else:
            return HttpResponseBadRequest("No data found for the specified PM.")


    return render(request, 'report_kartel/update_pm.html', {
        'pms_with_regions': pms_with_regions,
        'dropdown_activity_type': dropdown_activity_type,
        'dropdown_activity_field': dropdown_activity_field,
        'dropdown_work_type': dropdown_work_type,
        'dropdown_customer': dropdown_customer,
    })


def get_error(request) -> JsonResponse:
    id = request.GET.get("id")
    try:
        main_model = MainModel.objects.using('report_kartel_db').get(id=id)
    except MainModel.DoesNotExist:
        return JsonResponse({
            "detail": f"MainModel with id {id} does not exists"
        }, status=404)
    
    return JsonResponse({
        "error": main_model.error()
    })


def update_plan_month(request):
    if request.method == 'POST':
        # Получаем сегодняшнюю дату
        today = now().date()
        # Получаем дату завтрашнего дня
        tomorrow = today + timedelta(days=1)
        # Обновляем данные в соответствии с условиями
        MainModel.objects.using('report_kartel_db').filter(
            plan_month__startswith=today.strftime('%Y-%m'),
            # work_status = 'Выполнено'
        ).update(plan_month=tomorrow.strftime('%Y-%m'))

        return render(request, 'report_kartel/display_table.html', {'info': 'Планы на месяц невыполненных заказов успешно изменено на новый месяц'})
    return render(request, 'report_kartel/display_table.html')



def export_to_excel(request):
   
    order_number = request.GET.get('order_number')
    activity_type = request.GET.get('activity_type')
    region = request.GET.get('region')
    on_who = request.GET.get('on_who')
    pm = request.GET.get('pm')
    
    queryset = MainModel.objects.using('report_kartel_db').all()

    if order_number:
        queryset = queryset.filter(order_number__icontains=order_number)
    if activity_type:
        queryset = queryset.filter(activity_type__icontains=activity_type)
    if region:
        queryset = queryset.filter(region__icontains=region)
    if on_who:
        queryset = queryset.filter(on_who__icontains=on_who)
    if pm:
        if pm != ' ':
            queryset = queryset.filter(pm=pm)

    # Создаем новый workbook
    wb = Workbook()
    ws = wb.active

    # Заполняем заголовки столбцов
    headers = [
        "Ключ поиска", "Номер заказа", "Дата внесения заказа в систему", "Дата внесения заказа в систему (месяц)",
        "Дата внесения заказа в систему (год)", "Дата передачи в подряд", "Дата передачи в подряд (МЕСЯЦ)",
        "Дата передачи в подряд (ГОД)", "№ заказа на работу", "Дата подписания заказа заказчиком",
        "Дата подписания заказа заказчиком (МЕСЯЦ)", "Дата подписания заказа заказчиком (ГОД)", "Срок выполнения работ, дней",
        "Дата выполнения работ по плану", "Дней осталось до планируемой даты", "Статус согласования", "Покупатель",
        "Поставщик", "ПМ", "Проект", "Приложение к договору", "Подразделение",
        "Группа проектов", "Регион", "Сфера деятельности", "Вид деятельности", "Номенклатура   услуг",
        "Комментарий", "№ Контракта (Номер ДС)", "Дата подписания контракта", "Сумма по заказу, с НДС", "План на месяц",
        "План на месяц (ГОД)", "Комментарий фин. отдела", "Комментарий ПМ", "№ счета авансового платежа, с НДС", "Дата выставления счёта ",
        "Дата выставления счёта авансового платежа (МЕСЯЦ)", "Дата выставления счёта авансового платежа (ГОД)", "Сумма по счёту авансового платежа, с НДС", "Номер счёта №", "№ счет фактуры 1С",
        "Дата выставления счёта", "Дата выставления счёта (МЕСЯЦ)", "Дата выставления счёта (ГОД)",
        "Сумма по счёту", "Тип проекта(работ)", "Статус заказа", "Статус работ", "Месяц выполнения работ",
        "Год выполнения работ", "Дата отправки заказа заказчику", "Дата отправки заказа", "Планируемая дата платежа счёт-фактуры  ",
        "Фактическая дата платежа счёт-фактуры", "Дата факторинга"
    ]
    ws.append(headers)

    def format_date(date_str):
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            return date.strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return date_str

    def format_float(value):
        try:
            return "{:,.2f}".format(round(float(value), 2)).replace(',', '').replace('.', ',')
        except (ValueError, TypeError):
            return value

    for obj in queryset:
        row = [
            f"{obj.order_number}{obj.pm}", obj.order_number,
            format_date(obj.order_entered_date) if obj.order_entered_date else None,
            obj.order_entered_date_month, obj.order_entered_date_year,
            format_date(obj.podryad_transfer) if obj.podryad_transfer else None,
            obj.podryad_transfer_month, obj.podryad_transfer_year, obj.order_number_for_work,
            format_date(obj.signed_order_date) if obj.signed_order_date else None,
            obj.signed_order_date_month, obj.signed_order_date_year, obj.work_period_days,
            obj.finish_date_plan, obj.left_days_to_finish, obj.agreement_status, obj.customer,
            obj.provider, obj.pm, obj.project, obj.agreement_attachment, obj.partition,
            obj.project_group, obj.region, obj.activity_field, obj.activity_type, obj.service_range,
            obj.comment, obj.contract_number,
            format_date(obj.contract_sign_date) if obj.contract_sign_date else None,
            format_float(obj.order_sum_vat) if obj.order_sum_vat else None,
            obj.plan_month, obj.plan_year, obj.comment_finance, obj.comment_pm,
            obj.account_number_avans,
            format_date(obj.date_invoice) if obj.date_invoice else None,
            obj.date_avans, obj.date_avans_year,
            format_float(obj.avans_sum) if obj.avans_sum else None,
            obj.account_number, obj.invoice_faktura_number,
            format_date(obj.date_invoice_release) if obj.date_invoice_release else None,
            obj.date_invoice_release_month, obj.date_invoice_release_year,
            format_float(obj.sum_by_invoice) if obj.sum_by_invoice else None,
            obj.work_type, obj.on_who, obj.work_status, obj.work_finish_month,
            obj.work_finish_year, obj.sign_status,
            format_date(obj.order_sent_date) if obj.order_sent_date else None,
            format_date(obj.invoice_payment_plan_date) if obj.invoice_payment_plan_date else None,
            format_date(obj.invoice_payment_real_date) if obj.invoice_payment_real_date else None,
            format_date(obj.date_factoring) if obj.date_factoring else None,
        ]
        ws.append(row)

    # Создаем ответ с типом содержимого excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=7.23(filled).xlsx'

    return response


# def export_to_excel(request):
    

#     order_number = request.GET.get('order_number')
#     activity_type = request.GET.get('activity_type')
#     region = request.GET.get('region')
#     on_who = request.GET.get('on_who')
#     pm = request.GET.get('pm')
    
#     queryset = MainModel.objects.using('report_kartel_db').all()

#     if order_number:
#         queryset = queryset.filter(order_number__icontains=order_number)
#     if activity_type:
#         queryset = queryset.filter(activity_type__icontains=activity_type)
#     if region:
#         queryset = queryset.filter(region__icontains=region)
#     if on_who:
#         queryset = queryset.filter(on_who__icontains=on_who)
#     if pm:
#         if pm == ' ':
#             queryset = queryset
#         else:
#             queryset = queryset.filter(pm=pm)    
    

#     wb = load_workbook('templates/7.23 template.xlsx')
#     ws = wb.active

#     # Счетчик рядов
#     row_count = 3

#     for obj in queryset:
#         ws.insert_rows(row_count)  # вставляем новую строку
#         combined_data = f"{obj.order_number}{obj.pm}" 
#         ws.cell(row=row_count, column=1, value=combined_data)
#         ws.cell(row=row_count, column=2, value=obj.order_number)

#         if obj.order_entered_date:
#             order_entered_date = datetime.datetime.strptime(obj.order_entered_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=3, value=order_entered_date.strftime("%d.%m.%Y"))
            
#         ws.cell(row=row_count, column=4, value=obj.order_entered_date_month)
#         ws.cell(row=row_count, column=5, value=obj.order_entered_date_year)
        
#         if obj.podryad_transfer:
#             podryad_transfer = datetime.datetime.strptime(obj.podryad_transfer, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=6, value=podryad_transfer.strftime("%d.%m.%Y"))

#         ws.cell(row=row_count, column=7, value=obj.podryad_transfer_month)
#         ws.cell(row=row_count, column=8, value=obj.podryad_transfer_year)
#         ws.cell(row=row_count, column=9, value=obj.order_number_for_work)
#         # signed_order_date
#         if obj.signed_order_date:
#             signed_order_date = datetime.datetime.strptime(obj.signed_order_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=10, value=signed_order_date.strftime("%d.%m.%Y"))

#         ws.cell(row=row_count, column=11, value=obj.signed_order_date_month)
#         ws.cell(row=row_count, column=12, value=obj.signed_order_date_year)
#         ws.cell(row=row_count, column=13, value=obj.work_period_days)
#         ws.cell(row=row_count, column=14, value=obj.finish_date_plan)
#         ws.cell(row=row_count, column=15, value=obj.left_days_to_finish)
#         ws.cell(row=row_count, column=16, value=obj.agreement_status)
#         ws.cell(row=row_count, column=17, value=obj.customer)
#         ws.cell(row=row_count, column=18, value=obj.provider)
#         ws.cell(row=row_count, column=19, value=obj.pm)
#         ws.cell(row=row_count, column=20, value=obj.project)
#         ws.cell(row=row_count, column=21, value=obj.agreement_attachment)
#         ws.cell(row=row_count, column=22, value=obj.partition)
#         ws.cell(row=row_count, column=23, value=obj.project_group)
#         ws.cell(row=row_count, column=24, value=obj.region)
#         ws.cell(row=row_count, column=25, value=obj.activity_field)
#         ws.cell(row=row_count, column=26, value=obj.activity_type)
#         ws.cell(row=row_count, column=27, value=obj.service_range)
#         ws.cell(row=row_count, column=28, value=obj.comment)
#         ws.cell(row=row_count, column=29, value=obj.contract_number)

#         if obj.contract_sign_date:
#             contract_sign_date = datetime.datetime.strptime(obj.contract_sign_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=30, value=contract_sign_date.strftime("%d.%m.%Y"))

#         if obj.order_sum_vat:
#             formatted_value = "{:,.2f}".format(round(float(obj.order_sum_vat), 2)).replace(',', '')
#             order_sum_vat = formatted_value.replace('.', ',')
#             ws.cell(row=row_count, column=31, value=order_sum_vat)
#         ws.cell(row=row_count, column=32, value=obj.plan_month)
#         ws.cell(row=row_count, column=33, value=obj.plan_year)
#         ws.cell(row=row_count, column=34, value=obj.comment_finance)
#         ws.cell(row=row_count, column=35, value=obj.comment_pm)
#         ws.cell(row=row_count, column=36, value=obj.account_number_avans)
#         # date_invoice
#         if obj.date_invoice:
#             date_invoice = datetime.datetime.strptime(obj.date_invoice, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=37, value=date_invoice.strftime("%d.%m.%Y"))

#         ws.cell(row=row_count, column=38, value=obj.date_avans)
#         ws.cell(row=row_count, column=39, value=obj.date_avans_year)

#         if obj.avans_sum:
#             formatted_value1 = "{:,.2f}".format(round(float(obj.avans_sum), 2)).replace(',', '')
#             avans_sum = formatted_value1.replace('.', ',')
#             ws.cell(row=row_count, column=40, value=avans_sum)
        
#         ws.cell(row=row_count, column=41, value=obj.account_number)
#         ws.cell(row=row_count, column=42, value=obj.invoice_faktura_number)
#         # date_invoice_release
#         if obj.date_invoice_release:
#             date_invoice_release = datetime.datetime.strptime(obj.date_invoice_release, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=43, value=date_invoice_release.strftime("%d.%m.%Y"))

#         ws.cell(row=row_count, column=44, value=obj.date_invoice_release_month)
#         ws.cell(row=row_count, column=45, value=obj.date_invoice_release_year)

#         if obj.sum_by_invoice:
#             formatted_value2 = "{:,.2f}".format(round(float(obj.sum_by_invoice), 2)).replace(',', '')
#             sum_by_invoice = formatted_value2.replace('.', ',')
#             ws.cell(row=row_count, column=46, value=sum_by_invoice)

#         ws.cell(row=row_count, column=47, value=obj.work_type)
#         ws.cell(row=row_count, column=48, value=obj.on_who)
#         ws.cell(row=row_count, column=49, value=obj.work_status)
#         ws.cell(row=row_count, column=50, value=obj.work_finish_month)
#         ws.cell(row=row_count, column=51, value=obj.work_finish_year)
#         ws.cell(row=row_count, column=52, value=obj.sign_status)
#         # order_sent_date
#         if obj.order_sent_date:
#             order_sent_date = datetime.datetime.strptime(obj.order_sent_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=53, value=order_sent_date.strftime("%d.%m.%Y"))
#         # invoice_payment_plan_date
#         if obj.invoice_payment_plan_date:
#             invoice_payment_plan_date = datetime.datetime.strptime(obj.invoice_payment_plan_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=54, value=invoice_payment_plan_date.strftime("%d.%m.%Y"))
#         # invoice_payment_real_date
#         if obj.invoice_payment_real_date:
#             invoice_payment_real_date = datetime.datetime.strptime(obj.invoice_payment_real_date, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=55, value=invoice_payment_real_date.strftime("%d.%m.%Y"))
#         # date_factoring
#         if obj.date_factoring:
#             date_factoring = datetime.datetime.strptime(obj.date_factoring, "%Y-%m-%d %H:%M:%S")
#             ws.cell(row=row_count, column=56, value=date_factoring.strftime("%d.%m.%Y"))

#         # продолжаем заполнять остальные столбцы...
#         row_count += 1
#     # Add data from the database to the Excel file starting from the third row
   
#     # Save the Excel file
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=7.23(filled).xlsx'
#     wb.save(response)

#     return response


def delete_data(request):
    # Ваша логика удаления данных из базы данных
    MainModel.objects.using('report_kartel_db').all().delete()
    
    # Перенаправить пользователя на другую страницу после удаления данных
    return HttpResponseRedirect(reverse('display_data'))


def sum_itogi(request):

    months = {
        1 : 'Январь',
        2 : 'Февраль',
        3 : 'Март',
        4 : 'Апрель',
        5 : 'Май',
        6 : 'Июнь',
        7 : 'Июль',
        8 : 'Август',
        9 : 'Сентябрь',
        10 : 'Октябрь',
        11 : 'Ноябрь',
        12 : 'Декабрь',
    }
    
    current_m = datetime.datetime.now().month
    current_year = datetime.datetime.now().year

    current_month = months[current_m]
    


    sum_1 = MainModel.objects.using('report_kartel_db').filter(plan_month_new=current_month, plan_year_new=current_year).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_by_plan = "{:,.2f}".format(sum_1).replace(',', ' ').replace('.', ',')
    # Вычисляем сумму по фильтру "Дата выставления счёта (МЕСЯЦ)" - текущий месяц и "Дата выставления счёта (ГОД)"
    sum_2 = MainModel.objects.using('report_kartel_db').filter(date_invoice_release_month=current_month, date_invoice_release_year=current_year).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_by_release = "{:,.2f}".format(sum_2).replace(',', ' ').replace('.', ',')
    
    sum_3 = MainModel.objects.using('report_kartel_db').filter(
        date_invoice_release_month=current_month,
        date_invoice_release_year=current_year
    ).exclude(
        Q(comment_finance='на проверке') | Q(comment_finance='к закрытию')
    ).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_plan_not = "{:,.2f}".format(sum_3).replace(',', ' ').replace('.', ',')
    # Вычисляем сумму для sum_4
    sum_4 = MainModel.objects.using('report_kartel_db').filter(
        date_invoice_release_month=current_month,
        date_invoice_release_year=current_year
    ).filter(
        Q(comment_finance='на проверке') | Q(comment_finance='к закрытию')
    ).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_plan_in = "{:,.2f}".format(sum_4).replace(',', ' ').replace('.', ',')

    sum_result = sum_3 + sum_4
    sum_overall = "{:,.2f}".format(sum_result).replace(',', ' ').replace('.', ',')
    
    if sum_4 != 0:
        sum_percent = sum_3 / sum_4
    else:
        sum_percent = 0 
    sum_percentage = "{:,.2f}".format(sum_percent).replace(',', ' ').replace('.' , ',')

    #_____________

    previous_month = months[current_m - 1]

     
    

    sum_1_prev = MainModel.objects.using('report_kartel_db').filter(plan_month_new=previous_month, plan_year_new=current_year).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_by_plan_prev = "{:,.2f}".format(sum_1_prev).replace(',', ' ').replace('.', ',')
    # Вычисляем сумму по фильтру "Дата выставления счёта (МЕСЯЦ)" - текущий месяц и "Дата выставления счёта (ГОД)"
    sum_2_prev = MainModel.objects.using('report_kartel_db').filter(date_invoice_release_month=previous_month, date_invoice_release_year=current_year).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_by_release_prev = "{:,.2f}".format(sum_2_prev).replace(',', ' ').replace('.', ',')
    
    sum_3_prev = MainModel.objects.using('report_kartel_db').filter(
        date_invoice_release_month=previous_month,
        date_invoice_release_year=current_year
    ).exclude(
        Q(comment_finance='на проверке') | Q(comment_finance='к закрытию')
    ).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_plan_not_prev = "{:,.2f}".format(sum_3_prev).replace(',', ' ').replace('.', ',')
    # Вычисляем сумму для sum_4
    sum_4_prev = MainModel.objects.using('report_kartel_db').filter(
        date_invoice_release_month=previous_month,
        date_invoice_release_year=current_year
    ).filter(
        Q(comment_finance='на проверке') | Q(comment_finance='к закрытию')
    ).aggregate(total_sum=Sum('order_sum_vat'))['total_sum'] or 0
    sum_plan_in_prev = "{:,.2f}".format(sum_4_prev).replace(',', ' ').replace('.', ',')

    sum_result_prev = sum_3_prev + sum_4_prev
    sum_overall_prev = "{:,.2f}".format(sum_result_prev).replace(',', ' ').replace('.', ',')
    
    if sum_4_prev != 0:
        sum_percent_prev = sum_3_prev / sum_4_prev
    else:
        sum_percent_prev = 0 
    sum_percentage_prev = "{:,.2f}".format(sum_percent_prev).replace(',', ' ').replace('.' , ',')



    context = {
        'current_month' : current_month, 
        'sum_by_plan': sum_by_plan,
        'sum_by_release': sum_by_release,
        'sum_plan_not': sum_plan_not,
        'sum_plan_in': sum_plan_in,
        'sum_overall': sum_overall,  # Сумма столбцов 3 и 4
        'sum_percentage': sum_percentage,
          
        'previous_month': previous_month, 
        'sum_by_plan_prev': sum_by_plan_prev,
        'sum_by_release_prev': sum_by_release_prev,
        'sum_plan_not_prev': sum_plan_not_prev,
        'sum_plan_in_prev': sum_plan_in_prev,
        'sum_overall_prev': sum_overall_prev,  # Сумма столбцов 3 и 4
        'sum_percentage_prev': sum_percentage_prev,
    }

    
    return render(request, 'report_kartel/sum_for_report.html', context)