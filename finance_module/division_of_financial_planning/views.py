import datetime
import decimal
import functools
import io
import itertools
import operator
import typing

import django.core.mail
import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from django.db import transaction
from django.db.models import Case, F, Model, Q, QuerySet, Sum, Value, When
from django.db.models.fields import DecimalField
from django.db.models.functions import Cast, Coalesce, Concat, TruncDate
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.http import FileResponse, HttpRequest
from django.utils import timezone
from django.views import generic
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response

from finance_module.services import common_service
from finance_module.services.smart_table_service import smart_table

from .. import models, serializers
from ..services import mandatory_payments_service
from .models import Cell, Row, Table

RESPONSIBLE_SEND_TEST_EMAIL: typing.Final[str | None] = "22293@avh.kz"  # None
BANK_TELECOM_EMAIL: typing.Final[str] = (
    "22293@avh.kz"  # Банк-Телеком <bank-telecom@avh.kz>
)
CONFIRMER_EMAIL: typing.Final[str] = "22293@avh.kz"  # Улар Токтар <21241@avh.kz>


class ConfirmationTemplateView(generic.TemplateView):
    template_name = "finance_module/division_of_financial_planning/confirmation.html"


class MonthlyPaymentsTemplateView(generic.TemplateView):
    template_name = (
        "finance_module/division_of_financial_planning/monthly_payments.html"
    )

    def get_context_data(self, **kwargs: typing.Any) -> dict[str, typing.Any]:
        mandatory_payments = list(models.MandatoryPayment.objects.all())
        project_regions = list(mandatory_payments_service.get_ordered_project_regions())
        categories, categories_colspan = mandatory_payments_service.get_categories(
            mandatory_payments
        )
        mandatory_payments_service.annotate_mandatory_payments_with_levels(
            mandatory_payments, categories
        )
        directors = []
        for key, group in itertools.groupby(
            project_regions, key=operator.attrgetter("director")
        ):
            group = list(group)
            directors.append(
                (
                    key,
                    (len(list(group)) - 1),
                    sum(project_region.percent for project_region in group),
                )
            )

        for mandatory_payment in mandatory_payments:
            payments = []
            for project_region in project_regions:
                payment = mandatory_payments_service.get_payment(
                    mandatory_payment=mandatory_payment,
                    project_region=project_region,
                    exclude_statuses=[
                        models.rejected_mandatory_payment_seizure_status(),
                        models.default_mandatory_payment_seizure_status(),
                    ],
                )
                setattr(
                    project_region,
                    "total_sum",
                    getattr(project_region, "total_sum", decimal.Decimal())
                    + payment.sum,
                )
                setattr(
                    mandatory_payment,
                    "total_sum",
                    getattr(mandatory_payment, "total_sum", decimal.Decimal())
                    + payment.sum,
                )
                payments.append(payment)
            mandatory_payment.payments = payments

        managers = [
            (key, len(list(group)))
            for key, group in itertools.groupby(
                project_regions, key=operator.attrgetter("manager")
            )
        ]
        mandatory_payments_total_sum = sum(
            getattr(mandatory_payment, "total_sum", decimal.Decimal())
            for mandatory_payment in mandatory_payments
        )
        disbursements_total_sum = sum(
            mandatory_payment.get_disbursements()
            for mandatory_payment in mandatory_payments
        )
        return super().get_context_data(
            disbursements_total_sum=disbursements_total_sum,
            mandatory_payments_total_sum=mandatory_payments_total_sum,
            categories_colspan=categories_colspan,
            directors=directors,
            managers=managers,
            mandatory_payments=mandatory_payments,
            project_regions=project_regions,
            **kwargs,
        )


class DailyTemplateView(generic.TemplateView):
    template_name = "finance_module/division_of_financial_planning/daily.html"


class TransfersTemplateView(generic.TemplateView):
    template_name = "finance_module/division_of_financial_planning/transfers.html"


@api_view(["GET"])
def get_project_regions(request: Request) -> Response:
    project_regions = models.ProjectRegion.objects.all()
    data = {"projectRegions": []}

    for project_region in project_regions:
        row = {
            "name": project_region.name,
            "projectManager": (
                project_region.project_manager.get_full_name()
                if project_region.project_manager
                else project_region.project_manager_display
            ),
            "director": (
                project_region.director.get_full_name()
                if project_region.director
                else project_region.director_display
            ),
        }
        mandatory_payment_seizures = (
            models.MandatoryPaymentSeizure.objects.exclude(
                status__in=[
                    models.default_mandatory_payment_seizure_status(),
                    models.rejected_mandatory_payment_seizure_status(),
                ]
            )
            .annotate(datetime_date=TruncDate("datetime"))
            .filter(datetime_date=timezone.localdate(), project_region=project_region)
        )
        mandatory_payment_seizures_sum = mandatory_payment_seizures.aggregate(
            mandatory_payment_seizures_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["mandatory_payment_seizures_sum"]
        inflows_sum = project_region.inflows.filter(
            date=timezone.localdate(), reserve_percent=10.7
        ).aggregate(inflows_sum=Coalesce(Sum("remainder"), decimal.Decimal()))[
            "inflows_sum"
        ]
        row["inflowsSum"] = inflows_sum
        row["mandatoryPaymentSeizuresSum"] = mandatory_payment_seizures_sum
        if inflows_sum:
            data["projectRegions"].append(row)
    return Response(data)


_Show: typing.TypeAlias = typing.Literal[
    "per-day", "per-month", "per-year", "during-period"
]


class AccountTableQuerySets(typing.NamedTuple):
    rows: QuerySet
    all_cells: QuerySet
    cells: QuerySet
    all_mandatory_payment_seizures: QuerySet
    mandatory_payment_seizures: QuerySet


def _get_account_table_querysets(
    show: _Show,
    date: str | datetime.date | None = None,
    month: str | None = None,
    year: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
):
    table = Table.objects.get_or_create(name="Расчётные счета")[0]

    all_cells, cells = daily_filter(
        Cell.objects.all(),
        "date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    all_mandatory_payment_seizures, mandatory_payment_seizures = daily_filter(
        models.MandatoryPaymentSeizure.objects.exclude(
            status__in=[
                models.default_mandatory_payment_seizure_status(),
                models.rejected_mandatory_payment_seizure_status(),
            ]
        ).annotate(datetime_date=TruncDate("datetime")),
        "datetime_date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    rows_q = daily_rows_q(
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )

    return AccountTableQuerySets(
        daily_filter_with_rows_q(Row.objects.filter(table=table), rows_q),
        all_cells,
        cells,
        all_mandatory_payment_seizures,
        mandatory_payment_seizures,
    )


def _get_account_table_row(
    account_row: Row,
    querysets: AccountTableQuerySets,
    cash_register_subdivision: bool = True,
) -> dict[str, typing.Any]:
    row = {
        "id": account_row.id,
        "name": account_row.name,
        "is_cash_register": False,
        "user_added": bool(
            account_row.properties and account_row.properties.get("user_added")
        ),
        "editable": True,
        "subdivision": None,
        "reserve": 0,
        "remainder": 0,
    }
    today = timezone.localdate()
    current_cells = querysets.cells.filter(row=account_row)
    account = None
    subdivision = None
    try:
        account = models.Account.objects.get(name=account_row.name)
    except models.Account.DoesNotExist:
        row["mandatory_payment_seizures_sum"] = (
            current_cells.filter(
                field="mandatory_payment_seizures_sum",
                type=Cell.Type.NUMBER,
                date=today,
            )
            .annotate(
                mandatory_payment_seizures_sum=Cast(
                    "value", DecimalField(max_digits=25, decimal_places=10)
                )
            )
            .aggregate(
                total_sum=Coalesce(
                    Sum("mandatory_payment_seizures_sum"), decimal.Decimal()
                )
            )["total_sum"]
        )
        row["remainder"] = (
            current_cells.filter(field="remainder", type=Cell.Type.NUMBER, date=today)
            .annotate(
                remainder=Cast("value", DecimalField(max_digits=25, decimal_places=10))
            )
            .aggregate(
                total_remainder_sum=Coalesce(Sum("remainder"), decimal.Decimal())
            )["total_remainder_sum"]
        )
    else:
        is_cash_register = account.is_cash_register
        row["is_cash_register"] = is_cash_register
        subdivision = (
            "Касса"
            if is_cash_register and cash_register_subdivision
            else account.subdivision.name
            if account.subdivision
            else None
        )
        mandatory_payment_seizures_sum = (
            querysets.mandatory_payment_seizures.filter(account=account).aggregate(
                mandatory_payment_seizures_sum=Coalesce(Sum("sum"), decimal.Decimal())
            )["mandatory_payment_seizures_sum"]
            + current_cells.filter(
                field="mandatory_payment_seizures_sum", type=Cell.Type.NUMBER
            )
            .annotate(
                mandatory_payment_seizures_sum=Cast(
                    "value", DecimalField(max_digits=25, decimal_places=10)
                )
            )
            .aggregate(
                total_sum=Coalesce(
                    Sum("mandatory_payment_seizures_sum"), decimal.Decimal()
                )
            )["total_sum"]
        )
        row["subdivision"] = subdivision
        row["mandatory_payment_seizures_sum"] = mandatory_payment_seizures_sum
        if subdivision == "74п":
            account_row.cells.get_or_create(
                date=today,
                field="remainder",
                type=Cell.Type.NUMBER,
                defaults={"value": common_service.get_account_balance(account)},
            )
            account_row.cells.get_or_create(
                date=today,
                field="reserve",
                type=Cell.Type.NUMBER,
                defaults={"value": common_service.get_account_balance(account)},
            )
        if subdivision == "7п":
            row["remainder"] = common_service.get_account_balance(
                account, irrelevant=True
            )
            row["reserve"] = account.inflows.aggregate(
                total_reserve_sum=Coalesce(Sum("reserve"), decimal.Decimal())
            )["total_reserve_sum"]
            row["editable"] = False
        else:
            row["remainder"] = (
                current_cells.filter(
                    field="remainder", type=Cell.Type.NUMBER, date=today
                )
                .annotate(
                    remainder=Cast(
                        "value", DecimalField(max_digits=25, decimal_places=10)
                    )
                )
                .aggregate(
                    total_remainder_sum=Coalesce(Sum("remainder"), decimal.Decimal())
                )["total_remainder_sum"]
            )
            row["reserve"] = (
                current_cells.filter(field="reserve", type=Cell.Type.NUMBER, date=today)
                .annotate(
                    reserve=Cast(
                        "value", DecimalField(max_digits=25, decimal_places=10)
                    )
                )
                .aggregate(
                    total_reserve_sum=Coalesce(Sum("reserve"), decimal.Decimal())
                )["total_reserve_sum"]
            )
    return row


@api_view(["GET"])
def get_accounts_table(request: Request) -> Response:
    data = {"rows": []}
    querysets = _get_account_table_querysets("per-day", date=timezone.localdate())
    for account_row in querysets.rows:
        data["rows"].append(_get_account_table_row(account_row, querysets))
    return Response(data)


class DailyTableQuerySets(typing.NamedTuple):
    rows: QuerySet
    cells: QuerySet
    all_cells: QuerySet
    mandatory_payment_seizures: QuerySet
    all_mandatory_payment_seizures: QuerySet
    administrative_transfers: QuerySet
    all_administrative_transfers: QuerySet
    inflows: QuerySet
    all_inflows: QuerySet


def daily_rows_q(
    show: _Show,
    date: str | datetime.date | None = None,
    month: str | None = None,
    year: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> Q | None:
    rows_q = None
    match show:
        case "per-day":
            if isinstance(date, str):
                try:
                    date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
                except ValueError:
                    return None
            elif isinstance(date, datetime.date):
                pass
            else:
                return None
            rows_q = (Q(created_date__isnull=True) | Q(created_date__lte=date)) & (
                Q(deleted_date__isnull=True) | Q(deleted_date__gt=date)
            )
        case "per-month":
            year, month = tuple(map(int, month.split("-")))
            rows_q = (
                Q(created_date__isnull=True)
                | Q(created_date__year__lte=year, created_date__month__lte=month)
            ) & (
                Q(deleted_date__isnull=True)
                | Q(deleted_date__year__gt=year, deleted_date__month__gt=month)
            )
        case "per-year":
            year = int(year)
            rows_q = (
                Q(created_date__isnull=True) | Q(created_date__year__lte=year)
            ) & (Q(deleted_date__isnull=True) | Q(deleted_date__year__gt=year))
        case "during-period":
            start_date = (
                start_date and datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            )
            end_date = (
                end_date and datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            )
            if start_date:
                rows_q = (
                    Q(created_date__isnull=True) | Q(created_date__lte=start_date)
                ) & (Q(deleted_date__isnull=True) | Q(deleted_date__gt=start_date))
    return rows_q


def daily_filter_with_rows_q(queryset: QuerySet, rows_q: Q | None = None) -> QuerySet:
    return queryset.filter(
        (rows_q if rows_q else Q(deleted_date__isnull=True)),
    )


def daily_filter(
    queryset: QuerySet,
    field: str,
    /,
    *,
    show: _Show,
    date: str | datetime.date | None = None,
    month: str | None = None,
    year: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> tuple[QuerySet, QuerySet]:
    rest = current = queryset
    match show:
        case "per-day":
            if isinstance(date, str):
                try:
                    date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
                except ValueError:
                    return None
            elif isinstance(date, datetime.date):
                pass
            else:
                return None
            current = current.filter(**{field: date})
            rest = rest.filter(**{f"{field}__lt": date})
        case "per-month":
            year, month = tuple(map(int, month.split("-")))
            current = current.filter(
                **{f"{field}__year": year, f"{field}__month": month}
            )
            rest = rest.filter(
                **{f"{field}__year__lte": year, f"{field}__month__lt": month}
            )
        case "per-year":
            year = int(year)
            current = current.filter(**{f"{field}__year": year})
            rest = rest.filter(**{f"{field}__year__lt": year})
        case "during-period":
            start_date = (
                start_date and datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            )
            end_date = (
                end_date and datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            )
            if start_date:
                current = current.filter(**{f"{field}__gte": start_date})
                rest = rest.filter(**{f"{field}__lt": start_date})
            else:
                rest = rest.none()

            if end_date:
                current = current.filter(**{f"{field}__lte": end_date})
    return [rest, current]


def _get_daily_table() -> Table:
    return Table.objects.get_or_create(name="Ежедневно")[0]


def _get_daily_table_querysets(
    *,
    show: _Show,
    date: str | datetime.date | None = None,
    month: str | None = None,
    year: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> DailyTableQuerySets | None:
    table = _get_daily_table()
    all_mandatory_payment_seizures = mandatory_payment_seizures = ()
    all_administrative_transfers = administrative_transfers = (
        models.AdministrativeTransfer.objects.exclude(
            status=models.rejected_administrative_transfer_status()
        )
    )
    rows_q = daily_rows_q(
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    all_mandatory_payment_seizures, mandatory_payment_seizures = daily_filter(
        models.MandatoryPaymentSeizure.objects.exclude(
            Q(
                status__in=[
                    models.rejected_mandatory_payment_seizure_status(),
                    models.default_mandatory_payment_seizure_status(),
                ]
            )
            | Q(imported_from_file=True)
        ).annotate(datetime_date=TruncDate("datetime")),
        "datetime_date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    all_administrative_transfers, administrative_transfers = daily_filter(
        models.AdministrativeTransfer.objects.exclude(
            status=models.rejected_administrative_transfer_status()
        ).annotate(created_at_date=TruncDate("created_at")),
        "created_at_date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    all_cells, cells = daily_filter(
        Cell.objects.all(),
        "date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    all_inflows, inflows = daily_filter(
        models.Inflow.objects.filter(reserve_percent=decimal.Decimal("10.7")),
        "date",
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )
    return DailyTableQuerySets(
        daily_filter_with_rows_q(Row.objects.filter(table=table), rows_q),
        cells,
        all_cells,
        mandatory_payment_seizures,
        all_mandatory_payment_seizures,
        administrative_transfers,
        all_administrative_transfers,
        inflows,
        all_inflows,
    )


def _get_expense_cells(queryset: QuerySet) -> QuerySet:
    return queryset.filter(field="expense", type=Cell.Type.NUMBER).annotate(
        expense=Cast("value", DecimalField(max_digits=25, decimal_places=10))
    )


def _get_daily_table_row(
    daily_row: Row, querysets: DailyTableQuerySets, show: _Show
) -> dict[str, typing.Any]:
    ULAR_LOGIC: bool = True
    row = {
        "id": daily_row.id,
        "name": daily_row.name,
        "user_added": bool(
            daily_row.properties and daily_row.properties.get("user_added")
        ),
        "created_date": daily_row.created_date,
    }

    current_cells = querysets.cells.filter(row=daily_row)
    current_all_cells = querysets.all_cells.filter(row=daily_row)
    exclude_expense_cell_ids = set()

    def get_total_debt_sum(
        debt_cells: QuerySet, expense_cells: QuerySet
    ) -> decimal.Decimal:
        total_debt = decimal.Decimal()
        for debt_cell in debt_cells.filter(
            field="debt", type=Cell.Type.NUMBER
        ).annotate(debt=Cast("value", DecimalField(max_digits=25, decimal_places=10))):
            debt = debt_cell.debt
            if not debt:
                continue
            for expense_cell in expense_cells.exclude(
                Q(id__in=exclude_expense_cell_ids) & ~Q(date=debt_cell.date)
            ).filter(date__gte=debt_cell.date):
                if debt < 0:
                    break
                debt -= expense_cell.expense
                exclude_expense_cell_ids.add(expense_cell.id)
            if debt > 0:
                total_debt += debt
        return total_debt

    if show == "per-day":
        comment_cell = current_cells.filter(
            field="comment", type=Cell.Type.STRING
        ).first()
        row["comment"] = comment_cell and comment_cell.value
        row["debt"] = (
            current_cells.filter(field="debt", type=Cell.Type.NUMBER)
            .annotate(
                debt=Cast("value", DecimalField(max_digits=25, decimal_places=10))
            )
            .aggregate(total_debt_sum=Coalesce(Sum("debt"), decimal.Decimal()))[
                "total_debt_sum"
            ]
        )
    else:
        row["total_debt"] = get_total_debt_sum(
            current_cells, _get_expense_cells(current_cells)
        )
    row["expense"] = (
        current_cells.filter(field="expense", type=Cell.Type.NUMBER)
        .annotate(expense=Cast("value", DecimalField(max_digits=25, decimal_places=10)))
        .aggregate(total_expense_sum=Coalesce(Sum("expense"), decimal.Decimal()))[
            "total_expense_sum"
        ]
    )

    total_current_parish_cells_sum = (
        current_cells.filter(field="parish", type=Cell.Type.NUMBER)
        .annotate(parish=Cast("value", DecimalField(max_digits=25, decimal_places=10)))
        .aggregate(total_parish_sum=Coalesce(Sum("parish"), decimal.Decimal()))[
            "total_parish_sum"
        ]
    )

    def get_parish(
        mandatory_payment_seizures: QuerySet, administrative_transfers: QuerySet
    ) -> decimal.Decimal:
        return (
            mandatory_payment_seizures.aggregate(
                total_parish_sum=Coalesce(Sum("sum"), decimal.Decimal())
            )["total_parish_sum"]
            + total_current_parish_cells_sum
            + administrative_transfers.aggregate(
                total_parish_sum=Coalesce(Sum("sum"), decimal.Decimal())
            )["total_parish_sum"]
        )

    reserve_percent = None
    match daily_row.properties:
        case {"reserve_percent": reserve_percent}:
            pass

    def get_reserve_sum(queryset: QuerySet) -> decimal.Decimal:
        if not reserve_percent:
            return decimal.Decimal()
        return queryset.annotate(
            reserve_=Cast(
                F("sum") * reserve_percent / 100,
                DecimalField(max_digits=25, decimal_places=10),
            )
        ).aggregate(reserve_sum=Coalesce(Sum("reserve_"), decimal.Decimal()))[
            "reserve_sum"
        ]

    row["parish"] = decimal.Decimal()
    if reserve_percent and querysets.inflows:
        row["subrows"] = []
        for project_region in models.ProjectRegion.objects.filter(
            id__in=querysets.inflows.values_list("project_region", flat=True).distinct()
        ):
            parish = get_reserve_sum(
                querysets.inflows.filter(project_region=project_region)
            )
            row["parish"] += parish
            row["subrows"].append({"name": project_region.name, "parish": parish})

    match daily_row.properties:
        case {"include": list(include)}:
            row["subrows"] = []
            for name in include:
                parish = get_parish(
                    querysets.mandatory_payment_seizures.filter(
                        mandatory_payment__name=name
                    ),
                    querysets.administrative_transfers.filter(name=name),
                )
                row["parish"] += parish
                row["subrows"].append({"name": name, "parish": parish})
            current_mandatory_payment_seizures = (
                querysets.mandatory_payment_seizures.filter(
                    mandatory_payment__name__in=include
                )
            )
            current_all_mandatory_payment_seizures = (
                querysets.all_mandatory_payment_seizures.filter(
                    mandatory_payment__name__in=include
                )
            )
            current_administrative_transfers = (
                querysets.administrative_transfers.filter(name__in=include)
            )
            current_all_administrative_transfers = (
                querysets.all_administrative_transfers.filter(name__in=include)
            )
        case _:
            current_mandatory_payment_seizures = (
                querysets.mandatory_payment_seizures.filter(
                    mandatory_payment__name=daily_row.name
                )
            )
            current_all_mandatory_payment_seizures = (
                querysets.all_mandatory_payment_seizures.filter(
                    mandatory_payment__name=daily_row.name
                )
            )
            current_administrative_transfers = (
                querysets.administrative_transfers.filter(name=daily_row.name)
            )
            current_all_administrative_transfers = (
                querysets.all_administrative_transfers.filter(name=daily_row.name)
            )
            row["parish"] += get_parish(
                current_mandatory_payment_seizures, current_administrative_transfers
            )
    current_all_expense_cells = _get_expense_cells(current_all_cells)
    if ULAR_LOGIC:
        row["parish"] = total_current_parish_cells_sum

    row["previous_debts"] = get_total_debt_sum(
        current_all_cells, current_all_expense_cells
    )

    total_incoming_balance_sum = (
        Cell.objects.filter(row=daily_row, date=daily_row.created_date)
        .filter(field="incoming_balance", type=Cell.Type.NUMBER)
        .annotate(
            incoming_balance=Cast(
                "value", DecimalField(max_digits=25, decimal_places=10)
            )
        )
        .aggregate(
            total_incoming_balance_sum=Coalesce(
                Sum("incoming_balance"), decimal.Decimal()
            )
        )["total_incoming_balance_sum"]
    )
    total_current_all_mandatory_payment_seizures_sum = (
        current_all_mandatory_payment_seizures.aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
    )
    total_current_all_parish_cells_sum = (
        current_all_cells.filter(field="parish", type=Cell.Type.NUMBER)
        .annotate(parish=Cast("value", DecimalField(max_digits=25, decimal_places=10)))
        .aggregate(total_parish_sum=Coalesce(Sum("parish"), decimal.Decimal()))[
            "total_parish_sum"
        ]
    )
    total_expense_sum = current_all_expense_cells.aggregate(
        total_expense_sum=Coalesce(Sum("expense"), decimal.Decimal())
    )["total_expense_sum"]
    total_current_all_administrative_transfers_sum = (
        current_all_administrative_transfers.aggregate(
            total_sum=Coalesce(Sum("sum"), decimal.Decimal())
        )["total_sum"]
    )
    total_reserve_sum = get_reserve_sum(querysets.all_inflows)
    row["incoming_balance"] = (
        total_incoming_balance_sum
        + total_current_all_mandatory_payment_seizures_sum
        + total_current_all_parish_cells_sum
        + total_reserve_sum
        + total_current_all_administrative_transfers_sum
        - total_expense_sum
    )
    if ULAR_LOGIC:
        row["incoming_balance"] = (
            total_incoming_balance_sum
            + total_current_all_parish_cells_sum
            - total_expense_sum
        )
    return row


@api_view(["GET"])
def get_daily_table(request: Request) -> Response:
    show = request.query_params.get("show")
    date = request.query_params.get("date")
    year = request.query_params.get("year")
    month = request.query_params.get("month")
    start_date = request.query_params.get("start-date")
    end_date = request.query_params.get("end-date")
    # table = _get_daily_table()
    # Row.objects.get_or_create(
    #     table=table, name="6% резерв", defaults={"properties": {"reserve_percent": 6.0}}
    # )
    # Row.objects.get_or_create(
    #     table=table, name="2% резерв", defaults={"properties": {"reserve_percent": 2.0}}
    # )
    # Row.objects.get_or_create(
    #     table=table,
    #     name="0,7% резерв (Холдинг)",
    #     defaults={"properties": {"reserve_percent": 0.7}},
    # )
    # Row.objects.get_or_create(
    #     table=table,
    #     name="2% резерв (Холдинг склад)",
    #     defaults={"properties": {"reserve_percent": 2.0}},
    # )
    querysets = _get_daily_table_querysets(
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )

    data = {"rows": []}
    for daily_row in querysets.rows:
        data["rows"].append(_get_daily_table_row(daily_row, querysets, show))
    return Response(data)


def _insert_daily_day_table(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    date: datetime.date,
    current: bool = False,
) -> int:
    import openpyxl
    import openpyxl.styles
    import openpyxl.styles.alignment

    daily_table_querysets = _get_daily_table_querysets(
        show="per-day",
        date=date,
    )
    accounts_table_querysets = _get_account_table_querysets("per-day", date=date)
    daily_table_rows_count = daily_table_querysets.rows.count()
    accounts_table_rows_count = accounts_table_querysets.rows.count()
    rows_count = max(daily_table_rows_count, accounts_table_rows_count)
    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=rows_count + row,
        end_column=1,
    )

    def calibri(font_size: int, /, *, bold: bool = False) -> openpyxl.styles.fonts.Font:
        return openpyxl.styles.fonts.Font(
            "Calibri",
            sz=font_size,
            b=bold,
        )

    def border() -> openpyxl.styles.borders.Border:
        return openpyxl.styles.borders.Border(
            left=openpyxl.styles.borders.Side(style="medium"),
            right=openpyxl.styles.borders.Side(style="medium"),
            top=openpyxl.styles.borders.Side(style="medium"),
            bottom=openpyxl.styles.borders.Side(style="medium"),
        )

    cell = worksheet.cell(row, 1, date)
    cell.font = calibri(22, bold=True)
    cell.alignment = openpyxl.styles.alignment.Alignment(
        horizontal="center", vertical="center", text_rotation=90
    )
    sorted_accounts_table_rows = sorted(
        (
            _get_account_table_row(
                account_row, accounts_table_querysets, cash_register_subdivision=False
            )
            for account_row in accounts_table_querysets.rows
        ),
        key=lambda account_row: account_row["subdivision"] or "",
        reverse=True,
    )
    accounts_table_rows = []
    for key, group in itertools.groupby(
        sorted_accounts_table_rows, lambda account_row: account_row["subdivision"]
    ):
        group = list(group)
        accounts_table_rows.append([key, len(group), group[0]])
        accounts_table_rows.extend(
            [key, None, account_row] for account_row in group[1:]
        )

    for index, (account, daily_row) in enumerate(
        itertools.zip_longest(
            accounts_table_rows,
            (
                _get_daily_table_row(daily_row, daily_table_querysets, "per-day")
                for daily_row in daily_table_querysets.rows
            ),
        ),
        row,
    ):
        subdivision = row_span = account_row = None
        if account:
            subdivision, row_span, account_row = account
        cell = worksheet.cell(index, 1)
        cell.border = border()
        cell = worksheet.cell(index, 2, date)
        cell.font = calibri(8)
        cell.alignment = openpyxl.styles.alignment.Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = border()
        cell = worksheet.cell(index, 3)
        cell.font = calibri(8)
        cell.alignment = openpyxl.styles.alignment.Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = border()
        if subdivision and row_span:
            cell.value = subdivision
            worksheet.merge_cells(
                start_row=index,
                start_column=3,
                end_row=index + row_span - 1,
                end_column=3,
            )
        cell = worksheet.cell(index, 4, account_row and account_row["name"])
        cell.font = calibri(10)
        cell.border = border()
        cell.alignment = openpyxl.styles.alignment.Alignment(wrap_text=True)

        cell = worksheet.cell(
            index,
            5,
            account_row and float(account_row["mandatory_payment_seizures_sum"]),
        )
        cell.font = calibri(11)
        cell.border = border()
        cell.number_format = "# ##0"
        cell = worksheet.cell(
            index,
            6,
            account_row
            and float(
                account_row["remainder"] - account_row["mandatory_payment_seizures_sum"]
            ),
        )
        cell.font = calibri(11)
        cell.border = border()
        cell.number_format = "# ##0"

        cell = worksheet.cell(index, 7, daily_row and daily_row["name"])
        cell.font = calibri(10)
        cell.border = border()
        cell.alignment = openpyxl.styles.alignment.Alignment(wrap_text=True)
        cell = worksheet.cell(index, 8, daily_row and float(daily_row["expense"]))
        cell.font = calibri(11)
        cell.border = border()
        cell.number_format = "# ##0"
        cell = worksheet.cell(index, 9, daily_row and daily_row["comment"])
        cell.font = calibri(9)
        cell.border = border()
        cell = worksheet.cell(index, 10, daily_row and float(daily_row["parish"]))
        cell.font = calibri(11)
        cell.border = border()
        cell.number_format = "# ##0"
    index += 1
    cell = worksheet.cell(index, 1)
    cell.border = border()
    cell = worksheet.cell(index, 2, date)
    cell.font = calibri(8)
    cell.alignment = openpyxl.styles.alignment.Alignment(
        horizontal="center", vertical="center"
    )
    cell.border = border()
    cell = worksheet.cell(index, 3, "Итого")
    cell.font = calibri(10, bold=True)
    cell.alignment = openpyxl.styles.alignment.Alignment(
        horizontal="center", vertical="center"
    )
    cell.border = border()
    cell = worksheet.cell(index, 4)
    cell.font = calibri(10)
    cell.border = border()
    cell.alignment = openpyxl.styles.alignment.Alignment(wrap_text=True)
    cell = worksheet.cell(index, 5, 0)
    cell.font = calibri(11)
    cell.border = border()
    cell.number_format = "# ##0"
    cell = worksheet.cell(index, 6, 0)
    cell.font = calibri(11)
    cell.border = border()
    cell.number_format = "# ##0"
    cell = worksheet.cell(index, 7)
    cell.font = calibri(10)
    cell.border = border()
    cell.alignment = openpyxl.styles.alignment.Alignment(wrap_text=True)
    cell = worksheet.cell(index, 8, 0)
    cell.font = calibri(11)
    cell.border = border()
    cell.number_format = "# ##0"
    cell = worksheet.cell(index, 9)
    cell.font = calibri(9)
    cell.border = border()
    cell = worksheet.cell(index, 10, 0)
    cell.font = calibri(11)
    cell.border = border()
    cell.number_format = "# ##0"
    return rows_count + 1


def unload_daily_table(request: HttpRequest) -> FileResponse:
    show = request.GET.get("show")
    date = request.GET.get("date")
    year = request.GET.get("year")
    month = request.GET.get("month")
    start_date = request.GET.get("start-date")
    end_date = request.GET.get("end-date")

    querysets = _get_daily_table_querysets(
        show=show,
        date=date,
        month=month,
        year=year,
        start_date=start_date,
        end_date=end_date,
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Ежедневно"
    worksheet.column_dimensions["A"].width = 6.45
    worksheet.column_dimensions["B"].width = 8.55
    worksheet.column_dimensions["D"].width = 25.82
    worksheet.column_dimensions["C"].width = 17.55
    worksheet.column_dimensions["F"].width = 17.36
    worksheet.column_dimensions["G"].width = 38.55
    worksheet.column_dimensions["H"].width = 12.91
    worksheet.column_dimensions["I"].width = 66.91
    worksheet.column_dimensions["J"].width = 12.55
    row = 2
    row += _insert_daily_day_table(worksheet, row, timezone.localdate(), current=True)
    bytes_io = io.BytesIO()
    openpyxl.styles.DEFAULT_FONT.name = "Calibri"
    workbook.save(bytes_io)
    bytes_io.seek(0)
    return FileResponse(bytes_io, as_attachment=True, filename="daily-table.xlsx")


def table_required():
    def decorator(function):
        @functools.wraps(function)
        def wrapper(*args, **kwargs):
            table_name = kwargs.get("table_name")
            try:
                table = Table.objects.get(name=table_name)
            except Table.DoesNotExist:
                return Response(
                    {"detail": f"Table with name {table_name} does not exists"},
                    status=404,
                )
            return function(*args, table=table, **kwargs)

        return wrapper

    return decorator


def default_date():
    def decorator(function):
        @functools.wraps(function)
        def wrapper(*args, **kwargs):
            kwargs.setdefault("date", timezone.localdate())
            return function(*args, **kwargs)

        return wrapper

    return decorator


def row_required():
    def decorator(function):
        @default_date()
        @table_required()
        @functools.wraps(function)
        def wrapper(*args, **kwargs):
            table: Table = kwargs.get("table")
            date: datetime.date = kwargs.get("date")
            row_id: int = kwargs.get("row_id")
            try:
                row = Row.objects.filter(
                    Q(deleted_date__isnull=True) | Q(deleted_date__gt=date),
                    Q(created_date__isnull=True) | Q(created_date__lte=date),
                ).get(table=table, id=row_id)
            except Row.DoesNotExist:
                return Response(
                    {"detail": f'Daily row with id "{row_id}" does not exists'},
                    status=404,
                )
            return function(*args, row=row, **kwargs)

        return wrapper

    return decorator


@api_view(["POST"])
@row_required()
def cells(request: Request, row: Row, date: datetime.date, **kwargs) -> Response:
    field = request.data.get("field")
    type = request.data.get("type")
    value = request.data.get("value")
    assert (
        field != "incoming_balance"
        or type != Cell.Type.NUMBER
        or date == row.created_date
    ), "The incoming balance can only be changed on the date the row was created"

    Cell.objects.update_or_create(
        row=row, date=date, field=field, type=type, defaults={"value": value}
    )
    return Response(
        {"detail": f'Field "{field}" with type "${type}" updated with value "${value}"'}
    )


@api_view(["POST"])
@default_date()
@table_required()
def rows(
    request: Request, table: Table, date: datetime.date | None = None, **kwargs
) -> Response:
    name = request.data.get("name")
    if not name:
        return Response({"detail": "Name is required"}, status=400)
    if request.method == "POST":
        row, created = Row.objects.filter(
            Q(deleted_date__isnull=True) | Q(deleted_date__gt=date),
            Q(created_date__isnull=True) | Q(created_date__lte=date),
        ).get_or_create(
            table=table,
            name=name,
            defaults={
                "created_date": date,
                "properties": {
                    "user_added": True,
                },
            },
        )

        if created:
            return Response({"detail": "Created", "id": row.id}, 201)
        return Response({"detail": f"Daily row with name {name} exists"}, status=409)


@api_view(["DELETE", "PATCH"])
@row_required()
def row(request: Request, row: Row, date: datetime.date, **kwargs):
    if request.method == "DELETE":
        row.deleted_date = date
        row.save()
        return Response({"detail": "Deleted"})
    elif request.method == "PATCH":
        name = request.data.get("name")
        if name:
            row.name = name
        row.save()
        return Response({"detail": "Updated"})


@smart_table(
    "mandatory-payment-seizures",
    models.MandatoryPaymentSeizure,
    serializers.MandatoryPaymentSeizureSerializer,
)
def mandatory_payment_seizures_smart_table(
    request: HttpRequest, queryset: QuerySet, field: str | None = None
) -> QuerySet:
    return queryset.exclude(imported_from_file=True).annotate(
        can_be_confirmed=Q(status=models.default_mandatory_payment_seizure_status())
    )


@smart_table(
    "administrative-transfers",
    models.AdministrativeTransfer,
    serializers.AdministrativeTransferSerializer,
)
def administrative_transfers_smart_table(
    request: HttpRequest, queryset: QuerySet, field: str | None = None
) -> QuerySet:
    return queryset


@smart_table(
    "chsi-groups",
    models.CHSIGroup,
    serializers.CHSIGroupSerializer,
)
def chsi_groups_smart_table(
    request: HttpRequest, queryset: QuerySet, field: str | None = None
) -> QuerySet:
    return queryset


@smart_table(
    "raschetnye",
    models.Raschetnye,
    serializers.RaschetnyeSerializer,
)
def raschetnye_smart_table(
    request: HttpRequest, queryset: QuerySet, field: str | None = None
) -> QuerySet:
    return queryset


@smart_table(
    "transfers",
    models.Transfer,
    serializers.TransferSerializer,
)
def transfers_smart_table(
    request: HttpRequest, queryset: QuerySet, field: str | None = None
) -> QuerySet:
    return queryset.annotate(
        responsible_name=Case(
            When(
                Q(responsible__isnull=False),
                then=Concat(
                    "responsible__first_name", Value(" "), "responsible__last_name"
                ),
            )
        )
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def confirmation(request: Request) -> Response:
    model: type[Model] = getattr(models, request.data.get("model"), None)
    if not model:
        return Response({}, status=404)

    for id in request.data.getlist("id"):
        instance = model.objects.get(id=id)
        confirmation = request.data.get(f"confirmation-{id}")
        if confirmation == "reject":
            reject_comment = request.data.get(f"reject-comment-{id}")
            instance.reject(responsible=request.user, reject_comment=reject_comment)
        else:
            instance.send_for_second_confirmation(responsible=request.user)
    return Response({"detail": "OK"})


@permission_classes([IsAuthenticated])
@api_view(["POST"])
def create_administrative_transfer(request: Request) -> Response:
    account_id = request.data["account-id"]
    project_region_id = request.data["project-region-id"]
    name = request.data["name"]
    note = request.data["note"]
    sum = request.data["sum"]
    try:
        account = models.Account.objects.get(id=account_id)
        project_region = models.ProjectRegion.objects.get(id=project_region_id)
        models.AdministrativeTransfer.objects.create(
            name=name,
            sum=sum,
            account=account,
            project_region=project_region,
            note=note or None,
            responsible=request.user,
        )
    except Exception as exception:
        return Response({"detail": str(exception)}, 400)

    return Response({"detail": "OK"})


@permission_classes([IsAuthenticated])
@api_view(["POST"])
def create_transfers(request: Request) -> Response:
    from_account_id = request.data["from-account-id"]
    from_whom_id = request.data["from-whom-id"]
    to_account_id = request.data["to-account-id"]
    to_whom_id = request.data["to-whom-id"]
    sum = request.data["sum"]
    try:
        from_account = models.Account.objects.get(id=from_account_id)
        from_whom = models.ProjectRegion.objects.get(id=from_whom_id)
        to_account = models.Account.objects.get(id=to_account_id)
        to_whom = models.ProjectRegion.objects.get(id=to_whom_id)
        models.Transfer.objects.create(
            sum=sum,
            from_account=from_account,
            from_whom=from_whom,
            to_account=to_account,
            to_whom=to_whom,
            responsible=request.user,
        )
    except Exception as exception:
        return Response({"detail": str(exception)}, 400)

    return Response({"detail": "OK"})


@permission_classes([IsAuthenticated])
@api_view(["POST"])
def create_raschetnye(request: Request) -> Response:
    account_id = request.data["account-id"]
    project_region_id = request.data["project-region-id"]
    name = request.data["name"]
    layoff_date = datetime.datetime.strptime(request.data["layoff-date"], "%Y-%m")
    raschetnye_by_1c = request.data["raschetnye-by-1c"]
    subreport = request.data["subreport"]
    percent_15 = request.data["percent-15"]
    try:
        account = models.Account.objects.get(id=account_id)
        project_region = models.ProjectRegion.objects.get(id=project_region_id)
        raschetnye = models.Raschetnye.objects.create(
            account=account,
            project_region=project_region,
            name=name,
            layoff_date=layoff_date,
            raschetnye_by_1c=raschetnye_by_1c,
            subreport=subreport,
            percent_15=percent_15,
            responsible=request.user,
        )
        models.AdministrativeTransfer.objects.create(
            account=account,
            project_region=project_region,
            name="20% за кэш",
            sum=percent_15,
            note="Расчётные",
            responsible=request.user,
            raschetnye=raschetnye,
        )
    except Exception as exception:
        return Response({"detail": str(exception)}, 400)
    return Response({"detail": "OK"})


@permission_classes([IsAuthenticated])
@api_view(["POST"])
def mail_raschetnye(request: Request) -> Response:
    MAIL = "22293@avh.kz"

    raschetnye = models.Raschetnye.objects.annotate(
        date=TruncDate("created_at")
    ).filter(date=timezone.localdate())
    email_message = django.core.mail.EmailMessage(
        subject="Расчётные",
        body="SWIFT",
        to=[MAIL],
    )
    email_message.send()
    return Response({"detail": f"Отправлен на почту {MAIL}"})


@receiver(post_save, sender=Cell)
def _(sender: type[Cell], instance: Cell, **kwargs) -> None:
    from finance_module.models import Debt

    if (
        instance.field == "expense"
        and instance.type == "number"
        and instance.row.name == "20% за кэш"
    ):
        Debt.objects.update_or_create(
            from_whom="Касса р/с", to_whom="Касса р/с", defaults={"sum": instance.value}
        )

    if (
        instance.field in ["remainder", "mandatory_payment_seizures_sum"]
        and instance.type == "number"
        and instance.row.name in ["касса АДМ 7п", "касса LVE офис (АДМ)"]
    ):
        today = timezone.localdate()
        a = (
            Cell.objects.filter(
                row__name="касса АДМ 7п",
                row__table__name="Расчётные счета",
                date=today,
                field="remainder",
                type=Cell.Type.NUMBER,
            )
            .annotate(
                remainder=Cast("value", DecimalField(max_digits=25, decimal_places=10))
            )
            .aggregate(
                total_remainder_sum=Coalesce(Sum("remainder"), decimal.Decimal())
            )["total_remainder_sum"]
        )
        b = (
            Cell.objects.filter(
                row__name="касса АДМ 7п",
                row__table__name="Расчётные счета",
                date=today,
                field="mandatory_payment_seizures_sum",
                type=Cell.Type.NUMBER,
            )
            .annotate(
                mandatory_payment_seizures_sum=Cast(
                    "value", DecimalField(max_digits=25, decimal_places=10)
                )
            )
            .aggregate(
                total_mandatory_payment_seizures_sum=Coalesce(
                    Sum("mandatory_payment_seizures_sum"), decimal.Decimal()
                )
            )["total_mandatory_payment_seizures_sum"]
        )
        c = (
            Cell.objects.filter(
                row__name="касса LVE офис (АДМ)",
                row__table__name="Расчётные счета",
                date=today,
                field="remainder",
                type=Cell.Type.NUMBER,
            )
            .annotate(
                remainder=Cast("value", DecimalField(max_digits=25, decimal_places=10))
            )
            .aggregate(
                total_remainder_sum=Coalesce(Sum("remainder"), decimal.Decimal())
            )["total_remainder_sum"]
        )
        d = (
            Cell.objects.filter(
                row__name="касса LVE офис (АДМ)",
                row__table__name="Расчётные счета",
                date=today,
                field="mandatory_payment_seizures_sum",
                type=Cell.Type.NUMBER,
            )
            .annotate(
                mandatory_payment_seizures_sum=Cast(
                    "value", DecimalField(max_digits=25, decimal_places=10)
                )
            )
            .aggregate(
                total_mandatory_payment_seizures_sum=Coalesce(
                    Sum("mandatory_payment_seizures_sum"), decimal.Decimal()
                )
            )["total_mandatory_payment_seizures_sum"]
        )
        Debt.objects.update_or_create(
            from_whom="Касса наличные",
            to_whom="Касса наличные",
            defaults={"sum": (c - d) + (a - b)},
        )
    if (
        (instance.field) in ["expense", "incoming_balance", "parish"]
        and instance.type == Cell.Type.NUMBER
        and instance.row.name
        in [
            "Резерв 2% от приходов",
            "Резерв 2% от приходов от 03.01.2024 на депозите",
            "Резерв 2% от приходов от 03.01.2024",
        ]
    ):
        querysets = _get_daily_table_querysets(
            show="per-day", date=timezone.localdate()
        )
        sum = decimal.Decimal()
        try:
            daily_row = Row.objects.get(
                table__name="Ежедневно", name="Резерв 2% от приходов"
            )
        except Row.DoesNotExist:
            pass
        else:
            row = _get_daily_table_row(daily_row, querysets, "per-day")
            sum += row["incoming_balance"] + row["parish"] - row["expense"]
        try:
            daily_row = Row.objects.get(
                table__name="Ежедневно",
                name="Резерв 2% от приходов от 03.01.2024 на депозите",
            )
        except Row.DoesNotExist:
            pass
        else:
            row = _get_daily_table_row(daily_row, querysets, "per-day")
            sum += row["incoming_balance"] + row["parish"] - row["expense"]
        try:
            daily_row = Row.objects.get(
                table__name="Ежедневно", name="Резерв 2% от приходов от 03.01.2024"
            )
        except Row.DoesNotExist:
            pass
        else:
            row = _get_daily_table_row(daily_row, querysets, "per-day")
            sum += row["incoming_balance"] + row["parish"] - row["expense"]
        Debt.objects.update_or_create(
            from_whom="Резерв",
            to_whom="Резерв",
            defaults={"sum": sum},
        )


@transaction.atomic
@api_view(["POST"])
def pay_chsi(request: Request) -> Response:
    rows = request.data.get("rows", [])
    project_region_id = request.data.get("projectRegionId")
    project_region = models.ProjectRegion.objects.get(id=project_region_id)
    date = request.data.get("date")
    if date:
        date = datetime.datetime.fromisoformat(date).date()
    account_id = request.data.get("accountId")
    account = models.Account.objects.get(id=account_id)
    chsi_group = models.CHSIGroup(
        account=account, project_region=project_region, date=date
    )
    chsi_instances = []
    for row in rows:
        executive_order_receipt_date = row["executiveOrderReceiptDate"]
        if executive_order_receipt_date:
            executive_order_receipt_date = datetime.datetime.fromisoformat(
                executive_order_receipt_date
            ).date()
        chsi_instances.append(
            models.CHSI(
                group=chsi_group,
                llc=row["llc"],
                recipient=row["recipient"],
                bin_or_iin=row["binOrIin"],
                iik=row["iik"],
                executive_inscription=row["executiveInscription"],
                retention_type=row["retentionType"],
                collaborator=row["collaborator"],
                iin=row["iin"],
                actual_retention_rate=row["actualRetentionRate"],
                sum=row["sum"],
                executive_order_receipt_date=executive_order_receipt_date,
            )
        )
    chsi_group.save()
    models.CHSI.objects.bulk_create(chsi_instances)
    return Response({"detail": "OK"})
