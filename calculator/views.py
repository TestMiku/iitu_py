import datetime
import functools
import math
import json
import traceback
import re
from typing import ParamSpec, Tuple, TypeVar, Final
import collections.abc
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import numbers
import gspread
import pandas as pd
from django.core.cache import cache
from django.db.models import Case, When, F, Value, FloatField, Sum, Count, IntegerField
from django.db.models.functions import Coalesce
from django.utils.timezone import now
from django.db import transaction, connection
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.core import serializers


from calculator.models import COrder, COrderPosition, COrderPositionPrice, HeaderPosition, Data2122, Data7112
from main.models import AvhUser
from reporter.models import Report
from .forms import DeleteOptionsForm, UploadFileForm
from calculator_emr.models import dataBS, data1922, dataTCP

from portal_avh.celery import app
# ?user_email=<?= $USER->GetEmail() ?>


def main(request):
    is_order_number = None
    data = {"table_show": False}

    user_email = request.GET.get("user_email")
    header = request.GET.get("header_page")
    bx = request.GET.get("bx")
    bs_name = request.GET.get("bs_name").strip() if request.GET.get("bs_name") else request.GET.get("bs_name")
    order_number = request.GET.get("order-number")
    if bs_name:
        is_order_number = re.match(r'П-\d{5,6}-\d{2}', bs_name)
    if user_email:
        try:
            user = AvhUser.objects.get(email=user_email)
            if user:
                data["user"] = user
                if header and not bx:
                    context = {
                        "data": data,
                        "data_BS": dataBS.objects.values('project').distinct()
                    }
                    if is_order_number:
                        context["show_smr"] = True
                        return render(request, "mp/calculator/page_for_headers.html", context)
                    elif order_number:
                        filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                        filtered_data_by_order = dataBS.objects.filter(order_number=order_number)
                        balance = filtered_data_by_order[0].order_sem_with_nds * 0.6
                        queryset = data1922.objects.filter(order_number__contains=order_number)
                        payamt1c_sum = queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] if \
                            queryset.aggregate(Sum('payamt1c'))[
                                'payamt1c__sum'] else 0
                        notpayamt1c_sum = queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] if \
                            queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] else 0
                        balance -= (payamt1c_sum + notpayamt1c_sum)

                        context["order_numer_"] = order_number
                        context["choised_order"] = filtered_data_by_order
                        context["balance"] = str(float('{:.2f}'.format(balance)))
                        context["amount_diff"] = str(float('{:.2f}'.format(payamt1c_sum + notpayamt1c_sum)))
                        context["dataToParse"] = serializers.serialize("json", (filtered_data_by_order))
                        context["data_tcp"] = dataTCP.objects.all()
                        context["data_BS"] = dataBS.objects.all()
                        context["filtered_data_bs"] = filtered_data_by_bs
                    elif bs_name:
                        filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                        context["filtered_data_bs"] = filtered_data_by_bs
                    return render(request, "mp/calculator/page_for_headers.html", context)
                elif bx:
                    context = {
                        "data": data,
                        "data_BS": dataBS.objects.values('project').distinct()
                    }
                    if is_order_number:
                        context["show_smr"] = True
                        return render(request, "mp/calculator/page_for_headers.html", context)
                    elif order_number:
                        filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                        filtered_data_by_order = dataBS.objects.filter(order_number=order_number)
                        balance = filtered_data_by_order[0].order_sem_with_nds * 0.6
                        queryset = data1922.objects.filter(order_number__contains=order_number)
                        payamt1c_sum = queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] if \
                            queryset.aggregate(Sum('payamt1c'))[
                                'payamt1c__sum'] else 0
                        notpayamt1c_sum = queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] if \
                            queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] else 0
                        balance -= (payamt1c_sum + notpayamt1c_sum)

                        context["order_numer_"] = order_number
                        context["choised_order"] = filtered_data_by_order
                        context["balance"] = str(float('{:.2f}'.format(balance)))
                        context["amount_diff"] = str(float('{:.2f}'.format(payamt1c_sum + notpayamt1c_sum)))
                        context["dataToParse"] = serializers.serialize("json", (filtered_data_by_order))
                        context["data_tcp"] = dataTCP.objects.all()
                        context["data_BS"] = dataBS.objects.all()
                        context["filtered_data_bs"] = filtered_data_by_bs
                    elif bs_name:
                        filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                        context["filtered_data_bs"] = filtered_data_by_bs
                    return render(request, "mp/calculator/page_for_headers_bx.html", context)
                return render(request, "mp/calculator/main_bx.html", data)
            # else:
            #     data["base_html_exist"] = True
            #     data[
            #         "message"
            #     ] = "Скорее всего у вас нет доступа к порталу <a href='http://portal.avh.kz/'>portal.avh.kz</a>. Пожалуйста, обратитесь к администратору (22050@avh.kz)."
            #     data["error_code"] = 403
            #     data["error_name"] = "Нет доступа"
            #     return render(request, "mp/calculator/404_out.html", data)
        except:
            if bx:
                return render(request, "mp/calculator/page_for_headers_bx.html", data)
            data["base_html_exist"] = True
            data[
                "message"
            ] = "Скорее всего у вас нет доступа к порталу <a href='http://portal.avh.kz/' target='_blank'>portal.avh.kz</a>. Пожалуйста, обратитесь к администратору (22050@avh.kz)."
            data["error_code"] = 403
            data["error_name"] = "Нет доступа"
            return render(request, "mp/calculator/404_out.html", data)
            # return render(request, "mp/calculator/404.html", data)
    elif request.user.is_authenticated:
        data["user"] = request.user
        if header and not bx:
            context = {
                "data": data,
                "data_BS": dataBS.objects.values('project').distinct()
            }
            if is_order_number:
                context["show_smr"] = True
                return render(request, "mp/calculator/page_for_headers.html", context)
            elif order_number:
                filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                filtered_data_by_order = dataBS.objects.filter(order_number=order_number)
                balance = filtered_data_by_order[0].order_sem_with_nds * 0.6
                queryset = data1922.objects.filter(order_number__contains=order_number)
                payamt1c_sum = queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] if \
                queryset.aggregate(Sum('payamt1c'))[
                    'payamt1c__sum'] else 0
                notpayamt1c_sum = queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] if \
                    queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] else 0
                balance -= (payamt1c_sum + notpayamt1c_sum)

                context["order_numer_"] = order_number
                context["choised_order"] = filtered_data_by_order
                context["balance"] = str(float('{:.2f}'.format(balance)))
                context["amount_diff"] = str(float('{:.2f}'.format(payamt1c_sum + notpayamt1c_sum)))
                context["dataToParse"] = serializers.serialize("json", (filtered_data_by_order))
                context["data_tcp"] = dataTCP.objects.all()
                context["data_BS"] = dataBS.objects.all()
                context["filtered_data_bs"] = filtered_data_by_bs
            elif bs_name:
                filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                context["filtered_data_bs"] = filtered_data_by_bs
            return render(request, "mp/calculator/page_for_headers.html", context)
        elif bx:
            context = {
                "data": data,
                "data_BS": dataBS.objects.values('project').distinct()
            }
            if is_order_number:
                context["show_smr"] = True
                return render(request, "mp/calculator/page_for_headers.html", context)
            elif order_number:
                filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                filtered_data_by_order = dataBS.objects.filter(order_number=order_number)
                balance = filtered_data_by_order[0].order_sem_with_nds * 0.6
                queryset = data1922.objects.filter(order_number__contains=order_number)
                payamt1c_sum = queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] if \
                queryset.aggregate(Sum('payamt1c'))[
                    'payamt1c__sum'] else 0
                notpayamt1c_sum = queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] if \
                    queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] else 0
                balance -= (payamt1c_sum + notpayamt1c_sum)

                context["order_numer_"] = order_number
                context["choised_order"] = filtered_data_by_order
                context["balance"] = str(float('{:.2f}'.format(balance)))
                context["amount_diff"] = str(float('{:.2f}'.format(payamt1c_sum + notpayamt1c_sum)))
                context["dataToParse"] = serializers.serialize("json", (filtered_data_by_order))
                context["data_tcp"] = dataTCP.objects.all()
                context["data_BS"] = dataBS.objects.all()
                context["filtered_data_bs"] = filtered_data_by_bs
            elif bs_name:
                filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
                context["filtered_data_bs"] = filtered_data_by_bs
            return render(request, "mp/calculator/page_for_headers_bx.html", context)
        return render(request, "mp/calculator/main.html", data)
    return render(request, "mp/calculator/404.html", data)



def get_header_data_for_multiple_orders(order_numbers: list) -> dict:
    # Фильтруем данные только по нужным номерам заказов
    header_data = HeaderPosition.objects.filter(order_number__in=order_numbers)

    # Группируем данные по необходимым полям и считаем суммы
    grouped_positions = header_data.values('order_number', 'idocumentno', 'taxincluded', 'iprovider', 'category', 'idescription', 'odescription', 'iagreement' 'region', 'project_manager', 'productname').annotate(
        total_buyigrandtotal=Coalesce(Sum('buyigrandtotal'), 0),
        total_refundamtonorder=Coalesce(Sum('refundamtonorder'), 0),
    )

    # Считаем сумму всех позиций, сумму по всем ТМЦ и по всем другим категориям для каждого заказа
    summ_dict = header_data.values('order_number').annotate(total_summ=Coalesce(Sum('summagreement'), 0)).order_by('order_number')
    tmc_summ_dict = header_data.filter(category="ТМЦ").values('order_number').annotate(tmc_summ=Coalesce(Sum(F('refundamtonorder') + F('refundamtonorder') * 0.12), 0)).order_by('order_number')
    other_summ_dict = header_data.exclude(category="ТМЦ").values('order_number').annotate(other_summ=Coalesce(Sum(F('refundamtonorder') + F('refundamtonorder') * 0.12), 0)).order_by('order_number')

    # Формируем словарь с данными для каждого заказа
    datas = {}
    for order_number in order_numbers:
        order_grouped_positions = grouped_positions.filter(order_number=order_number)
        summ = summ_dict.filter(order_number=order_number).first()['total_summ']
        tmc_summ = tmc_summ_dict.filter(order_number=order_number).first()['tmc_summ']
        other_summ = other_summ_dict.filter(order_number=order_number).first()['other_summ']
        total_totallines_7_15 = order_grouped_positions.aggregate(total_totallines_7_15=Coalesce(Sum('totallines_7_15'), 0))['total_totallines_7_15']
        nds_coming = summ * 12 / 112
        nds_expense = (other_summ + tmc_summ) * 12 / 112
        nds_due = nds_coming - nds_expense
        gross_profitability = summ - other_summ - tmc_summ
        CIT_20 = (gross_profitability - (gross_profitability * 12 / 112)) * 0.2
        flow_rate_avh_5 = (summ - (summ * 12 / 112)) * 0.05
        net_profit = gross_profitability - CIT_20 - nds_due - flow_rate_avh_5
        morhz = net_profit / total_totallines_7_15 * 100 if total_totallines_7_15 else 0

        datas[order_number] = {
            "total_totallines_7_15": total_totallines_7_15,
            "order_number": order_number,
            "header_data": order_grouped_positions, 
            "summ": summ, 
            "tmc_summ": tmc_summ, 
            "other_summ": other_summ,
            "nds_coming": nds_coming,
            "nds_expense": nds_expense,
            "nds_due": nds_due,
            "gross_profitability": gross_profitability,
            "CIT_20": CIT_20,
            "flow_rate_avh_5": flow_rate_avh_5,
            "net_profit": net_profit,
            "morhz": round(morhz, 2),
        }

    return datas

def calculate_morzh(summ, other_summ, tmc_summ, total_totallines_7_15) -> Tuple[float]:
    temp_summ = summ
    if total_totallines_7_15:
        temp_summ = total_totallines_7_15
    other_summ = other_summ if other_summ else 0
    tmc_summ = tmc_summ if tmc_summ else 0
    nds_coming = temp_summ*12/112
    nds_expense = (other_summ+tmc_summ)*12/112
    nds_due = (nds_coming-nds_expense)
    gross_profitability = temp_summ-other_summ-tmc_summ
    CIT_20 = (gross_profitability-(gross_profitability*12/112))*0.2
    if CIT_20 < 0:
        CIT_20 = 0
    flow_rate_АВХ_5 = (temp_summ-(temp_summ*12/112))*0.15
    net_profit = gross_profitability-CIT_20-nds_due-flow_rate_АВХ_5
    morhz = net_profit/temp_summ*100 if temp_summ else 0
    return morhz, other_summ, tmc_summ, nds_coming, nds_expense, nds_due, gross_profitability, CIT_20, flow_rate_АВХ_5, net_profit

def get_header_data_summs(header_model: HeaderPosition) -> Tuple[float, float]:
    header_model
    tmc_summ = header_model.annotate(
        adjusted_refundamtonorder=Case(
            When(taxincluded="БЕЗ НДС", then=F("refundamtonorder") + F("refundamtonorder") * 0.12),
            default=F("refundamtonorder"),
            output_field=FloatField()
        )
    ).filter(category="ТМЦ").aggregate(Sum("adjusted_refundamtonorder"))["adjusted_refundamtonorder__sum"]
    
    tmc_summ = tmc_summ if tmc_summ else 0
    other_summ = header_model.annotate(
        adjusted_refundamtonorder=Case(
            When(taxincluded="БЕЗ НДС", then=F("refundamtonorder") + F("refundamtonorder") * 0.12),
            default=F("refundamtonorder"),
            output_field=FloatField()
        )
    ).exclude(category="ТМЦ").aggregate(Sum("adjusted_refundamtonorder"))["adjusted_refundamtonorder__sum"]
    return tmc_summ, other_summ

def get_header_data(order_number: str) -> dict:
    header_model = HeaderPosition.objects.filter(order_number=order_number)
    header_data = header_model.order_by("iprovider")
    
    grouped_positions = header_data.values('idocumentno', 'taxincluded', 'iprovider', 'category', 'idescription', 'odescription', 'iagreement', 'region', 'project_manager', 'productname').annotate(
        total_buyigrandtotal=Sum('buyigrandtotal'),
        total_refundamtonorder=Sum('refundamtonorder'),
        # total_totallines_7_15=Sum('totallines_7_15'),
    ).order_by('category')
    
    unique_categories = header_data.values('category').annotate(
    total_buyigrandtotal=Sum('buyigrandtotal'),
    total_refundamtonorder=Sum('refundamtonorder'),
    count=Case(
        When(category__isnull=True, then=Value(1)),
        When(category='', then=Value(1)),
        default=Count('category'),
        output_field=IntegerField()
    )
    ).order_by('category')
    # unique_categories = grouped_positions

    summ, summ_without_nds = 0, 0
    if header_model:
        summ = header_model[0].summagreement if header_model[0].summagreement else 0
    summ_without_nds = summ - summ * 0.12
    all_items_summ = header_model.aggregate(Sum("refundamtonorder"))["refundamtonorder__sum"]
    total_totallines_7_15 = header_model.aggregate(Sum("totallines_7_15"))["totallines_7_15__sum"]
    
    tmc_summ, other_summ = get_header_data_summs(header_model)
    morhz, other_summ, tmc_summ, nds_coming, nds_expense, nds_due, gross_profitability, CIT_20, flow_rate_АВХ_5, net_profit = calculate_morzh(summ, other_summ, tmc_summ, total_totallines_7_15)
    other_summ_without_nds = other_summ - other_summ * 0.12
    tmc_summ_without_nds = tmc_summ - tmc_summ * 0.12

    gross_profitability_new = summ_without_nds - other_summ_without_nds - tmc_summ_without_nds
    CIT_20_new = gross_profitability_new / 1.12 * 0.2
    flow_rate_АВХ_5_new = summ_without_nds * 0.15
    net_profit_new = gross_profitability_new - CIT_20_new - flow_rate_АВХ_5_new
    morhz_new = net_profit_new / summ_without_nds * 100
    fact_morzh_new = net_profit_new / total_totallines_7_15 * 100

    bs_name = grouped_positions[0]['odescription'].replace('\n', '').split(',')[0]

    clonned_unique_categories = []
    indexes = []
    for i, item in enumerate(unique_categories):
        temp_header = HeaderPosition.objects.filter(category=item["category"], order_number=order_number)
        temp_totalline = temp_header.aggregate(Sum("totallines_7_15"))["totallines_7_15__sum"]
        item['morzh'] = calculate_morzh(summ, *get_header_data_summs(temp_header), temp_totalline)[0].__round__(2)
        if i == 0:
            item['index'] = 0
        else:
            item['index'] = unique_categories[i-1]["index"] + unique_categories[i-1]["count"]
        for count in range(item["count"]):
            clonned_unique_categories.append(item)
        indexes.append(item["index"])
    # print(header_data.values())
    datas = {
        "indexes": indexes,
        "clonned_unique_categories": clonned_unique_categories,
        "total_totallines_7_15": total_totallines_7_15,
        "all_items_summ": all_items_summ,
        "order_number": order_number,
        "bs_name": bs_name,
        "header_data": grouped_positions, 
        "summ": summ,
        "summ_without_nds": summ_without_nds,
        "tmc_summ": tmc_summ,
        "tmc_summ_without_nds": tmc_summ_without_nds,
        "other_summ": other_summ,
        "other_summ_without_nds": other_summ_without_nds,
        "nds_coming":nds_coming,
        "nds_expense":nds_expense,
        "nds_due":nds_due,
        "gross_profitability":gross_profitability,
        "gross_profitability_new": gross_profitability_new,
        "CIT_20":CIT_20,
        "CIT_20_new": CIT_20_new,
        "flow_rate_АВХ_5":flow_rate_АВХ_5,
        "flow_rate_АВХ_5_new": flow_rate_АВХ_5_new,
        "net_profit":net_profit,
        "net_profit_new": net_profit_new,
        "morhz":morhz.__round__(2),
        "morhz_new": morhz_new.__round__(2),
        "fact_morzh_new": fact_morzh_new.__round__(2),
        "filtered_data_2122": Data2122.objects.filter(order_number=order_number),
             }
    return datas


def find_order(request):
    data = {}
    if request.method == "GET":
        order_number = request.GET.get("bs_name").strip() if request.GET.get("bs_name") else request.GET.get("order_number")
        user_email = request.GET.get("user_email")
        header = request.GET.get("header_page")
        from_bx24 = False
        try:
            user = get_object_or_404(AvhUser, email=user_email)
            from_bx24 = True
        except:
            user = request.user

        orders = COrder.objects.filter(number=order_number)
        order = None
        if orders.exists():
            order = orders.first()

        # order = COrder.objects.get(number=order_number)
        if order:
            data["order"] = order
            if user_email and user:
                data["user"] = user

        else:
            if from_bx24:
                data["base_html_exist"] = True
                data["message"] = "Заказ не найден"
                data["error_code"] = 404
                data["error_name"] = "Заказ не найден"
                return render(request, "mp/calculator/404_out.html", data)
                    
        try: 
            newreport = Report.objects.create(
                text=f"Поиск заказа с номером {order_number} с помощью find_order скрипта (BX24)",
                process="calculator - Поиск заказа",
                responsible=data['user'],
            )
            newreport.save()
        except: pass
        if header:
            datas = get_header_data(order_number)
            return render(request, "mp/calculator/table_for_headers.html", datas)
        return render(request, "mp/calculator/only_table.html", data)


    try: 
        if not from_bx24:
            newreport = Report.objects.create(
                text=f"Поиск заказа с номером {order_number} с помощью find_order скрипта",
                process="calculator - Поиск заказа",
                responsible=data['user'],
            )
            newreport.save()
    except: pass

    if header:
        datas = get_header_data(order_number)
        return render(request, "mp/calculator/table_for_headers.html", datas)
    return render(request, "mp/calculator/only_table.html", data)


def dom_deleter(x: str, /) -> str:
    return x.encode().decode("utf-8-sig")

T = TypeVar("T")
P = ParamSpec("P")


def time_limit(min_hour: int = 10, max_hour: int = 18):
    def decorator(
        view: collections.abc.Callable[P, T]
    ) -> collections.abc.Callable[P, T]:
        @functools.wraps(view)
        def wrapper(
            request: HttpRequest, *args: P.args, **kwargs: P.kwargs
        ) -> T | HttpResponse:
            
            if min_hour <= datetime.datetime.now().hour+6 <= max_hour and request.user.is_authenticated and not request.user.is_superuser :
                return render(
                    request,
                    "mp/calculator/404_out.html",
                    {
                        "base_html_exist": True,
                        "message": f"Время работы этой страницы: с {max_hour}:00 сегодняшнего дня до {min_hour}:00 следующего дня.",
                        "error_code": 403,
                        "error_name": "Нет доступа",
                    },
                )
            return view(request, *args, **kwargs)

        return wrapper

    return decorator


@login_required
@time_limit()
def set_prices(request: HttpRequest):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            if not request.POST.get("order_project"):
                return render(
                    request,
                    "mp/calculator/set_prices.html",
                    {
                        "form": form,
                        "import_type": "Импорт цен ТЦП",
                        "custom_error": f"Был упущен проект. Выберите проект из списка",
                    },
                )
            order_project = request.POST.get("order_project")
            position_price = COrderPositionPrice.objects.filter(
                customer_name=order_project
            )
            if position_price:
                position_price.delete()
                cache.clear()
            excel_file = request.FILES["file"]

            data = pd.read_excel(excel_file)
            column_headers = data.columns.tolist()

            TELE2_PRICES = [
                "Южные_регионы",
                "Северо_Восточные_регионы",
                "Западные_регионы",
            ]

            DEFAULT_FIELDS_FROM_FILE = [
                "Заказчик",
                "Пункт ТЦП Заказчика",
                "Пункт ТЦП с подрядчиком",
                "Ключ поиска в Адеме",
            ]

            for i in DEFAULT_FIELDS_FROM_FILE:
                if i not in column_headers:
                    return render(
                        request,
                        "mp/calculator/set_prices.html",
                        {
                            "form": form,
                            "import_type": "Импорт цен ТЦП",
                            "custom_error": f"{i} - нет такого столбца в файле.",
                        },
                    )

            for i in TELE2_PRICES:
                if (
                    (order_project == "Мобайл Телеком - Сервис, ТОО" or order_project == "Хуавей Текнолоджиз Казахстан, ТОО")
                    and i not in column_headers
                ):
                    return render(
                        request,
                        "mp/calculator/set_prices.html",
                        {
                            "form": form,
                            "import_type": "Импорт цен ТЦП",
                            "custom_error": f"{i} - нет такого столбца в файле. (Tele 2)",
                        },
                    )
            if (
                order_project != "Мобайл Телеком - Сервис, ТОО" 
                and order_project != "Хуавей Текнолоджиз Казахстан, ТОО"
                and "Цена" not in column_headers
            ):
                return render(
                    request,
                    "mp/calculator/set_prices.html",
                    {
                        "form": form,
                        "import_type": "Импорт цен ТЦП",
                        "custom_error": f"Цена - нет такого столбца в файле.",
                    },
                )

            # Пройдитесь по данным и сохраните их в модели YourModel
            ERROR_ROWS = []

            for index, row in data.iterrows():
                customer_name = row.get("Заказчик", order_project)
                note = ""
                if (
                    customer_name != "Мобайл Телеком - Сервис, ТОО"
                    and customer_name != "Хуавей Текнолоджиз Казахстан, ТОО"
                    and row["Цена"]
                    and f'{type(row["Цена"])}' != "<class 'int'>"
                    and f'{type(row["Цена"])}' != "<class 'float'>"
                ):
                    note = row["Цена"]
                    row["Цена"] = 0
                    if note == "nan":
                        note = "-"

                elif customer_name == "Мобайл Телеком - Сервис, ТОО" or customer_name == "Хуавей Текнолоджиз Казахстан, ТОО":
                    if row["Западные_регионы"] and not isinstance(
                        row["Западные_регионы"], (int, float)
                    ):
                        note = row["Западные_регионы"]
                        row["Западные_регионы"] = 0

                    if (
                        row["Южные_регионы"]
                        and f'{type(row["Южные_регионы"])}' != "<class 'int'>"
                        and f'{type(row["Южные_регионы"])}' != "<class 'float'>"
                    ):
                        note = row["Южные_регионы"]
                        row["Южные_регионы"] = 0
                    if (
                        row["Северо_Восточные_регионы"]
                        and f'{type(row["Северо_Восточные_регионы"])}'
                        != "<class 'int'>"
                        and f'{type(row["Северо_Восточные_регионы"])}'
                        != "<class 'float'>"
                    ):
                        note = row["Северо_Восточные_регионы"]
                        row["Северо_Восточные_регионы"] = 0

                if customer_name != "Мобайл Телеком - Сервис, ТОО" and customer_name != "Хуавей Текнолоджиз Казахстан, ТОО":
                    try:
                        COrderPositionPrice.objects.create(
                            find_key=row["Ключ поиска в Адеме"],
                            contractor=dom_deleter(row["Пункт ТЦП с подрядчиком"]),
                            customer=dom_deleter(row["Пункт ТЦП Заказчика"]),
                            customer_name=customer_name,
                            price=row["Цена"],
                            price_first=0,
                            price_second=0,
                            price_third=0,
                            notes=note,
                            minimum_price=80,
                            maximum_price=80,
                            default_price=80,
                        )
                        print("Created")
                    except:
                        ERROR_ROWS.append(
                            f"Ошибка в строке:{index+1} ({traceback.format_exc()})"
                        )

                else:
                    try:
                        position_price = COrderPositionPrice.objects.create(
                            find_key=row["Ключ поиска в Адеме"],
                            contractor=dom_deleter(row["Пункт ТЦП с подрядчиком"]),
                            customer=dom_deleter(row["Пункт ТЦП Заказчика"]),
                            customer_name=customer_name,
                            price=row["Западные_регионы"],
                            price_first=row["Южные_регионы"],
                            price_second=row["Северо_Восточные_регионы"],
                            price_third=row["Западные_регионы"],
                            notes=note,
                            minimum_price=80,
                            maximum_price=80,
                            default_price=80,
                        )
                        print(f'find_key = "{position_price.find_key}";')
                    except:
                        ERROR_ROWS.append(f"Ошибка в строке:{index+1}")

            logs = (
                "<br>".join(ERROR_ROWS)
                if ERROR_ROWS
                else "Все строки успешно импортировались"
            )

            try: 
                newreport = Report.objects.create(
                    text=logs,
                    process="calculator - Загрузка цен",
                    responsible=request.user,
                )
                newreport.save()
            except: pass

            return render(
                request,
                "mp/calculator/set_prices.html",
                {"form": form, "import_type": "Импортировать цены ТЦП", "logs": logs},
            )
    else:
        form = UploadFileForm()

    return render(
        request,
        "mp/calculator/set_prices.html",
        {"form": form, "import_type": "Импортировать цены ТЦП"},
    )


@time_limit()
@login_required
def set_orders(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["file"]

            # old_orders = COrder.objects.all()
            # if old_orders:
            #     old_orders.delete()

            data = pd.read_excel(excel_file, sheet_name="детально")
            column_headers = data.columns.tolist()

            DEFAULT_FIELDS_FROM_FILE = [
                "Номер заказа на продажу",
                "Регион",
                "Комментарий",
                "Кол-во",
                "Заказчик",
            ]

            for i in DEFAULT_FIELDS_FROM_FILE:
                if i not in column_headers:
                    return render(
                        request,
                        "mp/calculator/set_prices.html",
                        {
                            "form": form,
                            "import_type": "Импортировать заказы с Adem",
                            "custom_error": f"{i} - нет такого столбца в файле.",
            "contractor_disabled": "disabled",
                        },
                    )

            ERROR_ROWS = []
            # Пройдитесь по данным и сохраните их в модели YourModel
            x = COrder.objects.all()
            order_len = len(x)
            x.delete()

            xx = COrder.objects.all()
            order_new_len = len(xx)

            myMessage = f"Удалено {order_len} заказов. Осталось {order_new_len}."
            for index, row in data.iterrows():
                    if not row["Номер заказа на продажу"]:
                        ERROR_ROWS.append(
                            f'В строке - {index+1} нет "Номер заказа на продажу"'
                        )
                        continue
                    if not row["Регион"]:
                        ERROR_ROWS.append(f'В строке - {index+1} нет "Регион"')
                        continue
                    if not row["Комментарий"]:
                        ERROR_ROWS.append(f'В строке - {index+1} нет "Комментарий"')
                        continue
                    if row["Кол-во"] is None:
                        ERROR_ROWS.append(f'В строке - {index+1} нет "Кол-во"')
                        continue
                    if row["Заказчик"] is None:
                        ERROR_ROWS.append(f'В строке - {index+1} нет "Заказчик"')
                        continue

                    order = COrder.objects.filter(number=row["Номер заказа на продажу"])

                    quantity = row.get("Кол-во", 0)
                    quantity = float(quantity)

                    if math.isnan(quantity):
                        continue

                    if len(order) > 0:
                        COrderPosition.objects.create(
                            order=order[0],
                            quantity=quantity,
                            comments=row["Комментарий"],
                            position_name=dom_deleter(row["Номенклатура"]),
                        )
                    else:
                        new_order = COrder.objects.create(
                            number=row["Номер заказа на продажу"],
                            region=row["Регион"],
                            comments=row["Комментарий"],
                            # customer_name = "КаР-Тел, ТОО",
                            customer_name=row["Заказчик"],
                        )

                        COrderPosition.objects.create(
                            order=new_order,
                            quantity=quantity,
                            comments=row["Комментарий"],
                            position_name=dom_deleter(row["Номенклатура"]),
                        )

            orders_len = 0
            try: orders_len = len(COrder.objects.all())
            except: pass
            
            orderpositions_len = 0
            try: orderpositions_len = len(COrderPosition.objects.all())
            except: pass
            

            logs = f"Сохранено {orders_len} заказов и {orderpositions_len} позиции"
            if ERROR_ROWS:
                logs = "<br>".join(ERROR_ROWS)

        logs = myMessage +"<hr>" + logs
        try: 
            newreport = Report.objects.create(
                text=logs,
                process="calculator - Загрузка заказов",
                responsible=request.user,
            )
            newreport.save()
        except: pass
        return render(
            request,
            "mp/calculator/set_prices.html",
            {"form": form, "import_type": "Импортировать заказы с Adem", "logs": logs,
            "contractor_disabled": "disabled",},
        )  # Перенаправьте на страницу успешного импорта
    else:
        form = UploadFileForm()


    return render(
        request,
        "mp/calculator/set_prices.html",
        {
            "form": form,
            "import_type": "Импортировать заказы с Adem",
            "contractor_disabled": "disabled",
        },
    )


@time_limit()
@login_required
def remove_orders(request):
    
    if request.method == "POST":
        form = DeleteOptionsForm(request.POST)
        if form.is_valid():
            # Получаем значение выбранной опции
            delete_option = form.cleaned_data["delete_option"]
            myMessage = ""
            # Здесь вы можете выполнить необходимые действия в зависимости от выбранной опции
            if delete_option == DeleteOptionsForm.DELETE_ORDERS:
                x = COrder.objects.all()
                order_len = len(x)
                x.delete()

                xx = COrder.objects.all()
                order_new_len = len(xx)

                myMessage = f"Удалено {order_len} заказов. Осталось {order_new_len}."
                # myMessage = f"Все заказы удалены. Удалено {order_len} заказов."
            elif delete_option == DeleteOptionsForm.DELETE_PRICES:
                z = COrderPositionPrice.objects.all()
                order_len = len(z)
                z.delete()
                myMessage = f"Все цены удалены. Удалено {order_len} цен."

            elif delete_option == DeleteOptionsForm.DELETE_BOTH:
                x = COrder.objects.all()
                z = COrderPositionPrice.objects.all()
                x_order_len = len(x)
                z_order_len = len(z)

                z.delete()
                x.delete()
                
                xx = COrder.objects.all()
                order_new_len = len(x)

                # myMessage = f"Удалено {order_len} заказов. Осталось {order_new_len}."
                myMessage = f"""Удалено {order_len} заказов. Осталось {order_new_len}..
                             <br>
                             Все цены удалены. Удалено {z_order_len} цен.
                             """

            cache.clear()
            form = DeleteOptionsForm()
            try: 
                newreport = Report.objects.create(
                    text="Все заказы удалены с помощью remove_orders скрипта",
                    process="calculator - Удаление заказов",
                    responsible=request.user,
                )
                newreport.save()
            except: pass
            return render(
                request,
                "mp/calculator/remove_orders_confirm.html",
                {"form": form, "logs": myMessage},
            )

    else:
        form = DeleteOptionsForm()
    return render(request, "mp/calculator/remove_orders_confirm.html", {"form": form})





# @time_limit()
@login_required
def set_order_fast(request):
    
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["file"]

            # old_orders = COrder.objects.all()
            # if old_orders:
            #     old_orders.delete()

            data = pd.read_excel(excel_file, header=1)
            column_headers = data.columns.tolist()
            for cycle in range(2):
                DEFAULT_FIELDS_FROM_FILE = [
                    "Номер заказа на продажу",
                    "Номенклатура",
                    "Комментарий",
                    "Кол-во",
                    "Регион",
                ]

                for i in DEFAULT_FIELDS_FROM_FILE:
                    if i not in column_headers and i != "Контрагент":
                        return render(
                            request,
                            "mp/calculator/set_prices.html",
                            {
                                "form": form,
                                "import_type": "Импортировать кусок заказов с Adem",
                                "custom_error": f"{i} - нет такого столбца в файле.",
                            },
                        )
                
                # if "Контрагент" in column_headers or "Заказчик" in column_headers:
                #     return render(
                #         request,
                #         "mp/calculator/set_prices.html",
                #         {
                #             "form": form,
                #             "import_type": "Импортировать кусок заказов с Adem",
                #             "custom_error": f"<Контрагент> или <Заказчик> - нет такого столбца в файле.",
                #         },
                #     )

                ERROR_ROWS = []
                DELETED_ORDERS = []
                # Пройдитесь по данным и сохраните их в модели YourModel
                for index, row in data.iterrows():
                    
                        if not row["Номер заказа на продажу"]:
                            ERROR_ROWS.append(
                                f'В строке - {index+1} нет "Номер заказа на продажу"'
                            )
                            continue
                        if not row["Регион"]:
                            ERROR_ROWS.append(f'В строке - {index+1} нет "Регион"')
                            continue
                        if not row["Комментарий"]:
                            ERROR_ROWS.append(f'В строке - {index+1} нет "Комментарий"')
                            continue
                        if row["Кол-во"] is None:
                            ERROR_ROWS.append(f'В строке - {index+1} нет "Кол-во"')
                            continue
                        if row["Контрагент"] is None:
                            if row["Заказчик"] is None:
                                ERROR_ROWS.append(f'В строке - {index+1} нет "Заказчик"')
                                continue
                            else: 
                                row["Контрагент"] = row["Заказчик"]

                        order : list[COrder] = COrder.objects.filter(number=row["Номер заказа на продажу"])

                        quantity = row.get("Кол-во", 0)
                        quantity = float(quantity)

                        if math.isnan(quantity):
                            continue
                        if len(order) > 0 and order[0].number not in DELETED_ORDERS:
                            DELETED_ORDERS.append(order[0].number)
                            order[0].delete()
                            order = []

                        if len(order) > 0:
                            COrderPosition.objects.create(
                                order=order[0],
                                quantity=quantity,
                                comments=row["Комментарий"],
                                position_name=dom_deleter(row["Номенклатура"]),
                            )
                        else:
                            new_order = COrder.objects.create(
                                number=row["Номер заказа на продажу"],
                                region=row["Регион"],
                                comments=row["Комментарий"],
                                # customer_name = "КаР-Тел, ТОО",
                                customer_name=row["Контрагент"],
                            )

                            COrderPosition.objects.create(
                                order=new_order,
                                quantity=quantity,
                                comments=row["Комментарий"],
                                position_name=dom_deleter(row["Номенклатура"]),
                            )
                logs = "Все заказы импортировались"
                if ERROR_ROWS:
                    logs += "<h4> Ошибка при импорте в заказах: </h4>"
                    logs = "<br>".join(ERROR_ROWS)
                if DELETED_ORDERS:
                    logs += "<h4> Обновлены эти заказы: </h4>"
                    logs = "<br>".join(DELETED_ORDERS)

                try: 
                    newreport = Report.objects.create(
                        text=logs,
                        process="calculator - Обновление нескольких заказов",
                        responsible=request.user,
                    )
                    newreport.save()
                except: pass
        return render(
            request,
            "mp/calculator/set_prices.html",
            {"form": form, "import_type": "Импортировать кусок заказов с Adem", "logs": logs},
        )  # Перенаправьте на страницу успешного импорта
    else:
        form = UploadFileForm()

    return render(
        request,
        "mp/calculator/set_prices.html",
        {
            "form": form,
            "import_type": "Импортировать кусок заказов с Adem",
            "contractor_disabled": "disabled",
        },
    )



@login_required
def change_nds(request):
    try: 
        newreport = Report.objects.create(
            text="Изменение НДС с помощью change_nds скрипта",
            process="calculator - Изменение НДС в HTML(JSON)",
            responsible=request.user,
        )
        newreport.save()
    except: pass
    return render(request, "mp/calculator/change_nds.html")



def test(request):
    pas = request.GET.get("pass")
    if pas == '77':
        return render(request, update_data_from_json())
    return HttpResponse("Error")
def test2(request):
    pas = request.GET.get("pass")
    if pas == '77':
        import_19_45()
        return HttpResponse("Success")
    return HttpResponse("Error")
def update_data_from_json():
     
    json_path = r'/mnt/adem-otchet/7_35.json'#'/mnt/adem-otchet/7_35.json'
    
    with open(json_path, 'r', encoding='utf-8') as f:
        # json_data = f.read().encode('UTF-8').decode('unicode_escape')
        # json_data = json.loads(rf'{json_data}'.encode('UTF-8').decode('UTF-8'))
        json_data = bytes(f.read(), 'utf-8').decode("unicode_escape")
        json_data = json.loads(json_data)
    old_orders = COrder.objects.all()
    if old_orders:
        old_orders.delete()
    ERROR_ROWS = []
    DELETED_ORDERS = []
    # Пройдитесь по данным и сохраните их в модели YourModel
    for value in json_data:
        
        
        if not value["odocumentno"]:
            ERROR_ROWS.append(
                f'В строке - {value} нет "Номер заказа на продажу"'
            )
            # continue
        if not value["region"]:
            encoded_dict = {key: str(v).encode('ISO-8859-1').decode('utf-8') for key, v in value.items()}
            ERROR_ROWS.append(f'В строке - {encoded_dict} нет "Регион"')
            # continue
        if not value["odescription"]:
            encoded_dict = {key: str(v).encode('ISO-8859-1').decode('utf-8') for key, v in value.items()}
            ERROR_ROWS.append(f'В строке - {encoded_dict} нет "Комментарий"')
            # continue
        if value["qtyentered"] is None:
            encoded_dict = {key: str(v).encode('ISO-8859-1').decode('utf-8') for key, v in value.items()}
            ERROR_ROWS.append(f'В строке - {encoded_dict} нет "Кол-во"')
            # continue
        if value["productname"] is None:
            if value["consumer"] is None:
                encoded_dict = {key: str(v).encode('ISO-8859-1').decode('utf-8') for key, v in value.items()}
                ERROR_ROWS.append(f'В строке - {encoded_dict} нет "Заказчик"')
                # continue
            # else: 
            #     value["Контрагент"] = value["Заказчик"]
        
        
        order : list[COrder] = COrder.objects.filter(number=value["odocumentno"].encode('ISO-8859-1').decode('utf-8'))


        quantity = value.get("qtyentered", 0)
        if not quantity:
            quantity = 0
        quantity = float(quantity)

        if math.isnan(quantity):
            # print(value["odescription"].encode('ISO-8859-1').decode('utf-8'))
            continue
        
        # if len(order) > 0 and order[0].number not in DELETED_ORDERS:
        #     DELETED_ORDERS.append(order[0].number)
        #     order[0].delete()
        #     order = []
        if len(order) > 0:
            COrderPosition.objects.create(
                order=order[0],
                quantity=quantity,
                comments=value["odescription"].encode('ISO-8859-1').decode('utf-8') if value["odescription"] else ' ',
                position_name=dom_deleter(value["productname"].encode('ISO-8859-1').decode('utf-8')) if value["productname"] else ' ',
                project_group=value["projectgroup"].encode('ISO-8859-1').decode('utf-8') if "projectgroup" in value and value["projectgroup"] else ' ',
                
            )
        else:
            new_order = COrder.objects.create(
                number=value["odocumentno"].encode('ISO-8859-1').decode('utf-8'),
                region=value["region"].encode('ISO-8859-1').decode('utf-8') if value["region"] else ' ',
                comments=value["odescription"].encode('ISO-8859-1').decode('utf-8') if value["odescription"] else ' ',
                customer_name=value["consumer"].encode('ISO-8859-1').decode('utf-8'),
            )

            COrderPosition.objects.create(
                order=new_order,
                quantity=quantity,
                comments=value["odescription"].encode('ISO-8859-1').decode('utf-8') if value["odescription"] else ' ',
                position_name=dom_deleter(value["productname"].encode('ISO-8859-1').decode('utf-8')) if value["productname"] else ' ',
                project_group=value["projectgroup"].encode('ISO-8859-1').decode('utf-8') if "projectgroup" in value and value["projectgroup"] else ' ',
            )
            
    logs = "Все заказы импортировались"
    if ERROR_ROWS:
        logs += "<h4> Ошибка при импорте в заказах: </h4>"
        logs = "<br>".join(ERROR_ROWS)
    if DELETED_ORDERS:
        logs += "<h4> Обновлены эти заказы: </h4>"
        logs = "<br>".join(DELETED_ORDERS)

    try: 
        newreport = Report.objects.create(
            text=logs,
            process="calculator - Обновление нескольких заказов",
            responsible="Автоматическое обновление",
        )
        print("Report created")
        newreport.save()
    except: pass
    return logs  # Перенаправьте на страницу успешного импорта



def zero_order(requests):
    key_word = requests.GET.get("key_word")
    projectgroup = requests.GET.get('projectgroup')
    max_sum = 1_000_000
    if key_word:
        key_word = key_word.lower()
        if "электромонтаж" in key_word:
            max_sum = 7_000_000
        elif "монтаж" in key_word:
            max_sum = 3_000_000
        else:
            max_sum = 10_000_000
    if not projectgroup:
        return render(requests, 'mp/calculator/main_zero_order.html', {'data_tcp': COrderPositionPrice.objects.filter(customer_name='КаР-Тел, ТОО'), "ms":max_sum})
    
    else: 
        tcp_for_252 = [{
            "id":1,
            "find_key": "ART1003275",
            "customer_name": "КаР-Тел, ТОО",
            "contractor":"Проект",
            "maximum_price":35000,
            "minimum_price":1,
            "default_price":35000,
            "price":35000,
            "customer":"Проект"
        },
        {
            "id":2,
            "find_key": "ART1003274",
            "customer_name": "КаР-Тел, ТОО",
            "contractor":"Обследование",
            "maximum_price":25000,
            "minimum_price":1,
            "default_price":25000,
            "price":25000,
            "customer":"Обследование"
            },
        {
            "id":3,
            "find_key": "ART1003276",
            "customer_name": "КаР-Тел, ТОО",
            "contractor":"Логистика",
            "maximum_price":25000,
            "minimum_price":1,
            "default_price":25000,
            "price":25000,
            "customer":"Логистика"
        },
        {
            "id":4,
            "find_key": "ART1001903",
            "customer_name": "КаР-Тел, ТОО",
            "contractor":"Строительно-монтажные работы",
            "maximum_price":250000,
            "minimum_price":1,
            "default_price":250000,
            "price":250000,
            "customer":"Строительно-монтажные работы"
        }
        ]
        return render(requests, 'mp/calculator/main_zero_order.html', {'data_tcp': tcp_for_252, "ms":max_sum, "limit": True})


from django.db import transaction

@transaction.atomic
def excel_import_19_45(request):
    if request.method == "POST":        
        excel_file = request.FILES['file_upload']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        order_positions_to_create = []
        order_positions_to_update = []

        COrderPosition.objects.update(used_count=0, used_summ=0, used_quantity=0)
        for row in ws.iter_rows(min_row=2):
            order_number = row[1].value if row[1].value else ''
            position_name = row[5].value if row[5].value else ''
            used_summ = float(row[6].value if row[6].value else 0)
            used_quantity = float(row[7].value if row[7].value else 0)
            try:
                order = COrder.objects.get(number=order_number)
                order_positions = COrderPosition.objects.filter(order=order, position_name=position_name)
                for order_position in order_positions:
                    order_position.used_count += 1
                    order_position.used_summ += used_summ
                    order_position.used_quantity += used_quantity
                    order_position.save()
            except COrder.DoesNotExist:
                pass
            # except COrderPosition.DoesNotExist:
            #     order_positions_to_create.append(
            #         COrderPosition(
            #             order=order,
            #             position_name=position_name,
            #             used_count=1,
            #             used_summ=used_summ,
            #             used_quantity=used_quantity
            #         )
            #     )
            except Exception as e:
                print(e)

        # COrderPosition.objects.bulk_update(order_positions_to_update, ['used_count', 'used_summ', 'used_quantity'])
        # COrderPosition.objects.bulk_create(order_positions_to_create)
        return render(request, 'mp/calculator/excel_import_19_45.html', {'success': True})
    else:
        return render(request, 'mp/calculator/excel_import_19_45.html')
    

def import_19_45():
    
    json_path = r'/mnt/adem-otchet/19_45.json'#'/mnt/adem-otchet/7_35.json'
    with open(json_path, 'r', encoding='utf-8') as f:
        # json_data = f.read().encode('UTF-8').decode('unicode_escape')
        # json_data = json.loads(rf'{json_data}'.encode('UTF-8').decode('UTF-8'))
        try:
            json_data = bytes(f.read(), 'utf-8').decode("unicode_escape")
            json_data = json.loads(json_data)
        except Exception as e:
            print(e)
            return 501
    #.encode('ISO-8859-1').decode('utf-8')
    COrderPosition.objects.update(used_count=0, used_summ=0, used_quantity=0)
    for value in json_data:
        order_number = value["orderno"].encode('ISO-8859-1').decode('utf-8') if value["orderno"] else ''  # Номер заказа из файла
        position_name = value["nomenclature"].encode('ISO-8859-1').decode('utf-8') if value["nomenclature"] else ''  # Наименование работ из файла
        used_summ = float(value["priceentered"] if value["priceentered"] else 0)  # Сумма использованных из файла
        used_quantity = float(value["qtyentered"] if value["qtyentered"] else 0)  # Количество использованных из файла
        
        # Находим соответствующий объект COrder
        try:
            order = COrder.objects.get(number=order_number)

            # Находим соответствующий объект COrderPosition
        
            order_position = COrderPosition.objects.get(order=order, position_name=position_name)

            # Обновляем данные
            order_position.used_count += 1
            order_position.used_summ += used_summ
            order_position.used_quantity += used_quantity
            order_position.save()
        except:
            pass
    return 200


def hand_update_7_11_2(request):
    update_data_7_11_2()
    return HttpResponse("OK")


def update_data_7_11_2():
    file_path = r'/mnt/adem-otchet/7_11_2.json'#'/mnt/adem-otchet/7_11_2.json'
    # batch_size=100
    with open(file_path, 'r', encoding='utf-8') as f:
        try:
            json_data = bytes(f.read(), 'utf-8').decode("unicode_escape")
            json_data = json.loads(json_data)
        except Exception as e:
            print(e)
            return 501

    order = COrder.objects.all()
    order.update(buyigrandtotal=0, money_wasted=0)

    HeaderPosition.objects.all().delete()
    Data7112.objects.all().delete()

    try:
        gc = gspread.service_account(filename="calculator/credentials.json")
        sh = gc.open_by_key('1evlFSaiWgkkRqDdN7XblJfzqtE9CF3tegkwYhH5esxQ')
        worksheet = sh.get_worksheet_by_id(1259358010)

        columns = ['Название БС', 'Номер заказа (Адем)', 'Дата передачи', 'Проект статус']

        all_values = worksheet.get_all_values()

        headers = all_values[1]
        data_rows = all_values[2:]

        column_indices = {header: index for index, header in enumerate(headers) if header in columns}

        for row in data_rows:
            row_data = {header: row[index] for header, index in column_indices.items()}
            if row_data.get('Проект статус') == 'Согласован':
                Data7112.objects.create(
                    odescription=row_data.get('Название БС').replace('\n', '').split(',')[0],
                    order_number=row_data.get('Номер заказа (Адем)').split(' ')[0],
                    odateordered=row_data.get('Дата передачи'),
                    project_manager='from google sheet'
                )
    except Exception as e:
        print(f'Error update_data_7_11_2: {e}''')

    for value in json_data:
        order_number = value["odocumentno"].encode('ISO-8859-1').decode('utf-8')
        # batch = []
        if not order_number:
            continue
        buyigrandtotal = value.get("refundamtonorder", 0)
        percentpayment = float(value["percentpayment"].encode('ISO-8859-1').decode('utf-8').replace('%', '').replace(',', '.')) / 100 if value["percentpayment"] else 0
        consumer = value['consumer'].encode('ISO-8859-1').decode('utf-8') if value['consumer'] else ''

        Data7112.objects.create(
            order_number=order_number,
            idocumentno=value['idocumentno'].encode('ISO-8859-1').decode('utf-8') if value['idocumentno'] else '-',
            iprovider=value['iprovider'].encode('ISO-8859-1').decode('utf-8') if value['iprovider'] else '-',
            iconfirmer=value['iconfirmer'].encode('ISO-8859-1').decode('utf-8') if value['iconfirmer'] else '-',
            taxincluded=value['taxincluded'].encode('ISO-8859-1').decode('utf-8') if value['taxincluded'] else '-',
            idescription=value['idescription'].encode('ISO-8859-1').decode('utf-8') if value['idescription'] else '-',
            odescription=value['odescription'].encode('ISO-8859-1').decode('utf-8').replace('\n', '').split(',')[0] if
            value['odescription'] else '-',
            odateordered=value['odateordered'].encode('ISO-8859-1').decode('utf-8') if value['odateordered'] else '-',
            category=value['category'].encode('ISO-8859-1').decode('utf-8') if value['category'] else '-',
            iagreement=value['iagreement'].encode('ISO-8859-1').decode('utf-8') if value['iagreement'] else '-',
            consumer=value['consumer'].encode('ISO-8859-1').decode('utf-8') if value['consumer'] else '-',
            region=value['region'].encode('ISO-8859-1').decode('utf-8') if value['region'] else '-',
            project_manager=value['project_manager'].encode('ISO-8859-1').decode('utf-8') if value[
                'project_manager'] else '-',
            productname=value['productname'].encode('ISO-8859-1').decode('utf-8') if value['productname'] else '-',
            summagreement=value.get('summagreement', 0),
            buyigrandtotal=value.get('buyigrandtotal', 0),
            totallines_7_15=value.get('totallines_7_15', 0),
            refundamtonorder=buyigrandtotal
        )

        if consumer == "КаР-Тел, ТОО":
            if not value['iprovider'] and not value['idescription']:
                continue

            HeaderPosition.objects.create(
                order_number=order_number,
                idocumentno=value['idocumentno'].encode('ISO-8859-1').decode('utf-8') if value['idocumentno'] else '-',
                iprovider=value['iprovider'].encode('ISO-8859-1').decode('utf-8') if value['iprovider'] else '-',
                iconfirmer=value['iconfirmer'].encode('ISO-8859-1').decode('utf-8') if value['iconfirmer'] else '-',
                taxincluded=value['taxincluded'].encode('ISO-8859-1').decode('utf-8') if value['taxincluded'] else '-',
                idescription=value['idescription'].encode('ISO-8859-1').decode('utf-8') if value[
                    'idescription'] else '-',
                odescription=value['odescription'].encode('ISO-8859-1').decode('utf-8') if value[
                    'odescription'] else '-',
                odateordered=value['odateordered'].encode('ISO-8859-1').decode('utf-8') if value[
                    'odateordered'] else '-',
                category=value['category'].encode('ISO-8859-1').decode('utf-8') if value['category'] else '-',
                iagreement=value['iagreement'].encode('ISO-8859-1').decode('utf-8') if value['iagreement'] else '-',
                consumer=value['consumer'].encode('ISO-8859-1').decode('utf-8') if value['consumer'] else '-',
                region=value['region'].encode('ISO-8859-1').decode('utf-8') if value['region'] else '-',
                project_manager=value['project_manager'].encode('ISO-8859-1').decode('utf-8') if value[
                    'project_manager'] else '-',
                productname=value['productname'].encode('ISO-8859-1').decode('utf-8') if value['productname'] else '-',
                summagreement=value.get('summagreement', 0),
                buyigrandtotal=value.get('buyigrandtotal', 0),
                totallines_7_15=value.get('totallines_7_15', 0),
                refundamtonorder=buyigrandtotal
            )
        # batch.append(model_instance)
        # if len(batch) >= batch_size:
                # HeaderPosition.objects.bulk_create(batch)
                # batch = []
        try:
            order = COrder.objects.get(number=order_number)
            if not order:
                continue
            money_wasted = buyigrandtotal * percentpayment

            if order.buyigrandtotal > 0:
                order.buyigrandtotal += buyigrandtotal
            else:
                order.buyigrandtotal = buyigrandtotal

            if order.money_wasted > 0:
                order.money_wasted += money_wasted
            else:
                order.money_wasted = money_wasted
            order.save()
        except Exception as e:
            pass
    return 200


_EXPORT_EXCEL_DATA_SQL: Final[str] = """
WITH sums AS (
    SELECT t1.order_number,
           COALESCE(SUM(CASE WHEN t1.category='ТМЦ' THEN 0 WHEN t1.taxincluded='БЕЗ НДС' THEN t1.refundamtonorder * 1.12 ELSE t1.refundamtonorder END), 0) AS tmc_summ,
           COALESCE(SUM(CASE WHEN t1.category!='ТМЦ' THEN 0 WHEN t1.taxincluded='БЕЗ НДС' THEN t1.refundamtonorder * 1.12 ELSE t1.refundamtonorder END), 0) AS other_summ,
           SUM(t1.summagreement) AS summ,
           SUM(t1.refundamtonorder) AS all_items_summ,
           SUM(t1.totallines_7_15) AS total_totallines_7_15
      FROM calculator_headerposition AS t1
     GROUP BY t1.order_number
), sums1 AS (
    SELECT t1.*, 
           t1.summ - t1.other_summ - t1.tmc_summ AS gross_profitability,
           t1.summ * 12 / 112 AS nds_coming,
           (t1.other_summ + t1.tmc_summ) * 12 / 112 AS nds_expense
      FROM sums AS t1
), sums2 AS (
    SELECT t1.*, 
           ABS((t1.gross_profitability - (t1.gross_profitability * 12 / 112)) * 0.2) AS CIT_20,
           ABS(t1.nds_coming - t1.nds_expense) AS nds_due
      FROM sums1 AS t1
), sums3 AS (
    SELECT t1.*, (t1.summ - (t1.summ * 12 / 112)) * 0.15 AS flow_rate_ABX_5
      FROM sums2 AS t1
), sums5 AS (
    SELECT t1.*, t1.gross_profitability - t1.CIT_20 - t1.nds_due - t1.flow_rate_ABX_5 AS net_profit
      FROM sums3 AS t1
), sums6 AS (
    SELECT t1.*, ROUND(CASE WHEN t1.total_totallines_7_15 <> 0 THEN t1.net_profit / t1.total_totallines_7_15 * 100 
                           WHEN t1.summ <> 0 THEN t1.net_profit / t1.summ * 100 ELSE 0 END::numeric, 2) AS morhz
      FROM sums5 AS t1
)
SELECT * FROM sums6 AS t1
"""

_EXPORT_EXCEL_DATA_SQL_NEW: Final[str] = """
WITH sums AS (
    SELECT t1.order_number,
           COALESCE(SUM(CASE WHEN t1.category='ТМЦ' THEN 0 WHEN t1.taxincluded='БЕЗ НДС' THEN t1.refundamtonorder * 1.12 ELSE t1.refundamtonorder END), 0) * 0.88 AS tmc_summ,
           COALESCE(SUM(CASE WHEN t1.category!='ТМЦ' THEN 0 WHEN t1.taxincluded='БЕЗ НДС' THEN t1.refundamtonorder * 1.12 ELSE t1.refundamtonorder END), 0) * 0.88 AS other_summ,
           SUM(t1.summagreement) * 0.88 AS summ,
           SUM(t1.refundamtonorder) AS all_items_summ,
           SUM(t1.totallines_7_15) AS total_totallines_7_15
      FROM calculator_headerposition AS t1
     GROUP BY t1.order_number
), sums1 AS (
    SELECT t1.*, 
           t1.summ - t1.other_summ - t1.tmc_summ AS gross_profitability,
           t1.summ * 12 / 112 AS nds_coming,
           (t1.other_summ + t1.tmc_summ) * 12 / 112 AS nds_expense
      FROM sums AS t1
), sums2 AS (
    SELECT t1.*, 
           ABS((t1.gross_profitability - (t1.gross_profitability * 12 / 112)) * 0.2) AS CIT_20,
           ABS(t1.nds_coming - t1.nds_expense) AS nds_due
      FROM sums1 AS t1
), sums3 AS (
    SELECT t1.*, summ * 0.15 AS flow_rate_ABX_5
      FROM sums2 AS t1
), sums5 AS (
    SELECT t1.*, t1.gross_profitability - t1.CIT_20 - t1.flow_rate_ABX_5 AS net_profit
      FROM sums3 AS t1
), sums6 AS (
    SELECT t1.*, ROUND(CASE WHEN t1.summ <> 0 THEN t1.net_profit / t1.summ * 100 ELSE 0 END::numeric, 2) AS morhz
      FROM sums5 AS t1
), sums7 AS (
    SELECT t1.*, ROUND(CASE WHEN t1.total_totallines_7_15 <> 0 THEN t1.net_profit / t1.total_totallines_7_15 * 100 ELSE 0 END::numeric, 2) AS fact_morhz
      FROM sums6 AS t1
)
SELECT * FROM sums7 AS t1
"""
def export_excel(request):
    if request.method == 'GET':
        rows = None
        with connection.cursor() as cursor:
            cursor.execute(_EXPORT_EXCEL_DATA_SQL_NEW)
            rows = [dict(zip((i[0] for i in cursor.description), row)) for row in cursor.fetchall()]
        # data = json.loads(request.body.decode('utf-8'))
        data = list(HeaderPosition.objects.values_list('order_number', flat=True).order_by("order_number").distinct())
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['№ Номера заказа', 'Сумма заказа', 'Сумма фактический закрытия', 'Сумма расхода общая', 'Чистая прибыль', 'Планируемая маржа', 'Фактическая маржа']
        ws.append(headers)
        
        # datas: dict[str, dict] = get_header_data(data)
        counter = 0
        # limit = 1000
        for row in rows:
            # header_data = get_header_data(order_number.strip())
            counter += 1
            ws.append([row['order_number'], row['summ'], row['all_items_summ'], row['tmc_summ'] + row['other_summ'], row['net_profit'], row['morhz'], row['fact_morhz']])
            # if not header_data:
                # continue
            # ws.append([header_data['order_number'], header_data['summ'], header_data['all_items_summ'], header_data['tmc_summ'] + header_data['other_summ'], header_data['net_profit'], header_data['morhz']])        
            # if counter == limit:
                # wb.save(f'data_{now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')
                # limit+=1000
                # print(limit)
        # for header_data in datas:
        #     ws.append([header_data['order_number'], header_data['summ'], header_data['all_items_summ'], header_data['tmc_summ'] + header_data['other_summ'], header_data['net_profit'], header_data['morhz']])        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="data_{now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx"'
        response.write
        wb.save(response)
        return response
    return HttpResponse(status=400)


def update_21_22(request):
    try:
        gc = gspread.service_account(filename="calculator/credentials.json")
        sh = gc.open_by_key('1iYNr4R-DZ13BJw44G37rfouk5to9__X2Yy0L7MpheWc')
        worksheet = sh.get_worksheet_by_id(834948387)

        columns = ['Номер заказа', 'Сфера деятельности', 'Вид деятельности', 'маржа']

        all_values = worksheet.get_all_values()
        Data2122.objects.all().delete()

        headers = all_values[1]
        data_rows = all_values[2:]

        column_indices = {header: index for index, header in enumerate(headers) if header in columns}

        for row in data_rows:
            row_data = {header: row[index] for header, index in column_indices.items()}
            margin = row_data.get('маржа').replace(',', '.').replace('%', '')

            Data2122.objects.create(
                order_number=row_data.get('Номер заказа'),
                activity_field=row_data.get('Сфера деятельности'),
                activity_kind=row_data.get('Вид деятельности'),
                margin=float(margin) if margin else 0
            )

        try:
            newreport = Report.objects.create(
                text="Обновлены данные из 21-22",
                process="calculator - Обновление данных",
                responsible=request.user,
            )
            newreport.save()
        except:
            pass
    except Exception as e:
        return e
    return HttpResponseRedirect(request.META['HTTP_REFERER'])


