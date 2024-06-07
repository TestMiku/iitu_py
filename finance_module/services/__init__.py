import dataclasses
import datetime
import decimal
import enum
import io
import json
import logging
import typing
from collections.abc import Callable, Iterator

import openpyxl
import requests
from django.http import JsonResponse
from django.utils import timezone
from openpyxl.worksheet.worksheet import Worksheet

from .. import models


def get_cash_register() -> models.CashRegister:
    return models.CashRegister.objects.get_or_create(id=1)[0]


def set_cash_register_sum(sum: decimal.Decimal) -> None:
    cash_register = get_cash_register()
    cash_register.sum = sum
    cash_register.save()


class ImportInflowsError(Exception):
    pass


def set_account_remainder(
    project_region: str | models.ProjectRegion,
    account: str | models.Account,
    sum: float | int | decimal.Decimal | str,
) -> models.Inflow:
    from .common_service import get_account_balance

    if isinstance(project_region, str):
        try:
            project_region = models.ProjectRegion.objects.get(name=project_region)
        except models.ProjectRegion.DoesNotExist:
            print(f'Проект регион "{project_region}" не найдено')
            return None
    if isinstance(account, str):
        try:
            account = models.Account.objects.get(name=account)
        except models.Account.DoesNotExist:
            print(f'Расчётный счёт "{account}" не найдено')
            return None
    if isinstance(sum, str):
        sum = "".join(i for i in sum if i in ["-", "."] or i.isdigit())
    try:
        sum = decimal.Decimal(sum)
    except decimal.InvalidOperation:
        print(f'Не удалось преоброзавать в число: "{sum}"')
        return None
    balance = get_account_balance(account, project_region)
    remainder = sum - balance
    if not remainder:
        print(
            "Уже есть такая сумма:\n\tСумма которая должна быть: {sum:,};\n\tСумма которая уже есть: {balance:,};"
        )
        return None

    print(
        f"Начисление:\n\tПроект регион: {project_region};\n\tРасчётный счёт: {account};\n\tСумма которая должна быть: {sum:,};\n\tСумма которая уже есть: {balance:,};\n\tСумма начисление: {remainder:,};"
    )
    return models.Inflow(account=account, project_region=project_region, sum=remainder)


def set_account_remainder_from_table(
    project_regions: list[str | models.ProjectRegion],
    accounts: list[list[str | models.Account | int | float | decimal.Decimal]],
) -> None:
    inflows = []
    for account, *sums in accounts:
        print(account, sums)
        for project_region, sum in zip(project_regions, sums):
            print(project_region, sum)
            inflow = set_account_remainder(project_region, account, sum)
            if inflow:
                inflows.append(inflow)
    print(inflows)
    models.Inflow.objects.bulk_create(inflows)


def get_inflows_for(
    *,
    sum: decimal.Decimal,
    project_region: models.ProjectRegion,
    account: models.Account,
    reserve_percent: decimal.Decimal = decimal.Decimal(),
    date: datetime.date | None = None,
    imported_from_file: bool = False,
) -> Iterator[models.Inflow]:
    inflow = models.Inflow(
        account=account,
        project_region=project_region,
        sum=sum,
        imported_from_file=imported_from_file,
        reserve_percent=reserve_percent,
    )
    if date:
        inflow.date = date
    yield inflow


def import_inflows_7(
    bytes_io: io.BytesIO, /, *, date: datetime.date | None = None
) -> None:
    workbook = openpyxl.load_workbook(bytes_io, data_only=True)
    try:
        worksheet = workbook["Данные"]
    except KeyError:
        raise ImportInflowsError('Отсутствует таблица "Данные"')
    today = timezone.localdate()
    if date is None:
        date = timezone.localdate() - datetime.timedelta(
            days=3 if today.weekday() == 0 else 2 if today.weekday() == 6 else 1
        )
    models.Inflow.objects.filter(date=today, imported_from_file=True).delete()
    project_regions = {
        project_region.name: project_region
        for project_region in models.ProjectRegion.objects.all()
    }
    inflows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        inflow_date = row[0].date() if isinstance(row[0], datetime.datetime) else row[0]
        project_region = row[3]
        sum_ = row[4]
        reserve_percent = row[6]
        account = row[11]
        account_number = row[12]
        if (
            inflow_date is None
            or inflow_date != date
            or account is None
            or project_region == "#N/A"
            or project_region.lower() == account.lower()
            or (project_region := project_regions.get(project_region)) is None
            or account_number is None
        ):
            continue
        account = models.Account.objects.get_or_create(
            number="".join(account_number.split()), name=" ".join(account.split())
        )[0]
        sum_ = decimal.Decimal(sum_)
        reserve_percent = round(decimal.Decimal(reserve_percent) * 100, 2)
        inflows.extend(
            get_inflows_for(
                sum=sum_,
                project_region=project_region,
                account=account,
                date=today,
                reserve_percent=reserve_percent,
                imported_from_file=True,
            )
        )
    models.Inflow.objects.bulk_create(inflows)


BALANCE_ROW_INDEX: typing.Final[int] = 76
DATA_START_INDEX: typing.Final[int] = 6
DIRECTORS_ROW_INDEX: typing.Final[int] = 2
PROJECT_END_INDEX: typing.Final[int] = 75


def import_unpaid_invoices():
    try:
        response = requests.post(
            "https://script.google.com/macros/s/AKfycbw_dGUfqg8Auwk9dnwEjsLcXBtS_vSHKsbccC-VDNn5KMzZTM0BxQctnIZMw2mKElmY/exec?get_data=1"
        )

        json_data = response.json()
        models.UnpaidInvoice.objects.all().delete()

        for data_values in json_data:
            try:
                date_string = data_values["Дата"]
                date_object = datetime.datetime.strptime(
                    date_string, "%Y-%m-%dT%H:%M:%S.%fZ"
                )
                formatted_date = date_object.strftime("%Y-%m-%d")
            except (KeyError, ValueError):
                formatted_date = None

            try:
                invoice_date_string = data_values["Дата счёта"]
                invoice_date_object = datetime.datetime.strptime(
                    invoice_date_string, "%Y-%m-%dT%H:%M:%S.%fZ"
                )
                formatted_invoice_date = invoice_date_object.strftime("%Y-%m-%d")
            except (KeyError, ValueError):
                formatted_invoice_date = None

            try:
                paid_amount_1c_value = data_values["Оплаченная сумма(1С)"]
                paid_amount_1c = (
                    float(paid_amount_1c_value) if paid_amount_1c_value else None
                )
            except (KeyError, ValueError):
                paid_amount_1c = None

            document_date_string = data_values.get("Дата документа", "")
            if document_date_string:
                try:
                    document_date_object = datetime.datetime.strptime(
                        document_date_string, "%m/%d/%Y"
                    )
                    formatted_document_date = document_date_object.strftime("%Y-%m-%d")
                except ValueError:
                    print(
                        f"Неверный формат даты для 'Дата документа': {document_date_string}"
                    )
                    formatted_document_date = None
            else:
                formatted_document_date = None

            try:
                due_date_string = data_values.get("Оплтить до(дата)", "")
                if due_date_string:
                    due_date_object = datetime.datetime.strptime(
                        due_date_string, "%m/%d/%Y"
                    )
                    formatted_due_date = due_date_object.strftime("%Y-%m-%d")
                else:
                    formatted_due_date = None
            except ValueError:
                formatted_due_date = None

            closing_document_amount_value = data_values.get(
                "Закрывающий документ представлен на сумму"
            )
            closing_document_amount = (
                float(closing_document_amount_value)
                if closing_document_amount_value
                else None
            )

            data_dict = {
                "number": data_values["ДО"],
                "date": formatted_date,
                "invoice_number": data_values["№ счёта"],
                "invoice_date": formatted_invoice_date,
                "project": data_values["Проект"],
                "responsible_user_id": data_values["Ответственный"],
                "approver": data_values["Утвердитель"],
                "llc": data_values["ТОО"],
                "contractor": data_values["Контрагент"],
                "comment": data_values["Комментарий"],
                "currency": data_values["Валюта"],
                "invoice_category": data_values["Категория счёта"],
                "revenue_expense_articles": data_values["Статьи доходов/расходов"],
                "sales_order": data_values["Заказ на продажу"],
                "bin_or_iin": data_values["БИН/ИИН"],
                "iic": data_values["Расчетный счет (ИИК)"],
                "contract_number": data_values["Фактический номер договора"],
                "invoice_amount": data_values["Сумма по счёту"],
                "paid_amount_1c": paid_amount_1c,
                "bank": data_values["Банк"],
                "payment_type": data_values["ТИП"],
                "status": data_values["Статус ДО"],
                "creator_user_id": data_values["Создатель счёта"],
                "department": data_values["Подразделение"],
                "due_date": formatted_due_date,
                "document_number": data_values["Номер документа"],
                "document_date": formatted_document_date,
                "document_amount": (
                    float(data_values.get("Сумма документа"))
                    if data_values.get("Сумма документа")
                    else None
                ),
                "closing_document_amount": closing_document_amount,
            }

            models.UnpaidInvoice.objects.create(**data_dict)
        import_unpaid_invoices_work_status()
        return JsonResponse({"success": True})
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"})
    except requests.RequestException as e:
        return JsonResponse({"error": f"Request error: {e}"})


def import_paid_invoices() -> None:
    try:
        logging.info("Начало импорта оплаченных счетов.")
        models.PaidInvoice.objects.all().delete()
        response = requests.post(
            "https://script.google.com/macros/s/AKfycbzo2-wKFJhxkBTNIAbv7ge37eCTyXTusQR7liK_IgKX4l2CnyNft0hhnRr8ykVst02G1A/exec?get_data=1"
        )
        json_data = response.json()
        for data in json_data:
            logging.info(data)
            if not data["ДО"] or not data["Оплачено"] or not data["Дата оплаты"]:
                continue
            models.PaidInvoice.objects.create(
                number=data["ДО"],
                sum=data["Оплачено"],
                at=datetime.datetime.strptime(
                    data["Дата оплаты"], "%Y-%m-%dT%H:%M:%S.%fZ"
                ),
            )
        logging.info("Данные импортировались.")
    except json.JSONDecodeError:
        logging.error("Invalid JSON format")
    except requests.RequestException as request_exception:
        logging.error(f"Request error: {request_exception}")


def import_unpaid_invoices_work_status() -> None:
    try:
        logging.info("Начало импорта статусов неоплаченных счетов.")
        response = requests.post(
            "https://script.google.com/macros/s/AKfycbw63gXKFMPd8oi80wDSaH1plSWCTANQlzzCYSDscKyN2bggy2T1oIrHspNBkmsLXnKl/exec?get_data=1"
        )
        json = response.json()
        for row in json:
            try:
                unpaid_invoice = models.UnpaidInvoice.objects.get(
                    number=row["Номер ДО"]
                )
            except models.UnpaidInvoice.DoesNotExist:
                continue
            else:
                unpaid_invoice.work_status = row["Статус ДО"]
                unpaid_invoice.save()
        logging.info("Данные импортировались.")
    except json.JSONDecodeError:
        logging.error("Invalid JSON format")
    except requests.RequestException as request_exception:
        logging.error(f"Request error: {request_exception}")


def find_BALANCE_ROW_INDEX(sheet: Worksheet) -> int:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Итого:":
                # Найдено значение, возвращаем номер строки
                return cell.row


# ---------------------------------------------


def get_projects_row_index(column: int, *args, **kwargs) -> int:
    return 5 if 6 <= column <= 48 else 4


def get_project_column_index(row: int, *args, **kwargs) -> int:
    return get_projects_row_index(row, *args, **kwargs) - 3


@dataclasses.dataclass(slots=True, unsafe_hash=True)
class NodeWithSum:
    node: "Node"
    sum: float


@dataclasses.dataclass(slots=True, unsafe_hash=True)
class Director:
    name: str
    offset: int
    width: int

    @property
    def end(self) -> int:
        return self.offset + self.width


@dataclasses.dataclass(slots=True, unsafe_hash=True)
class Node:
    balance: float
    name: str
    director: Director
    recipients: dict[str, NodeWithSum] = dataclasses.field(
        default_factory=dict, compare=False, hash=False, repr=False
    )
    creditors: dict[str, NodeWithSum] = dataclasses.field(
        default_factory=dict, compare=False, hash=False, repr=False
    )

    def add_recipient(self, node: "Node", sum: float) -> None:
        self.recipients[node.name] = NodeWithSum(node, sum)
        node.creditors[self.name] = NodeWithSum(self, sum)

    def get_plus_balance_recipients(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, "NodeWithSum"]:
        return {
            node_name: node_with_sum
            for node_name, node_with_sum in self.recipients.items()
            if node_with_sum.node.balance > 0
            and (with_director is None or node_with_sum.node.director is with_director)
        }

    def get_minus_balance_recipients(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, "NodeWithSum"]:
        return {
            node_name: node_with_sum
            for node_name, node_with_sum in self.recipients.items()
            if node_with_sum.node.balance < 0
            and (with_director is None or node_with_sum.node.director is with_director)
        }

    def get_plus_balance_creditors(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, "NodeWithSum"]:
        return {
            node_name: node_with_sum
            for node_name, node_with_sum in self.creditors.items()
            if node_with_sum.node.balance > 0
            and (with_director is None or node_with_sum.node.director is with_director)
        }

    def get_minus_balance_creditors(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, "NodeWithSum"]:
        return {
            node_name: node_with_sum
            for node_name, node_with_sum in self.creditors.items()
            if node_with_sum.node.balance < 0
            and (with_director is None or node_with_sum.node.director is with_director)
        }


class LoadError(Exception):
    pass


class DirectorSelectingOptions(str, enum.Enum):
    GENERAL = "под одним директором"
    ALL = "под всеми директорами"


@dataclasses.dataclass(slots=True, frozen=True)
class WithRest:
    class Type(str, enum.Enum):
        PLUS_MINUS = "Прямой долг"
        MINUS_PLUS = "Погашение долга"

    type: Type
    who: Node
    to_whom: Node
    rest: list[Node]


@dataclasses.dataclass(slots=True, frozen=True)
class WithSlave:
    class Type(str, enum.Enum):
        PLUS_MINUS_PLUS = "Взаимозачет, когда \плюсовой\ ПМ дает в долг \плюсовому\ ПМ"
        MINUS_PLUS_MINUS = (
            "Взаимозачет, когда \минусовой\ ПМ дает в долг \минусовому\ ПМ"
        )
        PLUS_NEW_MINUS_TO_PLUS = (
            "Взаимозачет, когда два \плюсовых\ ПМ и образуется новая ячейка"
        )
        MINUS_NEW_PLUS_TO_MINUS = (
            "Взаимозачет, когда два \минусовых\ ПМ и образуется новая ячейка"
        )

    type: Type
    director_selecting_option: DirectorSelectingOptions
    who: Node
    slave: Node
    to_whom: Node
    director: Director | None = None


@dataclasses.dataclass(slots=True)
class Loader:
    directors: list[Director] = dataclasses.field(default_factory=list)
    nodes: dict[str, Node] = dataclasses.field(default_factory=dict)

    def get_plus_balance_nodes(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, Node]:
        return {
            node_name: node
            for node_name, node in self.nodes.items()
            if node.balance > 0
            and (with_director is None or node.director is with_director)
        }

    def get_minus_balance_nodes(
        self, /, *, with_director: Director | None = None
    ) -> dict[str, Node]:
        return {
            node_name: node
            for node_name, node in self.nodes.items()
            if node.balance < 0
            and (with_director is None or node.director is with_director)
        }

    @property
    def last_director(self) -> Director | None:
        try:
            return self.directors[-1]
        except IndexError:
            return None

    def load_bytes_io(
        self,
        bytes_io: io.BytesIO,
        get_projects_row_index: Callable[..., int] = get_projects_row_index,
        get_project_column_index: Callable[..., int] = get_project_column_index,
    ) -> None:
        workbook = openpyxl.load_workbook(bytes_io, data_only=True)
        worksheet: Worksheet | None = typing.cast(
            Worksheet | None, workbook["Долги между подразделениями (К)"]
        )
        if worksheet is None:
            raise LoadError('Лист "Долги между подразделениями (К)" не найден')
        recipients: list[tuple[str, str, float]] = []
        for row in worksheet.iter_rows(
            min_row=DIRECTORS_ROW_INDEX,
            max_row=DIRECTORS_ROW_INDEX,
            min_col=DATA_START_INDEX,
            max_col=PROJECT_END_INDEX,
        ):
            for cell in row:
                if self.last_director is None:
                    if cell.value is None or not isinstance(cell.value, str):
                        raise LoadError("Директор не найден F2")
                    self.directors.append(Director(cell.value, DATA_START_INDEX, 1))
                else:
                    if cell.column != PROJECT_END_INDEX and (
                        cell.value is None or cell.value == self.last_director.name
                    ):
                        self.last_director.width += 1
                    else:
                        if cell.column == PROJECT_END_INDEX:
                            self.last_director.width += 1
                        projects_row_index = get_projects_row_index(
                            column=cell.column - 1
                        )
                        for project_row in worksheet.iter_rows(
                            min_row=projects_row_index,
                            max_row=projects_row_index,
                            min_col=self.last_director.offset,
                            max_col=self.last_director.end - 1,
                        ):
                            for project_cell in project_row:
                                project_name = str(project_cell.value)
                                project_balance = float(
                                    str(
                                        worksheet.cell(
                                            row=find_BALANCE_ROW_INDEX(sheet=worksheet),
                                            column=project_cell.column,
                                        ).value
                                    )
                                )  # BALANCE_ROW_INDEX
                                node = Node(
                                    project_balance, project_name, self.last_director
                                )
                                self.nodes[project_name] = node
                                for recipient_row in worksheet.iter_rows(
                                    min_row=project_cell.column + 1,
                                    max_row=PROJECT_END_INDEX,
                                    min_col=project_cell.column,
                                    max_col=project_cell.column,
                                ):
                                    value = recipient_row[0].value
                                    if value is None:
                                        continue
                                    sum = float(str(value))
                                    if -1 < sum < 1:
                                        continue
                                    who = project_name
                                    to_whom = str(
                                        worksheet.cell(
                                            column=get_project_column_index(
                                                recipient_row[0].row
                                            ),
                                            row=recipient_row[0].row,
                                        ).value
                                    )
                                    if sum < 0:
                                        who, to_whom = to_whom, who
                                    recipients.append((who, to_whom, abs(sum)))
                        if cell.column != PROJECT_END_INDEX:
                            if not isinstance(cell.value, str):
                                raise LoadError("Директор не найден")
                            self.directors.append(
                                Director(cell.value, self.last_director.end, 1)
                            )

        for who, to_whom, sum in recipients:
            self.nodes[who].add_recipient(self.nodes[to_whom], sum)

    def load_data(self, data: str | dict) -> None:
        if isinstance(data, str):
            data = json.loads(data)

        directors: dict[str, Director] = {}
        for director in data["directors"]:
            director = Director(director["name"], director["offset"], director["width"])
            self.directors.append(director)
            directors[director.name] = director

        for project in data["projects"].values():
            self.nodes[project["name"]] = Node(
                project["balance"], project["name"], directors[project["director"]]
            )

        for recipient in data["recipients"]:
            self.nodes[recipient["who"]].add_recipient(
                self.nodes[recipient["toWhom"]], recipient["sum"]
            )

    @property
    def data(self) -> dict:
        return {
            "directors": [
                {
                    "name": director.name,
                    "offset": director.offset,
                    "width": director.width,
                }
                for director in self.directors
            ],
            "projects": {
                node_name: {
                    "name": node_name,
                    "balance": node.balance,
                    "director": node.director.name,
                }
                for node_name, node in self.nodes.items()
            },
            "recipients": [
                {"who": node_name, "toWhom": recipient_node_name, "sum": recipient.sum}
                for node_name, node in self.nodes.items()
                for recipient_node_name, recipient in node.recipients.items()
            ],
        }

    @property
    def plus_balance_nodes(self) -> dict[str, Node]:
        return self.get_plus_balance_nodes()

    @property
    def minus_balance_nodes(self) -> dict[str, Node]:
        return self.get_minus_balance_nodes()

    def parent_proccess(
        self,
        who: str,
        to_whom: str,
        /,
        *,
        director_selecting_option: DirectorSelectingOptions = DirectorSelectingOptions.GENERAL,
    ) -> list[WithSlave] | None:
        if (
            who == to_whom
            or who not in self.plus_balance_nodes
            or to_whom not in self.plus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.plus_balance_nodes[who],
            self.plus_balance_nodes[to_whom],
        )
        director = None
        if director_selecting_option == DirectorSelectingOptions.GENERAL:
            if who_node.director is to_whom_node.director:
                director = who_node.director
            else:
                director_selecting_option = DirectorSelectingOptions.ALL
        nodes = set(
            who_node.get_minus_balance_recipients(with_director=director)
        ) & set(to_whom_node.get_minus_balance_recipients(with_director=director))
        if not nodes:
            if director_selecting_option == DirectorSelectingOptions.GENERAL:
                return self.parent_proccess(
                    who, to_whom, director_selecting_option=DirectorSelectingOptions.ALL
                )
            return None
        return [
            WithSlave(
                WithSlave.Type.PLUS_MINUS_PLUS,
                director_selecting_option,
                who_node,
                self.nodes[node],
                to_whom_node,
                director,
            )
            for node in nodes
        ]

    def orphan_parent_proccess(
        self,
        who: str,
        to_whom: str,
        /,
        *,
        director_selecting_option: DirectorSelectingOptions = DirectorSelectingOptions.GENERAL,
    ) -> list[WithSlave] | None:
        if (
            who == to_whom
            or who not in self.minus_balance_nodes
            or to_whom not in self.minus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.minus_balance_nodes[who],
            self.minus_balance_nodes[to_whom],
        )
        director = None
        if director_selecting_option == DirectorSelectingOptions.GENERAL:
            if who_node.director is to_whom_node.director:
                director = who_node.director
            else:
                director_selecting_option = DirectorSelectingOptions.ALL

        nodes = set(who_node.get_plus_balance_creditors(with_director=director)) & set(
            to_whom_node.get_plus_balance_creditors(with_director=director)
        )
        if not nodes:
            if director_selecting_option == DirectorSelectingOptions.GENERAL:
                return self.orphan_parent_proccess(
                    who, to_whom, director_selecting_option=DirectorSelectingOptions.ALL
                )
            return None
        return [
            WithSlave(
                WithSlave.Type.MINUS_PLUS_MINUS,
                director_selecting_option,
                who_node,
                self.nodes[node],
                to_whom_node,
                director,
            )
            for node in nodes
        ]

    def adoption_process(
        self,
        who: str,
        to_whom: str,
        /,
        *,
        director_selecting_option: DirectorSelectingOptions = DirectorSelectingOptions.GENERAL,
    ) -> list[WithSlave] | None:
        if (
            who == to_whom
            or who not in self.plus_balance_nodes
            or to_whom not in self.plus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.plus_balance_nodes[who],
            self.plus_balance_nodes[to_whom],
        )
        director = None
        if director_selecting_option == DirectorSelectingOptions.GENERAL:
            if who_node.director is to_whom_node.director:
                director = who_node.director
            else:
                director_selecting_option = DirectorSelectingOptions.ALL

        search_items = []

        for node_name, recipient in who_node.get_minus_balance_recipients(
            with_director=director
        ).items():
            if (
                node_name not in to_whom_node.recipients
                and to_whom_node.name not in recipient.node.recipients
            ):
                search_items.append(
                    WithSlave(
                        WithSlave.Type.PLUS_NEW_MINUS_TO_PLUS,
                        director_selecting_option,
                        who_node,
                        recipient.node,
                        to_whom_node,
                        director,
                    )
                )

        if not search_items:
            if director_selecting_option == DirectorSelectingOptions.GENERAL:
                return self.adoption_process(
                    who, to_whom, director_selecting_option=DirectorSelectingOptions.ALL
                )
            return None
        return search_items

    def oraphan_adoption_process(
        self,
        who: str,
        to_whom: str,
        /,
        *,
        director_selecting_option: DirectorSelectingOptions = DirectorSelectingOptions.GENERAL,
    ) -> list[WithSlave] | None:
        if (
            who == to_whom
            or who not in self.minus_balance_nodes
            or to_whom not in self.minus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.minus_balance_nodes[who],
            self.minus_balance_nodes[to_whom],
        )
        director = None
        if director_selecting_option == DirectorSelectingOptions.GENERAL:
            if who_node.director is to_whom_node.director:
                director = who_node.director
            else:
                director_selecting_option = DirectorSelectingOptions.ALL
        nodes = []
        for creditor in who_node.get_plus_balance_creditors(
            with_director=director
        ).values():
            if to_whom_node.name not in creditor.node.recipients:
                nodes.append(creditor.node)
        if not nodes:
            if director_selecting_option == DirectorSelectingOptions.GENERAL:
                return self.oraphan_adoption_process(
                    who, to_whom, director_selecting_option=DirectorSelectingOptions.ALL
                )
            return None
        return [
            WithSlave(
                WithSlave.Type.MINUS_NEW_PLUS_TO_MINUS,
                director_selecting_option,
                who_node,
                node,
                to_whom_node,
                director,
            )
            for node in nodes
        ]

    def on_credit(self, who: str, to_whom: str) -> WithRest:  #
        if (
            who == to_whom
            or who not in self.plus_balance_nodes
            or to_whom not in self.minus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.plus_balance_nodes[who],
            self.minus_balance_nodes[to_whom],
        )

        return WithRest(
            WithRest.Type.PLUS_MINUS,
            who_node,
            to_whom_node,
            [
                recipient.node
                for recipient in to_whom_node.get_plus_balance_creditors().values()
            ],
        )

    def debt_closure(self, who: str, to_whom: str) -> bool:
        if (
            who == to_whom
            or who not in self.minus_balance_nodes
            or to_whom not in self.plus_balance_nodes
        ):
            return None
        who_node, to_whom_node = (
            self.minus_balance_nodes[who],
            self.plus_balance_nodes[to_whom],
        )

        return WithRest(
            WithRest.Type.MINUS_PLUS,
            who_node,
            to_whom_node,
            [
                recipient.node
                for recipient in to_whom_node.get_minus_balance_recipients().values()
            ],
        )

    def get_processes_for(
        self, who: str, to_whom: str
    ) -> Iterator[WithRest | list[WithSlave]]:
        for method in [
            self.parent_proccess,
            self.oraphan_adoption_process,
            self.orphan_parent_proccess,
            self.adoption_process,
            self.on_credit,
            self.debt_closure,
        ]:
            result = method(who, to_whom)
            if result:
                yield result
