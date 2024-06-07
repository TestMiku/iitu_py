import json
from datetime import datetime
from django.conf import settings
from django.db import connection
from django.http import FileResponse, HttpRequest, HttpResponse, HttpResponseNotFound
from django.shortcuts import redirect, render
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.decorators import api_view
from designer_requests_for_equipment import models as m
from .serializers import allDataByEquip, allSubDataByEquip
from reporter.models import Report

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
def main(request: HttpRequest) -> HttpResponse:
    # data = serializers.serialize('json', m.kcellTransitionData.objects.all())
    Report.objects.create(responsible=request.user, process="kcell_constructor - Открыт главная страница", text="kcell_constructor.open")
    
    return render(request, "designer_requests_for_equipment/index.html", {'equip_btns': m.kcellEquipTitle.objects.all()})

def junk_page(request: HttpRequest) -> HttpResponse:
    return render(request, "designer_requests_for_equipment/first_lvl_import.html")
    

def import_from_file(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        path = request.FILES['file']
        # path = r'C:\Users\22279\Desktop\PythonProject\import_kcell_data.xlsx'
        m.equipmentData.objects.all().delete()
        m.kcellSubEquipName.objects.all().delete()
        m.kcellTransitionData.objects.all().delete()
        m.kcellEquipTitle.objects.all().delete()
        with connection.cursor() as cursor:
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'designer_requests_for_equipment_equipmentdata'")
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'designer_requests_for_equipment_kcellequiptitle'")
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'designer_requests_for_equipment_kcellsubequipname'")
            cursor.execute("UPDATE sqlite_sequence SET seq = 0 WHERE name = 'designer_requests_for_equipment_kcelltransitiondata'")
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2):
            if row[0].value:
                title_name = m.kcellEquipTitle.objects.create(title_name=row[0].value)
            
            if row[1].value:
                base_equipment_name = m.kcellSubEquipName.objects.create(
                    base_equipment_name = row[1].value,
                    title_name=title_name
                )
            
            if row[2].value:
                transition_name = m.kcellTransitionData.objects.create(
                    transition_name=row[2].value,
                    base_equipment_name=base_equipment_name
                )
            m.equipmentData.objects.create(
                sap=row[3].value,
                description=row[4].value,
                product_code=row[5].value,
                unit=row[6].value,
                q_ty_in_set=row[7].value,
                transition_name=transition_name
            )
            
        return redirect('designer_requests_for_equipment')

@api_view(['GET'])
def get_data_by_equip(request: Request) -> Response:
    category = request.GET.get("category")
    equip = request.GET.get("equip")
    id = request.GET.get("id")
    filters = {}
    if equip:
        filters['base_equipment_name__base_equipment_name'] = equip
        filters['base_equipment_name'] = id
        data = m.kcellTransitionData.objects.filter(**filters)
        return Response(allDataByEquip(data, many=True).data)
    elif category:
        filters['base_equipment_name__title_name__title_name'] = category
        data = m.kcellTransitionData.objects.filter(**filters)
        return Response(allDataByEquip(data, many=True).data)
        
    return Response({"error": "Please provide 'id' parameter in the URL."}, status=400)

@api_view(['GET'])
def get_sub_data_by_id(request: Request) -> Response:
    id = request.GET.get("id")
    transition_name = request.GET.get("tr_name")
    category = request.GET.get("category")
    subcategory = request.GET.get("subcategory")
    if id:
        data = m.equipmentData.objects.filter(transition_name=id)
        return Response(allSubDataByEquip(data, many=True).data)
    if transition_name:
        filters = {}
        filters['transition_name__transition_name'] = transition_name
        filters['transition_name__base_equipment_name__base_equipment_name'] = subcategory
        filters['transition_name__base_equipment_name__title_name__title_name'] = category
        data = m.equipmentData.objects.filter(**filters)
        return Response(allSubDataByEquip(data, many=True).data)
    return Response({"error": "Please provide 'id' parameter in the URL."}, status=400)

def check_data_base():
    # datas = m.equipmentData.objects.get(id=1)
    # transition_name = datas.transition_name.transition_name
    # base_equipment_name = datas.transition_name.base_equipment_name.base_equipment_name
    # title_name = datas.transition_name.base_equipment_name.title_name.title_name
    
    transition_name = "Kathrein 84510865"  # Замените на фактическое название

    # Получаем все записи equipmentData с заданным transition_name
    datas = m.equipmentData.objects.filter(transition_name__transition_name=transition_name)

    # Перебираем все найденные записи
    for data in datas:
        # Обращаемся к необходимым полям через связи и выводим их значения
        print("Transition Name:", data.transition_name.transition_name)
        print("Base Equipment Name:", data.transition_name.base_equipment_name.base_equipment_name)
        print("Title Name:", data.transition_name.base_equipment_name.title_name.title_name)


def cut_and_paste(template_sheet: openpyxl.Workbook.worksheets, max_len_table: int) -> None:
    start_row = 233
    end_row = 246
    start_col = 1
    end_col = 7
    

    # Определяем место, куда нужно вставить данные
    target_row = 4 + max_len_table
    target_col = 1

    

    # Копирование данных
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell_copy = template_sheet.cell(row=row_idx, column=col_idx)
            cell_target = template_sheet.cell(row=target_row + row_idx - start_row, column=target_col + col_idx - start_col)
            
            # Копирование значений для верхнего левого угла каждой объединенной области
            merged_ranges = template_sheet.merged_cells.ranges
            for merged_range in merged_ranges:
                min_row, min_col, max_row, max_col = merged_range.bounds
                if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
                    cell_target.value = cell_copy.value
                    # Устанавливаем выравнивание текста в центр
                    cell_target.alignment = Alignment(horizontal='center', vertical='center')
                    break
            else:
                cell_target.value = cell_copy.value

            # Копирование стилей
            if cell_copy.has_style:
                cell_target.font = cell_copy.font.copy()
                cell_target.border = cell_copy.border.copy()
                cell_target.fill = cell_copy.fill.copy()

    # Очистка данных в заданном диапазоне
    for row in template_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            try:
                cell.value = None
            except:pass
            cell.font = openpyxl.styles.Font()
            cell.border = openpyxl.styles.Border()
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
    
    # Копирование объединенных ячеек
    for merged_range in template_sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if start_row <= min_row <= end_row and start_col <= min_col <= end_col:
            for row_offset in range(min_row, max_row + 1):
                for col_offset in range(min_col, max_col + 1):
                    cell_copy = template_sheet.cell(row=row_offset, column=col_offset)
                    cell_target = template_sheet.cell(row=target_row + row_offset - start_row, column=target_col + col_offset - start_col)
                    # Копирование значений
                    cell_target.value = cell_copy.value
                    # Устанавливаем выравнивание текста в центр
                    cell_target.alignment = Alignment(horizontal='center', vertical='center')
            # Очистка данных в объединенной ячейке
            for row_offset in range(min_row, max_row + 1):
                for col_offset in range(min_col, max_col + 1):
                    cell = template_sheet.cell(row=row_offset, column=col_offset)
                    cell.value = None
                    cell.font = openpyxl.styles.Font()
                    cell.border = openpyxl.styles.Border()
                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)

    
    template_sheet.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=3)
    template_sheet.merge_cells(start_row=target_row+1, start_column=1, end_row=target_row+1, end_column=3)
    template_sheet.merge_cells(start_row=target_row+4, start_column=1, end_row=target_row+4, end_column=3)
    template_sheet.merge_cells(start_row=target_row+5, start_column=1, end_row=target_row+5, end_column=3)

    template_sheet.merge_cells(start_row=target_row+2, start_column=4, end_row=target_row+2, end_column=6)
    template_sheet.merge_cells(start_row=target_row+6, start_column=4, end_row=target_row+6, end_column=6)
    
    template_sheet.merge_cells(start_row=target_row+8, start_column=1, end_row=target_row+8, end_column=7)
    template_sheet.merge_cells(start_row=target_row+9, start_column=1, end_row=target_row+9, end_column=7)

    # Очистка объединенных ячеек
    for merged_range in template_sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if start_row <= min_row <= end_row and start_col <= min_col <= end_col:
            template_sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

border_style = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
            )


def title_line(teamplate_sheet: openpyxl.Workbook.worksheets, row_idx: int, site_name: str, address: str) -> None:
    yellow_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    teamplate_sheet.cell(row=row_idx, column=1).value = site_name
    teamplate_sheet.cell(row=row_idx, column=4).value = address
    teamplate_sheet.cell(row=row_idx, column=1).fill = yellow_fill
    teamplate_sheet.cell(row=row_idx, column=4).fill = yellow_fill
    teamplate_sheet.cell(row=row_idx, column=1).font = bold_font
    teamplate_sheet.cell(row=row_idx, column=4).font = bold_font
    teamplate_sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    teamplate_sheet.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=7)
    for col in range(1, 8):  # Цикл по всем столбцам
        cell = teamplate_sheet.cell(row=row_idx, column=col)
        cell.border = border_style

def export_to_excel(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        data: dict[str, str|list] = json.loads(request.body)
        file_path = (
            settings.BASE_DIR / f"designer_requests_for_equipment/templates/designer_requests_for_equipment/Шаблон экспорта.xlsx"
        )

        template_wb = openpyxl.load_workbook(file_path)
        templte_sh = template_wb.active
        main_site_name = data['main_site_name']
        main_address = data['main_address']
        jr_number = data["all_data"][0]['jr_number']
        templte_sh['D12'].value = f'{main_site_name} {jr_number}'
        templte_sh['F21'].value = f'{jr_number}'
        templte_sh['F22'].value = f'{jr_number}'
        today = datetime.now().strftime("%d.%m.%Y")
        templte_sh['G21'].value = f'for {today}'
        templte_sh['G22'].value = f'for {today}'
        templte_sh['D15'].value = f'{main_address}'
        current_row = 24
        count = 1
        for dict_ in data['all_data']:
            site_name = dict_['site_name']
            address = dict_['address']
            if len(data['all_data']) > 1:
                title_line(templte_sh, current_row, site_name, address)
                current_row+=1
                count = 1
            sap_counts = {}
            for equip_list in dict_['table_data']:
                for equip_item in equip_list:
                    sap = equip_item['sap']
                    set_code = equip_item['set_code']
                    description = equip_item['description']
                    product_code = equip_item['product_code']
                    unit = equip_item['unit']
                    q_ty_in_set = equip_item['q_ty_in_set']
                    
                    if sap in sap_counts:
                        # Если да, увеличиваем количество на q_ty_in_set
                        sap_counts[sap][0] += q_ty_in_set
                        templte_sh[f'G{sap_counts[sap][1]}'].value = sap_counts[sap][0]
                        # current_row+=1
                        continue
                    else:
                        # Если нет, добавляем SAP в словарь и устанавливаем количество q_ty_in_set
                        sap_counts[sap] = [q_ty_in_set, current_row]
                    
                    templte_sh[f'A{current_row}'].value = count
                    templte_sh[f'B{current_row}'].value = set_code
                    templte_sh[f'C{current_row}'].value = sap
                    templte_sh[f'D{current_row}'].value = description
                    templte_sh[f'E{current_row}'].value = product_code
                    templte_sh[f'F{current_row}'].value = unit
                    templte_sh[f'G{current_row}'].value = q_ty_in_set
                    # Применяем стиль рамок ко всем ячейкам в строке
                    for col in range(1, 8):  # Цикл по всем столбцам
                        cell = templte_sh.cell(row=current_row, column=col)
                        cell.border = border_style
                        
                    current_row+=1
                    count+=1
        cut_and_paste(templte_sh, current_row)
        result_file_path = settings.BASE_DIR / f"designer_requests_for_equipment/templates/designer_requests_for_equipment/export.xlsx"
        template_wb.save(result_file_path)
        try:
            # Отправляем файл пользователю
            Report.objects.create(responsible=request.user, process="kcell_constructor - Скачен экспорт файл", text="kcell_constructor.download_to_excel")
            
            return FileResponse(open(result_file_path, "rb"), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except FileNotFoundError:
            # Возвращаем 404, если файл не найден
            return HttpResponseNotFound()
    else:
        return HttpResponse(status=405)
    
    
    


def import_consrtuctor(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        data: dict[str, str|list] = json.loads(request.body)
        if 'equipments' in data:
            category: str = data['category']
            subcategory: str = data['subcategory']
            package: str = data['package']
            equipments: list[dict[str, str|None]] = json.loads(data['equipments'])
            
            
            try:
                category_instance = m.kcellEquipTitle.objects.get(title_name=category)
                subcategory_instance = m.kcellSubEquipName.objects.get(base_equipment_name=subcategory, title_name=category_instance)
                package_instance = m.kcellTransitionData.objects.get(transition_name=package, base_equipment_name=subcategory_instance)
                for equipment in equipments:
                    m.equipmentData.objects.create(
                        set_code = equipment['set_code'],
                        sap = equipment['sap'],
                        description = equipment['description'],
                        product_code = equipment['product_code'],
                        unit = equipment['unit'],
                        q_ty_in_set = equipment['q_ty_in_set'],
                        transition_name = package_instance,
                    )
            except:
                title_name, title_name_created = m.kcellEquipTitle.objects.get_or_create(title_name=category)
                base_equipment_name, base_equipment_name_created = m.kcellSubEquipName.objects.get_or_create(
                        base_equipment_name = subcategory,
                        title_name=title_name
                    )
                transition_name, transition_name_created = m.kcellTransitionData.objects.get_or_create(
                        transition_name=package,
                        base_equipment_name=base_equipment_name
                    )
                for equipment in equipments:
                        m.equipmentData.objects.create(
                            set_code = equipment['set_code'],
                            sap = equipment['sap'],
                            description = equipment['description'],
                            product_code = equipment['product_code'],
                            unit = equipment['unit'],
                            q_ty_in_set = equipment['q_ty_in_set'],
                            transition_name = transition_name,
                        )
            Report.objects.create(responsible=request.user, process="kcell_constructor - Добавленые новые данные", text="kcell_constructor.update_data")            
            return HttpResponse(status=200)
        
        elif "subcategory_for_delete" in data:
            category = data['category']
            subcategory = data['subcategory_for_delete']
            obj = m.kcellSubEquipName.objects.filter(
                title_name__title_name=category,
                base_equipment_name=subcategory
            )
            obj.delete()
            
        elif "package_for_delete" in data:
            category = data['category']
            subcategory = data['subcategory']
            package_for_delete = data['package_for_delete']
            obj = m.kcellTransitionData.objects.filter(
                transition_name = package_for_delete,
                base_equipment_name__title_name__title_name=category,
                base_equipment_name__base_equipment_name=subcategory,
            )
            obj.delete()
            
        else:
            category: str = data['category']
            subcategory: str = data['subcategory']
            package: str = data['package']
            package_title: str = data['package_title']
            data_to_add: list[dict[str, str|None]] = json.loads(data['data'])
            data_to_delete: list[dict[str, str|None]] = json.loads(data['data_to_delete'])
            data_to_update: list[dict[str, str|None]] = json.loads(data['data_to_update'])
            
            if data_to_delete:
                for item in data_to_delete:
                    obj = m.equipmentData.objects.get(id=item['id'])
                    obj.delete()
                    
                    
            if data_to_add:
                for item in data_to_add:
                    category = m.kcellEquipTitle.objects.get(title_name=category)
                    subcategory = m.kcellSubEquipName.objects.get(base_equipment_name=subcategory)
                    package = m.kcellTransitionData.objects.get(transition_name=package)
                    
                    # Создаем новый объект equipmentData и сохраняем его
                    new_equipment_data = m.equipmentData.objects.create(
                        set_code=item['set_code'],
                        sap=item['sap'],
                        description=item['description'],
                        product_code=item['product_code'],
                        unit=item['unit'],
                        q_ty_in_set=item['q_ty_in_set'],
                        transition_name=package
                    )
                    
                    # Устанавливаем связи с категорией и подкатегорией для нового объекта
                    new_equipment_data.transition_name.base_equipment_name = subcategory
                    new_equipment_data.transition_name.base_equipment_name.title_name = category
                    new_equipment_data.transition_name.base_equipment_name.save()
                
            if data_to_update:
                for item in data_to_update:
                    equipment_instance = m.equipmentData.objects.get(id=item['id'])
                    equipment_instance.q_ty_in_set = item['q_ty_in_set']
                    equipment_instance.set_code = item['set_code']
                    equipment_instance.save()

            
            if package_title:
                package_instance = m.kcellTransitionData.objects.filter(transition_name=package).update(transition_name=package_title)
            Report.objects.create(responsible=request.user, process="kcell_constructor - Изменено данные", text="kcell_constructor.change_data")    
            return HttpResponse(status=212)
    Report.objects.create(responsible=request.user, process="kcell_constructor - Открыт страница изменения данных", text="kcell_constructor.open_update_page")
    return render(request, 'designer_requests_for_equipment/import_constructor.html', 
                  {
                    'cotegories': m.kcellEquipTitle.objects.all(), 
                    'subcategories': m.kcellSubEquipName.objects.all(),
                    'packages': m.kcellTransitionData.objects.all()
                    })