import json
import os
import re
import traceback
import zipfile
from collections import defaultdict
from datetime import datetime

from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.core.serializers import serialize
from django.db import transaction
from django.db.models import Count
from django.http import (FileResponse, HttpResponse, HttpResponseBadRequest,
                         HttpResponseServerError, JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic.edit import DeleteView
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from main.models import AvhUser

from . import services
from .forms import DocumentForm, DocumentRentForm, RequestForm
from .models import ( DocType, Document, DocumentRent, Project, Region,
                     RejectedDocument, Request, WorkRentType, WorkType)


class ProjectList(generic.FormView):
    template_name = "document_management/project_list.html"
    form_class = DocumentForm

    def get_context_data(self, **kwargs):
        request_number = self.kwargs.get("request_number")
        documents = Document.objects.all()
        document_rents = DocumentRent.objects.all()

        if request_number:
            documents = documents.filter(request__request_number=request_number)
            document_rents = document_rents.filter(
                request__request_number=request_number
            )

        return super().get_context_data(
            documents=documents, document_rents=document_rents, **kwargs
        )


@login_required
@csrf_exempt
def upload_document(request):
    if request.method == "POST":
        try:
            project = request.POST["project"]
            doc_type = request.POST["doc_type"]
            total_size = 0
            print(request.FILES)
            for file in request.FILES:
                if match := re.match(r"work_type_(\d+)_documents", file):
                    work_type_id = match[1]

                    # год, месяц, день, час, минута, секунда
                    # current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
                    original_file_name, file_extension = os.path.splitext(
                        request.FILES[file].name
                    )
                    unique_file_name = f"{original_file_name}{file_extension}"

                    existing_document = Document.objects.filter(
                        project_id=project,
                        work_type_id=work_type_id,
                        doc_type_id=doc_type,
                        document=unique_file_name,
                        request_id=request.POST["request"],
                    ).first()

                    if existing_document:
                        return render(
                            request,
                            "document_management/error.html",
                            {"message": "Документ уже загружен"},
                        )
                    else:
                        for file_ in request.FILES.getlist(file):
                            total_size += file_.size

                            # Check if the total size exceeds 15MB
                            if total_size > settings.MAX_UPLOAD_SIZE:
                                return render(
                                    request,
                                    "document_management/error.html",
                                    {
                                        "message": f"Документ превышает 15мб. Вы можете сжать по ссылке https://www.ilovepdf.com/ru/compress_pdf"
                                    },
                                )

                            # file_.name = unique_file_name
                            Document.objects.create(
                                project_id=project,
                                work_type_id=work_type_id,
                                doc_type_id=doc_type,
                                request_id=request.POST["request"],
                                document=file_,
                                modified_by=request.user,
                            )

            if total_size == 0:
                Request.objects.filter(request_number=request.POST["request"]).delete()

            status = request.POST.get("status", "pending")

            request_instance = Request.objects.get(
                request_number=request.POST["request"]
            )
            request_instance.status = status
            request_instance.save(user=request.user)

            return redirect("project_list")
        except Exception as e:
            return HttpResponseServerError(
                f"Произошла ошибка: {traceback.format_exc()}"
            )
    else:
        form = DocumentForm()
        request_form = RequestForm()

    def get_order_date(r: Request):
        return r.order_date

    request_dates = list(map(get_order_date, Request.objects.all()))

    context = {
        "form": form,
        "projects": Project.objects.all(),
        "doc_types": DocType.objects.all(),
        "work_types": WorkType.objects.all(),
        "request_form": request_form,
        "order_dates": request_dates,
        "form_project_choices": Project.objects.all().values_list("id", "name"),
    }

    return render(request, "document_management/upload_document.html", context)


def upload_additional_document(request, request_id):
    request_obj = get_object_or_404(Request, request_number=request_id)
    document = request_obj.document_set.first()
    work_types = WorkType.objects.filter(doc_type=document.doc_type)

    if request.method == "POST":
        # current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        # Loop through work types to handle multiple files
        for work_type in work_types:
            files_key = f"documents_{work_type.id}"
            file = request.FILES.get(files_key)

            if file:
                document_name = file.name
                unique_file_name = f"{document_name}"

                data = {
                    **request.POST,
                    "project": document.project,
                    "request": request_obj,
                    "doc_type": document.doc_type,
                    "work_type": work_type.id,
                }

                request_obj.comment_reject = ""

                form = DocumentForm(data, {"document": file})

                if form.is_valid():
                    document = form.save(commit=False)
                    document.document.name = unique_file_name
                    document.save()

                    status = request.POST.get("status")
                    if status == "pending2":
                        request_obj.status = "pending2"
                    elif status == "pending3":
                        request_obj.status = "pending3"
                else:
                    return HttpResponseBadRequest(
                        "Form is not valid. Check your input."
                    )
        request_obj.save(user=request.user)

    # else:
    form = DocumentForm()

    return render(
        request,
        "document_management/upload_additional_documents.html",
        {"request": request_obj, "form": form, "work_types": work_types},
    )


@csrf_exempt
def download_document(request, document_id):
    document = get_object_or_404(Document, pk=document_id)
    file_path = document.document.path
    response = FileResponse(open(file_path, "rb"), as_attachment=True)
    return response


class DocumentDelete(DeleteView):
    model = Document
    template_name = "document_management/document_confirm_delete.html"
    success_url = reverse_lazy("project_list")

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)


@csrf_exempt
def get_work_types(request):
    doc_type_id = request.GET.get("doc_type_id")
    if doc_type_id:
        work_types = WorkType.objects.filter(doc_type=doc_type_id).values("id", "name")
    else:
        work_types = WorkType.objects.none()

    return JsonResponse(list(work_types), safe=False)


# Добавил документ для отдела по аренде и выкупа
@csrf_exempt
def get_workRent_types(request):
    doc_type_id = request.GET.get("doc_type_id")
    if doc_type_id:
        workRent_types = WorkRentType.objects.filter(doc_type=doc_type_id).values(
            "id", "name"
        )
    else:
        workRent_types = WorkRentType.objects.none()

    return JsonResponse(list(workRent_types), safe=False)


@csrf_exempt
def get_doc_types(request):
    project_id = request.GET.get("project_id")
    if project_id:
        doc_types = DocType.objects.filter(project=project_id).values("id", "name")
    else:
        doc_types = DocType.objects.none()

    return JsonResponse(list(doc_types), safe=False)


def user_financial_otdel(user):
    return user.is_authenticated and user.role_id == 5 or user.id == 1 or user.id == 3


@user_passes_test(user_financial_otdel)
@csrf_exempt
def document_list(request):
    project = request.GET.get("project")
    request_number = request.GET.get("request_number")
    doc_types = DocType.objects.annotate(num_documents=Count("document")).filter(
        num_documents__gt=0
    )
    documents = Document.objects.all()

    requests = Request.objects.all()
    paginator = Paginator(requests, 200)
    page = request.GET.get("page")

    try:
        req = paginator.page(page)
    except PageNotAnInteger:
        req = paginator.page(1)
    except EmptyPage:
        req = paginator.page(paginator.num_pages)

    sorted_documents = sorted(
        documents, key=lambda x: x.request.request_number, reverse=True
    )
    if project is not None:
        doc_types = doc_types.filter(project__pk=project)

    selected_request = None
    document_rents = DocumentRent.objects.none()

    if request_number:
        selected_request = get_object_or_404(Request, request_number=request_number)
        doc_types = doc_types.filter(document__request=selected_request)
        document_rents = DocumentRent.objects.filter(request=selected_request)

    context = {
        "doc_types": doc_types,
        "projects": Project.objects.all(),
        "requests": req,
        "documents": sorted_documents,
        "selected_request": selected_request,
        "document_rents": document_rents,
        "rejected_documents": RejectedDocument.objects.all(),
    }
    return render(request, "document_management/document_list.html", context)


def view_rejected_documents(request):
    rejected_documents = RejectedDocument.objects.order_by("-request_number")

    context = {
        "rejected_documents": rejected_documents,
    }

    return render(request, "document_management/rejected_documents.html", context)


@csrf_exempt
def confirm_document(request, document_id):
    document = get_object_or_404(Document, pk=document_id)

    if request.method == "POST":
        document.status = "approved"
        document.comment = ""
        document.save(user=request.user)

        # Обновление статуса связанной заявки
        request_obj = document.request
        update_request_status(request_obj, request)

    return JsonResponse(
        {
            "status": document.get_status_display(),
            "request_status": request_obj.get_status_display(),
        }
    )


@csrf_exempt
def add_comment(request, document_id):
    document = get_object_or_404(Document, pk=document_id)

    if request.method == "POST":
        comment = request.POST.get("comment")
        document.comment = comment
        document.comment_added = True
        document.save()

        services.add_comment(
            entity="Request",
            entity_id=document.request.request_number,
            user=request.user,
            comment=comment,
            attached_entity="Document",
            attached_entity_id=document.id,
            request=request,
        )
        return redirect("document_list")

    return 0


@csrf_exempt
def confirm_document_rent(request, document_rent_id):
    document_rent = get_object_or_404(DocumentRent, pk=document_rent_id)

    if request.method == "POST":
        document_rent.status = "approved"
        document_rent.comment = ""
        document_rent.save(user=request.user)

        # Обновление статуса связанной заявки
        request_obj = document_rent.request
        update_request_status(request_obj, request)

    return JsonResponse(
        {
            "status": document_rent.get_status_display(),
            "request_status": request_obj.get_status_display(),
        }
    )


@csrf_exempt
def reject_document(request, document_id):
    document = get_object_or_404(Document, pk=document_id)

    if request.method == "POST":
        comment = request.POST.get("comment")

        if not comment:
            return HttpResponseBadRequest(
                "Комментарий обязателен при отклонении документа."
            )

        request_obj = document.request
        if request_obj.status != "rejected_docs":
            request_obj.status = "rejected_docs"

            request_obj.save(user=request.user)

        document.status = "rejected_docs"
        document.comment = comment
        document.save(user=request.user)

        RejectedDocument.objects.create(
            document=document,
            rejection_reason=comment,
            request_number=request_obj.request_number,
        )

        check_rejected_documents(request_obj)

    return JsonResponse(
        {
            "status": document.get_status_display(),
            "comment": document.comment,
            "request_status": request_obj.get_status_display(),
            "history_count": request_obj.history_count,
        }
    )


@csrf_exempt
def reject_document_rent(request, document_rent_id):
    document_rent = get_object_or_404(DocumentRent, pk=document_rent_id)

    if request.method == "POST":
        comment = request.POST.get("comment")

        if not comment:
            return HttpResponseBadRequest(
                "Комментарий обязателен при отклонении документа."
            )

        request_obj = document_rent.request
        if request_obj.status != "rejected_rent":
            request_obj.status = "rejected_rent"

            request_obj.save(user=request.user)

        document_rent.status = "rejected_rent"
        document_rent.comment = comment
        document_rent.save(user=request.user)

        RejectedDocument.objects.create(
            documentRent=document_rent,
            rejection_reason=comment,
            request_number=request_obj.request_number,
        )

        check_rejected_documents(request_obj)

    return JsonResponse(
        {
            "status": document_rent.get_status_display(),
            "comment": document_rent.comment,
            "request_status": request_obj.get_status_display(),
            "history_count": request_obj.history_count,
        }
    )


from django.shortcuts import get_object_or_404
from django.http import JsonResponse


@csrf_exempt
def increase_history_count(request, request_number):
    request_obj = get_object_or_404(Request, request_number=request_number)
    request_obj.history_count += 1
    request_obj.save()

    comments = request.POST.getlist('comments')  # Получаем комментарии из POST-запроса

    for comment in comments:
        id, value = comment.split(':')  # Разделяем строку по символу ':' на id и значение
        print(f"id: {id}, value: {value}")  # Печатаем id и значение каждого комментария
        if not value:
            continue
        services.add_comment(
            entity="Request",
            entity_id=request_obj.request_number,
            user=request.user,
            comment=value,
            attached_entity="Document",
            attached_entity_id=id,
            request=request,
        )

    return JsonResponse({"status": "success"})



def round_to_nearest_5_seconds(timestamp):
    rounded_seconds = (timestamp.second // 5) * 5
    return timestamp.replace(second=rounded_seconds, microsecond=0)


def request_history(request, request_number):
    request_instance = Request.objects.get(request_number=request_number)
    history_entries = request_instance.history.all()
    latest_entries_dict = defaultdict(dict)

    for entry in history_entries:
        rounded_timestamp = round_to_nearest_5_seconds(entry.modified_at)
        timestamp_str = rounded_timestamp.strftime("%Y-%m-%d %H:%M:%S")

        if (
            "latest_entry" not in latest_entries_dict[timestamp_str]
            or entry.modified_at
            > latest_entries_dict[timestamp_str]["latest_entry"].modified_at
        ):
            latest_entries_dict[timestamp_str]["latest_entry"] = entry

    latest_entries = [value["latest_entry"] for value in latest_entries_dict.values()]

    return render(
        request,
        "document_management/request_history_view.html",
        {"request_instance": request_instance, "latest_entries": latest_entries},
    )


def download_all_documents(request, request_number):
    documents = Document.objects.filter(request__request_number=request_number)
    documentRents = DocumentRent.objects.filter(request__request_number=request_number)
    zip_filename = f"all_documents_request_{request_number}.zip"

    with zipfile.ZipFile(zip_filename, "w") as zip_file:
        for document in documents:
            zip_file.write(document.document.path, arcname=document.document.name)
        for documentRent in documentRents:
            zip_file.write(
                documentRent.documentRent.path, arcname=documentRent.documentRent.name
            )

    response = HttpResponse(
        open(zip_filename, "rb").read(), content_type="application/zip"
    )
    response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
    os.remove(zip_filename)

    return response


def check_rejected_documents(request_obj):
    if DocumentRent.objects.filter(
        request=request_obj, status="rejected_rent"
    ).exists():
        if Document.objects.filter(
            request=request_obj, status="rejected_docs"
        ).exists():
            request_obj.status = "rejected_both"
            request_obj.save()
        else:
            request_obj.status = "rejected_rent"
            request_obj.save()
    if Document.objects.filter(request=request_obj, status="rejected_docs").exists():
        if DocumentRent.objects.filter(
            request=request_obj, status="rejected_rent"
        ).exists():
            request_obj.status = "rejected_both"
            request_obj.save()
        else:
            request_obj.status = "rejected_docs"
            request_obj.save()


# Я знаю что это костыыыль и можно сделать в разу лучше, так что сорян и улыбнись))
def update_request_status(request_obj, request):
    documents_rent = DocumentRent.objects.filter(request=request_obj)
    documents = Document.objects.filter(request=request_obj)

    doc_statuses = set(documents.values_list("status", flat=True))
    doc_rent_statuses = set(documents_rent.values_list("status", flat=True))

    if "rejected" in request_obj.status:
        if documents_rent.exists():
            if "rejected_docs" in doc_statuses and "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_both"
            elif "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            elif "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_rent"
            else:
                reject_request(request, request_obj.request_number)
                increase_history_count(request, request_obj.request_number)
                request_obj.status = "rejected"
        else:
            if "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            else:
                reject_request(request, request_obj.request_number)
                increase_history_count(request, request_obj.request_number)
                request_obj.status = "rejected"
    elif "pending3" in request_obj.status:
        if not documents_rent.exists():
            if "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            elif "pending2" in doc_statuses:
                request_obj.status = "pending3"
            elif "rejected" in request_obj.status:
                increase_history_count(request, request_obj.request_number)
                request_obj.status = "rejected"
            elif "approved" in doc_statuses:
                request_obj.status = "approved_full"
        else:
            if "rejected_docs" in doc_statuses and "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_both"
            elif "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            elif "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_rent"
            elif "approved" in doc_statuses:
                if (
                    "pending3" in doc_rent_statuses
                    or "pending2" in doc_rent_statuses
                    or "pending2" in doc_statuses
                ):
                    request_obj.status = "pending3"
                else:
                    request_obj.status = "approved_full"
            else:
                request_obj.status = "pending3"
    elif "pending2" in request_obj.status:
        if documents_rent.exists():
            if "rejected_docs" in doc_statuses and "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_both"
            elif "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            elif "rejected_rent" in doc_rent_statuses:
                request_obj.status = "rejected_rent"
            elif "pending2" in doc_statuses:
                request_obj.status = "pending2"
            elif "approved" in doc_statuses:
                if "pending2" in doc_rent_statuses:
                    request_obj.status = "pending2"
                elif "approved" in doc_statuses:
                    request_obj.status = "approved"
                    send_notification_partially_closed(request_obj, request.user)
                    request_obj.isapproved = True
            else:
                request_obj.status = "pending2"
        else:
            if "rejected_docs" in doc_statuses:
                request_obj.status = "rejected_docs"
            elif "pending2" in doc_statuses:
                request_obj.status = "pending2"
            elif "rejected" in request_obj.status:
                increase_history_count(request, request_obj.request_number)
                request_obj.status = "rejected"
            elif "approved" in doc_statuses:
                request_obj.status = "approved"
                send_notification_partially_closed(request_obj, request.user)
                request_obj.isapproved = True

    request_obj.save(user=request.user)


@csrf_exempt
def bulk_confirm_documents(request, pk):
    if request.method == "POST":
        documents = Document.objects.filter(status="pending", doc_type__pk=pk)
        for document in documents:
            document.status = "approved"
            document.save()
    return redirect("document_list")


@csrf_exempt
def bulk_reject_documents(request, pk):
    if request.method == "POST":
        comment = request.POST.get("bulk_comment", "")
        documents = Document.objects.filter(status="pending", doc_type__pk=pk)
        for document in documents:
            document.status = "rejected"
            document.comment = comment
            document.save()
    return redirect("document_list")


@csrf_exempt
def view_request(request, request_number):
    request_obj = get_object_or_404(Request, request_number=request_number)
    documents = Document.objects.filter(request=request_obj)
    documents_rent = DocumentRent.objects.filter(request=request_obj)
    work_rent_types = WorkRentType.objects.none()

    if documents.exists():
        work_rent_types = WorkRentType.objects.filter(
            doc_type=documents.first().doc_type
        )

    if request.method == "POST":
        # Удаляем предыдущий документ, если он существует
        DocumentRent.objects.filter(request=request_obj).delete()

        for work_rent_type in work_rent_types:
            uploaded_file = request.FILES.get(f"documentRent_{work_rent_type.pk}")

            if uploaded_file is not None:
                document_rent = DocumentRent(
                    workRent_type=work_rent_type, request=request_obj
                )
                document_rent.documentRent.save(uploaded_file.name, uploaded_file)
                submit_button = request.POST.get("submit_button")
                if submit_button == "pending2":
                    request_obj.status = "pending2"
                elif submit_button == "pending3":
                    request_obj.status = "pending3"

                request_obj.save(user=request.user)

        update_request_status(request_obj, request)

    return render(
        request,
        "document_management/rent_document.html",
        {
            "request": request_obj,
            "documents": documents,
            "documents_rent": documents_rent,
            "work_rent_types": work_rent_types,
        },
    )


@transaction.atomic
def edit_request_documents(request, request_number):
    request_obj = get_object_or_404(Request, request_number=request_number)
    documents = Document.objects.filter(request=request_obj, status="rejected_docs")

    if request.method == "POST":
        documents_to_delete = request.POST.getlist("documents_to_delete")
        if documents_to_delete:
            Document.objects.filter(id__in=documents_to_delete).delete()

        for file in request.FILES:
            if match := re.match(r"document_(\d+)", file):
                document_id = match[1]

                # Получаем существующий документ
                existing_document = Document.objects.filter(id=document_id).first()

                if existing_document:
                    # current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
                    original_file_name, file_extension = os.path.splitext(
                        request.FILES[file].name
                    )
                    unique_file_name = f"{original_file_name}{file_extension}"

                    existing_document.document = request.FILES[file]
                    existing_document.document.name = unique_file_name

                    # Проверяем статус documentRent
                    if existing_document.request.documentrent_set.filter(
                        status="approved"
                    ).exists():
                        existing_document.status = "pending2"

                    else:
                        existing_document.status = "pending2"

                    existing_document.save(user=request.user)

        status = request.POST.get("status")
        if status == "pending2":
            request_obj.status = "pending2"
        elif status == "pending3":
            request_obj.status = "pending3"

        request_obj.save(user=request.user)
        update_request_status(request_obj, request)
        return redirect("view_upload_documents")
    else:
        existing_documents = [
            {"id": doc.id, "document": doc.document, "comment": doc.comment}
            for doc in documents
        ]

    return render(
        request,
        "document_management/edit_document.html",
        {
            "existing_documents": existing_documents,
            "request": request_obj,
        },
    )


def edit_document_rent(request, request_number):
    request_obj = get_object_or_404(Request, request_number=request_number)
    documents_rent = DocumentRent.objects.filter(
        request=request_obj, status="rejected_rent"
    )

    if request.method == "POST":
        documents_rent_to_delete = request.POST.getlist("documents_rent_to_delete")
        if documents_rent_to_delete:
            DocumentRent.objects.filter(id__in=documents_rent_to_delete).delete()

        for file in request.FILES:
            if match := re.match(r"documentRent_(\d+)", file):
                document_rent_id = match[1]

                existing_document_rent = DocumentRent.objects.filter(
                    id=document_rent_id
                ).first()

                if existing_document_rent:
                    # current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
                    original_file_name, file_extension = os.path.splitext(
                        request.FILES[file].name
                    )
                    unique_file_name = f"{original_file_name}{file_extension}"

                    existing_document_rent.documentRent = request.FILES[file]
                    existing_document_rent.documentRent.name = unique_file_name
                    existing_document_rent.status = "pending2"
                    existing_document_rent.save(user=request.user)

        status = request.POST.get("status")
        if status == "pending2":
            request_obj.status = "pending2"
        elif status == "pending3":
            request_obj.status = "pending3"

        update_request_status(request_obj, request)

        return redirect("view_requests")
    else:
        existing_documents_rent = [
            {"id": doc.id, "document_rent": doc.documentRent, "comment": doc.comment}
            for doc in documents_rent
        ]
        document_rent_form = DocumentRentForm()

    return render(
        request,
        "document_management/edit_rent_document.html",
        {
            "existing_document_rents": existing_documents_rent,
            "request": request_obj,
            "document_rent_form": document_rent_form,
        },
    )


@csrf_exempt
def view_requests(request):
    requests = Request.objects.all()
    documents = Document.objects.all()

    paginator = Paginator(requests, 200)
    page = request.GET.get("page")

    try:
        req = paginator.page(page)
    except PageNotAnInteger:
        req = paginator.page(1)
    except EmptyPage:
        req = paginator.page(paginator.num_pages)

    return render(
        request,
        "document_management/rent_documents.html",
        {
            "requests": req,
            "documents": documents,
        },
    )


def user_documentaions_otdel(user):
    return user.is_authenticated and (user.role_id == 4 or user.id == 1 or user.id == 3)


@user_passes_test(user_documentaions_otdel)
def view_upload_document(request):
    requests = Request.objects.order_by("-request_number")
    documents = Document.objects.all()

    # Paginate the requests
    paginator = Paginator(requests, 200)
    page = request.GET.get("page")

    try:
        req = paginator.page(page)
    except PageNotAnInteger:
        req = paginator.page(1)
    except EmptyPage:
        req = paginator.page(paginator.num_pages)

    return render(
        request,
        "document_management/upload_documents.html",
        {
            "requests": req,
            "documents": documents,
        },
    )


def download_crm_excel(request):
    wb = Workbook()

    ws_all_documents = wb.active
    ws_all_documents.title = "Документы"
    header_font = Font(bold=True)
    headers = [
        "Номер заявки",
        "Проект",
        "Вид работ",
        "Регион",
        "Дата и время создания заявки",
        "Номер заказа",
        "Статус",
        "Кол-во возвращений",
        "Название документа",
    ]
    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws_all_documents[f"{col_letter}1"] = column_title
        ws_all_documents[f"{col_letter}1"].font = header_font

    documents = Document.objects.all()
    for row_num, document in enumerate(documents, 2):
        created_at_str = document.request.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ws_all_documents.cell(
            row=row_num, column=1, value=document.request.request_number
        )
        ws_all_documents.cell(row=row_num, column=2, value=document.project.name)
        ws_all_documents.cell(row=row_num, column=3, value=document.doc_type.name)
        ws_all_documents.cell(row=row_num, column=4, value=document.request.region)
        ws_all_documents.cell(row=row_num, column=5, value=created_at_str)
        ws_all_documents.cell(
            row=row_num, column=6, value=document.request.order_number
        )
        ws_all_documents.cell(
            row=row_num, column=7, value=document.request.get_status_display()
        )
        ws_all_documents.cell(
            row=row_num, column=8, value=document.request.history_count
        )
        ws_all_documents.cell(row=row_num, column=9, value=document.document.name)

    ws_requests_only = wb.create_sheet(title="Заявки")
    request_headers = [
        "Номер заявки",
        "Проект",
        "Вид работ",
        "Регион",
        "Дата и время создания заявки",
        "Номер заказа",
        "Статус",
        "Кол-во возвращений",
    ]
    for col_num, column_title in enumerate(request_headers, 1):
        col_letter = get_column_letter(col_num)
        ws_requests_only[f"{col_letter}1"] = column_title
        ws_requests_only[f"{col_letter}1"].font = header_font

    requests = Request.objects.all()
    for row_num, req in enumerate(requests, 2):
        created_at_str = req.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ws_requests_only.cell(row=row_num, column=1, value=req.request_number)
        ws_requests_only.cell(row=row_num, column=2, value=req.doc_type.project.name)
        ws_requests_only.cell(row=row_num, column=3, value=req.doc_type.name)
        ws_requests_only.cell(row=row_num, column=4, value=req.region)
        ws_requests_only.cell(row=row_num, column=5, value=created_at_str)
        ws_requests_only.cell(row=row_num, column=6, value=req.order_number)
        ws_requests_only.cell(row=row_num, column=7, value=req.get_status_display())
        ws_requests_only.cell(row=row_num, column=8, value=req.history_count)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=exported_table.xlsx"

    wb.save(response)

    return response


@csrf_exempt
def approve_request(request, request_number):
    try:
        selected_request = Request.objects.get(request_number=request_number)
        selected_request.status = "approved"
        selected_request.save()
        return JsonResponse({"status": "success"})
    except Request.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Request not found"})


@csrf_exempt
def approve_request_full(request, request_number):
    try:
        selected_request = Request.objects.get(request_number=request_number)
        selected_request.status = "approved_full"
        selected_request.save()
        return JsonResponse({"status": "success"})
    except Request.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Request not found"})


def close_request_full(request, request_number):
    try:
        selected_request = Request.objects.get(request_number=request_number)
        selected_request.status = "approved_full"
        selected_request.save()

        return JsonResponse({"status": "success"})
    except Request.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Request not found"})


@csrf_exempt
def reject_request(request, request_number):
    request_obj = Request.objects.get(request_number=request_number)
    if request.method == "POST":
        comment_reject = request.POST.get("comment_reject", "")
        try:
            request_obj.comment_reject = comment_reject
            services.add_comment(
                entity="Request",
                entity_id=request_obj.request_number,
                user=request.user,
                comment=comment_reject,
                attached_entity=None,
                attached_entity_id=None,
                request=request
            )
            request_obj.status = "rejected"
            request_obj.save(user=request.user)

            return JsonResponse({"status": "success"})
        except Request.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Request not found"})

    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"})


@login_required
@require_http_methods(["POST"])
def create_request(request):
    form = RequestForm(request.POST)

    if form.is_valid():
        instance = form.save()
        instance.created_by = request.user
        instance.save()
        services.add_comment(
            entity="Request",
            entity_id=instance.request_number,
            user=request.user,
            comment=request.POST.get("comment"),
            attached_entity=None,
            attached_entity_id=None,
            request=request,
        )
        update_request_status(instance, request)

        send_notification_to_region_users(instance, request.user)

        return JsonResponse({"requestNumber": instance.request_number})
    else:
        print("Здесь возвращает ошибку", form.errors)
        return JsonResponse({"error_messages": form.errors})


@csrf_exempt
def close_request(request, request_number):
    try:
        selected_request = Request.objects.get(request_number=request_number)
        selected_request.status = "close"
        selected_request.save()
        return JsonResponse({"status": "success"})
    except Request.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Request not found"})


def send_notification_to_region_users(request_instance, current_user):
    region = Region.objects.get(name=request_instance.region)
    subject = f"К закрытию заявки №{request_instance.request_number}"
    message = (
        f'<strong style="font-size: 18px;">Системное уведомление портала https://portal.avh.kz/p2/projlis/{request_instance.request_number}</strong><br><br>'
        f"К закрытию {request_instance.doc_type} БС {request_instance.bis_name} №{request_instance.order_number}"
    )
    for project_manager in region.users.all():
        # send_mail(subject, message, 'portal@avh.kz', [project_manager.email], html_message=message)
        try:
            x = y
        except:
            print("отправка уведомления")
            traceback.print_exc()
        pass


def send_notification_partially_closed(request_instance, current_user):
    region = Region.objects.get(name=request_instance.region)
    subject = f"Заявка №{request_instance.request_number} частично завершена"
    message = (
        f'<strong style="font-size: 18px;">Системное уведомление портала https://portal.avh.kz/p2/projlis/{request_instance.request_number}</strong><br><br>'
        f" {request_instance.doc_type} БС {request_instance.bis_name} №{request_instance.order_number}"
    )
    for project_manager in region.users.all():
        send_mail(
            subject,
            message,
            "portal@avh.kz",
            [project_manager.email],
            html_message=message,
        )


def distributor(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect("project_list")
        elif request.user.role.id == 4:
            return redirect("view_upload_documents")
        elif request.user.role.id == 1:
            return redirect("view_requests")
        elif request.user.role.id == 5:
            return redirect("document_list")
        elif request.user.role.id == 7:
            return redirect("rejected_documents")
        else:
            return redirect("project_list")
    else:
        return redirect("login")


@csrf_exempt
def comments(request, entity, entity_id):
    if request.method == "POST":
        data = {}
        try:
            data = json.loads(request.body)
        except:
            try:
                data = request.POST
                comment = data["comment"]
            except:
                return JsonResponse({"result": "Некорректные данные"}, status=400)


        comment = data.get("comment")
        if not comment:
            return JsonResponse({"result": "Комментарий обязателен"}, status=400)
        
        attached_entity = data.get("attached_entity")
        attached_entity_id = data.get("attached_entity_id")

        services.add_comment(
            entity=entity,
            entity_id=entity_id,
            user=request.user,
            comment=comment,
            attached_entity=attached_entity,
            attached_entity_id=attached_entity_id,
            request=request,
        )
        comments = [comment for comment in Comment.objects.filter(entity=entity, entity_id=entity_id)]
        return render(request, "document_management/tmpls/comments.html", {"comments": comments})
    elif request.method == "GET":
        comments = [comment for comment in Comment.objects.filter(entity=entity, entity_id=entity_id)]
        attached_entity = request.GET.get("attached_entity")
        attached_entity_id = request.GET.get("attached_entity_id")
        if attached_entity and attached_entity_id:
            attached_entity
            attached_entity_id
            comments = [comment for comment in Comment.objects.filter(entity=entity, entity_id=entity_id, attached_entity=attached_entity,attached_entity_id=attached_entity_id)]

        if not comments:
            return JsonResponse({"result": "Нет комментариев"}, status=404)
        return render(request, "document_management/tmpls/comments.html", {"comments": comments})

    return JsonResponse({"result": "Метод не поддерживается"}, status=405)
