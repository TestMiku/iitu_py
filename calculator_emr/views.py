from datetime import datetime
import json
import openpyxl
from django.db import connection
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from reporter.models import Report

from main.models import AvhUser
from django.core import serializers
from django.http import FileResponse, HttpResponseServerError, HttpResponseNotFound, HttpResponse
from django.conf import settings
from django.contrib import messages
from django.core import serializers
from django.db import connection
from django.http import FileResponse, HttpResponseNotFound, HttpResponseServerError
from django.shortcuts import redirect, render
from django.db.models import Sum

from .models import dataBS, dataTCP, data1922


def main(request):
    
    user_email = request.GET.get("user_email")
    user = request.user

    if user_email:
        try:
            user = get_object_or_404(AvhUser, email=user_email)
        except AvhUser.DoesNotExist:
            # Обработка ситуации, когда пользователь не найден
            pass
        
        
    html_file = 'index'
    if 'bitrix' in request.path:
        Report.objects.create(responsible=request.user, process="calculator_emr - Открыт в Битриксе", text="calculator_emr.open.bitrix")
        html_file = 'bitrix'
    else:
        Report.objects.create(responsible=request.user, process="calculator_emr - Открыт главная страница", text="calculator_emr.open")
    bs_name = request.GET.get("bs_name")
    order_numer = request.GET.get("order-number")
    if order_numer:
        filtered_data_by_order = dataBS.objects.filter(order_number=order_numer)
        balance = filtered_data_by_order[0].order_sem_with_nds * 0.6
        queryset = data1922.objects.filter(order_number__contains=order_numer)
        payamt1c_sum = queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] if queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'] else 0
        notpayamt1c_sum = queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] if queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'] else 0
        balance -= (payamt1c_sum + notpayamt1c_sum)
        return render(request, f"calculator_emr/{html_file}.html", {"order_numer_": order_numer, "choised_order": filtered_data_by_order, "balance": str(float('{:.2f}'.format(balance))), "amount_diff": str(float('{:.2f}'.format(payamt1c_sum + notpayamt1c_sum))), "dataToParse": serializers.serialize("json", (filtered_data_by_order)), "data_tcp": dataTCP.objects.all(), "data_BS": dataBS.objects.all(), "Xuser": user})
    elif bs_name:
        filtered_data_by_bs = dataBS.objects.filter(project=bs_name)
        return render(request, f"calculator_emr/{html_file}.html", {"filtered_data_bs": filtered_data_by_bs, "Xuser": user})
    return render(request, f"calculator_emr/{html_file}.html", {"data_BS": dataBS.objects.values('project').distinct(), "Xuser": user}, )



def import_tcp(request):
    Report.objects.create(responsible=request.user, process="calculator_emr - Открыт импорт ТЦП", text="calculator_emr.import_tcp.open")
    if request.method == "POST":
        try:
            rows_imported = 0
            uploaded_file = request.FILES["file_upload"]
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            dataTCP.objects.all().delete()
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers = {}
            for cell in header_row:
                headers[cell.value] = cell.column
            for row in ws.iter_rows(min_row=2):
                dataTCP.objects.create(
                    name=ws.cell(row=row[0].row, column=headers["Название"]).value,
                    unit=ws.cell(row=row[0].row, column=headers["Ед.изм."]).value,
                    max_sum=ws.cell(row=row[0].row, column=headers["Max"]).value,
                    search_key=ws.cell(
                        row=row[0].row, column=headers["Ключ поиска"]
                    ).value,
                )
                rows_imported += 1
            messages.success(
                request,
                "Импорт успешно выполнен. Количество строк: {}".format(rows_imported),
            )
            Report.objects.create(responsible=request.user, process="calculator_emr - Импортирован ТЦП", text="calculator_emr.import_tcp.imported")
            return redirect("import_for_emr_tcp")
        
        except Exception as e:
            messages.error(
                request,
                f"При импорте произошла ошибка: {e}. Проверьте корректность файла.",
            )
            print(e)
            return redirect("import_for_emr_tcp")
    return render(request, "calculator_emr/import_tcp.html")


def import_bs(request):
    Report.objects.create(responsible=request.user, process="calculator_emr - Открыт импорт БС", text="calculator_emr.import_bs.open")
    if request.method == "POST":
        try:
            rows_imported = 0
            uploaded_file = request.FILES["file_upload"]
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            dataBS.objects.all().delete()
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers = {}
            for cell in header_row:
                headers[cell.value] = cell.column
            
            min_row = 2
            if ws['A2'].value == "Итого":
                    min_row = 3
            
            for row in ws.iter_rows(min_row=min_row):
                sum_ = ws.cell(
                    row=row[0].row, column=headers["Сумма по заказу, с НДС"]
                ).value
                if not sum_:
                    sum_ = 0

                dataBS.objects.create(
                    order_number=ws.cell(
                        row=row[0].row, column=headers["Номер заказа"]
                    ).value,
                    order_sem_with_nds=sum_,
                    order_sem_without_nds=round((sum_ / 1.12), 2),
                    kind_of_activity=ws.cell(
                        row=row[0].row, column=headers["Вид деятельности"]
                    ).value,
                    field_of_activity=ws.cell(
                        row=row[0].row, column=headers["Сфера деятельности"]
                    ).value,
                    region=ws.cell(row=row[0].row, column=headers["Регион"]).value,
                    project=ws.cell(row=row[0].row, column=headers["Проект"]).value,
                    customer=ws.cell(
                        row=row[0].row, column=headers["Покупатель"]
                    ).value,
                    order_entry_date=ws.cell(
                        row=row[0].row, column=headers["Дата внесения заказа в систему"]
                    ).value,
                )
                rows_imported += 1
            messages.success(
                request,
                "Импорт успешно выполнен. Количество строк: {}".format(rows_imported),
            )
            Report.objects.create(responsible=request.user, process="calculator_emr - Импортирован БС", text="calculator_emr.import_bs.imported")
            return redirect("import_for_emr_bs")
        except Exception as e:
            messages.error(
                request,
                f"При импорте произошла ошибка: {e}. Проверьте корректность файла.",
            )
            return redirect("import_for_emr_bs")
    return render(request, "calculator_emr/import_bs.html")


def import_1922(request):
    Report.objects.create(responsible=request.user, process="calculator_emr - Открыт импорт 19.22",
                          text="calculator_emr.import_1922.open")
    if request.method == "POST":
        try:
            uploaded_file = request.FILES["file_upload"]
            print(uploaded_file)

        except Exception as e:
            messages.error(
                request,
                f"При импорте произошла ошибка: {e}. Проверьте корректность файла.",
            )
            return redirect("import_for_emr_1922")
    return render(request, "calculator_emr/import_1922.html")

def find_order_form(request):
    print(request)
    return render(request, "calculator_emr/index.html")


def download_template(request, filename):
    print(filename)
    file_path = (
        settings.BASE_DIR / f"calculator_emr/templates/calculator_emr/{filename}.xlsx"
    )
    try:
        # Отправляем файл пользователю
        response = FileResponse(open(file_path, "rb"), filename=f"{filename}.xlsx")
        return response
    except FileNotFoundError:
        # Возвращаем 404, если файл не найден
        return HttpResponseNotFound()
    except Exception as e:
        print(f"Error: {e}")
        # Возвращаем 500 в случае других ошибок
        return HttpResponseServerError()


def hand_update(request):
    pas = request.GET.get("pas")
    if pas == '12345':
        auto_import_bs("4pos.json")
    auto_import_bs()
    return HttpResponse("OK")


def hand_update_1922(request):
    auto_import_1922()
    return HttpResponse("OK")

def auto_import_bs(temp_file_name: str=''):
    try:
        uploaded_file = '/mnt/adem-otchet/7_23.json'
        if temp_file_name:
            uploaded_file = temp_file_name
            
        # uploaded_file = r"C:\Users\0000\Downloads\7_23.json"
        
        with open(uploaded_file, 'r', encoding='utf-8') as f:
            try:
                json_data = bytes(f.read(), 'utf-8').decode("unicode_escape")
                json_data = json.loads(json_data)
            except Exception as e:
                print(e)
                return 501
        
        dataBS.objects.all().delete()

        for row in json_data:
            sum_ = row["summagreement"] if row["summagreement"] else 0

            dataBS.objects.create(
                order_number=row["documentno"].encode('ISO-8859-1').decode('utf-8') if row["documentno"] else "Номер заказа не указан",
                order_sem_with_nds=sum_,
                order_sem_without_nds=round((sum_ / 1.12), 2),
                kind_of_activity=row["atname"].encode('ISO-8859-1').decode('utf-8') if row["atname"] else None,
                field_of_activity=row["afname"].encode('ISO-8859-1').decode('utf-8') if row["afname"] else None,
                region=row["region"].encode('ISO-8859-1').decode('utf-8') if row["region"] else None,
                project=row["project"].encode('ISO-8859-1').decode('utf-8') if row["project"] else None,
                customer=row["consumer"].encode('ISO-8859-1').decode('utf-8') if row["consumer"] else "Покупатель не указан",
                order_entry_date=datetime.strptime(row["dateordered"].encode('ISO-8859-1').decode('utf-8'), '%Y-%m-%dT%H:%M:%S') if row["dateordered"] else "Дата не указана",
            )
        Report.objects.create(responsible="Автоматическое обновление", process="calculator_emr - Импортирован БС", text="calculator_emr.import_bs.imported")
        return 200
    except Exception as e:
        return e


def auto_import_1922():
    try:
        uploaded_file = '/mnt/adem-otchet/19_22.json'
        # uploaded_file = r"C:\Users\22486\Downloads\19_22.json"

        with open(uploaded_file, 'r', encoding='utf-8') as f:
            try:
                json_data = bytes(f.read(), 'utf-8').decode("unicode_escape")
                json_data = json.loads(json_data)
            except Exception as e:
                print(e)
                return 501
        data1922.objects.all().delete()

        for row in json_data:
            data1922.objects.create(
                order_number=row["doc_number"].encode('ISO-8859-1').decode('utf-8') if row["doc_number"] else None,
                doc_number=row["documentno"].encode('ISO-8859-1').decode('utf-8') if row["documentno"] else None,
                account_number=row["nscheta"].encode('ISO-8859-1').decode('utf-8') if row["nscheta"] else None,
                name=row["name"].encode('ISO-8859-1').decode('utf-8') if row["name"] else None,
                supplier=row["postavshik"].encode('ISO-8859-1').decode('utf-8') if row["postavshik"] else None,
                comment=row["coment"].encode('ISO-8859-1').decode('utf-8') if row["coment"] else None,
                payamt1c=row["payamt1c"] if row["payamt1c"] else 0,
                notpayamt1c=row["notpayamt1c"] if row["notpayamt1c"] else 0,
                totallines=row["totallines"] if row["totallines"] else 0,
                paid1c=row["paid1c"] if row["paid1c"] else 0,
                notpaid1c=row["notpaid1c"] if row["notpaid1c"] else 0,
                projects_group=row["gruppa_proekrov"].encode('ISO-8859-1').decode('utf-8') if row["gruppa_proekrov"] else None,
            )
        Report.objects.create(responsible="Автоматическое обновление", process="calculator_emr - Импортирован 19.12",
                              text="calculator_emr.import_1912.imported")
        return 200
    except Exception as e:
        print(e)
        return e


def opened_accounts(request, order_num):
    html_file = ''
    if 'bitrix' in request.path:
        html_file = 'bitrix_'
    queryset = data1922.objects.filter(order_number__contains=order_num).order_by('order_number')
    context = {
        "order_num": order_num,
        "orders": queryset,
        "payamt1c_sum": queryset.aggregate(Sum('payamt1c'))['payamt1c__sum'],
        "notpayamt1c_sum": queryset.aggregate(Sum('notpayamt1c'))['notpayamt1c__sum'],
        "totallines_sum": queryset.aggregate(Sum('totallines'))['totallines__sum'],
        "paid1c_sum": queryset.aggregate(Sum('paid1c'))['paid1c__sum'],
        "notpaid1c_sum": queryset.aggregate(Sum('notpaid1c'))['notpaid1c__sum'],

    }
    return render(request, f"calculator_emr/{html_file}opened_accounts.html", context)
